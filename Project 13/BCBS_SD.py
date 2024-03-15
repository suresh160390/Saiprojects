from tkinter import BOTH, LEFT, TOP, Button, Entry, Frame, Label, PhotoImage, StringVar, Tk,Radiobutton,StringVar,IntVar,filedialog
from idlelib.tooltip import Hovertip
import sys
import os
from operator import itemgetter
import time
from tkinter import messagebox
from os import listdir
from os.path import isfile, join
from datetime import datetime, date
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, StaleElementReferenceException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import openpyxl
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from urllib.parse import urlparse
import warnings
import numpy as np
from openpyxl.utils import get_column_letter,column_index_from_string
from datetime import datetime
from pytz import timezone
import requests
from zipfile import ZipFile
import subprocess

warnings.filterwarnings("ignore")

global rows
global xpath
global heding
global status
global key
global nme
global element_1  
global j

def browse():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
        image_path2 = os.path.join(sys._MEIPASS, 'Static', 'Mapping.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')
        image_path2 = os.path.join(os.getcwd(), 'Static', 'Mapping.png')

    root.title("Process File Picker")
    root.resizable(False,False)

    root.title("Process File Picker")
    root.resizable(False,False)
   
    w = 500
    h = 160
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Browse Process - File Picker",font=("Calibri",18,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
          
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=50)
    
    answer=StringVar()
    answer.set("")
       
    def browse_button():
        global filename
        answer.set("")  
        filename = filedialog.askopenfilename()  
        txt.delete(0, 'end') 
        txt.insert(0, filename)

    def Click_Done():
        global ans        
        
        ans=txt.get()        
        
        if ans=="":
           answer.set("File Path Fields Empty Is Not Allowed...")        
        else:      
            root.destroy()          
            return ans
        
    txt=Entry(Frame2,font=("Calibri",12,"bold","italic"),width=50,justify="left")
    txt.grid(row=0,column=0,padx=5,pady=10,sticky="E")    

    photo1 = PhotoImage(file=image_path2)
            
    btn1=Button(Frame2,text="Browse",command=browse_button,image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=0,column=1,padx=3,pady=0,sticky="W")     
                
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=3,y=80,width=490,height=25)
    
    title_3=Label(Frame3,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=1,padx=150,pady=0,sticky="E")

    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=3,y=105,width=500,height=50)
    
    photo2 = PhotoImage(file=image_path1)

    btn2=Button(Frame4,command=Click_Done,text="Done",image=photo2,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=3,column=0,padx=125,pady=0,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo3 = PhotoImage(file=image_path)   

    btn3=Button(Frame4,command=Close,text="Exit",image=photo3,borderwidth=0,bg="#2c3e50")
    btn3.grid(row=3,column=1,padx=15,pady=0,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn2,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn3,'Click to Exit Process',hover_delay=1000)
    myTip2 = Hovertip(btn1,'Click to Pick File',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def process():        
    global element_1    
    global ans1
    global ans2 
    global driver
    global j
    global dob
    global mid
    global frm_dt
    global to_dt
    global cpt

    # user_pass()
    browse()

    fil=ans               
    # user = ans1
    # password = ans2

    try:
        options = Options()            
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--ignore-ssl-errors')
        options.add_argument("--disable-popup-blocking")
        driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options)
        driver.maximize_window()
        driver.get('https://welcome.wellmark.com/Authentication/Login.aspx')
    except Exception as e:                        
        try:
            options = Options()
            options.add_argument('--ignore-certificate-errors')
            options.add_argument('--ignore-ssl-errors')
            options.add_argument("--disable-popup-blocking")
            # response = requests.get('https://chromedriver.storage.googleapis.com/LATEST_RELEASE')
            # latest_version = response.text.strip()
            # chrome_driver_url = f'https://chromedriver.storage.googleapis.com/{latest_version}/chromedriver_win32.zip'
            # response = requests.get(chrome_driver_url)
            # with open('chromedriver_win32.zip', 'wb') as f:
            #     f.write(response.content)
            # with ZipFile('chromedriver_win32.zip', 'r') as zip_ref:
            #     zip_ref.extractall('.')
            driver_path = os.path.abspath('chromedriver.exe')
            driver = webdriver.Chrome(executable_path=driver_path,options=options)
            driver.maximize_window()
            driver.get('https://welcome.wellmark.com/Authentication/Login.aspx')
        except Exception as e:           
            messagebox.showinfo("Driver Version Problem","Pls Check Your Chrome Driver Version")
            sys.exit(0)
        
    def text_box(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:   
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding, status)
            sys.exit(0)        
    
    def text_box_key(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:                   
                input_field =WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath)))
                input_field.send_keys(Keys.BACKSPACE * len(input_field))                
                input_field.send_keys(key)               
                # element.send_keys(Keys.TAB)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding, status)
            sys.exit(0)    

    def text_box_js(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:
                js_code = "document.querySelector('" + xpath + "').setAttribute('value', '" + key + "');"
                WebDriverWait(driver, 0).until(lambda driver: driver.execute_script(js_code))               
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo(heding, status)
            sys.exit(0)       

    def click(xpath,heding,status):
        counter = 0
        while counter < 15:
            try:             
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding,status)
            sys.exit(0)                                

    def count(xpath,heding,status):
        global rows
        counter = 0
        while counter < 15:
            try:             
                rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding,status)
            sys.exit(0)                         

    def Alert():
        counter = 0
        while counter < 15:
            try:             
                WebDriverWait(driver, 0).until (EC.alert_is_present())
                a=driver.switch_to.alert
                a.accept()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo('Alert','Alert Not Present')                                                                
                
    messagebox.showinfo('Waiting','Authentication Waiting')        
    
    # driver.switch_to.window(driver.window_handles[1])
    
    file=pd.read_excel(fil,sheet_name='BCBS SD Claim Status BOT',header=0)
    
    for index, row in file.iterrows():                                             
        fnm = row[3]            
        mem_id=row[5]
        frm_dt=row[6]
        to_dt = row[7]                 
        cpt_cd=row[8]
        
        xpath= "/html/body/table[3]/tbody/tr/td[2]/table[1]/tbody/tr[3]/td[2]/a"
        heding="Claims Click"
        status="Claims Click Button Not Found"
        click(xpath,heding,status)   
        
        xpath= "/html/body/table[3]/tbody/tr/td[2]/span[1]/table[1]/tbody/tr[1]/td[2]/a"
        heding="Claims Click"
        status="Claims Click Button Not Found"
        click(xpath,heding,status)   

        xpath= "/html/body/form/div[5]/table/tbody/tr/td/div[2]/table/tbody/tr/td[1]/table[1]/tbody/tr/td/div[2]/div[1]/input"
        heding="Member Number"
        status="Member Number Field Not Found"
        key=str(mem_id)
        text_box(xpath,heding,status,key)  
        
        xpath= "/html/body/form/div[5]/table/tbody/tr/td/div[2]/table/tbody/tr/td[1]/div[1]/table/tbody/tr/td[1]/div[1]/input"
        heding="First Name"
        status="First Name Field Not Found"
        key=fnm.upper()
        text_box(xpath,heding,status,key)  
        
        date_object_1 = datetime.strptime(str(frm_dt), "%Y-%m-%d %H:%M:%S")
        dob3 = date_object_1.strftime("%m/%d/%Y")            
        
        xpath= "/html/body/form/div[5]/table/tbody/tr/td/div[2]/table/tbody/tr/td[1]/div[1]/table/tbody/tr/td[1]/div[2]/input[1]"
        heding="From Date"
        status="From Date Field Not Found"
        key=dob3
        text_box(xpath,heding,status,key)

        date_object_1 = datetime.strptime(str(to_dt), "%Y-%m-%d %H:%M:%S")
        dob3 = date_object_1.strftime("%m/%d/%Y")            
        
        xpath= "/html/body/form/div[5]/table/tbody/tr/td/div[2]/table/tbody/tr/td[1]/div[1]/table/tbody/tr/td[1]/div[2]/input[2]"
        heding="To Date"
        status="To Date Field Not Found"
        key=dob3
        text_box(xpath,heding,status,key)
       
        xpath= "/html/body/form/div[5]/table/tbody/tr/td/div[2]/table/tbody/tr/td[1]/table[2]/tbody/tr/td[1]/div/input[1]"
        heding="Claim Search"
        status="Claim Search Button Not Found"
        click(xpath,heding,status)   

        counter = 0
        while counter < 15:
            try:                   
                st = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[4]/table[2]/tbody/tr/td/table/tbody/tr/td/div/table/tbody/tr[1]/td'))).text                                  
                break
            except Exception as e:
                try:
                    st = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[5]/table/tbody/tr/td/div[3]/table/tbody/tr[3]/td[1]/h3'))).text                                  
                    break
                except Exception as e:                
                    time.sleep(1)
                    counter += 1
        else:
            messagebox.showinfo(heding, status)
            sys.exit(0)
        
        if st.lstrip().rstrip()=='Search Results':            
            
            xpath= "/html/body/form/div[5]/table/tbody/tr/td/div[3]/div/table[3]/tbody/tr"
            heding="Search Result Table Count"
            status="Search Result Table Count Not Found"        
            count(xpath,heding,status) 
                        
            j=2
            while j<rows+1:
                if rows==3:
                    lst=[]    
                    lst1=[]
                    cn = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[5]/table/tbody/tr/td/div[3]/div/table[3]/tbody/tr[2]/td[1]'))).text

                    if cn==' ' or cn=='':
                        cn='N/A' 
                        lst1.append(cn)   
                    else:
                        lst1.append(cn)   

                    lt=len(lst1)
                    if lt==0:                           
                        lst.append('N/A') 
                    else:
                        vr=lst1[0]
                        lst.append(vr)

                    lst2=[]
                    st = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[5]/table/tbody/tr/td/div[3]/div/table[3]/tbody/tr[2]/td[2]'))).text

                    if st==' ' or st=='':
                        st='N/A' 
                        lst2.append(st)   
                    else:
                        lst2.append(st)   

                    lt=len(lst2)
                    if lt==0:                           
                        lst.append('N/A') 
                    else:
                        vr=lst2[0]
                        lst.append(vr)

                    lst3=[]
                    dt = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[5]/table/tbody/tr/td/div[3]/div/table[3]/tbody/tr[2]/td[6]'))).text

                    if dt==' ' or dt=='':
                        dt='N/A' 
                        lst3.append(dt)   
                    else:
                        lst3.append(dt)   

                    lt=len(lst3)
                    if lt==0:                           
                        lst.append('N/A')
                    else:
                        vr=lst3[0]
                        lst.append(vr)

                    lst4=[]
                    samt = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[5]/table/tbody/tr/td/div[3]/div/table[3]/tbody/tr[2]/td[7]'))).text

                    if samt==' ' or samt=='':
                        samt='N/A' 
                        lst4.append(samt)   
                    else:
                        lst4.append(samt)   

                    lt=len(lst4)
                    if lt==0:                           
                        lst.append('N/A')
                    else:
                        vr=lst4[0]
                        lst.append(vr)

                    lst5=[]
                    mows = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[5]/table/tbody/tr/td/div[3]/div/table[3]/tbody/tr[2]/td[9]'))).text

                    if mows==' ' or mows=='':
                        mows='N/A' 
                        lst5.append(mows)   
                    else:
                        lst5.append(mows)   

                    lt=len(lst5)
                    if lt==0:                           
                        lst.append('N/A')          
                    else:
                        vr=lst5[0]
                        lst.append(vr)
                    
                    xpath= "/html/body/form/div[5]/table/tbody/tr/td/div[3]/div/table[3]/tbody/tr[2]/td[1]/a"
                    heding="Search Tabel"
                    status="Claim Number Link Not Found"
                    click(xpath,heding,status)                     
                    break                
                
                elif j==rows-1:
                
                    lst=[]    
                    lst1=[]
                    cn = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[5]/table/tbody/tr/td/div[3]/div/table[3]/tbody/tr[{}]/td[1]'.format(j)))).text

                    if cn==' ' or cn=='':
                        cn='N/A' 
                        lst1.append(cn)   
                    else:
                        lst1.append(cn)   

                    lt=len(lst1)
                    if lt==0:                           
                        lst.append('N/A') 
                    else:
                        vr=lst1[0]
                        lst.append(vr)

                    lst2=[]
                    st = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[5]/table/tbody/tr/td/div[3]/div/table[3]/tbody/tr[{}]/td[2]'.format(j)))).text

                    if st==' ' or st=='':
                        st='N/A' 
                        lst2.append(st)   
                    else:
                        lst2.append(st)   

                    lt=len(lst2)
                    if lt==0:                           
                        lst.append('N/A') 
                    else:
                        vr=lst2[0]
                        lst.append(vr)

                    lst3=[]
                    dt = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[5]/table/tbody/tr/td/div[3]/div/table[3]/tbody/tr[{}]/td[6]'.format(j)))).text

                    if dt==' ' or dt=='':
                        dt='N/A' 
                        lst3.append(dt)   
                    else:
                        lst3.append(dt)   

                    lt=len(lst3)
                    if lt==0:                           
                        lst.append('N/A')
                    else:
                        vr=lst3[0]
                        lst.append(vr)

                    lst4=[]
                    samt = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[5]/table/tbody/tr/td/div[3]/div/table[3]/tbody/tr[{}]/td[7]'.format(j)))).text

                    if samt==' ' or samt=='':
                        samt='N/A' 
                        lst4.append(samt)   
                    else:
                        lst4.append(samt)   

                    lt=len(lst4)
                    if lt==0:                           
                        lst.append('N/A')
                    else:
                        vr=lst4[0]
                        lst.append(vr)

                    lst5=[]
                    mows = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[5]/table/tbody/tr/td/div[3]/div/table[3]/tbody/tr[{}]/td[9]'.format(j)))).text

                    if mows==' ' or mows=='':
                        mows='N/A' 
                        lst5.append(mows)   
                    else:
                        lst5.append(mows)   

                    lt=len(lst5)
                    if lt==0:                           
                        lst.append('N/A')          
                    else:
                        vr=lst5[0]
                        lst.append(vr)

                    xpath= "/html/body/form/div[5]/table/tbody/tr/td/div[3]/div/table[3]/tbody/tr[{}]/td[1]/a".format(j)
                    heding="Search Tabel"
                    status="Claim Number Link Not Found"
                    click(xpath,heding,status)                     
                    break         
                j=j+1

            xpath= "/html/body/form/div[4]/table[2]/tbody/tr/td/table/tbody/tr/td/table/tbody/tr[3]/td[2]/a"
            heding="Claim Summary View"
            status="Claim Summary View Link Not Found"
            click(xpath,heding,status) 

            xpath= "/html/body/form/div[5]/div[2]/div[3]/table/tbody/tr"
            heding="Claim Lines Table Count"
            status="Claim Lines Table Count Not Found"        
            count(xpath,heding,status) 
                        
            lst6=[]                        
            j=3
            while j<rows+1: 
                xpath= "/html/body/form/div[5]/div[2]/div[3]/table/tbody/tr[{}]/td[3]"                                
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text     
                if cnm==str(cpt_cd):
                    cpt = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[5]/div[2]/div[3]/table/tbody/tr[{}]/td[9]'.format(j)))).text
                    if cpt==' ' or cpt=='':
                        cpt='N/A' 
                        lst6.append(cpt)   
                    else:
                        lst6.append(cpt)   
                    break
                j=j+1

            lt=len(lst6)
            if lt==0:                           
                lst.append('N/A')    
            else:
                vr=lst6[0]
                lst.append(vr)

            lst8=[]                        
            j=3
            while j<rows+1: 
                xpath= "/html/body/form/div[5]/div[2]/div[3]/table/tbody/tr[{}]/td[3]"                                
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text     
                if cnm==str(cpt_cd):                                        
                    try:
                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/form/div[5]/div[2]/div[3]/table/tbody/tr[{}]/td[2]/a'.format(j)))).click()                    
                        # all_handles = driver.window_handles
                        # for handle in all_handles:
                        #     driver.switch_to.window(handle)
                        #     print("Window Title:", driver.title)                    
                        # WebDriverWait(driver, 5).until(EC.title_contains('Claim Message'))
                        WebDriverWait(driver, 10).until(lambda driver: len(driver.window_handles) > 1)
                        
                        for handle in driver.window_handles:
                            driver.switch_to.window(handle)
                            if 'Claim Message' in driver.title:                            
                                des=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/form/div[3]/table[1]/tbody/tr[2]'))).text                             
                                if des==' ' or des=='':
                                    des='N/A' 
                                    lst8.append(des)   
                                else:
                                    lst8.append(des)  
                                driver.close()
                                break
                        break
                    except Exception as e:
                        pass
                j=j+1

            lt=len(lst8)
            if lt==0:                           
                lst.append('N/A')    
            else:
                vr=lst8[0]
                lst.append(vr)
            
            try:
                main_window_handle = driver.window_handles[0]
                driver.switch_to.window(main_window_handle)
            except Exception as e:
                pass                    

            lst7=[] 
            try:                       
                pmt_num = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[5]/div[2]/div[4]/table/tbody/tr[2]/td[1]'))).text
                if pmt_num==' ' or pmt_num=='':
                    pmt_num='N/A' 
                    lst7.append(pmt_num)   
                else:
                    lst7.append(pmt_num)   
            except Exception as e:
                pass

            lt=len(lst7)
            if lt==0:                           
                lst.append('N/A')
            else:
                vr=lst7[0]
                lst.append(vr)

            xpath= "/html/body/form/div[4]/table[1]/tbody/tr/td/div/table/tbody/tr/td[3]/a"
            heding="Secure provider home"
            status="Secure provider home Link Not Found"
            click(xpath,heding,status)            

            wb1=load_workbook(fil)
            sheet = wb1['BCBS SD Claim Status BOT']
            column_letter = 'T'  
            column_cells = sheet[column_letter]
            last_row = None
            for cell in reversed(column_cells):
                if cell.value:
                    last_row = cell.row
                    break              
            start_column = 'L'
            current_column_index = openpyxl.utils.column_index_from_string(start_column)
            current_row = last_row + 1

            for value in lst:
                current_column = openpyxl.utils.get_column_letter(current_column_index)
                sheet[current_column + str(current_row)] = value
                current_column_index += 1                
            sheet['T' + str(int(last_row + 1))]='Done'
            wb1.save(filename=fil)
            wb1.close()                                 
           
        elif st.lstrip().rstrip()=='Message': 
            
            xpath= "/html/body/form/div[4]/table[1]/tbody/tr/td/div/table/tbody/tr/td[3]/a"
            heding="Secure provider home"
            status="Secure provider home Link Not Found"
            click(xpath,heding,status)            

            wb1=load_workbook(filename=fil)
            sheet = wb1['BCBS SD Claim Status BOT']
            column_letter = 'T'  
            column_cells = sheet[column_letter]
            last_row = None
            for cell in reversed(column_cells):
                if cell.value:
                    last_row = cell.row
                    break                            
            sheet['T' + str(int(last_row + 1))]='No Claim Found'   
            wb1.save(fil)
            wb1.close()  

        else:

            xpath= "/html/body/form/div[4]/table[1]/tbody/tr/td/div/table/tbody/tr/td[3]/a"
            heding="Secure provider home"
            status="Secure provider home Link Not Found"
            click(xpath,heding,status)            

            wb1=load_workbook(filename=fil)
            sheet = wb1['BCBS SD Claim Status BOT']
            column_letter = 'T'  
            column_cells = sheet[column_letter]
            last_row = None
            for cell in reversed(column_cells):
                if cell.value:
                    last_row = cell.row
                    break                            
            sheet['T' + str(int(last_row + 1))]='Error'   
            wb1.save(fil)
            wb1.close()                                     

    driver.quit()
    messagebox.showinfo("Process Status", "Process Completed")
    sys.exit(0)    

if __name__=="__main__":        
    process()