from tkinter import BOTH, LEFT, TOP, Button, Entry, Frame, Label, PhotoImage, StringVar, Tk,Radiobutton,StringVar,IntVar,filedialog
from idlelib.tooltip import Hovertip
import sys
import os
from operator import itemgetter
import time
from tkinter import messagebox
from os import listdir
from os.path import isfile, join
from datetime import datetime, date
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, StaleElementReferenceException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import openpyxl
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from urllib.parse import urlparse
import warnings
import numpy as np
from openpyxl.utils import get_column_letter,column_index_from_string
from datetime import datetime
from pytz import timezone
import requests
from zipfile import ZipFile
import subprocess

warnings.filterwarnings("ignore")

global rows
global xpath
global heding
global status
global key
global nme
global element_1  
global j

def browse():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
        image_path2 = os.path.join(sys._MEIPASS, 'Static', 'Mapping.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')
        image_path2 = os.path.join(os.getcwd(), 'Static', 'Mapping.png')

    root.title("Process File Picker")
    root.resizable(False,False)

    root.title("Process File Picker")
    root.resizable(False,False)
   
    w = 500
    h = 160
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Browse Process - File Picker",font=("Calibri",18,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
          
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=50)
    
    answer=StringVar()
    answer.set("")
       
    def browse_button():
        global filename
        answer.set("")  
        filename = filedialog.askopenfilename()  
        txt.delete(0, 'end') 
        txt.insert(0, filename)

    def Click_Done():
        global ans        
        
        ans=txt.get()        
        
        if ans=="":
           answer.set("File Path Fields Empty Is Not Allowed...")        
        else:      
            root.destroy()          
            return ans
        
    txt=Entry(Frame2,font=("Calibri",12,"bold","italic"),width=50,justify="left")
    txt.grid(row=0,column=0,padx=5,pady=10,sticky="E")    

    photo1 = PhotoImage(file=image_path2)
            
    btn1=Button(Frame2,text="Browse",command=browse_button,image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=0,column=1,padx=3,pady=0,sticky="W")     
                
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=3,y=80,width=490,height=25)
    
    title_3=Label(Frame3,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=1,padx=150,pady=0,sticky="E")

    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=3,y=105,width=500,height=50)
    
    photo2 = PhotoImage(file=image_path1)

    btn2=Button(Frame4,command=Click_Done,text="Done",image=photo2,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=3,column=0,padx=125,pady=0,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo3 = PhotoImage(file=image_path)   

    btn3=Button(Frame4,command=Close,text="Exit",image=photo3,borderwidth=0,bg="#2c3e50")
    btn3.grid(row=3,column=1,padx=15,pady=0,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn2,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn3,'Click to Exit Process',hover_delay=1000)
    myTip2 = Hovertip(btn1,'Click to Pick File',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def process():        
    global element_1    
    global ans1
    global ans2 
    global driver
    global j
    global dob
    global mid
    global frm_dt
    global to_dt
    global cpt
    global rows1

    # user_pass()
    browse()   

    fil=ans 

    # file=pd.read_excel(fil,sheet_name='MCR Claim Submission',header=0)
    # file1 = file.iloc[:, 1].drop_duplicates().reset_index(drop=True)   # Using column index            

    # for pt_Acc_num in file1:                                                                
    #     filtered_file = pd.DataFrame()
    #     condition = file['Patient Account Number'] == pt_Acc_num                
    #     filtered_file = pd.concat([filtered_file, file[condition]])
    #     filtered_file = filtered_file.reset_index(drop=True)
    #     cnt=len(filtered_file)     
    #     mc_num=filtered_file.iloc[0, 1]
    #     print(mc_num)

    # user = ans1
    # password = ans2

    try:
        options = Options()            
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--ignore-ssl-errors')
        options.add_argument("--disable-popup-blocking")
        driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options)
        driver.maximize_window()
        driver.get('http://www.ngsmedicare.com/')
    except Exception as e:                        
        try:
            options = Options()
            options.add_argument('--ignore-certificate-errors')
            options.add_argument('--ignore-ssl-errors')
            options.add_argument("--disable-popup-blocking")
            # response = requests.get('https://chromedriver.storage.googleapis.com/LATEST_RELEASE')
            # latest_version = response.text.strip()
            # chrome_driver_url = f'https://chromedriver.storage.googleapis.com/{latest_version}/chromedriver_win32.zip'
            # response = requests.get(chrome_driver_url)
            # with open('chromedriver_win32.zip', 'wb') as f:
            #     f.write(response.content)
            # with ZipFile('chromedriver_win32.zip', 'r') as zip_ref:
            #     zip_ref.extractall('.')
            driver_path = os.path.abspath('chromedriver.exe')
            driver = webdriver.Chrome(executable_path=driver_path,options=options)
            driver.maximize_window()
            driver.get('http://www.ngsmedicare.com/')
        except Exception as e:           
            messagebox.showinfo("Driver Version Problem","Pls Check Your Chrome Driver Version")
            sys.exit(0)
        
    def text_box(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:   
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()                
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)               
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding, status)
            sys.exit(0)        
    
    def text_box_key(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:                   
                input_field =WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath)))
                input_field.send_keys(Keys.CONTROL,'a')                
                input_field.send_keys(key)               
                input_field.send_keys(Keys.TAB)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding, status)
            sys.exit(0)    

    def text_box_js(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:
                js_code = "document.querySelector('" + xpath + "').setAttribute('value', '" + key + "');"
                WebDriverWait(driver, 0).until(lambda driver: driver.execute_script(js_code))               
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo(heding, status)
            sys.exit(0)       

    def click(xpath,heding,status):
        counter = 0
        while counter < 15:
            try:             
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:            
            # raise e
            messagebox.showinfo(heding,status)
            sys.exit(0)                                

    def count(xpath,heding,status):
        global rows
        counter = 0
        while counter < 15:
            try:             
                rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding,status)
            sys.exit(0)                         
    
    def count_1(xpath,heding,status):
        global rows1
        counter = 0
        while counter < 15:
            try:             
                rows1=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding,status)
            sys.exit(0)   

    def Alert():
        counter = 0
        while counter < 15:
            try:             
                WebDriverWait(driver, 0).until (EC.alert_is_present())
                a=driver.switch_to.alert
                a.accept()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo('Alert','Alert Not Present')                                                                
                
    messagebox.showinfo('Waiting','Authentication Waiting')        
            
    file=pd.read_excel(fil,sheet_name='MCR Claim Submission',header=0)
    file1 = file.iloc[:, 1].drop_duplicates().reset_index(drop=True)   # Using column index            

    for pt_Acc_num in file1:                                                                
        filtered_file = pd.DataFrame()
        condition = file['Patient Account Number'] == pt_Acc_num                
        filtered_file = pd.concat([filtered_file, file[condition]])
        filtered_file = filtered_file.reset_index(drop=True)
        cnt=len(filtered_file)        

        mc_num = filtered_file.iloc[0, 0]
        Acc_num = filtered_file.iloc[0, 1]
        fn=filtered_file.iloc[0, 2]
        ln = filtered_file.iloc[0, 3]
        dob = filtered_file.iloc[0, 4]
        sex = filtered_file.iloc[0, 5]
        add = filtered_file.iloc[0, 6]
        cty = filtered_file.iloc[0, 7]
        st = filtered_file.iloc[0, 8]
        zip_cod = filtered_file.iloc[0, 9]
        a = filtered_file.iloc[0, 10]
        b = filtered_file.iloc[0, 11]
        c = filtered_file.iloc[0, 12]
        d = filtered_file.iloc[0, 13]
        frm_dob = filtered_file.iloc[0, 14]
        to_dob = filtered_file.iloc[0, 15]
        pos = filtered_file.iloc[0, 16]
        prd_cod = filtered_file.iloc[0, 17]
        amt = filtered_file.iloc[0, 18]
        dys_unt = filtered_file.iloc[0, 19]
        mod1 = filtered_file.iloc[0, 20]
        mod2 = filtered_file.iloc[0, 21]
        mod3 = filtered_file.iloc[0, 22]
        mod4 = filtered_file.iloc[0, 23]
        dog1 = filtered_file.iloc[0, 24]
        dog2 = filtered_file.iloc[0, 25]
        dog3 = filtered_file.iloc[0, 26]
        dog4 = filtered_file.iloc[0, 27]
        npi = filtered_file.iloc[0, 28]
        ptan = filtered_file.iloc[0, 29]
               
        scroll_element = driver.find_element(By.TAG_NAME, "body")
        scroll_element.send_keys(Keys.PAGE_DOWN)

        wait = WebDriverWait(driver, 5)
        scroll_to_top_script = "arguments[0].scrollTo(0,0);"
        wait.until(EC.element_to_be_clickable((By.TAG_NAME, "body")))
        driver.execute_script(scroll_to_top_script, scroll_element)          

        time.sleep(1)
        
        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div/div[3]/div/button'            
        heding='Initiate Claim'
        status='Initiate Claim Button Not Found'
        click(xpath,heding,status)
        
        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/form/div/div[2]/div[2]/div/div/button[1]'
        heding='Initiate a New Claim'
        status='Initiate a New Claim Yes Button Not Found'
        click(xpath,heding,status)

        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[1]/div/div[2]/div[2]/form/div[2]/div[2]/div/div/div/button[1]'
        heding='Federal Tax ID Identifer - EIN'
        status='Federal Tax ID Identifer - EIN Button Not Found'
        click(xpath,heding,status)        

        # xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[1]/div/div[2]/div[2]/form/div[5]/div[1]/span/input'
        # heding='Use Billing NPI/PTAN as Rendering NPI/PTAN?'
        # status='Use Billing NPI/PTAN as Rendering NPI/PTAN? Button Not Found'
        # click(xpath,heding,status)

        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[1]/div/div[2]/div[2]/form/div[5]/div[2]/div/span/input'
        heding='Provider Signature on File?'
        status='Provider Signature on File? Button Not Found'
        click(xpath,heding,status)

        scroll_element = driver.find_element(By.TAG_NAME, "body")
        scroll_element.send_keys(Keys.PAGE_DOWN)

        wait = WebDriverWait(driver, 5)
        scroll_to_top_script = "arguments[0].scrollTo(0,0);"
        wait.until(EC.element_to_be_clickable((By.TAG_NAME, "body")))
        driver.execute_script(scroll_to_top_script, scroll_element)          

        time.sleep(2)

        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[2]/div/div/div[2]/div/button[2]'
        heding='Billing Provider Details'
        status='Billing Provider Details Next Button Not Found'
        click(xpath,heding,status)
    
        counter = 0
        while counter < 15:
            try:             
                xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[1]/div/div[2]/div[2]/form/div[3]/div[3]/span/span'
                ck3=WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH,xpath))).text   

                if ck3.lstrip().rstrip()=='Billing Zip Code must be 9 numeric characters.':     
                    scroll_element = driver.find_element(By.TAG_NAME, "body")
                    scroll_element.send_keys(Keys.PAGE_DOWN)

                    wait = WebDriverWait(driver, 5)
                    scroll_to_top_script = "arguments[0].scrollTo(0,0);"
                    wait.until(EC.element_to_be_clickable((By.TAG_NAME, "body")))
                    driver.execute_script(scroll_to_top_script, scroll_element)          

                    time.sleep(2)

                    xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[2]/div/div/div[2]/div/button[2]'
                    heding='Billing Provider Details'
                    status='Billing Provider Details Next Button Not Found'
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                    break
            except Exception as e:
                try:
                    xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[1]/div/div[2]/div[2]/form/div[1]/div[1]/span/input'
                    WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath)))
                    break
                except Exception as e:                    
                    time.sleep(1)
                    counter += 1
        else:                       
            messagebox.showinfo(heding,status)
            sys.exit(0)      
        
            # try:                 
            #     xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[1]/div/div[2]/div[2]/form/div[3]/div[3]/span/span'
            #     ck3=WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH,xpath))).text   

            #     if ck3.lstrip().rstrip()=='Billing Zip Code must be 9 numeric characters.':            
            #         xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[2]/div/div/div[2]/div/button[2]'
            #         heding='Billing Provider Details'
            #         status='Billing Provider Details Next Button Not Found'
            #         click(xpath,heding,status)
            # except Exception as e:
            #     pass
        
        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[1]/div/div[2]/div[2]/form/div[1]/div[1]/span/input'
        heding="Medicare Number"
        status="Medicare Number Field Not Found"
        key=mc_num
        text_box(xpath,heding,status,key)
                    
        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[1]/div/div[2]/div[2]/form/div[1]/div[2]/span/input'               
        heding="Patient Account Number"
        status="Patient Account Number Field Not Found"
        key=str(Acc_num)
        text_box(xpath,heding,status,key)

        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[1]/div/div[2]/div[2]/form/div[1]/div[3]/span/input'
        heding="First Name"
        status="First Name Field Not Found"
        key=fn
        text_box(xpath,heding,status,key)

        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[1]/div/div[2]/div[2]/form/div[2]/div[2]/span/input'
        heding="Last Name"
        status="Last Name Field Not Found"
        key=ln
        text_box(xpath,heding,status,key)

        date_object_1 = datetime.strptime(str(dob), "%Y-%m-%d %H:%M:%S")
        dob3 = date_object_1.strftime("%m/%d/%Y") 

        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[1]/div/div[2]/div[2]/form/div[3]/div[1]/span/input'
        heding="Date of Birth"
        status="Date of Birth Field Not Found"
        key=dob3
        text_box(xpath,heding,status,key)

        xpath= "/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[1]/div/div[2]/div[2]/form/div[3]/div[2]/div/select/option"
        heding="Sex Count"
        status="Sex Count Not Found"        
        count(xpath,heding,status) 

        j=1
        while j<rows+1:             
            xpath= "/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[1]/div/div[2]/div[2]/form/div[3]/div[2]/div/select/option[{}]"                                
            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
            if cnm.lower().lstrip().rstrip()==sex.lower().lstrip().rstrip():                            
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                        
                break
            j=j+1

        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[1]/div/div[2]/div[2]/form/div[3]/div[3]/span/input'
        heding="Street Address"
        status="Street Address Field Not Found"
        key=add
        text_box(xpath,heding,status,key)

        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[1]/div/div[2]/div[2]/form/div[4]/div[1]/span/input'
        heding="City"
        status="City Field Not Found"
        key=cty
        text_box(xpath,heding,status,key)

        xpath= "/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[1]/div/div[2]/div[2]/form/div[4]/div[2]/div/select/option"
        heding="State Count"
        status="State Count Not Found"        
        count(xpath,heding,status)
        
        j=1
        while j<rows+1:             
            xpath= "/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[1]/div/div[2]/div[2]/form/div[4]/div[2]/div/select/option[{}]"                                
            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
            if cnm.lower().lstrip().rstrip()==st.lower().lstrip().rstrip():                            
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                        
                break
            j=j+1

        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[1]/div/div[2]/div[2]/form/div[4]/div[3]/span/input'
        heding="Zip Code"
        status="Zip Code Field Not Found"
        key=str(zip_cod)
        text_box(xpath,heding,status,key)
        
        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[1]/div/div[2]/div[2]/form/div[5]/div/div/span/input'
        heding='Beneficiary Signature On File'
        status='Beneficiary Signature On File Button Not Found'
        click(xpath,heding,status)

        scroll_element = driver.find_element(By.TAG_NAME, "body")
        scroll_element.send_keys(Keys.PAGE_DOWN)

        wait = WebDriverWait(driver, 5)
        scroll_to_top_script = "arguments[0].scrollTo(0,0);"
        wait.until(EC.element_to_be_clickable((By.TAG_NAME, "body")))
        driver.execute_script(scroll_to_top_script, scroll_element)          

        time.sleep(1)

        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[2]/div/div/div[2]/div/button[2]'
        heding='Beneficiary Details Next Button'
        status='Beneficiary Details Next Button Not Found'
        click(xpath,heding,status)
        
        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[1]/div/div[2]/div[2]/form/div[1]/div/div/div/button[1]'
        heding='Do you agree to Accept Assignment?'
        status='Do you agree to Accept Assignment? Button Not Found'
        click(xpath,heding,status)

        if pd.isnull(a):
            a=''

        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[1]/div/div[2]/div[2]/form/div[3]/div[1]/div[1]/span/input'
        heding="Claim Header Information"
        status="Claim Header Information A Field Not Found"
        key=a
        text_box(xpath,heding,status,key)

        if pd.isnull(b):
            b=''

        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[1]/div/div[2]/div[2]/form/div[3]/div[1]/div[2]/span/input'
        heding="Claim Header Information"
        status="Claim Header Information B Field Not Found"
        key=b
        text_box(xpath,heding,status,key)

        if pd.isnull(c):
            c=''

        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[1]/div/div[2]/div[2]/form/div[3]/div[1]/div[3]/span/input'
        heding="Claim Header Information"
        status="Claim Header Information C Field Not Found"
        key=c
        text_box(xpath,heding,status,key)
        
        if pd.isnull(d):
            d=''

        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[1]/div/div[2]/div[2]/form/div[3]/div[1]/div[4]/span/input'
        heding="Claim Header Information"
        status="Claim Header Information D Field Not Found"
        key=d
        text_box(xpath,heding,status,key)

        scroll_element = driver.find_element(By.TAG_NAME, "body")
        scroll_element.send_keys(Keys.PAGE_DOWN)

        wait = WebDriverWait(driver, 5)
        scroll_to_top_script = "arguments[0].scrollTo(0,0);"
        wait.until(EC.element_to_be_clickable((By.TAG_NAME, "body")))
        driver.execute_script(scroll_to_top_script, scroll_element) 

        time.sleep(1)

        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[2]/div/div/div[2]/div/button[2]'
        heding="Claim Header Information Next Button"
        status="Claim Header Information Next Button Not Found"        
        click(xpath,heding,status)
        
        scroll_element = driver.find_element(By.TAG_NAME, "body")
        scroll_element.send_keys(Keys.PAGE_DOWN)

        wait = WebDriverWait(driver, 5)
        scroll_to_top_script = "arguments[0].scrollTo(0,0);"
        wait.until(EC.element_to_be_clickable((By.TAG_NAME, "body")))
        driver.execute_script(scroll_to_top_script, scroll_element) 

        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[1]/div/div[2]/div[2]/div[3]/div[2]/a'
        heding="Add Claim Line"
        status="Add Claim Line Field Not Found"        
        click(xpath,heding,status)

        date_object_2 = datetime.strptime(str(frm_dob), "%Y-%m-%d %H:%M:%S")
        dob1 = date_object_2.strftime("%m/%d/%Y") 

        xpath='/html/body/div[3]/div/div/div/div/form/div[2]/div[1]/span/input'
        heding="From Service Date"
        status="From Service Date Field Not Found"
        key=dob1
        text_box(xpath,heding,status,key)

        date_object_3 = datetime.strptime(str(to_dob), "%Y-%m-%d %H:%M:%S")
        dob2 = date_object_3.strftime("%m/%d/%Y") 

        xpath='/html/body/div[3]/div/div/div/div/form/div[2]/div[2]/span/input'
        heding="To Service Date"
        status="To Service Date Field Not Found"
        key=dob2
        text_box(xpath,heding,status,key)

        xpath='/html/body/div[3]/div/div/div/div/form/div[2]/div[3]/span/input'
        heding="Place Of Service"
        status="Place Of Service Field Not Found"
        key=str(pos)
        text_box(xpath,heding,status,key)

        xpath='/html/body/div[3]/div/div/div/div/form/div[3]/div[1]/span/input'
        heding="Procedure Code"
        status="Procedure Code Field Not Found"
        key=str(prd_cod)
        text_box(xpath,heding,status,key)

        xpath='/html/body/div[3]/div/div/div/div/form/div[4]/div[1]/span/input'
        heding="Charges"
        status="Charges Field Not Found"
        key=str(amt)       
        text_box_key(xpath,heding,status,key)                

        xpath='/html/body/div[3]/div/div/div/div/form/div[4]/div[2]/span/input'
        heding="Days Or Units"
        status="Days Or Units Field Not Found"
        key=str(dys_unt)
        text_box_key(xpath,heding,status,key)        

        if pd.isnull(mod1):
            pass
        else:
            xpath='/html/body/div[3]/div/div/div/div/form/div[5]/div[1]/span/input'
            heding="Modifier 1"
            status="Modifier 1 Field Not Found"
            key=str(mod1)
            text_box(xpath,heding,status,key)

        if pd.isnull(mod2):
            pass
        else:
            xpath='/html/body/div[3]/div/div/div/div/form/div[5]/div[2]/span/input'
            heding="Modifier 2"
            status="Modifier 2 Field Not Found"
            key=str(mod2)
            text_box(xpath,heding,status,key)

        if pd.isnull(mod3):
            pass
        else:
            xpath='/html/body/div[3]/div/div/div/div/form/div[5]/div[3]/span/input'
            heding="Modifier 3"
            status="Modifier 3 Field Not Found"
            key=str(mod3)
            text_box(xpath,heding,status,key)

        if pd.isnull(mod4):
            pass
        else:
            xpath='/html/body/div[3]/div/div/div/div/form/div[5]/div[4]/span/input'
            heding="Modifier 4"
            status="Modifier 4 Field Not Found"
            key=str(mod4)
            text_box(xpath,heding,status,key)        
        
        if pd.isnull(dog1):
            pass
        else:
            xpath= "/html/body/div[3]/div/div/div/div/form/div[6]/div[1]/div/select/option"
            heding="Diagnosis Pointer 1 Count"
            status="Diagnosis Pointer 1 Count Not Found"        
            count(xpath,heding,status) 
            
            dog1 = dog1 + ' - '

            j=1
            while j<rows+1:             
                xpath= "/html/body/div[3]/div/div/div/div/form/div[6]/div[1]/div/select/option[{}]"                               
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                if dog1 in cnm:                           
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                        
                    break
                j=j+1
        
        if pd.isnull(dog2):
            pass
        else:
            xpath= "/html/body/div[3]/div/div/div/div/form/div[6]/div[2]/div/select/option"
            heding="Diagnosis Pointer 2 Count"
            status="Diagnosis Pointer 2 Count Not Found"        
            count(xpath,heding,status)
       
            dog2 = dog2 + ' - '
        
            j=1
            while j<rows+1:             
                xpath= "/html/body/div[3]/div/div/div/div/form/div[6]/div[2]/div/select/option[{}]"                               
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                if dog2 in cnm:                           
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                        
                    break
                j=j+1        

        if pd.isnull(dog3):
            pass
        else:
            xpath= "/html/body/div[3]/div/div/div/div/form/div[6]/div[3]/div/select/option"
            heding="Diagnosis Pointer 3 Count"
            status="Diagnosis Pointer 3 Count Not Found"        
            count(xpath,heding,status)
            
            dog3 = dog3 + ' - '
            j=1
            while j<rows+1:             
                xpath= "/html/body/div[3]/div/div/div/div/form/div[6]/div[3]/div/select/option[{}]"                               
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                if dog3 in cnm:                           
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                        
                    break
                j=j+1        
        
        if pd.isnull(dog4):
            pass
        else:
            xpath= "/html/body/div[3]/div/div/div/div/form/div[6]/div[4]/div/select/option"
            heding="Diagnosis Pointer 4 Count"
            status="Diagnosis Pointer 4 Count Not Found"        
            count(xpath,heding,status)
            
            dog4 = dog4 + ' - '
            j=1
            while j<rows+1:             
                xpath= "/html/body/div[3]/div/div/div/div/form/div[6]/div[4]/div/select/option[{}]"                               
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                if dog4 in cnm:                           
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                        
                    break
                j=j+1
        
        xpath='/html/body/div[3]/div/div/div/div/form/div[8]/div[1]/span/input'
        heding="Rendering Provider NPI"
        status="Rendering Provider NPI Field Not Found"
        key=str(npi)
        text_box(xpath,heding,status,key)        
                     
        xpath='/html/body/div[3]/div/div/div/div/form/div[8]/div[2]/div/span/input'
        heding="Rendering Provider PTAN"
        status="Rendering Provider PTAN Field Not Found"
        key=str(ptan)
        text_box(xpath,heding,status,key) 

        xpath='/html/body/div[3]/div/div/div/div/form/div[9]/button'
        heding="Claim Line Save Button"
        status="Claim Line Save Button Not Found"        
        click(xpath,heding,status)

        try:
            xpath='/html/body/div[3]/div/div/div/div/form/div[9]/button'
            WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
        except Exception as e:
            pass

        if 1<cnt:
            i=2
            while i<=cnt:                 
                frm_dob = filtered_file.iloc[i-1, 14]
                to_dob = filtered_file.iloc[i-1, 15]
                pos = filtered_file.iloc[i-1, 16]
                prd_cod = filtered_file.iloc[i-1, 17]
                amt = filtered_file.iloc[i-1, 18]
                dys_unt = filtered_file.iloc[i-1, 19]
                mod1 = filtered_file.iloc[i-1, 20]
                mod2 = filtered_file.iloc[i-1, 21]
                mod3 = filtered_file.iloc[i-1, 22]
                mod4 = filtered_file.iloc[i-1, 23]
                dog1 = filtered_file.iloc[i-1, 24]
                dog2 = filtered_file.iloc[i-1, 25]
                dog3 = filtered_file.iloc[i-1, 26]
                dog4 = filtered_file.iloc[i-1, 27]
                npi = filtered_file.iloc[0, 28]
                ptan = filtered_file.iloc[0, 29]

                scroll_element = driver.find_element(By.TAG_NAME, "body")
                scroll_element.send_keys(Keys.PAGE_DOWN)

                wait = WebDriverWait(driver, 5)
                scroll_to_top_script = "arguments[0].scrollTo(0,0);"
                wait.until(EC.element_to_be_clickable((By.TAG_NAME, "body")))
                driver.execute_script(scroll_to_top_script, scroll_element) 

                xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[1]/div/div[2]/div[2]/div[3]/div[2]/a'
                heding="Add Claim Line"
                status="Add Claim Line Field Not Found"        
                click(xpath,heding,status)

                counter1 = 0
                while counter1 < 2:
                    try:   
                        xpath='/html/body/div[3]/div/div/div/div/div/div[3]/button[1]'          
                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                        break
                    except Exception as e:
                        time.sleep(1)
                        counter1 += 1                
            
                # xpath='/html/body/div[3]/div/div/div/div/div/div[3]/button[1]'
                # heding="Add Claim Alert"
                # status="Add Claim Alert Not Found"        
                # click(xpath,heding,status)                

                date_object_2 = datetime.strptime(str(frm_dob), "%Y-%m-%d %H:%M:%S")
                dob1 = date_object_2.strftime("%m/%d/%Y") 

                xpath='/html/body/div[3]/div/div/div/div/form/div[2]/div[1]/span/input'
                heding="From Service Date"
                status="From Service Date Field Not Found"
                key=dob1
                text_box(xpath,heding,status,key)

                date_object_3 = datetime.strptime(str(to_dob), "%Y-%m-%d %H:%M:%S")
                dob2 = date_object_3.strftime("%m/%d/%Y") 

                xpath='/html/body/div[3]/div/div/div/div/form/div[2]/div[2]/span/input'
                heding="To Service Date"
                status="To Service Date Field Not Found"
                key=dob2
                text_box(xpath,heding,status,key)

                xpath='/html/body/div[3]/div/div/div/div/form/div[2]/div[3]/span/input'
                heding="Place Of Service"
                status="Place Of Service Field Not Found"
                key=str(pos)
                text_box(xpath,heding,status,key)

                xpath='/html/body/div[3]/div/div/div/div/form/div[3]/div[1]/span/input'
                heding="Procedure Code"
                status="Procedure Code Field Not Found"
                key=str(prd_cod)
                text_box(xpath,heding,status,key)

                xpath='/html/body/div[3]/div/div/div/div/form/div[4]/div[1]/span/input'
                heding="Charges"
                status="Charges Field Not Found"
                key=str(amt)
                text_box_key(xpath,heding,status,key)

                xpath='/html/body/div[3]/div/div/div/div/form/div[4]/div[2]/span/input'
                heding="Days Or Units"
                status="Days Or Units Field Not Found"
                key=str(dys_unt)
                text_box_key(xpath,heding,status,key)

                if pd.isnull(mod1):
                    pass
                else:
                    xpath='/html/body/div[3]/div/div/div/div/form/div[5]/div[1]/span/input'
                    heding="Modifier 1"
                    status="Modifier 1 Field Not Found"
                    key=str(mod1)
                    text_box(xpath,heding,status,key)

                if pd.isnull(mod2):
                    pass
                else:
                    xpath='/html/body/div[3]/div/div/div/div/form/div[5]/div[2]/span/input'
                    heding="Modifier 2"
                    status="Modifier 2 Field Not Found"
                    key=str(mod2)
                    text_box(xpath,heding,status,key)

                if pd.isnull(mod3):
                    pass
                else:
                    xpath='/html/body/div[3]/div/div/div/div/form/div[5]/div[3]/span/input'
                    heding="Modifier 3"
                    status="Modifier 3 Field Not Found"
                    key=str(mod3)
                    text_box(xpath,heding,status,key)

                if pd.isnull(mod4):
                    pass
                else:
                    xpath='/html/body/div[3]/div/div/div/div/form/div[5]/div[4]/span/input'
                    heding="Modifier 4"
                    status="Modifier 4 Field Not Found"
                    key=str(mod4)
                    text_box(xpath,heding,status,key)                
                
                if pd.isnull(dog1):
                    pass
                else:
                    xpath= "/html/body/div[3]/div/div/div/div/form/div[6]/div[1]/div/select/option"
                    heding="Diagnosis Pointer 1 Count"
                    status="Diagnosis Pointer 1 Count Not Found"        
                    count(xpath,heding,status) 
                    
                    dog1 = dog1 + ' - '
                    j=1
                    while j<rows+1:             
                        xpath= "/html/body/div[3]/div/div/div/div/form/div[6]/div[1]/div/select/option[{}]"                               
                        cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                        if dog1 in cnm:                           
                            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                        
                            break
                        j=j+1
            
                if pd.isnull(dog2):
                    pass
                else:
                    xpath= "/html/body/div[3]/div/div/div/div/form/div[6]/div[2]/div/select/option"
                    heding="Diagnosis Pointer 2 Count"
                    status="Diagnosis Pointer 2 Count Not Found"        
                    count(xpath,heding,status)
                    
                    dog2 = dog2 + ' - '
                    j=1
                    while j<rows+1:             
                        xpath= "/html/body/div[3]/div/div/div/div/form/div[6]/div[2]/div/select/option[{}]"                               
                        cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                        if dog2 in cnm:                           
                            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                        
                            break
                        j=j+1
            
                if pd.isnull(dog3):
                    pass
                else:
                    xpath= "/html/body/div[3]/div/div/div/div/form/div[6]/div[3]/div/select/option"
                    heding="Diagnosis Pointer 3 Count"
                    status="Diagnosis Pointer 3 Count Not Found"        
                    count(xpath,heding,status)
                
                    dog3 = dog3 + ' - '
                    j=1
                    while j<rows+1:             
                        xpath= "/html/body/div[3]/div/div/div/div/form/div[6]/div[3]/div/select/option[{}]"                               
                        cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                        if dog3 in cnm:                           
                            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                        
                            break
                        j=j+1
                            
                if pd.isnull(dog4):
                    pass
                else:
                    xpath= "/html/body/div[3]/div/div/div/div/form/div[6]/div[4]/div/select/option"
                    heding="Diagnosis Pointer 4 Count"
                    status="Diagnosis Pointer 4 Count Not Found"        
                    count(xpath,heding,status)
                    
                    dog4 = dog4 + ' - '
                    j=1
                    while j<rows+1:             
                        xpath= "/html/body/div[3]/div/div/div/div/form/div[6]/div[4]/div/select/option[{}]"                               
                        cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                        if dog4 in cnm:                           
                            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                        
                            break
                        j=j+1                                
                     
                xpath='/html/body/div[3]/div/div/div/div/form/div[8]/div[1]/span/input'
                heding="Rendering Provider NPI"
                status="Rendering Provider NPI Field Not Found"
                key=str(npi)
                text_box(xpath,heding,status,key)        
                                     
                xpath='/html/body/div[3]/div/div/div/div/form/div[8]/div[2]/div/span/input'
                heding="Rendering Provider PTAN"
                status="Rendering Provider PTAN Field Not Found"
                key=str(ptan)
                text_box(xpath,heding,status,key) 

                xpath='/html/body/div[3]/div/div/div/div/form/div[9]/button'
                heding="Claim Line Save Button"
                status="Claim Line Save Button Not Found"        
                click(xpath,heding,status)

                i += 1

        scroll_element = driver.find_element(By.TAG_NAME, "body")
        scroll_element.send_keys(Keys.PAGE_DOWN)

        wait = WebDriverWait(driver, 5)
        scroll_to_top_script = "arguments[0].scrollTo(0,0);"
        wait.until(EC.element_to_be_clickable((By.TAG_NAME, "body")))
        driver.execute_script(scroll_to_top_script, scroll_element) 

        time.sleep(1)

        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[2]/div/div/div[2]/div/button[2]'
        heding="Claim Lines Information Next Button"
        status="Claim Lines Information Next Button Not Found"        
        click(xpath,heding,status)
        
        scroll_element = driver.find_element(By.TAG_NAME, "body")
        scroll_element.send_keys(Keys.PAGE_DOWN)

        wait = WebDriverWait(driver, 5)
        scroll_to_top_script = "arguments[0].scrollTo(0,0);"
        wait.until(EC.element_to_be_clickable((By.TAG_NAME, "body")))
        driver.execute_script(scroll_to_top_script, scroll_element) 

        time.sleep(1)

        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[2]/div/div/div[2]/div/button[2]'
        heding="Service Specific Information Next Button"
        status="Service Specific Information Next Button Not Found"        
        click(xpath,heding,status)

        time.sleep(1)

        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[2]/div/div/div[2]/div/div/div/button'
        heding="Submit Claim"
        status="Submit Claim Button Not Found"        
        click(xpath,heding,status)
        
        time.sleep(2)
        
        ck=WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/div/div'))).text

        if ck.lstrip().rstrip()=='No Claim Lines Found for Claim Submission.':
            xpath='/html/body/div[1]/div/i'
            heding="No Claim Found Close Button"
            status="No Claim Found Close Button - Not Found"        
            click(xpath,heding,status)
            
            try:
                xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[3]/div/a'
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
            except Exception as e:                
                xpath='/html/body/div[2]/div/div/div/div/div/div/div/div/div[2]/div/div/div[3]/div/a'
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()               
        else:        
            # /html/body/div[1]/div
            ck=WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[1]/div/div[2]/div[1]/div'))).text   

            if ck.lstrip().rstrip()=='Submitted':            
                xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[2]/div/div/div[2]/div/button'
                heding="Close"
                status="Close Button Not Found"        
                click(xpath,heding,status)
            elif ck.lstrip().rstrip()=='Ready To Submit?':
                xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[2]/div/div/div[2]/div/div/div/button'
                heding="Submit Claim"
                status="Submit Claim Button Not Found"        
                click(xpath,heding,status)
                
                ck=WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[1]/div/div[2]/div[1]/div'))).text

                xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[4]/div/div/div[2]/div/div/div[2]/div/button'
                heding="Close"
                status="Close Button Not Found"        
                click(xpath,heding,status)
            else:            
                xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div/div[3]/div/a'
                heding="Close"
                status="Close Button Not Found"        
                click(xpath,heding,status)

        filtered_file['Status']=ck
        
        wb1=load_workbook(fil)
        sheet = wb1['Final']
        column_letter = 'A'  
        column_cells = sheet[column_letter]
        last_row = None
        for cell in reversed(column_cells):
            if cell.value:
                last_row = cell.row
                break              
        start_row = last_row + 1
        
        for index, row in filtered_file.iterrows():
            for col_num, value in enumerate(row, start=1):
                sheet.cell(row=start_row + index, column=col_num, value=value)
        
        wb1.save(fil)
        
    driver.quit()
    messagebox.showinfo("Process Status", "Process Completed")
    sys.exit(0)    

if __name__=="__main__":        
    process()