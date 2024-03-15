from tkinter import BOTH, LEFT, TOP, Button, Entry, Frame, Label, PhotoImage, StringVar, Tk,Radiobutton,StringVar,IntVar,filedialog
from idlelib.tooltip import Hovertip
import sys
import os
from operator import itemgetter
import time
from tkinter import messagebox
from os import listdir
from os.path import isfile, join
from datetime import datetime, date
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, StaleElementReferenceException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import openpyxl
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from urllib.parse import urlparse
import warnings
import numpy as np
from openpyxl.utils import get_column_letter,column_index_from_string
from datetime import datetime
from pytz import timezone
import requests
from zipfile import ZipFile
import subprocess

warnings.filterwarnings("ignore")

global rows
global xpath
global heding
global status
global key
global nme
global element_1  
global j

def browse():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
        image_path2 = os.path.join(sys._MEIPASS, 'Static', 'Mapping.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')
        image_path2 = os.path.join(os.getcwd(), 'Static', 'Mapping.png')

    root.title("Process File Picker")
    root.resizable(False,False)

    root.title("Process File Picker")
    root.resizable(False,False)
   
    w = 500
    h = 160
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Browse Process - File Picker",font=("Calibri",18,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
          
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=50)
    
    answer=StringVar()
    answer.set("")
       
    def browse_button():
        global filename
        answer.set("")  
        filename = filedialog.askopenfilename()  
        txt.delete(0, 'end') 
        txt.insert(0, filename)

    def Click_Done():
        global ans        
        
        ans=txt.get()        
        
        if ans=="":
           answer.set("File Path Fields Empty Is Not Allowed...")        
        else:      
            root.destroy()          
            return ans
        
    txt=Entry(Frame2,font=("Calibri",12,"bold","italic"),width=50,justify="left")
    txt.grid(row=0,column=0,padx=5,pady=10,sticky="E")    

    photo1 = PhotoImage(file=image_path2)
            
    btn1=Button(Frame2,text="Browse",command=browse_button,image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=0,column=1,padx=3,pady=0,sticky="W")     
                
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=3,y=80,width=490,height=25)
    
    title_3=Label(Frame3,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=1,padx=150,pady=0,sticky="E")

    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=3,y=105,width=500,height=50)
    
    photo2 = PhotoImage(file=image_path1)

    btn2=Button(Frame4,command=Click_Done,text="Done",image=photo2,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=3,column=0,padx=125,pady=0,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo3 = PhotoImage(file=image_path)   

    btn3=Button(Frame4,command=Close,text="Exit",image=photo3,borderwidth=0,bg="#2c3e50")
    btn3.grid(row=3,column=1,padx=15,pady=0,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn2,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn3,'Click to Exit Process',hover_delay=1000)
    myTip2 = Hovertip(btn1,'Click to Pick File',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def process():        
    global element_1    
    global ans1
    global ans2 
    global driver
    global j
    global dob
    global mid
    global frm_dt
    global to_dt
    global cpt
    global rows1

    # user_pass()
    browse()

    fil=ans               
    # user = ans1
    # password = ans2

    try:
        options = Options()            
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--ignore-ssl-errors')
        options.add_argument("--disable-popup-blocking")
        driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options)
        driver.maximize_window()
        driver.get('http://www.ngsmedicare.com/')
    except Exception as e:                        
        try:
            options = Options()
            options.add_argument('--ignore-certificate-errors')
            options.add_argument('--ignore-ssl-errors')
            options.add_argument("--disable-popup-blocking")
            # response = requests.get('https://chromedriver.storage.googleapis.com/LATEST_RELEASE')
            # latest_version = response.text.strip()
            # chrome_driver_url = f'https://chromedriver.storage.googleapis.com/{latest_version}/chromedriver_win32.zip'
            # response = requests.get(chrome_driver_url)
            # with open('chromedriver_win32.zip', 'wb') as f:
            #     f.write(response.content)
            # with ZipFile('chromedriver_win32.zip', 'r') as zip_ref:
            #     zip_ref.extractall('.')
            driver_path = os.path.abspath('chromedriver.exe')
            driver = webdriver.Chrome(executable_path=driver_path,options=options)
            driver.maximize_window()
            driver.get('http://www.ngsmedicare.com/')
        except Exception as e:           
            messagebox.showinfo("Driver Version Problem","Pls Check Your Chrome Driver Version")
            sys.exit(0)
        
    def text_box(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:   
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding, status)
            sys.exit(0)        
    
    def text_box_key(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:                   
                input_field =WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath)))
                input_field.send_keys(Keys.BACKSPACE * len(input_field))                
                input_field.send_keys(key)               
                # element.send_keys(Keys.TAB)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding, status)
            sys.exit(0)    

    def text_box_js(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:
                js_code = "document.querySelector('" + xpath + "').setAttribute('value', '" + key + "');"
                WebDriverWait(driver, 0).until(lambda driver: driver.execute_script(js_code))               
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo(heding, status)
            sys.exit(0)       

    def click(xpath,heding,status):
        counter = 0
        while counter < 15:
            try:             
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding,status)
            sys.exit(0)                                

    def count(xpath,heding,status):
        global rows
        counter = 0
        while counter < 15:
            try:             
                rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding,status)
            sys.exit(0)                         
    
    def count_1(xpath,heding,status):
        global rows1
        counter = 0
        while counter < 15:
            try:             
                rows1=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding,status)
            sys.exit(0)   

    def Alert():
        counter = 0
        while counter < 15:
            try:             
                WebDriverWait(driver, 0).until (EC.alert_is_present())
                a=driver.switch_to.alert
                a.accept()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo('Alert','Alert Not Present')                                                                
                
    messagebox.showinfo('Waiting','Authentication Waiting')        
            
    file=pd.read_excel(fil,sheet_name='MCR Connex BOT',header=0)
    
    for index, row in file.iterrows():                                             
        fnm = row[2]
        lnm = row[3]
        pol_id=row[4]        
        dob = row[5]                         
        
        try:                             
            xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[2]/div/div/div/div/div/div/div[2]/div/div/form/div[2]/div/div[1]/div/span/input'
            element = WebDriverWait(driver, 0).until(EC.presence_of_element_located((By.XPATH, xpath)))
        except Exception as e:                 
            xpath='/html/body/div[2]/div/div/div/div/div/div/div/div/div[2]/div/div[2]/div/div/div/div/div/div/div[2]/div/div/form/div[2]/div/div[1]/div/span/input'
            element = WebDriverWait(driver, 0).until(EC.presence_of_element_located((By.XPATH, xpath)))
        
        heding="Medicare Number"
        status="Medicare Number Field Not Found"
        key=pol_id
        text_box(xpath,heding,status,key)
        
        try:
            xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[2]/div/div/div/div/div/div/div[2]/div/div/form/div[2]/div/div[2]/div/span/input'
            element = WebDriverWait(driver, 0).until(EC.presence_of_element_located((By.XPATH, xpath)))
        except Exception as e:
            xpath='/html/body/div[2]/div/div/div/div/div/div/div/div/div[2]/div/div[2]/div/div/div/div/div/div/div[2]/div/div/form/div[2]/div/div[2]/div/span/input'
            element = WebDriverWait(driver, 0).until(EC.presence_of_element_located((By.XPATH, xpath)))

        heding="Last Name"
        status="Last Name Field Not Found"
        key=lnm
        text_box(xpath,heding,status,key)

        try:
            xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[2]/div/div/div/div/div/div/div[2]/div/div/form/div[2]/div/div[3]/div/span/input'
            element = WebDriverWait(driver, 0).until(EC.presence_of_element_located((By.XPATH, xpath)))
        except Exception as e:
            xpath='/html/body/div[2]/div/div/div/div/div/div/div/div/div[2]/div/div[2]/div/div/div/div/div/div/div[2]/div/div/form/div[2]/div/div[3]/div/span/input'
            element = WebDriverWait(driver, 0).until(EC.presence_of_element_located((By.XPATH, xpath)))

        heding="First Name"
        status="First Name Field Not Found"
        key=fnm
        text_box(xpath,heding,status,key)

        date_object_1 = datetime.strptime(str(dob), "%Y-%m-%d %H:%M:%S")
        dob3 = date_object_1.strftime("%m/%d/%Y")  

        try:
            xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[2]/div/div/div/div/div/div/div[2]/div/div/form/div[2]/div/div[4]/div/span/input'
            element = WebDriverWait(driver, 0).until(EC.presence_of_element_located((By.XPATH, xpath)))
        except Exception as e:
            xpath='/html/body/div[2]/div/div/div/div/div/div/div/div/div[2]/div/div[2]/div/div/div/div/div/div/div[2]/div/div/form/div[2]/div/div[4]/div/span/input'
            element = WebDriverWait(driver, 0).until(EC.presence_of_element_located((By.XPATH, xpath)))
            
        heding="Date of Birth"
        status="Date of Birth Field Not Found"
        key=dob3
        text_box(xpath,heding,status,key)

        try:              
            xpath= "/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[2]/div/div/div/div/div/div/div[2]/div/div/form/div[2]/div/div[5]/div/div/div/div[1]/div/button"
            element = WebDriverWait(driver, 0).until(EC.presence_of_element_located((By.XPATH, xpath)))
        except Exception as e:
            xpath= "/html/body/div[2]/div/div/div/div/div/div/div/div/div[2]/div/div[2]/div/div/div/div/div/div/div[2]/div/div/form/div[2]/div/div[5]/div/div/div/div[1]/div/button"
            element = WebDriverWait(driver, 0).until(EC.presence_of_element_located((By.XPATH, xpath)))

        heding="Submit Button"
        status="Submit Button Not Found"
        click(xpath,heding,status)
       
        try:            
            err = WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="feedbackMessageContainer"]/div/div'))).text    
            wb1=load_workbook(filename=fil)
            sheet = wb1['MCR Connex BOT']
            column_letter = 'S'  
            column_cells = sheet[column_letter]
            last_row = None
            for cell in reversed(column_cells):
                if cell.value:
                    last_row = cell.row
                    break                            
            sheet['S' + str(int(last_row + 1))]=err
            wb1.save(fil)
            wb1.close()  
        except Exception as e:

            xpath= "/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[2]/form/div[3]/div/div[1]/div/div/button"
            heding="Search Button"
            status="Search Button Not Found"
            click(xpath,heding,status)
            
            try:               
                err = WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="feedbackMessageContainer"]/div/div'))).text    
                wb1=load_workbook(filename=fil)
                sheet = wb1['MCR Connex BOT']
                column_letter = 'S'  
                column_cells = sheet[column_letter]
                last_row = None
                for cell in reversed(column_cells):
                    if cell.value:
                        last_row = cell.row
                        break                            
                sheet['S' + str(int(last_row + 1))]=err
                wb1.save(fil)
                wb1.close()  

                try:             
                    xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[2]/div/div/div/div/div/div/div[1]/div[1]/div/div/div[2]/div/div/div[4]/div/a'
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                except Exception as e:
                    xpath= "/html/body/div[2]/div/div/div/div/div/div/div/div/div[2]/div/div[2]/div/div/div/div/div/div/div[1]/div[1]/div/div/div[2]/div/div/div[4]/div"
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()                   

            except Exception as e:
                try:
                    lst=[]
                    
                    counter = 0
                    while counter < 15:
                        try:             
                            xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[1]/article/div/div/div/div[2]/div/div/div/div[2]/div[1]/div/div/div[2]/div[3]/div/span/input'
                            element = WebDriverWait(driver, 0).until(EC.presence_of_element_located((By.XPATH, xpath)))
                            dod = element.get_attribute("value")
                            break
                        except Exception as e:
                            time.sleep(1)
                            counter += 1
                    else:                   
                        messagebox.showinfo('Beneficiary Eligibility','Beneficiary Eligibility Field Not Found')
                        sys.exit(0)   
                    
                    lst1=[]
                    if dod==' ' or dod=='':
                        dod='N/A' 
                        lst1.append(dod)
                    else:
                        lst1.append(dod)  

                    lt=len(lst1)
                    if lt==0:                           
                        lst.append('N/A')
                    else:
                        vr=lst1[0]
                        lst.append(vr)
                                    
                    xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[1]/article/div/div/div/div[2]/div/div/div/div[2]/div[1]/div/div/div[3]/div[2]/div/span/input'
                    element = WebDriverWait(driver, 0).until(EC.presence_of_element_located((By.XPATH, xpath)))
                    add1 = element.get_attribute("value")

                    xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[1]/article/div/div/div/div[2]/div/div/div/div[2]/div[1]/div/div/div[3]/div[3]/div/span/input'
                    element = WebDriverWait(driver, 0).until(EC.presence_of_element_located((By.XPATH, xpath)))
                    add2 = element.get_attribute("value")

                    xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[1]/article/div/div/div/div[2]/div/div/div/div[2]/div[1]/div/div/div[4]/div[1]/div/span/input'
                    element = WebDriverWait(driver, 0).until(EC.presence_of_element_located((By.XPATH, xpath)))
                    ct = element.get_attribute("value")

                    xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[1]/article/div/div/div/div[2]/div/div/div/div[2]/div[1]/div/div/div[4]/div[2]/div/span/input'
                    element = WebDriverWait(driver, 0).until(EC.presence_of_element_located((By.XPATH, xpath)))
                    st = element.get_attribute("value")

                    xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[1]/article/div/div/div/div[2]/div/div/div/div[2]/div[1]/div/div/div[4]/div[3]/div/span/input'
                    element = WebDriverWait(driver, 0).until(EC.presence_of_element_located((By.XPATH, xpath)))
                    pc = element.get_attribute("value")

                    fin_add = add1 + ' ' + add2 + ':' + ct + ':' + st + ':' + pc
                    fin_add1 = fin_add.lstrip().rstrip()
                    
                    fin_add2 = fin_add1.replace(':', '')

                    lst2=[]
                    if fin_add2==' ' or fin_add2=='':
                        fin_add='N/A' 
                        lst2.append(fin_add)
                    else:
                        lst2.append(fin_add)  

                    lt=len(lst2)
                    if lt==0:                           
                        lst.append('N/A')
                    else:
                        vr=lst2[0]
                        lst.append(vr)

                    xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[1]/article/div/div/div/div[2]/div/div/div/div[2]/div[2]/div/div/div[3]/div[2]/div/span/input'
                    element = WebDriverWait(driver, 0).until(EC.presence_of_element_located((By.XPATH, xpath)))
                    en_dt = element.get_attribute("value")

                    lst3=[]
                    if en_dt==' ' or en_dt=='':
                        en_dt='N/A' 
                        lst3.append(en_dt)
                    else:
                        lst3.append(en_dt)  

                    lt=len(lst3)
                    if lt==0:                           
                        lst.append('N/A')
                    else:
                        vr=lst3[0]
                        lst.append(vr)

                    xpath= "/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div"
                    heding="Left Table Count"
                    status="Left Table Count Not Found"        
                    count(xpath,heding,status) 

                    lst4=[]
                    j=1
                    while j<rows+1:  
                        try:
                            xpath= "/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div[{}]/div/button/div/span"                                
                            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                            if cnm.lstrip().rstrip()=='Medicare Advantage':                            
                                xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div[{}]/div/button/div/i'   
                                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j))))                               
                                lst4.append('N/A')                            
                                break
                        except Exception as e:
                            xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div[{}]/div/button'.format(j)
                            heding="Medicare Advantage"
                            status="Medicare Advantage Click Not Found"
                            click(xpath,heding,status)
                            
                            xpath= "/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[3]/article/div/div/div/div[2]/div/form/div[1]/div/table/tbody/tr"
                            heding="Medicare Advantage Table Count"
                            status="Medicare Advantage Table Count Not Found"        
                            count(xpath,heding,status) 
                            
                            l=1
                            while l<rows+1:                              
                                xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[3]/article/div/div/div/div[2]/div/form/div[1]/div/table/tbody/tr[{}]/td[1]'
                                sdt=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(l)))).text
                                    
                                xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[3]/article/div/div/div/div[2]/div/form/div[1]/div/table/tbody/tr[{}]/td[2]'
                                edt=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(l)))).text
                                
                                xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[3]/article/div/div/div/div[2]/div/form/div[1]/div/table/tbody/tr[{}]/td[4]'
                                inm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(l)))).text
                                
                                if sdt==' ' or sdt=='':
                                    sdt='N/A'
                                if edt==' ' or edt=='':
                                    edt='N/A'      
                                if inm==' ' or inm=='':
                                    inm='N/A'  
                                                        
                                if l==1:
                                    fin_mda = f"{sdt}:{edt}:{inm}"
                                else:
                                    fin_mda += f":{sdt}:{edt}:{inm}" 
                                                            
                                l=l+1

                            lst4.append(fin_mda)
                            break
                        j=j+1

                    lt=len(lst4)
                    if lt==0:                           
                        lst.append('N/A')
                    else:
                        vr=lst4[0]
                        lst.append(vr)

                    xpath= "/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div"
                    heding="Left Table Count"
                    status="Left Table Count Not Found"        
                    count(xpath,heding,status) 

                    lst5=[]
                    j=1
                    while j<rows+1:  
                        try:
                            xpath= "/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div[{}]/div/button/div/span"                                
                            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                            if cnm.lstrip().rstrip()=='Medicare Secondary Payer':                            
                                xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div[{}]/div/button/div/i'   
                                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j))))                               
                                lst5.append('N/A')                            
                                break
                        except Exception as e:                        
                            xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div[{}]/div/button'.format(j)
                            heding="Medicare Secondary Payer"
                            status="Medicare Secondary Payer Click Not Found"
                            click(xpath,heding,status)                                              
                            
                            try:                            
                                xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[4]/article/div/div/div/div[2]/div/form/div[2]/div/table/tbody/tr'                            
                                rows1=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))

                                i=1                            
                                while i<rows1+1:
                                    try:
                                        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[4]/article/div/div/div/div[2]/div/form/div[2]/div/table/tbody/tr/td[1]'                         
                                        eff_dt=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text  
                                    except Exception as e:
                                        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[4]/article/div/div/div/div[2]/div/form/div[2]/div/table/tbody/tr[{}]/td[1]'.format(i)                         
                                        eff_dt=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text                                      
                                    try:
                                        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[4]/article/div/div/div/div[2]/div/form/div[2]/div/table/tbody/tr/td[2]'                         
                                        ter_dt=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text  
                                    except Exception as e:
                                        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[4]/article/div/div/div/div[2]/div/form/div[2]/div/table/tbody/tr[{}]/td[2]'.format(i)                         
                                        ter_dt=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text  
                                    try:
                                        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[4]/article/div/div/div/div[2]/div/form/div[2]/div/table/tbody/tr/td[3]'                         
                                        lst_dt=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text  
                                    except Exception as e:
                                        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[4]/article/div/div/div/div[2]/div/form/div[2]/div/table/tbody/tr[{}]/td[3]'.format(i)                         
                                        lst_dt=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text  
                                    try:
                                        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[4]/article/div/div/div/div[2]/div/form/div[2]/div/table/tbody/tr/td[4]'                         
                                        ind=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text  
                                    except Exception as e:
                                        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[4]/article/div/div/div/div[2]/div/form/div[2]/div/table/tbody/tr[{}]/td[4]'.format(i)                         
                                        ind=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text         
                                    try:
                                        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[4]/article/div/div/div/div[2]/div/form/div[2]/div/table/tbody/tr/td[5]'                         
                                        typ=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text  
                                    except Exception as e:
                                        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[4]/article/div/div/div/div[2]/div/form/div[2]/div/table/tbody/tr[{}]/td[5]'.format(i)                         
                                        typ=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text         
                                    try:
                                        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[4]/article/div/div/div/div[2]/div/form/div[2]/div/table/tbody/tr/td[6]'                         
                                        orm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text  
                                    except Exception as e:
                                        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[4]/article/div/div/div/div[2]/div/form/div[2]/div/table/tbody/tr[{}]/td[6]'.format(i)                         
                                        orm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text
                                    try:
                                        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[4]/article/div/div/div/div[2]/div/form/div[2]/div/table/tbody/tr/td[7]'                         
                                        py_num=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text  
                                    except Exception as e:
                                        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[4]/article/div/div/div/div[2]/div/form/div[2]/div/table/tbody/tr[{}]/td[7]'.format(i)                         
                                        py_num=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text
                                    try:
                                        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[4]/article/div/div/div/div[2]/div/form/div[2]/div/table/tbody/tr/td[8]'                         
                                        grp_num=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text  
                                    except Exception as e:
                                        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[4]/article/div/div/div/div[2]/div/form/div[2]/div/table/tbody/tr[{}]/td[8]'.format(i)                         
                                        grp_num=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text                                                                
                                    
                                    element_xpath ='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[4]/article/div/div/div/div[2]/div/form/div[2]/div'
                                    element = driver.find_element("xpath", element_xpath)                                
                                    driver.execute_script("arguments[0].scrollBy(1000, 0);", element)

                                    try:
                                        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[4]/article/div/div/div/div[2]/div/form/div[2]/div/table/tbody/tr/td[9]'                         
                                        rel_dec=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text  
                                    except Exception as e:
                                        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[4]/article/div/div/div/div[2]/div/form/div[2]/div/table/tbody/tr[{}]/td[9]'.format(i)                         
                                        rel_dec=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text
                                    try:                                    
                                        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[4]/article/div/div/div/div[2]/div/form/div[2]/div/table/tbody/tr/td[10]'                         
                                        sec_dec=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text  
                                    except Exception as e:
                                        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[4]/article/div/div/div/div[2]/div/form/div[2]/div/table/tbody/tr[{}]/td[10]'.format(i)                         
                                        sec_dec=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text
                                    try:
                                        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[4]/article/div/div/div/div[2]/div/form/div[2]/div/table/tbody/tr/td[11]'                         
                                        rdc=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text  
                                    except Exception as e:
                                        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[4]/article/div/div/div/div[2]/div/form/div[2]/div/table/tbody/tr[{}]/td[11]'.format(i)                         
                                        rdc=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text
                                    try:
                                        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[4]/article/div/div/div/div[2]/div/form/div[2]/div/table/tbody/tr/td[12]'                         
                                        ins_nm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text  
                                    except Exception as e:
                                        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[4]/article/div/div/div/div[2]/div/form/div[2]/div/table/tbody/tr[{}]/td[12]'.format(i)                         
                                        ins_nm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text
                                                                
                                    if eff_dt==' ' or eff_dt=='':
                                        lst5.append('N/A')
                                    else:
                                        lst5.append(eff_dt)

                                    if ter_dt==' ' or ter_dt=='':
                                        lst5.append('N/A')
                                    else:
                                        lst5.append(ter_dt)
                                    
                                    if lst_dt==' ' or lst_dt=='':
                                        lst5.append('N/A')
                                    else:
                                        lst5.append(lst_dt)

                                    if ind==' ' or ind=='':
                                        lst5.append('N/A')
                                    else:
                                        lst5.append(ind)

                                    if typ==' ' or typ=='':
                                        lst5.append('N/A')
                                    else:
                                        lst5.append(typ)

                                    if orm==' ' or orm=='':
                                        lst5.append('N/A')
                                    else:
                                        lst5.append(orm)

                                    if py_num==' ' or py_num=='':
                                        lst5.append('N/A')
                                    else:
                                        lst5.append(py_num)

                                    if grp_num==' ' or grp_num=='':
                                        lst5.append('N/A')
                                    else:
                                        lst5.append(grp_num)

                                    if rel_dec==' ' or rel_dec=='':
                                        lst5.append('N/A')
                                    else:
                                        lst5.append(rel_dec)

                                    if sec_dec==' ' or sec_dec=='':
                                        lst5.append('N/A')
                                    else:
                                        lst5.append(sec_dec)

                                    if rdc==' ' or rdc=='':
                                        lst5.append('N/A')
                                    else:
                                        lst5.append(rdc)

                                    if ins_nm==' ' or ins_nm=='':
                                        lst5.append('N/A')
                                    else:
                                        lst5.append(ins_nm)

                                    lst5.append(':')
                                            
                                    i=i+1        
                            except Exception as e:
                                pass                       
                        j=j+1
                    
                    result_string = ','.join(lst5)

                    dt1 = result_string.replace(':', '').replace('N/A', '').replace(',', '')                                   
                    if dt1.lstrip().rstrip()==' ' or dt1.lstrip().rstrip()=='':
                        result_string="'N/A','N/A','N/A','N/A','N/A','N/A','N/A','N/A','N/A','N/A','N/A','N/A'" 

                    lst.append(result_string)              

                    xpath= "/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div"
                    heding="Left Table Count"
                    status="Left Table Count Not Found"        
                    count(xpath,heding,status) 

                    lst6=[]
                    j=1
                    while j<rows+1:  
                        try:
                            xpath= "/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div[{}]/div/button/div/span"                                
                            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                            if cnm.lstrip().rstrip()=='Crossover':                            
                                xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div[{}]/div/button/div/i'   
                                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j))))                               
                                lst6.append('N/A')                            
                                break
                        except Exception as e:
                            xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div[{}]/div/button'.format(j)
                            heding="Crossover"
                            status="Crossover Click Not Found"
                            click(xpath,heding,status)
                            
                            xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[5]/div/article/div/div/div/div[2]/div/form/div[1]/div/table/tbody/tr[1]/td[3]'
                            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text
                            if cnm==' ' or cnm=='':
                                lst6.append('N/A')
                            else:
                                lst6.append(cnm)
                            break
                        j=j+1

                    lt=len(lst6)
                    if lt==0:                           
                        lst.append('N/A')
                    else:
                        vr=lst6[0]
                        lst.append(vr)

                    xpath= "/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div"
                    heding="Left Table Count"
                    status="Left Table Count Not Found"        
                    count(xpath,heding,status) 

                    lst7=[]
                    j=1
                    while j<rows+1:  
                        try:
                            xpath= "/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div[{}]/div/button/div/span"                                
                            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                            if cnm.lstrip().rstrip()=='Hospice Benefit Periods':                            
                                xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div[{}]/div/button/div/i'   
                                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j))))                               
                                lst7.append('N/A')                            
                                break
                        except Exception as e:
                            xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div[{}]/div/button'.format(j)
                            heding="Hospice Benefit Periods"
                            status="Hospice Benefit Periods Click Not Found"
                            click(xpath,heding,status)
                                                    
                            xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[9]/div/article/div/div/div/div[2]/div/form[2]/div[1]/div/table/tbody/tr/td[1]'
                            std=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text                        
                            
                            xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[9]/div/article/div/div/div/div[2]/div/form[2]/div[1]/div/table/tbody/tr/td[2]'
                            end=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text
                            
                            dt='Start Date: ' + std + ' End Date: ' + end
                            
                            dt1 = dt.replace('Start Date: ', '').replace('End Date: ', '')

                            if dt1.lstrip().rstrip()==' ' or dt1.lstrip().rstrip()=='':
                                lst7.append('N/A')
                            else:
                                lst7.append(dt)

                            break
                        j=j+1

                    lt=len(lst7)
                    if lt==0:                           
                        lst.append('N/A')
                    else:
                        vr=lst7[0]
                        lst.append(vr)

                    xpath= "/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div"
                    heding="Left Table Count"
                    status="Left Table Count Not Found"        
                    count(xpath,heding,status) 

                    lst8=[]
                    j=1
                    while j<rows+1:
                        try:                         
                            xpath= "/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div[{}]/div/button/div/span"                                
                            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                            if cnm.lstrip().rstrip()=='Inpatient/SNF Spell History':                            
                                xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div[{}]/div/button/div/i'   
                                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j))))                               
                                lst8.append('N/A')                       
                                break
                        except Exception as e:
                            xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div[{}]/div/button'.format(j)
                            heding="Inpatient/SNF Spell History"
                            status="Inpatient/SNF Spell History Click Not Found"
                            click(xpath,heding,status)
                            
                            try:
                                xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[10]/div/article/div/div/div/div[2]/div/form/div[1]/div/table/tbody/tr'                            
                                rows1=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))

                                i=1
                                dt3 = "" 
                                while i<rows1+1:
                                    xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[10]/div/article/div/div/div/div[2]/div/form/div[1]/div/table/tbody/tr[{}]/td[2]'                         
                                    ck=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(i)))).text  

                                    if ck.lstrip().rstrip()=='Inpatient':
                                        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[10]/div/article/div/div/div/div[2]/div/form/div[1]/div/table/tbody/tr[{}]/td[3]'.format(i)
                                        std=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text                        
                                        
                                        xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[10]/div/article/div/div/div/div[2]/div/form/div[1]/div/table/tbody/tr[{}]/td[4]'.format(i)
                                        end=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text

                                        dt='Start Date: ' + std + ' End Date: ' + end                        
                                        dt1 = dt.replace('Start Date: ', '').replace('End Date: ', '')                                

                                        if dt1.lstrip().rstrip()==' ' or dt1.lstrip().rstrip()=='':
                                            pass    
                                        else:
                                            dt3 += dt + ','     
                                    i=i+1
                            except Exception as e:
                                pass

                            try:
                                dt3 = dt3.rstrip(',')
                            except Exception as e:
                                pass

                            try:                        
                                if dt3.lstrip().rstrip()==' ' or dt3.lstrip().rstrip()=='':
                                    lst8.append('N/A')
                                else:
                                    lst8.append(dt3)
                            except Exception as e:
                                lst8.append('N/A')

                            break
                        j=j+1

                    lt=len(lst8)
                    if lt==0:                           
                        lst.append('N/A')
                    else:
                        vr=lst8[0]
                        lst.append(vr)
                        
                    xpath= "/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div"
                    heding="Left Table Count"
                    status="Left Table Count Not Found"        
                    count(xpath,heding,status) 

                    lst9=[]
                    j=1
                    while j<rows+1:  
                        try:                             
                            xpath= "/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div[{}]/div/button/div/span"                                
                            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                            if cnm.lstrip().rstrip()=='Qualified Medicare Beneficiary':                            
                                xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div[{}]/div/button/div/i'   
                                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j))))                               
                                lst9.append('N/A')                            
                                break
                        except Exception as e:
                            xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div[{}]/div/button'.format(j)
                            heding="Qualified Medicare Beneficiary"
                            status="Qualified Medicare Beneficiary Click Not Found"
                            click(xpath,heding,status)
                                                        
                            xpath= "/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[6]/div/article/div/div/div/div[2]/div/form/div[1]/div/table/tbody/tr"
                            heding="Qualified Medicare Beneficiary Table Count"
                            status="Qualified Medicare Beneficiary Table Count Not Found"        
                            count(xpath,heding,status) 
                            
                            l=1
                            while l<rows+1:                              
                                xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[6]/div/article/div/div/div/div[2]/div/form/div[1]/div/table/tbody/tr[{}]/td[1]'
                                sdt=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(l)))).text
                                    
                                xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[6]/div/article/div/div/div/div[2]/div/form/div[1]/div/table/tbody/tr[{}]/td[2]'
                                edt=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(l)))).text
                                                            
                                if sdt==' ' or sdt=='':
                                    sdt='N/A'
                                if edt==' ' or edt=='':
                                    edt='N/A'                                  
                                                        
                                if l==1:
                                    fin_mda = f"{sdt}:{edt}"
                                else:
                                    fin_mda += f":{sdt}:{edt}" 
                                                            
                                l=l+1

                            lst9.append(fin_mda)
                            break
                        j=j+1

                    lt=len(lst9)
                    if lt==0:                           
                        lst.append('N/A')
                    else:
                        vr=lst9[0]
                        lst.append(vr)

                    xpath= "/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div"
                    heding="Left Table Count"
                    status="Left Table Count Not Found"        
                    count(xpath,heding,status) 

                    lst10=[]
                    j=1
                    while j<rows+1:  
                        try:                             
                            xpath= "/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div[{}]/div/button/div/span"                                
                            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                            if cnm.lstrip().rstrip()=='Part B Deductibles':                                                            
                                xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div[{}]/div/button/div/i'   
                                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j))))                               
                                lst10.append('N/A')                            
                                break
                        except Exception as e:
                            xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div[{}]/div/button'.format(j)
                            heding="Qualified Medicare Beneficiary"
                            status="Qualified Medicare Beneficiary Click Not Found"
                            click(xpath,heding,status)                                                                                    
                            
                            try:
                                xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[2]/div/article/div/div/div/div[2]/div/form/div/div[1]/div[1]/div/span/input'
                                yer_1=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath)))                                                          
                                
                                yer = yer_1.get_attribute('value')

                                if yer==' ' or yer=='':
                                    yer='N/A'                                
                                else:
                                    lst10.append(yer)                                                                                                                         
                                break
                            except Exception as e:
                                lst10.append('N/A')
                                break
                        j=j+1

                    lt=len(lst10)
                    if lt==0:                           
                        lst.append('N/A')
                    else:
                        vr=lst10[0]
                        lst.append(vr)

                    lst11=[]
                    j=1
                    while j<rows+1:  
                        try:                             
                            xpath= "/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div[{}]/div/button/div/span"                                
                            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                            if cnm.lstrip().rstrip()=='Part B Deductibles':                                                            
                                xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div[{}]/div/button/div/i'   
                                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j))))                               
                                lst11.append('N/A')                            
                                break
                        except Exception as e:
                            xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div[{}]/div/button'.format(j)
                            heding="Qualified Medicare Beneficiary"
                            status="Qualified Medicare Beneficiary Click Not Found"
                            click(xpath,heding,status)                                                                                    
                            
                            try:
                                xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[2]/div/article/div/div/div/div[2]/div/form/div/div[1]/div[2]/div/span/input'
                                amt_yer_1=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath)))                                                         
                                
                                amt_yer = amt_yer_1.get_attribute('value')

                                if amt_yer==' ' or amt_yer=='':
                                    amt_yer='N/A'                                
                                else:
                                    lst11.append(amt_yer)                                                                                                                         
                                break
                            except Exception as e:
                                lst11.append('N/A')
                                break
                        j=j+1

                    lt=len(lst11)
                    if lt==0:                           
                        lst.append('N/A')
                    else:
                        vr=lst11[0]
                        lst.append(vr)

                    lst12=[]
                    j=1
                    while j<rows+1:  
                        try:                             
                            xpath= "/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div[{}]/div/button/div/span"                                
                            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                            if cnm.lstrip().rstrip()=='Part B Deductibles':                                                            
                                xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div[{}]/div/button/div/i'   
                                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j))))                               
                                lst12.append('N/A')                            
                                break
                        except Exception as e:
                            xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/header/div[1]/div[{}]/div/button'.format(j)
                            heding="Qualified Medicare Beneficiary"
                            status="Qualified Medicare Beneficiary Click Not Found"
                            click(xpath,heding,status)                                                                                    
                            
                            try:                                     
                                xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[3]/div/div[3]/div[2]/div/section/section/div/div[2]/div/article/div/div/div/div[2]/div/form/div/div[1]/div[3]/div/span/input'
                                rem_amt_1=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath)))                                                       
                                
                                rem_amt = rem_amt_1.get_attribute('value')

                                if rem_amt==' ' or rem_amt=='':
                                    rem_amt='N/A'                                
                                else:
                                    lst12.append(rem_amt)                                                                                                                         
                                break
                            except Exception as e:
                                lst12.append('N/A')
                                break
                        j=j+1

                    lt=len(lst12)
                    if lt==0:                           
                        lst.append('N/A')
                    else:
                        vr=lst12[0]
                        lst.append(vr)

                    wb1=load_workbook(fil)
                    sheet = wb1['MCR Connex BOT']
                    column_letter = 'S'  
                    column_cells = sheet[column_letter]
                    last_row = None
                    for cell in reversed(column_cells):
                        if cell.value:
                            last_row = cell.row
                            break              
                    start_column = 'G'
                    current_column_index = openpyxl.utils.column_index_from_string(start_column)
                    current_row = last_row + 1

                    for value in lst:
                        current_column = openpyxl.utils.get_column_letter(current_column_index)
                        sheet[current_column + str(current_row)] = value
                        current_column_index += 1                
                    sheet['S' + str(int(last_row + 1))]='Done'
                    wb1.save(filename=fil)
                    wb1.close()     
                except Exception as e:
                    wb1=load_workbook(filename=fil)
                    sheet = wb1['MCR Connex BOT']
                    column_letter = 'S'
                    column_cells = sheet[column_letter]
                    last_row = None
                    for cell in reversed(column_cells):
                        if cell.value:
                            last_row = cell.row
                            break
                    sheet['S' + str(int(last_row + 1))]='Error'
                    wb1.save(fil)
                    wb1.close()
                
                try:             
                    xpath='/html/body/div[1]/div/div/div/div/div/div/div/div/div[2]/div/div[2]/div/div/div/div/div/div/div[1]/div[1]/div/div/div[2]/div/div/div[4]/div/a'
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                except Exception as e:
                    xpath= "/html/body/div[2]/div/div/div/div/div/div/div/div/div[2]/div/div[2]/div/div/div/div/div/div/div[1]/div[1]/div/div/div[2]/div/div/div[4]/div"
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()                

    driver.quit()
    messagebox.showinfo("Process Status", "Process Completed")
    sys.exit(0)    

if __name__=="__main__":        
    process()