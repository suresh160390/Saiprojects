from tkinter import BOTH, LEFT, TOP, Button, Entry, Frame, Label, PhotoImage, StringVar, Tk,Radiobutton,StringVar,IntVar,filedialog
from idlelib.tooltip import Hovertip
import sys
import os
from operator import itemgetter
import time
from tkinter import messagebox
from os import listdir
from os.path import isfile, join
from datetime import datetime, date
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, StaleElementReferenceException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import openpyxl
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from urllib.parse import urlparse
import warnings
import numpy as np
from openpyxl.utils import get_column_letter,column_index_from_string
from datetime import datetime
from pytz import timezone
import requests
from zipfile import ZipFile
import subprocess

warnings.filterwarnings("ignore")

global rows
global xpath
global heding
global status
global key
global nme
global element_1  
global j

def browse():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
        image_path2 = os.path.join(sys._MEIPASS, 'Static', 'Mapping.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')
        image_path2 = os.path.join(os.getcwd(), 'Static', 'Mapping.png')

    root.title("Process File Picker")
    root.resizable(False,False)

    root.title("Process File Picker")
    root.resizable(False,False)
   
    w = 500
    h = 160
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Browse Process - File Picker",font=("Calibri",18,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
          
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=50)
    
    answer=StringVar()
    answer.set("")
       
    def browse_button():
        global filename
        answer.set("")  
        filename = filedialog.askopenfilename()  
        txt.delete(0, 'end') 
        txt.insert(0, filename)

    def Click_Done():
        global ans        
        
        ans=txt.get()        
        
        if ans=="":
           answer.set("File Path Fields Empty Is Not Allowed...")        
        else:      
            root.destroy()          
            return ans
        
    txt=Entry(Frame2,font=("Calibri",12,"bold","italic"),width=50,justify="left")
    txt.grid(row=0,column=0,padx=5,pady=10,sticky="E")    

    photo1 = PhotoImage(file=image_path2)
            
    btn1=Button(Frame2,text="Browse",command=browse_button,image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=0,column=1,padx=3,pady=0,sticky="W")     
                
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=3,y=80,width=490,height=25)
    
    title_3=Label(Frame3,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=1,padx=150,pady=0,sticky="E")

    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=3,y=105,width=500,height=50)
    
    photo2 = PhotoImage(file=image_path1)

    btn2=Button(Frame4,command=Click_Done,text="Done",image=photo2,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=3,column=0,padx=125,pady=0,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo3 = PhotoImage(file=image_path)   

    btn3=Button(Frame4,command=Close,text="Exit",image=photo3,borderwidth=0,bg="#2c3e50")
    btn3.grid(row=3,column=1,padx=15,pady=0,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn2,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn3,'Click to Exit Process',hover_delay=1000)
    myTip2 = Hovertip(btn1,'Click to Pick File',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def process():        
    global element_1    
    global ans1
    global ans2 
    global driver
    global j
    global dob
    global mid
    global frm_dt
    global to_dt
    global cpt

    # user_pass()
    browse()

    fil=ans               
    # user = ans1
    # password = ans2

    try:
        options = Options()            
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--ignore-ssl-errors')
        options.add_argument("--disable-popup-blocking")
        driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options)
        driver.maximize_window()
        driver.get('https://mypres.phs.org/Pages/default.aspx')
    except Exception as e:                        
        try:
            options = Options()
            options.add_argument('--ignore-certificate-errors')
            options.add_argument('--ignore-ssl-errors')
            options.add_argument("--disable-popup-blocking")
            # response = requests.get('https://chromedriver.storage.googleapis.com/LATEST_RELEASE')
            # latest_version = response.text.strip()
            # chrome_driver_url = f'https://chromedriver.storage.googleapis.com/{latest_version}/chromedriver_win32.zip'
            # response = requests.get(chrome_driver_url)
            # with open('chromedriver_win32.zip', 'wb') as f:
            #     f.write(response.content)
            # with ZipFile('chromedriver_win32.zip', 'r') as zip_ref:
            #     zip_ref.extractall('.')
            driver_path = os.path.abspath('chromedriver.exe')
            driver = webdriver.Chrome(executable_path=driver_path,options=options)
            driver.maximize_window()
            driver.get('https://mypres.phs.org/Pages/default.aspx')
        except Exception as e:           
            messagebox.showinfo("Internet Problem","Pls Check Your Internet Connection")
            sys.exit(0)
        
    def text_box(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:   
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding, status)
            sys.exit(0)        
    
    def text_box_key(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:                   
                input_field =WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath)))
                input_field.send_keys(Keys.BACKSPACE * len(input_field))                
                input_field.send_keys(key)               
                # element.send_keys(Keys.TAB)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding, status)
            sys.exit(0)    

    def text_box_js(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:
                js_code = "document.querySelector('" + xpath + "').setAttribute('value', '" + key + "');"
                WebDriverWait(driver, 0).until(lambda driver: driver.execute_script(js_code))               
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo(heding, status)
            sys.exit(0)       

    def click(xpath,heding,status):
        counter = 0
        while counter < 15:
            try:             
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding,status)
            sys.exit(0)                                

    def count(xpath,heding,status):
        global rows
        counter = 0
        while counter < 15:
            try:             
                rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding,status)
            sys.exit(0)                         

    def Alert():
        counter = 0
        while counter < 15:
            try:             
                WebDriverWait(driver, 0).until (EC.alert_is_present())
                a=driver.switch_to.alert
                a.accept()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo('Alert','Alert Not Present')                                                                
                
    messagebox.showinfo('Waiting','Authentication Waiting')        
    
    driver.switch_to.window(driver.window_handles[1])
    
    file=pd.read_excel(fil,sheet_name='PRES Claim Status BOT',header=0)
    
    for index, row in file.iterrows():                                             
        mid = row[4]            
        frm_dt=row[5]
        to_dt=row[6]
        cpt = row[7]                 
        
        xpath= "/html/body/div[1]/div/header/nav/div/nav/ul/li"
        heding="Home"
        status="Home Count Not Found"        
        count(xpath,heding,status)

        j=1
        while j<rows+1:
            hm = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/header/nav/div/nav/ul/li[{}]'.format(j)))).text
            if hm.lstrip().rstrip()=='HOME':               
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[1]/div/header/nav/div/nav/ul/li[{}]/a'.format(j)))).click()                                
                break  
            j=j+1     

        xpath= "/html/body/div[1]/div/div/div[3]/form/div[3]/div[2]/div[1]/div[3]/div/div/div/div/div/ul/li"
        heding="Verify"
        status="Verify Count Not Found"        
        count(xpath,heding,status)

        j=1
        while j<rows+1:
            vf = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]/form/div[3]/div[2]/div[1]/div[3]/div/div/div/div/div/ul/li[{}]/a/div[2]/p/span'.format(j)))).text
            if vf.lstrip().rstrip()=='Verify Claims':                
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[1]/div/div/div[3]/form/div[3]/div[2]/div[1]/div[3]/div/div/div/div/div/ul/li[{}]/a'.format(j)))).click()                                                
                break
            j=j+1     

        xpath= "/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[1]/div/ul/li"
        heding="Claim Table"
        status="Claim Table Count Not Found"        
        count(xpath,heding,status)
                        
        j=1
        while j<rows+1:
            cn_st = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[1]/div/ul/li[{}]'.format(j)))).text
            if cn_st.lstrip().rstrip()=='Claims':
                try:  
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[1]/div/ul/li[{}]/a'.format(j)))).click()                                
                    break  
                except Exception as e:
                    break
            j=j+1          

        xpath= "/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[3]/div[2]/input[1]"
        heding="Member ID"
        status="Member ID Field Not Found"
        key=str(mid)
        text_box(xpath,heding,status,key)  
        
        date_object_1 = datetime.strptime(str(frm_dt), "%Y-%m-%d %H:%M:%S")
        dob3 = date_object_1.strftime("%m/%d/%Y")            
        
        xpath= "/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[3]/div[3]/input[1]"
        heding="From Date"
        status="From Date Field Not Found"
        key=dob3

        counter = 0
        while counter < 15:
            try:   
                input_field =WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath)))
                input_field.send_keys(Keys.BACKSPACE * len(input_field.get_attribute("value")))
                time.sleep(1)
                input_field.send_keys(key)                  
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo(heding, status)
            sys.exit(0)
        
        date_object_1 = datetime.strptime(str(to_dt), "%Y-%m-%d %H:%M:%S")
        dob3 = date_object_1.strftime("%m/%d/%Y")            
        
        xpath= "/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[3]/div[3]/input[2]"
        heding="To Date"
        status="To Date Field Not Found"
        key=dob3

        counter = 0
        while counter < 15:
            try:   
                input_field =WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath)))
                input_field.send_keys(Keys.BACKSPACE * len(input_field.get_attribute("value")))
                time.sleep(1)
                input_field.send_keys(key)   
                input_field.send_keys(Keys.TAB)                 
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo(heding, status)
            sys.exit(0)
                
        xpath= "/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[4]/input"
        heding="Search"
        status="Search Button Not Found"
        click(xpath,heding,status)            

        wait = WebDriverWait(driver, 20)
        wait.until(lambda driver: driver.execute_script("return document.readyState === 'complete';"))
        
        counter = 0
        while counter < 15:                
            try:                                                                                                      
                element_1 = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[5]/span[1]'))).text                                               
                if element_1 !='':                            
                        break
            except Exception as e:    
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo('Search', 'Search Result Not Found')
            sys.exit(0)        

        if element_1.lstrip().rstrip() == 'No claims found.':                              
            wb1=load_workbook(filename=fil)
            sheet = wb1['PRES Claim Status BOT']
            column_letter = 'V'  
            column_cells = sheet[column_letter]
            last_row = None
            for cell in reversed(column_cells):
                if cell.value:
                    last_row = cell.row
                    break                            
            sheet['V' + str(int(last_row + 1))]=element_1   
            wb1.save(fil)
            wb1.close()  
                       
        else:            
            
            lst=[]

            xpath= "/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[6]/table/tbody/tr"
            heding="Claim Table Count"
            status="Claim Table Count Not Found"        
            count(xpath,heding,status)
            
            j=2
            while j<rows+1:
                if rows==2:                                                                                                                                                                
                    cn = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[6]/table/tbody/tr[2]/td[1]'))).text

                    if cn==' ' or cn=='':
                        cn='N/A' 
                        lst.append(cn)   
                    else:
                        lst.append(cn)   

                    lt=len(lst)
                    if lt==0:                           
                        lst.append('N/A') 

                    st = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[6]/table/tbody/tr[2]/td[4]'))).text

                    if st==' ' or st=='':
                        st='N/A' 
                        lst.append(st)   
                    else:
                        lst.append(st)   

                    lt=len(lst)
                    if lt==0:                           
                        lst.append('N/A') 

                    xpath= "/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[6]/table/tbody/tr[2]/td[1]/a"
                    heding="Claim Number"
                    status="Claim Number Link Not Found"
                    click(xpath,heding,status)
                    
                    break
                
                elif j==rows:
                                                                                                  
                    cn = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[6]/table/tbody/tr[{}]/td[1]'.format(j)))).text

                    if cn==' ' or cn=='':
                        cn='N/A' 
                        lst.append(cn)   
                    else:
                        lst.append(cn)   

                    lt=len(lst)
                    if lt==0:                           
                        lst.append('N/A') 

                    st = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[6]/table/tbody/tr[{}]/td[4]'.format(j)))).text

                    if st==' ' or st=='':
                        st='N/A' 
                        lst.append(st)   
                    else:
                        lst.append(st)   

                    lt=len(lst)
                    if lt==0:                           
                        lst.append('N/A') 

                    xpath= "/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[6]/table/tbody/tr[{}]/td[1]/a".format(j)
                    heding="Claim Number"
                    status="Claim Number Link Not Found"
                    click(xpath,heding,status)
                    break
                j=j+1

            counter = 0
            while counter < 15:                
                try:
                    ck_tab = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[6]/h3[1]'))).text                    
                    if ck_tab.lstrip().rstrip()=='Claim Summary':                            
                        break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo('Claim Summary', 'Claim Summary Not Found')
                sys.exit(0)

            xpath= "/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[6]/table[1]/tbody/tr"
            heding="Claim Summary Table"
            status="Claim Summary Table Count Not Found"        
            count(xpath,heding,status)
            
            lst1=[]
            j=1
            while j<rows+1:
                cup_dt = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[6]/table[1]/tbody/tr[{}]/td[1]'.format(j)))).text
                if cup_dt.lstrip().rstrip()=='Claim Data Updated:':
                    cup_dt_1 = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[6]/table[1]/tbody/tr[{}]/td[2]'.format(j)))).text
                    if cup_dt_1==' ' or cup_dt_1=='':
                        cup_dt_1='N/A' 
                        lst1.append(cup_dt_1)   
                    else:
                        lst1.append(cup_dt_1)                       
                    break  
                j=j+1  

            lt=len(lst1) 
            if lt==0:                                    
                lst.append('N/A')   
            else:
                vr=lst1[0]
                lst.append(vr)

            xpath= "/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[6]/table[2]/tbody/tr"
            heding="Payment Details Table"
            status="Payment Details Table Count Not Found"        
            count(xpath,heding,status)

            lst2=[]
            j=1
            while j<rows+1:
                ck = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[6]/table[2]/tbody/tr[{}]/td[1]'.format(j)))).text
                if ck.lstrip().rstrip()=='Check #:':
                    ck_1 = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[6]/table[2]/tbody/tr[{}]/td[2]'.format(j)))).text
                    if ck_1==' ' or ck_1=='':
                        ck_1='N/A' 
                        lst2.append(ck_1)   
                    else:
                        lst2.append(ck_1)                       
                    break  
                j=j+1  
            
            lt=len(lst2) 
            if lt==0:                                    
                lst.append('N/A')   
            else:
                vr=lst2[0]
                lst.append(vr)

            lst3=[]
            j=1
            while j<rows+1:
                ck_amt = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[6]/table[2]/tbody/tr[{}]/td[1]'.format(j)))).text
                if ck_amt.lstrip().rstrip()=='Total:':
                    ck_amt_1 = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[6]/table[2]/tbody/tr[{}]/td[2]'.format(j)))).text
                    if ck_amt_1==' ' or ck_amt_1=='':
                        ck_amt_1='N/A' 
                        lst3.append(ck_amt_1)   
                    else:
                        lst3.append(ck_amt_1)                       
                    break  
                j=j+1 
            
            lt=len(lst3) 
            if lt==0:                                    
                lst.append('N/A')   
            else:
                vr=lst3[0]
                lst.append(vr)

            lst4=[]
            j=1
            while j<rows+1:
                ck_dt = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[6]/table[2]/tbody/tr[{}]/td[1]'.format(j)))).text
                if ck_dt.lstrip().rstrip()=='Date Claim was Paid On:':
                    ck_dt_1 = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[6]/table[2]/tbody/tr[{}]/td[2]'.format(j)))).text
                    if ck_dt_1==' ' or ck_dt_1=='':
                        ck_dt_1='N/A' 
                        lst4.append(ck_dt_1)   
                    else:
                        lst4.append(ck_dt_1)                      
                    break  
                j=j+1 
            
            lt=len(lst4) 
            if lt==0:                                    
                lst.append('N/A')   
            else:
                vr=lst4[0]
                lst.append(vr)
                  
            xpath= "/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[6]/table[3]/tbody/tr"
            heding="Claim Items Table"
            status="Claim Items Table Count Not Found"        
            count(xpath,heding,status)

            lst5=[]
            j=2
            while j<rows+1:                
                pro_cd = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[6]/table[3]/tbody/tr[{}]/td[2]'.format(j)))).text
                if pro_cd==str(cpt):
                    amt = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[6]/table[3]/tbody/tr[{}]/td[7]'.format(j)))).text                
                    if amt==' ' or amt=='':
                        amt='N/A' 
                        lst5.append(amt)                           
                    else:
                        lst5.append(amt)
                    break
                j=j+1 

            lt=len(lst5) 
            if lt==0:                                    
                lst.append('N/A')   
            else:
                vr=lst5[0]
                lst.append(vr)

            lst6=[]
            j=2
            while j<rows+1:                
                pro_cd = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[6]/table[3]/tbody/tr[{}]/td[2]'.format(j)))).text
                if pro_cd==str(cpt):
                    det = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[6]/table[3]/tbody/tr[{}]/td[8]'.format(j)))).text
                    if det==' ' or det=='':
                        det='N/A' 
                        lst6.append(det)   
                    else:
                        lst6.append(det)   
                    break
                j=j+1 

            lt=len(lst6) 
            if lt==0:                                    
                lst.append('N/A')   
            else:
                vr=lst6[0]
                lst.append(vr)

            lst7=[]
            j=2
            while j<rows+1:                
                pro_cd = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[6]/table[3]/tbody/tr[{}]/td[2]'.format(j)))).text
                if pro_cd==str(cpt):
                    ins = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[6]/table[3]/tbody/tr[{}]/td[9]'.format(j)))).text
                    if ins==' ' or ins=='':
                        ins='N/A' 
                        lst7.append(ins)   
                    else:
                        lst7.append(ins)
                    break
                j=j+1

            lt=len(lst7) 
            if lt==0:                                    
                lst.append('N/A')   
            else:
                vr=lst7[0]
                lst.append(vr)

            lst8=[]
            j=2
            while j<rows+1:                
                pro_cd = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[6]/table[3]/tbody/tr[{}]/td[2]'.format(j)))).text
                if pro_cd==str(cpt):
                    pln_pd = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[6]/table[3]/tbody/tr[{}]/td[10]'.format(j)))).text
                    if pln_pd==' ' or pln_pd=='':
                        pln_pd='N/A' 
                        lst8.append(pln_pd)   
                    else:
                        lst8.append(pln_pd)                
                    break
                j=j+1 

            lt=len(lst8) 
            if lt==0:                                    
                lst.append('N/A')   
            else:
                vr=lst8[0]
                lst.append(vr)           
            
            dec = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]/form/div[3]/div/div[2]/div[2]/div/div[2]/div/div[6]/table[4]/tbody'))).text

            if dec==' ' or dec=='':
                dec='N/A' 
                lst.append(dec)   
            else:
                lst.append(dec)   

            lt=len(lst)
            if lt==0:                           
                lst.append('N/A')        

            wb1=load_workbook(fil)
            sheet = wb1['PRES Claim Status BOT']
            column_letter = 'V'  
            column_cells = sheet[column_letter]
            last_row = None
            for cell in reversed(column_cells):
                if cell.value:
                    last_row = cell.row
                    break              
            start_column = 'K'
            current_column_index = openpyxl.utils.column_index_from_string(start_column)
            current_row = last_row + 1

            for value in lst:
                current_column = openpyxl.utils.get_column_letter(current_column_index)
                sheet[current_column + str(current_row)] = value
                current_column_index += 1                
            sheet['V' + str(int(last_row + 1))]='Done'
            wb1.save(filename=fil)
            wb1.close()                                 

    driver.quit()
    messagebox.showinfo("Process Status", "Process Completed")
    sys.exit(0)    

if __name__=="__main__":        
    process()
    