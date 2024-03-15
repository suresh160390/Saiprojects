from tkinter import BOTH, LEFT, TOP, Button, Entry, Frame, Label, PhotoImage, StringVar, Tk,Radiobutton,StringVar,IntVar,filedialog
from idlelib.tooltip import Hovertip
import sys
import os
from operator import itemgetter
import time
from tkinter import messagebox
from os import listdir
from os.path import isfile, join
from datetime import datetime, date
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, StaleElementReferenceException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import openpyxl
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from urllib.parse import urlparse
import warnings
import numpy as np
from openpyxl.utils import get_column_letter,column_index_from_string
from datetime import datetime
from pytz import timezone
import requests
from zipfile import ZipFile
import subprocess

warnings.filterwarnings("ignore")

global rows
global xpath
global heding
global status
global key
global nme
global ans
element_1 = None
global j

def radio():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')

    # root.title("Sharepoint - User Login And File Details")
    # root.resizable(False,False)

    root.title("Veterans Affairs All Process Report")
    root.resizable(False,False)
   
    w = 650
    h = 300
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Please Select Veterans Affairs - Claim Submission (OR) Eligibility (OR) Referrals (OR) Claim Status",font=("Calibri",11,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
    
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=300)
    
    title1=Label(Frame2,text="Claim Submission Report :",font=("Calibri",17,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title1.grid(row=0,column=0,padx=45,pady=5,sticky="W")

    title2=Label(Frame2,text="Eligibility Report :",font=("Calibri",17,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title2.grid(row=1,column=0,padx=45,pady=5,sticky="W")

    title3=Label(Frame2,text="Referrals Report :",font=("Calibri",17,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title3.grid(row=2,column=0,padx=45,pady=5,sticky="W")

    title4=Label(Frame2,text="Claim Status Report :",font=("Calibri",17,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title4.grid(row=3,column=0,padx=45,pady=5,sticky="W")

    answer=StringVar()
    answer.set("")

    title_3=Label(Frame2,text=answer.get(),textvariable=answer,font=("Calibri",12,"bold","italic"),bg="#2c3e50",fg="Red",justify="center",width=68)
    title_3.grid(row=4,column=0,columnspan=2,padx=0,pady=0,sticky="W")

    def Radio(*event):
         answer.set("")

    global var

    var = IntVar()

    R1 = Radiobutton(Frame2,variable=var, value=1,command=lambda: Radio())
    R1.grid(row=0,column=1,padx=0,pady=5,sticky="W")

    R2 = Radiobutton(Frame2,variable=var,value=2,command=lambda: Radio())
    R2.grid(row=1,column=1,padx=0,pady=5,sticky="W")

    R3 = Radiobutton(Frame2,variable=var,value=3,command=lambda: Radio())
    R3.grid(row=2,column=1,padx=0,pady=5,sticky="W")
    
    R4 = Radiobutton(Frame2,variable=var,value=4,command=lambda: Radio())
    R4.grid(row=3,column=1,padx=0,pady=5,sticky="W")
    
    def Click_Done():
        selection = str(var.get())

        if selection==str(1):
           answer.set("")
           root.destroy() 
           process()
        elif selection==str(2):
            answer.set("")
            root.destroy() 
            eligibility()           
        elif selection==str(3):
            answer.set("")
            root.destroy() 
            referral()           
        elif selection==str(4):
            answer.set("")
            root.destroy()            
        else:    
            answer.set("Please Select Any One Option...")            
    
    photo1 = PhotoImage(file=image_path1)

    btn=Button(Frame2,command=Click_Done,text="Done",image=photo1,borderwidth=0,bg="#2c3e50")
    btn.grid(row=4,column=0,padx=125,pady=20,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo = PhotoImage(file=image_path)   

    btn1=Button(Frame2,command=Close,text="Exit",image=photo,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=4,column=1,padx=0,pady=20,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn1,'Click to Exit Process',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def browse():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
        image_path2 = os.path.join(sys._MEIPASS, 'Static', 'Mapping.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')
        image_path2 = os.path.join(os.getcwd(), 'Static', 'Mapping.png')

    root.title("Process File Picker")
    root.resizable(False,False)

    root.title("Process File Picker")
    root.resizable(False,False)
   
    w = 500
    h = 160
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Browse Process - File Picker",font=("Calibri",18,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
          
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=50)
    
    answer=StringVar()
    answer.set("")
       
    def browse_button():
        global filename
        answer.set("")  
        filename = filedialog.askopenfilename()  
        txt.delete(0, 'end') 
        txt.insert(0, filename)

    def Click_Done():
        global ans        
        
        ans=txt.get()        
        
        if ans=="":
           answer.set("File Path Fields Empty Is Not Allowed...")        
        else:      
            root.destroy()          
            return ans
        
    txt=Entry(Frame2,font=("Calibri",12,"bold","italic"),width=50,justify="left")
    txt.grid(row=0,column=0,padx=5,pady=10,sticky="E")    

    photo1 = PhotoImage(file=image_path2)
            
    btn1=Button(Frame2,text="Browse",command=browse_button,image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=0,column=1,padx=3,pady=0,sticky="W")     
                
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=3,y=80,width=490,height=25)
    
    title_3=Label(Frame3,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=1,padx=150,pady=0,sticky="E")

    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=3,y=105,width=500,height=50)
    
    photo2 = PhotoImage(file=image_path1)

    btn2=Button(Frame4,command=Click_Done,text="Done",image=photo2,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=3,column=0,padx=125,pady=0,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo3 = PhotoImage(file=image_path)   

    btn3=Button(Frame4,command=Close,text="Exit",image=photo3,borderwidth=0,bg="#2c3e50")
    btn3.grid(row=3,column=1,padx=15,pady=0,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn2,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn3,'Click to Exit Process',hover_delay=1000)
    myTip2 = Hovertip(btn1,'Click to Pick File',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def process():        
    global element_1    
    global ans1
    global ans2 
    global driver
    global j
    global dob
    global mid
    global frm_dt
    global to_dt
    global cpt
    global rows
    global rows1
    
    browse()
      
    fil=ans               

    try:
        options = Options()            
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--ignore-ssl-errors')
        options.add_argument("--disable-popup-blocking")
        prefs ={"profile.password_manager_enabled": False}      
        options.add_experimental_option("prefs",prefs)  
        options.add_argument("--disable-blink-features=AutomationControlled")
        driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options)
        driver.maximize_window()
        driver.get('https://www.myvaccn.com/site/vaccn/main/public/login#/home')
    except Exception as e:                        
        try:
            options = Options()
            options.add_argument('--ignore-certificate-errors')
            options.add_argument('--ignore-ssl-errors')
            options.add_argument("--disable-popup-blocking")
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            prefs ={"profile.password_manager_enabled": False}      
            options.add_experimental_option("prefs",prefs) 
            driver_path = os.path.abspath('chromedriver.exe')
            driver = webdriver.Chrome(executable_path=driver_path,options=options)
            driver.maximize_window()
            driver.get('https://www.myvaccn.com/site/vaccn/main/public/login#/home')
        except Exception as e:           
            messagebox.showinfo("Driver Problem","Pls Check Your Chrome Driver Version")
            sys.exit(0)
        
    def text_box(xpath,heding,status,key):                
        counter = 0
        while counter < 5:
            try:   
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()                     
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding, status)
            sys.exit(0)        
    
    def text_box_key(xpath,heding,status,key):                
        counter = 0
        while counter < 5:
            try:                   
                element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, xpath)))
                element.clear()                
                element.send_keys(key)                
                # element.send_keys(Keys.TAB)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding, status)
            sys.exit(0)    

    def text_box_js(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:
                js_code = "document.querySelector('" + xpath + "').setAttribute('value', '" + key + "');"
                WebDriverWait(driver, 0).until(lambda driver: driver.execute_script(js_code))               
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo(heding, status)
            sys.exit(0)       

    def click(xpath,heding,status):
        counter = 0
        while counter < 5:
            try:             
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding,status)
            sys.exit(0)                                

    def count(xpath,heding,status):
        global rows
        counter = 0
        while counter < 5:
            try:             
                rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding,status)
            sys.exit(0)                         

    def count_1(xpath,heding,status):
        global rows1
        counter = 0
        while counter < 5:
            try:             
                rows1=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding,status)
            sys.exit(0)        

    def Alert():
        counter = 0
        while counter < 15:
            try:             
                WebDriverWait(driver, 0).until (EC.alert_is_present())
                a=driver.switch_to.alert
                a.accept()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo('Alert','Alert Not Present')                                                                
    
    messagebox.showinfo('Waiting','Authentication Waiting')        
   
    file=pd.read_excel(fil,sheet_name='VA Claims',header=0)
    
    for index, row in file.iterrows():                                      
        fnm = row[1]  
        lnm=row[2]
        dob = row[3]
        meb_id=row[4]
        gen=row[5]
        acc_num=row[6]
        add=row[7]
        cty=row[8]
        st=row[9]
        zip_cod=row[10]
        pls_ser=row[11]
        dig1=row[12]
        dig2=row[13]
        dig3=row[14]
        dig4=row[15]                     
        dos=row[16]
        cpt=row[17]
        unit=row[18]
        amt=row[19]
        ref_num=row[20]
        
        try:
            driver.switch_to.frame("component1_ssoFrame")
        except Exception as e:
            pass

        xpath= "/html/body/div[6]/div[3]/table/tbody[2]/tr"
        heding="MELEECIA TUCKER MSW Table Count"        
        status="MELEECIA TUCKER MSW Table Count Not Found"                    
        count(xpath,heding,status)
        
        j=1
        while j<rows+1:   
            ck=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[6]/div[3]/table/tbody[2]/tr[{}]/td[1]'.format(j)))).text
            if ck.lstrip().rstrip()=='MELEECIA TUCKER MSW':                            
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[6]/div[3]/table/tbody[2]/tr[{}]/td[1]/div/a'.format(j)))).click()                            
                break           
            j=j+1

        xpath= "/html/body/div[4]/form/table[1]/tbody/tr[6]/td[1]/input"
        heding="First Name"
        status=" First Name Not Found"
        key=fnm.lstrip().rstrip()
        text_box(xpath,heding,status,key)

        xpath= "/html/body/div[4]/form/table[1]/tbody/tr[6]/td[3]/input"
        heding="Last Name"
        status=" Last Name Not Found"
        key=lnm.lstrip().rstrip()
        text_box(xpath,heding,status,key)

        original_datetime = datetime.strptime(str(dob), '%Y-%m-%d %H:%M:%S')
        dob1 = original_datetime.strftime('%#m/%#d/%Y')
       
        lst=dob1.split('/')
        mn=lst[0]
        dt=lst[1]
        yy=lst[2]                       

        xpath= "/html/body/div[4]/form/table[1]/tbody/tr[6]/td[4]/input[1]"
        heding="Date of Birth "        
        status="Date of Birth Month Field Not Found"
        key=mn
        text_box(xpath,heding,status,key)        

        xpath= "/html/body/div[4]/form/table[1]/tbody/tr[6]/td[4]/input[2]"
        heding="Date of Birth"        
        status="Date of Birth Date Field Not Found"
        key=dt
        text_box(xpath,heding,status,key)        

        xpath= "/html/body/div[4]/form/table[1]/tbody/tr[6]/td[4]/input[3]"
        heding="Date of Birth"        
        status="Date of Birth Year Field Not Found"
        key=yy
        text_box(xpath,heding,status,key)        
        
        xpath='/html/body/div[4]/form/table[2]/tbody/tr[3]/td[1]/table/tbody/tr/td[1]/span/input[1]'
        heding="MVI ICN"
        status="MVI ICN Not Found"        
        click(xpath,heding,status)
              
        xpath= "/html/body/div[4]/form/table[2]/tbody/tr[3]/td[1]/table/tbody/tr/td[3]/input"
        heding="MemberID"        
        status="MemberID Field Not Found"
        key=meb_id
        text_box(xpath,heding,status,key)

        xpath='/html/body/div[4]/form/table[2]/tbody/tr[3]/td[3]/input'
        heding="Submit"
        status="Submit Button Not Found"        
        click(xpath,heding,status)
        
        counter = 0
        while counter < 5:
            try:            
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/table/tbody/tr[1]/td[2]/div')))                               
                err=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/table/tbody'))).text

                wb1=load_workbook(filename=fil)
                sheet = wb1['VA Claims']
                column_letter = 'V'  
                column_cells = sheet[column_letter]
                last_row = None
                for cell in reversed(column_cells):
                    if cell.value:
                        last_row = cell.row
                        break                            
                sheet['V' + str(int(last_row + 1))]=err
                wb1.save(fil)
                wb1.close()               
               
                break
            except Exception as e:            
                try:                        
                    xpath= "/html/body/form/table[1]/tbody/tr[5]/td[2]/input"
                    heding="Gender Count"        
                    status="Gender Count Not Found"                    
                    count(xpath,heding,status)
                    
                    gen=gen[0]
                    
                    j=1
                    while j<rows+1:                                                                                            
                        ck1=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/form/table[1]/tbody/tr[5]/td[2]/input[{}]'.format(j))))
                        ck = ck1.get_attribute('value')
                        if ck.lstrip().rstrip().lower()==gen.lstrip().rstrip().lower():                            
                            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/form/table[1]/tbody/tr[5]/td[2]/input[{}]'.format(j)))).click()                            
                            break
                        elif ck.lstrip().rstrip().lower()==gen.lstrip().rstrip().lower():                            
                            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/form/table[1]/tbody/tr[5]/td[2]/input[{}]'.format(j)))).click()                            
                            break
                        elif ck.lstrip().rstrip().lower()=='u':        
                            if gen.lstrip().rstrip().lower()=='p':                    
                                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/form/table[1]/tbody/tr[5]/td[2]/input[{}]'.format(j)))).click()                            
                                break
                        j=j+1
                          
                    xpath= "/html/body/form/table[1]/tbody/tr[7]/td[2]/input"
                    heding="Account Number"        
                    status="Account Number Field Not Found"
                    key=acc_num
                    text_box(xpath,heding,status,key)

                    xpath= "/html/body/form/table[1]/tbody/tr[8]/td[2]/input"
                    heding="Address"        
                    status="Address Field Not Found"
                    key=add
                    text_box(xpath,heding,status,key)

                    xpath= "/html/body/form/table[1]/tbody/tr[10]/td[2]/input"
                    heding="City"        
                    status="City Field Not Found"
                    key=cty
                    text_box(xpath,heding,status,key)

                    xpath= "/html/body/form/table[1]/tbody/tr[11]/td[2]/select/option"
                    heding="State Count"        
                    status="State Count Not Found"                    
                    count(xpath,heding,status)

                    j=1
                    while j<rows+1:   
                        ck=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/form/table[1]/tbody/tr[11]/td[2]/select/option[{}]'.format(j)))).text
                        if ck.lstrip().rstrip().lower()==st.lstrip().rstrip().lower():
                            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/form/table[1]/tbody/tr[11]/td[2]/select/option[{}]'.format(j)))).click()                                                     
                            break
                        j=j+1

                    xpath= "/html/body/form/table[1]/tbody/tr[12]/td[2]/input[1]"
                    heding="ZIP Code"        
                    status="ZIP Code Field Not Found"
                    key=zip_cod
                    text_box(xpath,heding,status,key)

                    xpath= "/html/body/form/table[1]/tbody/tr[24]/td[2]/select/option"
                    heding="Place of Service Count"        
                    status="Place of Service Count Not Found"                    
                    count(xpath,heding,status)

                    j=1
                    while j<rows+1:   
                        ck=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/form/table[1]/tbody/tr[24]/td[2]/select/option[{}]'.format(j)))).text
                        if pls_ser.lstrip().rstrip().lower() in ck.lstrip().rstrip().lower():
                            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/form/table[1]/tbody/tr[24]/td[2]/select/option[{}]'.format(j)))).click()                                                     
                            break
                        j=j+1

                    xpath='/html/body/form/table[2]/tbody/tr/td/input[1]'
                    heding="Submit"
                    status="Submit Button Not Found"        
                    click(xpath,heding,status)
                    
                    cnt=0
                    
                    if pd.isnull(dig1):
                        pass
                    else:
                        cnt=cnt+1

                        lst=dig1.split('.')
                        cd=lst[0]
                        nm=lst[1]
                    
                        xpath='/html/body/form/div[2]/table[2]/tbody/tr[2]/td[2]/input[1]'
                        heding="Diagnosis 1"
                        status="Diagnosis 1 Field Not Found"
                        key=str(cd)
                        text_box(xpath,heding,status,key)

                        xpath='/html/body/form/div[2]/table[2]/tbody/tr[2]/td[2]/input[2]'
                        heding="Diagnosis 1"
                        status="Diagnosis 1 Field Not Found"
                        key=str(nm)
                        text_box(xpath,heding,status,key)
                    

                    if pd.isnull(dig2):
                        pass
                    else:
                        cnt=cnt+1

                        lst=dig2.split('.')
                        cd=lst[0]
                        nm=lst[1]
                    
                        xpath='/html/body/form/div[2]/table[2]/tbody/tr[2]/td[4]/input[1]'
                        heding="Diagnosis 2"
                        status="Diagnosis 2 Field Not Found"
                        key=str(cd)
                        text_box(xpath,heding,status,key)

                        xpath='/html/body/form/div[2]/table[2]/tbody/tr[2]/td[4]/input[2]'
                        heding="Diagnosis 2"
                        status="Diagnosis 2 Field Not Found"
                        key=str(nm)
                        text_box(xpath,heding,status,key)

                    if pd.isnull(dig3):
                        pass
                    else:
                        cnt=cnt+1

                        lst=dig3.split('.')
                        cd=lst[0]
                        nm=lst[1]
                    
                        xpath='/html/body/form/div[2]/table[2]/tbody/tr[2]/td[6]/input[1]'
                        heding="Diagnosis 3"
                        status="Diagnosis 3 Field Not Found"
                        key=str(cd)
                        text_box(xpath,heding,status,key)

                        xpath='/html/body/form/div[2]/table[2]/tbody/tr[2]/td[6]/input[2]'
                        heding="Diagnosis 3"
                        status="Diagnosis 3 Field Not Found"
                        key=str(nm)
                        text_box(xpath,heding,status,key)

                    if pd.isnull(dig4):
                        pass
                    else:
                        cnt=cnt+1

                        lst=dig4.split('.')
                        cd=lst[0]
                        nm=lst[1]
                    
                        xpath='/html/body/form/div[2]/table[2]/tbody/tr[2]/td[8]/input[1]'
                        heding="Diagnosis 4"
                        status="Diagnosis 4 Field Not Found"
                        key=str(cd)
                        text_box(xpath,heding,status,key)

                        xpath='/html/body/form/div[2]/table[2]/tbody/tr[2]/td[8]/input[2]'
                        heding="Diagnosis 4"
                        status="Diagnosis 4 Field Not Found"
                        key=str(nm)
                        text_box(xpath,heding,status,key)
                    
                    original_datetime = datetime.strptime(str(dos), '%Y-%m-%d %H:%M:%S')
                    dos1 = original_datetime.strftime('%#m/%#d/%Y')
                
                    lst=dos1.split('/')
                    mn=lst[0]
                    dt=lst[1]
                    yy=lst[2]                       

                    xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td/table/tbody/tr[1]/td[1]/input[1]"
                    heding="From Date of service"        
                    status="From Date of service Month Field Not Found"
                    key=mn
                    text_box(xpath,heding,status,key)     

                    xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td/table/tbody/tr[1]/td[1]/input[2]"
                    heding="From Date of service"        
                    status="From Date of service Date Field Not Found"
                    key=dt
                    text_box(xpath,heding,status,key)   
                    
                    xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td/table/tbody/tr[1]/td[1]/input[3]"
                    heding="From Date of service"        
                    status="From Date of service Year Field Not Found"
                    key=yy
                    text_box(xpath,heding,status,key)   

                    xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td/table/tbody/tr[1]/td[1]/input[4]"
                    heding="To Date of service"        
                    status="To Date of service Month Field Not Found"
                    key=mn
                    text_box(xpath,heding,status,key)     

                    xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td/table/tbody/tr[1]/td[1]/input[5]"
                    heding="To Date of service"        
                    status="To Date of service Date Field Not Found"
                    key=dt
                    text_box(xpath,heding,status,key)   
                    
                    xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td/table/tbody/tr[1]/td[1]/input[6]"
                    heding="To Date of service"        
                    status="To Date of service Year Field Not Found"
                    key=yy
                    text_box(xpath,heding,status,key)   
                    
                    xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td/table/tbody/tr[1]/td[3]/nobr/input[1]"
                    heding="Procedure"        
                    status="Procedure Field Not Found"
                    key=cpt
                    text_box(xpath,heding,status,key)

                    if cnt==1:
                        xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td/table/tbody/tr[1]/td[5]/nobr/input[1]"
                        heding="Diagnosis Count 1"        
                        status="Diagnosis Count 1 Field Not Found"
                        key=str(1)
                        text_box(xpath,heding,status,key)
                       
                    if cnt==2:
                        xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td/table/tbody/tr[1]/td[5]/nobr/input[1]"
                        heding="Diagnosis Count 1"        
                        status="Diagnosis Count 1 Field Not Found"
                        key=str(1)
                        text_box(xpath,heding,status,key)

                        xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td/table/tbody/tr[1]/td[5]/nobr/input[2]"
                        heding="Diagnosis Count 2"        
                        status="Diagnosis Count 2 Field Not Found"
                        key=str(2)
                        text_box(xpath,heding,status,key)

                    if cnt==3:
                        xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td/table/tbody/tr[1]/td[5]/nobr/input[1]"
                        heding="Diagnosis Count 1"        
                        status="Diagnosis Count 1 Field Not Found"
                        key=str(1)
                        text_box(xpath,heding,status,key)

                        xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td/table/tbody/tr[1]/td[5]/nobr/input[2]"
                        heding="Diagnosis Count 2"        
                        status="Diagnosis Count 2 Field Not Found"
                        key=str(2)
                        text_box(xpath,heding,status,key)

                        xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td/table/tbody/tr[1]/td[5]/nobr/input[3]"
                        heding="Diagnosis Count 3"        
                        status="Diagnosis Count 3 Field Not Found"
                        key=str(3)
                        text_box(xpath,heding,status,key)

                    if cnt==4:
                        xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td/table/tbody/tr[1]/td[5]/nobr/input[1]"
                        heding="Diagnosis Count 1"        
                        status="Diagnosis Count 1 Field Not Found"
                        key=str(1)
                        text_box(xpath,heding,status,key)

                        xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td/table/tbody/tr[1]/td[5]/nobr/input[2]"
                        heding="Diagnosis Count 2"        
                        status="Diagnosis Count 2 Field Not Found"
                        key=str(2)
                        text_box(xpath,heding,status,key)

                        xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td/table/tbody/tr[1]/td[5]/nobr/input[3]"
                        heding="Diagnosis Count 3"        
                        status="Diagnosis Count 3 Field Not Found"
                        key=str(3)
                        text_box(xpath,heding,status,key)

                        xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td/table/tbody/tr[1]/td[5]/nobr/input[4]"
                        heding="Diagnosis Count 4"        
                        status="Diagnosis Count 4 Field Not Found"
                        key=str(4)
                        text_box(xpath,heding,status,key)

                    xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td/table/tbody/tr[1]/td[6]/nobr/input[1]"
                    heding="Charge"        
                    status="Charge Field Not Found"
                    key=str(amt)
                    text_box(xpath,heding,status,key)
                
                    xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td/table/tbody/tr[1]/td[7]/input[1]"
                    heding="Units"        
                    status="Units Field Not Found"
                    key=str(unit)
                    text_box(xpath,heding,status,key)

                    xpath='/html/body/form/div[4]/table[2]/tbody/tr[5]/td/table/tbody/tr/td[1]/input[1]'
                    heding="Continue with XPressClaim"
                    status="Continue with XPressClaim Button Not Found"        
                    click(xpath,heding,status)
                    
                    xpath='/html/body/form/div[5]/center/nobr/input[2]'
                    heding="No, I have supplemental claim and/or line data to enter."
                    status="No, I have supplemental claim and/or line data to enter. Button Not Found"        
                    click(xpath,heding,status)                   

                    xpath= "/html/body/form/fieldset[1]/span/span/span/table/tbody/tr"
                    heding="Prior authorization or referral number Table Count"        
                    status="Prior authorization or referral number Table Count Not Found"                    
                    count(xpath,heding,status)

                    flage=False

                    j=1
                    while j<rows+1:
                        xpath= "/html/body/form/fieldset[1]/span/span/span/table/tbody/tr[{}]/td".format(j)
                        heding="Prior authorization or referral number Table Count"
                        status="Prior authorization or referral number Table Field Count Not Found"                  
                        count_1(xpath,heding,status)
                        
                        i=1
                        while i<rows1+1:
                            ck=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/form/fieldset[1]/span/span/span/table/tbody/tr[{}]/td[{}]'.format(j,i)))).text
                            if ck.lstrip().rstrip()=='Prior authorization or referral number':
                                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/form/fieldset[1]/span/span/span/table/tbody/tr[{}]/td[{}]/input'.format(j,i)))).click()                                                     
                                flage=True
                                break                                                 
                            i=i+1

                        if flage:
                            break
                        
                        j=j+1
                    
                    xpath= "/html/body/form/fieldset[23]/span/span[1]/span/input"
                    heding="Prior authorization number"        
                    status="Prior authorization number Field Not Found"
                    key=str(ref_num)
                    text_box(xpath,heding,status,key)

                    xpath='/html/body/form/fieldset[31]/table/tbody/tr/td/div/table/tbody/tr/td[2]/input[1]'
                    heding="Continue with XPressClaim"
                    status="Continue with XPressClaim Button Not Found"        
                    click(xpath,heding,status)

                    xpath='/html/body/form/table[2]/tbody/tr/td/div/table/tbody/tr/td[2]/input[1]'
                    heding="Final Continue with XPressClaim"
                    status="Final Continue with XPressClaim Button Not Found"        
                    click(xpath,heding,status)

                    xpath='/html/body/div[4]/center/nobr/input[1]'
                    heding="Yes, submit this claim."
                    status="Yes, submit this claim. Button Not Found"
                    click(xpath,heding,status)
                    
                    counter = 0
                    while counter < 10:
                        try:
                            cn=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/table/tbody/tr[5]'))).text
                            break
                        except Exception as e:
                            time.sleep(1)
                            counter += 1

                    wb1=load_workbook(filename=fil)
                    sheet = wb1['VA Claims']
                    column_letter = 'V'
                    column_cells = sheet[column_letter]
                    last_row = None
                    for cell in reversed(column_cells):
                        if cell.value:
                            last_row = cell.row
                            break                            
                    sheet['V' + str(int(last_row + 1))]=cn
                    wb1.save(fil)
                    wb1.close()  
                    
                    xpath='/html/body/div[2]/form/table/tbody/tr[3]/td/a'
                    heding="Same Location"
                    status="Same Location Link Not Found"        
                    click(xpath,heding,status)                      

                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1             
        else:
            wb1=load_workbook(filename=fil)
            sheet = wb1['VA Claims']
            column_letter = 'V'  
            column_cells = sheet[column_letter]
            last_row = None
            for cell in reversed(column_cells):
                if cell.value:
                    last_row = cell.row
                    break                            
            sheet['V' + str(int(last_row + 1))]='Error'   
            wb1.save(fil)
            wb1.close()                 
    
    driver.quit()
    messagebox.showinfo("Process Status", "Process Completed")
    sys.exit(0)    

def eligibility():        
    global element_1    
    global ans    
    global driver            
    global rows
    global rows1
    
    browse()
      
    fil=ans

    try:
        options = Options()            
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--ignore-ssl-errors')
        options.add_argument("--disable-popup-blocking")
        prefs ={"profile.password_manager_enabled": False}      
        options.add_experimental_option("prefs",prefs)  
        options.add_argument("--disable-blink-features=AutomationControlled")
        driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options)
        driver.maximize_window()
        driver.get('https://www.myvaccn.com/site/vaccn/main/public/login#/home')
    except Exception as e:                        
        try:
            options = Options()
            options.add_argument('--ignore-certificate-errors')
            options.add_argument('--ignore-ssl-errors')
            options.add_argument("--disable-popup-blocking")
            prefs ={"profile.password_manager_enabled": False}      
            options.add_experimental_option("prefs",prefs)  
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_experimental_option("useAutomationExtension", False)
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            driver_path = os.path.abspath('chromedriver.exe')
            driver = webdriver.Chrome(executable_path=driver_path,options=options)
            driver.maximize_window()
            driver.get('https://www.myvaccn.com/site/vaccn/main/public/login#/home')
        except Exception as e:           
            messagebox.showinfo("Driver Problem","Pls Check Your Chrome Driver Version")
            sys.exit(0)
        
    def text_box(xpath,heding,status,key):                
        counter = 0
        while counter < 5:
            try:   
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()                     
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:           
            messagebox.showinfo(heding, status)
            sys.exit(0)        
    
    def text_box_key(xpath,heding,status,key):                
        counter = 0
        while counter < 5:
            try:                   
                element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, xpath)))
                element.clear()                
                element.send_keys(key)                
                # element.send_keys(Keys.TAB)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:            
            messagebox.showinfo(heding, status)
            sys.exit(0)    

    def text_box_js(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:
                js_code = "document.querySelector('" + xpath + "').setAttribute('value', '" + key + "');"
                WebDriverWait(driver, 0).until(lambda driver: driver.execute_script(js_code))               
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo(heding, status)
            sys.exit(0)       

    def click(xpath,heding,status):
        counter = 0
        while counter < 5:
            try:             
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:            
            messagebox.showinfo(heding,status)
            sys.exit(0)                                

    def count(xpath,heding,status):
        global rows
        counter = 0
        while counter < 5:
            try:             
                rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:            
            messagebox.showinfo(heding,status)
            sys.exit(0)                         

    def count_1(xpath,heding,status):
        global rows1
        counter = 0
        while counter < 5:
            try:             
                rows1=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:            
            messagebox.showinfo(heding,status)
            sys.exit(0)        

    def Alert():
        counter = 0
        while counter < 15:
            try:             
                WebDriverWait(driver, 0).until (EC.alert_is_present())
                a=driver.switch_to.alert
                a.accept()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo('Alert','Alert Not Present')                                                                
    
    messagebox.showinfo('Waiting','Authentication Waiting')        
   
    file=pd.read_excel(fil,sheet_name='VA EV',header=0)
    
    for index, row in file.iterrows():                                              
        meb_id=row[0]
        ssn=row[1]        
        
        if not pd.isnull(meb_id):
            key = meb_id
            xpath= '/html/body/div[1]/div[1]/div[2]/div[1]/div[1]/div/div/div/div/div[2]/form/div[2]/div[1]/div/div[1]/fieldset/div/div[2]/label'
            heding="MVI ICN"
            status="MVI ICN Not Found"
            click(xpath,heding,status)

            xpath= "/html/body/div[1]/div[1]/div[2]/div[1]/div[1]/div/div/div/div/div[2]/form/div[2]/div[1]/div/div[3]/div/div/input"
            heding="MVI ICN or SSN"
            status="MVI ICN or SSN Field Not Found"
            key=key.lstrip().rstrip()
            text_box(xpath,heding,status,key)      
            
        else:
            key = ssn           

            xpath= "/html/body/div[1]/div[1]/div[2]/div[1]/div[1]/div/div/div/div/div[2]/form/div[2]/div[1]/div/div[2]/div/div/input"
            heding="MVI ICN or SSN"
            status="MVI ICN or SSN Field Not Found"
            key=key.lstrip().rstrip()
            text_box(xpath,heding,status,key)            
           

        xpath= '/html/body/div[1]/div[1]/div[2]/div[1]/div[1]/div/div/div/div/div[2]/form/div[3]/div/button'
        heding="Submit"
        status="Submit Button Not Found"
        click(xpath,heding,status)                
                
        counter = 0
        while counter < 5:
            try:                                                          
                err=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/div[1]/div[2]/div[1]/div[1]/div/div/div/div/div[2]/form/div[1]/div/div[2]/span/div/p'))).text

                wb1=load_workbook(filename=fil)
                sheet = wb1['VA EV']
                column_letter = 'H'  
                column_cells = sheet[column_letter]
                last_row = None
                for cell in reversed(column_cells):
                    if cell.value:
                        last_row = cell.row
                        break                            
                sheet['H' + str(int(last_row + 1))]=err
                wb1.save(fil)
                wb1.close()               
               
                try:
                    xpath= '/html/body/div[1]/div[1]/div[2]/div[1]/div[1]/div/div/div/div/div[2]/form/div[2]/div[1]/div/div[1]/fieldset/div/div[1]/label'
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()                    
                except Exception as e:
                    pass

                break
            except Exception as e:            
                try:                        
                    lst=[]

                    xpath= "/html/body/div[1]/div[1]/div[2]/div[1]/div[1]/div/div/div/div/div[2]/div/div[1]/div[1]/fieldset[1]"
                    ef=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text                    
                    if ef==' ' or ef=='':
                        ef='N/A'
                        lst.append(ef)   
                    else:
                        lst.append(ef) 
                    
                    lt=len(lst)
                    if lt==0:                           
                        lst.append('N/A') 
                    
                    lst1 = []
                    xpath= "/html/body/div[1]/div[1]/div[2]/div[1]/div[1]/div/div/div/div/div[2]/div/div[1]/div[1]/fieldset[4]"
                    bdt=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text                    
                    if bdt==' ' or bdt=='':
                        bdt='N/A'
                        lst1.append(bdt)   
                    else:
                        lst1.append(bdt) 

                    lt=len(lst1)
                    if lt==0:                           
                        lst.append('N/A')
                    else:
                        vr=lst1[0]
                        lst.append(vr)

                    lst2 = []
                    xpath= "/html/body/div[1]/div[1]/div[2]/div[1]/div[1]/div/div/div/div/div[2]/div/div[1]/div[2]/fieldset[3]"
                    icn=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text                    
                    if icn==' ' or icn=='':
                        icn='N/A'
                        lst2.append(icn)   
                    else:
                        lst2.append(icn) 

                    lt=len(lst2)
                    if lt==0:                           
                        lst.append('N/A')
                    else:
                        vr=lst2[0]
                        lst.append(vr)

                    lst3 = []
                    xpath= "/html/body/div[1]/div[1]/div[2]/div[1]/div[1]/div/div/div/div/div[2]/div/div[1]/div[2]/fieldset[4]"
                    dob=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text                    
                    if dob==' ' or dob=='':
                        dob='N/A'
                        lst3.append(dob)   
                    else:
                        lst3.append(dob) 

                    lt=len(lst3)
                    if lt==0:                           
                        lst.append('N/A')
                    else:
                        vr=lst3[0]
                        lst.append(vr)
                    
                    lst4 = []
                    xpath= "/html/body/div[1]/div[1]/div[2]/div[1]/div[1]/div/div/div/div/div[2]/div/div[1]/div[2]/fieldset[5]"
                    add=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text                    
                    if add==' ' or add=='':
                        add='N/A'
                        lst4.append(add)   
                    else:
                        lst4.append(add) 

                    lt=len(lst4)
                    if lt==0:                           
                        lst.append('N/A')
                    else:
                        vr=lst4[0]
                        lst.append(vr)                                                                            

                    wb1=load_workbook(fil)
                    sheet = wb1['VA EV']
                    column_letter = 'H'  
                    column_cells = sheet[column_letter]
                    last_row = None
                    for cell in reversed(column_cells):
                        if cell.value:
                            last_row = cell.row
                            break
                    start_column = 'C'
                    current_column_index = openpyxl.utils.column_index_from_string(start_column)
                    current_row = last_row + 1

                    for value in lst:
                        current_column = openpyxl.utils.get_column_letter(current_column_index)
                        sheet[current_column + str(current_row)] = value
                        current_column_index += 1                
                    sheet['H' + str(int(last_row + 1))]='Done'
                    wb1.save(filename=fil)
                    wb1.close() 
                    
                    xpath='/html/body/div[1]/div[1]/div[2]/div[1]/div[1]/div/div/div/div/div[2]/div/div[2]/div/button'
                    heding="Back"
                    status="Back Button Not Found"        
                    click(xpath,heding,status)                      

                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1             
        else:
            wb1=load_workbook(filename=fil)
            sheet = wb1['VA EV']
            column_letter = 'H'  
            column_cells = sheet[column_letter]
            last_row = None
            for cell in reversed(column_cells):
                if cell.value:
                    last_row = cell.row
                    break                            
            sheet['H' + str(int(last_row + 1))]='Error'   
            wb1.save(fil)
            wb1.close()                 

    driver.quit()
    messagebox.showinfo("Process Status", "Process Completed")
    sys.exit(0)    

def referral():        
    global element_1    
    global ans    
    global driver            
    global rows
    global rows1
    
    browse()
      
    fil=ans

    try:
        options = Options()            
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--ignore-ssl-errors')
        options.add_argument("--disable-popup-blocking")
        prefs ={"profile.password_manager_enabled": False}      
        options.add_experimental_option("prefs",prefs)  
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_experimental_option("useAutomationExtension", False)
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options)
        driver.maximize_window()
        driver.get('https://www.myvaccn.com/site/vaccn/main/public/login#/home')
    except Exception as e:                        
        try:
            options = Options()
            options.add_argument('--ignore-certificate-errors')
            options.add_argument('--ignore-ssl-errors')
            options.add_argument("--disable-popup-blocking")
            prefs ={"profile.password_manager_enabled": False}      
            options.add_experimental_option("prefs",prefs)  
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            driver_path = os.path.abspath('chromedriver.exe')
            driver = webdriver.Chrome(executable_path=driver_path,options=options)
            driver.maximize_window()
            driver.get('https://www.myvaccn.com/site/vaccn/main/public/login#/home')
        except Exception as e:           
            messagebox.showinfo("Driver Problem","Pls Check Your Chrome Driver Version")
            sys.exit(0)
    
    def text_box(xpath,heding,status,key):                
        counter = 0
        while counter < 5:
            try:   
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()                     
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:           
            messagebox.showinfo(heding, status)
            sys.exit(0)        
    
    def text_box_key(xpath,heding,status,key):                
        counter = 0
        while counter < 5:
            try:                   
                element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, xpath)))
                element.clear()                
                element.send_keys(key)                
                # element.send_keys(Keys.TAB)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:            
            messagebox.showinfo(heding, status)
            sys.exit(0)    

    def text_box_js(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:
                js_code = "document.querySelector('" + xpath + "').setAttribute('value', '" + key + "');"
                WebDriverWait(driver, 0).until(lambda driver: driver.execute_script(js_code))               
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo(heding, status)
            sys.exit(0)       

    def click(xpath,heding,status):
        counter = 0
        while counter < 5:
            try:             
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:            
            messagebox.showinfo(heding,status)
            sys.exit(0)                                

    def count(xpath,heding,status):
        global rows
        counter = 0
        while counter < 5:
            try:             
                rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:            
            messagebox.showinfo(heding,status)
            sys.exit(0)                         

    def count_1(xpath,heding,status):
        global rows1
        counter = 0
        while counter < 5:
            try:             
                rows1=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:            
            messagebox.showinfo(heding,status)
            sys.exit(0)        

    def Alert():
        counter = 0
        while counter < 15:
            try:             
                WebDriverWait(driver, 0).until (EC.alert_is_present())
                a=driver.switch_to.alert
                a.accept()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo('Alert','Alert Not Present')                                                                
    
    messagebox.showinfo('Waiting','Authentication Waiting')        
   
    file=pd.read_excel(fil,sheet_name='VA Referrals',header=0)
    
    for index, row in file.iterrows():                                              
        meb_id=row[0]
        ssn=row[1]        
        
        if not pd.isnull(meb_id):
            key = meb_id                  
            xpath= '/html/body/div[1]/div/div[2]/div/div[1]/div[1]/div/div/div/div/div[3]/div[2]/div[5]/div[1]/div/div[1]/fieldset/div/div[2]/label'
            heding="MVI ICN"
            status="MVI ICN Not Found"
            click(xpath,heding,status)

            xpath= "/html/body/div[1]/div/div[2]/div/div[1]/div[1]/div/div/div/div/div[3]/div[2]/div[5]/div[1]/div/div[2]/input[2]"
            heding="MVI ICN or SSN"
            status="MVI ICN or SSN Field Not Found"
            key=key.lstrip().rstrip()
            text_box(xpath,heding,status,key)      
            
        else:
            key = ssn           

            xpath= "/html/body/div[1]/div/div[2]/div/div[1]/div[1]/div/div/div/div/div[3]/div[2]/div[5]/div[1]/div/div[2]/input[1]"
            heding="MVI ICN or SSN"
            status="MVI ICN or SSN Field Not Found"
            key=key.lstrip().rstrip()
            text_box(xpath,heding,status,key)            
        
        xpath= "/html/body/div[1]/div/div[2]/div/div[1]/div[1]/div/div/div/div/div[3]/div[2]/div[6]/div[1]/fieldset/div/select/option"
        heding="Search by"        
        status="Search by Count Not Found"                    
        count(xpath,heding,status)

        j=1
        while j<rows+1:
            xpath= "/html/body/div[1]/div/div[2]/div/div[1]/div[1]/div/div/div/div/div[3]/div[2]/div[6]/div[1]/fieldset/div/select/option[{}]"
            ck=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
            if ck.lstrip().rstrip()=='All referrals':                            
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                            
                break           
            j=j+1

        xpath= '/html/body/div[1]/div/div[2]/div/div[1]/div[1]/div/div/div/div/div[3]/div[2]/div[7]/div/button'
        heding="Submit"
        status="Submit Button Not Found"
        click(xpath,heding,status)                
                
        counter = 0
        while counter < 5:
            try:                                                          
                err=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/div/div[2]/div/div[1]/div[1]/div/div/div/div/div[3]/div[2]/div[4]/div/div/div/div/span'))).text

                wb1=load_workbook(filename=fil)
                sheet = wb1['VA Referrals']
                column_letter = 'E'  
                column_cells = sheet[column_letter]
                last_row = None
                for cell in reversed(column_cells):
                    if cell.value:
                        last_row = cell.row
                        break
                sheet['E' + str(int(last_row + 1))]=err
                wb1.save(fil)
                wb1.close()                                           

                break
            except Exception as e:            
                try:                        
                    lst=[]                    
                    xpath= "/html/body/div[1]/div/div[2]/div/div[1]/div[1]/div/div/div/div/div[4]/table/tbody/tr/td[1]"
                    ref_num=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text                    
                    if ref_num==' ' or ref_num=='':
                        ref_num='N/A'
                        lst.append(ref_num)   
                    else:
                        lst.append(ref_num) 
                    
                    lt=len(lst)
                    if lt==0:                           
                        lst.append('N/A') 
                    
                    lst1 = []
                    xpath= "/html/body/div[1]/div/div[2]/div/div[1]/div[1]/div/div/div/div/div[4]/table/tbody/tr/td[2]"
                    dos=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text                    
                    if dos==' ' or dos=='':
                        dos='N/A'
                        lst1.append(dos)   
                    else:
                        lst1.append(dos) 

                    lt=len(lst1)
                    if lt==0:                           
                        lst.append('N/A')
                    else:
                        vr=lst1[0]
                        lst.append(vr)
                    
                    wb1=load_workbook(fil)
                    sheet = wb1['VA Referrals']
                    column_letter = 'E'  
                    column_cells = sheet[column_letter]
                    last_row = None
                    for cell in reversed(column_cells):
                        if cell.value:
                            last_row = cell.row
                            break
                    start_column = 'C'
                    current_column_index = openpyxl.utils.column_index_from_string(start_column)
                    current_row = last_row + 1

                    for value in lst:
                        current_column = openpyxl.utils.get_column_letter(current_column_index)
                        sheet[current_column + str(current_row)] = value
                        current_column_index += 1                
                    sheet['E' + str(int(last_row + 1))]='Done'
                    wb1.save(filename=fil)
                    wb1.close() 
                    
                    xpath='/html/body/div[1]/div/div[2]/div/div[1]/div[1]/div/div/div/div/div[5]/div/div[2]/button'
                    heding="New Search"
                    status="New Search Button Not Found"        
                    click(xpath,heding,status)                      

                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1             
        else:
            wb1=load_workbook(filename=fil)
            sheet = wb1['VA Referrals']
            column_letter = 'E'  
            column_cells = sheet[column_letter]
            last_row = None
            for cell in reversed(column_cells):
                if cell.value:
                    last_row = cell.row
                    break                            
            sheet['E' + str(int(last_row + 1))]='Error'   
            wb1.save(fil)
            wb1.close()                 

    driver.quit()
    messagebox.showinfo("Process Status", "Process Completed")
    sys.exit(0)    
  

if __name__=="__main__":        
    radio()
    