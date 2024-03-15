from tkinter import BOTH, LEFT, TOP, Button, Entry, Frame, Label, PhotoImage, StringVar, Tk,Radiobutton,StringVar,IntVar,filedialog
from idlelib.tooltip import Hovertip
import sys
import os
from operator import itemgetter
import time
from tkinter import messagebox
from os import listdir
from os.path import isfile, join
from datetime import datetime, date
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, StaleElementReferenceException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import openpyxl
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from urllib.parse import urlparse
import warnings
import numpy as np
from openpyxl.utils import get_column_letter,column_index_from_string
from datetime import datetime
from pytz import timezone
import requests
from zipfile import ZipFile
import subprocess

warnings.filterwarnings("ignore")

global rows
global xpath
global heding
global status
global key
global nme
global element_1  
global j

def browse():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
        image_path2 = os.path.join(sys._MEIPASS, 'Static', 'Mapping.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')
        image_path2 = os.path.join(os.getcwd(), 'Static', 'Mapping.png')

    root.title("Process File Picker")
    root.resizable(False,False)

    root.title("Process File Picker")
    root.resizable(False,False)
   
    w = 500
    h = 160
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Browse Process - File Picker",font=("Calibri",18,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
          
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=50)
    
    answer=StringVar()
    answer.set("")
       
    def browse_button():
        global filename
        answer.set("")  
        filename = filedialog.askopenfilename()  
        txt.delete(0, 'end') 
        txt.insert(0, filename)

    def Click_Done():
        global ans        
        
        ans=txt.get()        
        
        if ans=="":
           answer.set("File Path Fields Empty Is Not Allowed...")        
        else:      
            root.destroy()          
            return ans
        
    txt=Entry(Frame2,font=("Calibri",12,"bold","italic"),width=50,justify="left")
    txt.grid(row=0,column=0,padx=5,pady=10,sticky="E")    

    photo1 = PhotoImage(file=image_path2)
            
    btn1=Button(Frame2,text="Browse",command=browse_button,image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=0,column=1,padx=3,pady=0,sticky="W")     
                
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=3,y=80,width=490,height=25)
    
    title_3=Label(Frame3,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=1,padx=150,pady=0,sticky="E")

    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=3,y=105,width=500,height=50)
    
    photo2 = PhotoImage(file=image_path1)

    btn2=Button(Frame4,command=Click_Done,text="Done",image=photo2,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=3,column=0,padx=125,pady=0,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo3 = PhotoImage(file=image_path)   

    btn3=Button(Frame4,command=Close,text="Exit",image=photo3,borderwidth=0,bg="#2c3e50")
    btn3.grid(row=3,column=1,padx=15,pady=0,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn2,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn3,'Click to Exit Process',hover_delay=1000)
    myTip2 = Hovertip(btn1,'Click to Pick File',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def process():        
    global element_1    
    global ans1
    global ans2 
    global driver
    global j
    global dob
    global mid
    global frm_dt
    global to_dt
    global cpt

    # user_pass()
    browse()

    fil=ans               
    # user = ans1
    # password = ans2

    try:
        options = Options()            
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--ignore-ssl-errors')
        options.add_argument("--disable-popup-blocking")
        driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options)
        driver.maximize_window()
        driver.get('http://www.ngsmedicare.com/')
    except Exception as e:                        
        try:
            options = Options()
            options.add_argument('--ignore-certificate-errors')
            options.add_argument('--ignore-ssl-errors')
            options.add_argument("--disable-popup-blocking")
            # response = requests.get('https://chromedriver.storage.googleapis.com/LATEST_RELEASE')
            # latest_version = response.text.strip()
            # chrome_driver_url = f'https://chromedriver.storage.googleapis.com/{latest_version}/chromedriver_win32.zip'
            # response = requests.get(chrome_driver_url)
            # with open('chromedriver_win32.zip', 'wb') as f:
            #     f.write(response.content)
            # with ZipFile('chromedriver_win32.zip', 'r') as zip_ref:
            #     zip_ref.extractall('.')
            driver_path = os.path.abspath('chromedriver.exe')
            driver = webdriver.Chrome(executable_path=driver_path,options=options)
            driver.maximize_window()
            driver.get('http://www.ngsmedicare.com/')
        except Exception as e:           
            messagebox.showinfo("Driver Version Problem","Pls Check Your Chrome Driver Version")
            sys.exit(0)
        
    def text_box(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:   
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding, status)
            sys.exit(0)        
    
    def text_box_key(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:                   
                input_field =WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath)))
                input_field.send_keys(Keys.BACKSPACE * len(input_field))                
                input_field.send_keys(key)               
                # element.send_keys(Keys.TAB)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding, status)
            sys.exit(0)    

    def text_box_js(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:
                js_code = "document.querySelector('" + xpath + "').setAttribute('value', '" + key + "');"
                WebDriverWait(driver, 0).until(lambda driver: driver.execute_script(js_code))               
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo(heding, status)
            sys.exit(0)       

    def click(xpath,heding,status):
        counter = 0
        while counter < 15:
            try:             
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding,status)
            sys.exit(0)                                

    def count(xpath,heding,status):
        global rows
        counter = 0
        while counter < 15:
            try:             
                rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding,status)
            sys.exit(0)                         

    def Alert():
        counter = 0
        while counter < 15:
            try:             
                WebDriverWait(driver, 0).until (EC.alert_is_present())
                a=driver.switch_to.alert
                a.accept()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo('Alert','Alert Not Present')                                                                
                
    messagebox.showinfo('Waiting','Authentication Waiting')        
            
    file=pd.read_excel(fil,sheet_name='MCR MBI Lookup BOT',header=0)
    
    for index, row in file.iterrows():                                             
        fnm = row[2]
        lnm = row[3]
        pri_sur=row[4]
        ssn=row[5]
        dob = row[6]                         

        try:
            xpath= "/html/body/div[3]/div/div/div/div/div[3]/div/div/div/form/div[1]/div/div[1]/div[1]/span/input"
            key=str(fnm)
            WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()
            WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)                   
        except Exception as e:
            xpath= "/html/body/div[4]/div/div/div/div/div[3]/div/div/div/form/div[1]/div/div[1]/div[1]/span/input"
            key=str(fnm)
            WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()
            WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)                        

        try:
            xpath= "/html/body/div[3]/div/div/div/div/div[3]/div/div/div/form/div[1]/div/div[2]/div[1]/span/input"
            key=str(lnm)
            WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()
            WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)                  
        except Exception as e:
            xpath= "/html/body/div[4]/div/div/div/div/div[3]/div/div/div/form/div[1]/div/div[2]/div[1]/span/input"
            key=str(lnm)
            WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()
            WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)                       

        if pd.isnull(pri_sur):
            pri_sur=''

        try:
            xpath= "/html/body/div[3]/div/div/div/div/div[3]/div/div/div/form/div[1]/div/div[3]/div/span/input"
            key=str(pri_sur)
            WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()
            WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)            
        except Exception as e:
            xpath= "/html/body/div[4]/div/div/div/div/div[3]/div/div/div/form/div[1]/div/div[3]/div/span/input"
            key=str(pri_sur)
            WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()
            WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)    

        try:
            xpath= "/html/body/div[3]/div/div/div/div/div[3]/div/div/div/form/div[1]/div/div[1]/div[2]/span/input"
            key=str(ssn)
            WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()
            WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)            
        except Exception as e:
            xpath= "/html/body/div[4]/div/div/div/div/div[3]/div/div/div/form/div[1]/div/div[1]/div[2]/span/input"
            key=str(ssn)
            WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()
            WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)                     

        date_object_1 = datetime.strptime(str(dob), "%Y-%m-%d %H:%M:%S")
        dob3 = date_object_1.strftime("%m/%d/%Y")            

        try:
            xpath= "/html/body/div[3]/div/div/div/div/div[3]/div/div/div/form/div[1]/div/div[2]/div[2]/span/input"
            key=dob3
            WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()
            WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
        except Exception as e:
            xpath= "/html/body/div[4]/div/div/div/div/div[3]/div/div/div/form/div[1]/div/div[2]/div[2]/span/input"
            key=dob3
            WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()
            WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
            
        # try:                
        #     xpath= "/html/body/div[3]/div/div/div/div/div[3]/div/div/div/form/div[1]/div/div[2]/div[2]/span/input"
        #     input_field = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, xpath)))
        # except Exception as e:
        #     xpath= "/html/body/div[4]/div/div/div/div/div[3]/div/div/div/form/div[1]/div/div[2]/div[2]/span/input"
        #     input_field = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, xpath)))

        # script = f"arguments[0].setAttribute('value', '{dob3}');"
        # driver.execute_script(script, input_field)

        try:
            xpath= "/html/body/div[3]/div/div/div/div/div[3]/div/div/div/form/div[2]/div[2]/div/div/button"            
            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()           
        except Exception as e:
            xpath= "/html/body/div[4]/div/div/div/div/div[3]/div/div/div/form/div[2]/div[2]/div/div/button"            
            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click() 
                
        counter = 0
        while counter < 15:
            try:                                                                                                             
                st = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[3]/div/div/div/div/div[2]/div/div/div/div[1]/span'))).text                     
                wb1=load_workbook(filename=fil)
                sheet = wb1['MCR MBI Lookup BOT']
                column_letter = 'H'  
                column_cells = sheet[column_letter]
                last_row = None
                for cell in reversed(column_cells):
                    if cell.value:
                        last_row = cell.row
                        break                            
                sheet['H' + str(int(last_row + 1))]=st
                wb1.save(fil)
                wb1.close()  

                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:            
            wb1=load_workbook(filename=fil)
            sheet = wb1['MCR MBI Lookup BOT']
            column_letter = 'H'  
            column_cells = sheet[column_letter]
            last_row = None
            for cell in reversed(column_cells):
                if cell.value:
                    last_row = cell.row
                    break                            
            sheet['H' + str(int(last_row + 1))]='Error'
            wb1.save(fil)
            wb1.close()  
                       
        
        counter = 0
        while counter < 15:
            try:                   
                xpath= "/html/body/div[3]/div/div/div/div/div[2]/div/div/div/div[5]/a/span"            
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click() 
                break          
            except Exception as e:
                try:
                    xpath= "/html/body/div[3]/div/div/div/div/div[2]/div/div/div/div[3]/a/span"            
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click() 
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
        else: 
            messagebox.showinfo('New MBI Lookup','New MBI Lookup - Not Found')
            sys.exit(0)     

    driver.quit()
    messagebox.showinfo("Process Status", "Process Completed")
    sys.exit(0)    

if __name__=="__main__":        
    process()