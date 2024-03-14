from tkinter import BOTH, LEFT, TOP, Button, Entry, Frame, Label, PhotoImage, StringVar, Tk
from idlelib.tooltip import Hovertip
import sys
import os
from operator import itemgetter
import time
from tkinter import messagebox
from os import listdir
from os.path import isfile, join
from datetime import datetime, date
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, StaleElementReferenceException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import openpyxl
import pandas as pd
from openpyxl import load_workbook
from urllib.parse import urlparse
import warnings
import numpy as np
import requests
from zipfile import ZipFile
from openpyxl.utils import get_column_letter,column_index_from_string

warnings.filterwarnings("ignore")

global rows
global xpath
global heding
global status
global key
global nme
element_1 = None


def path_sam():
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')

    root.title("Sharepoint - User Login And File Details")
    root.resizable(False,False)

    w = 700
    h = 250
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)

    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="User Name, Password And File Details...",font=("Calibri",20,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack() 
 
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=698,height=150)
    
    title1=Label(Frame2,text="File Path :",font=("Calibri",11,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title1.grid(row=0,column=0,padx=5,pady=5,sticky="W")
    
    txt=Entry(Frame2,font=("Calibri",11,"bold","italic"),width=81,justify="left")
    txt.grid(row=0,column=1,padx=30,pady=5,sticky="E")

    title2=Label(Frame2,text="File Name :",font=("Calibri",11,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title2.grid(row=1,column=0,padx=5,pady=5,sticky="W")
    
    txt1=Entry(Frame2,font=("Calibri",11,"bold","italic"),width=81,justify="left")
    txt1.grid(row=1,column=1,padx=30,pady=5,sticky="E")
    
    title3=Label(Frame2,text="User Name :",font=("Calibri",11,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title3.grid(row=2,column=0,padx=5,pady=5,sticky="W")
    
    txt2=Entry(Frame2,font=("Calibri",11,"bold","italic"),width=50,justify="left")
    txt2.grid(row=2,column=1,padx=30,pady=5,sticky="W")
    
    title4=Label(Frame2,text="Password :",font=("Calibri",11,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title4.grid(row=3,column=0,padx=5,pady=5,sticky="W")
    
    txt3=Entry(Frame2,font=("Calibri",11,"bold","italic"),width=50,justify="left")
    txt3.grid(row=3,column=1,padx=30,pady=5,sticky="W")

    answer=StringVar()
    answer.set("")

    def Click_Done():
        global ans
        global ans1
        global ans2
        global ans3

        ans=txt.get()
        ans1=txt1.get()
        ans2=txt2.get()
        ans3=txt3.get()

        if ans=="":
           answer.set("File Path Fields Empty Is Not Allowed...")
        elif ans1=="":
            answer.set("File Name Field Empty Is Not Allowed...")
        elif ans2=="":
            answer.set("User Name Field Empty Is Not Allowed...")
        elif ans3=="":
            answer.set("Password Field Empty Is Not Allowed...")
        else:    
            root.destroy()
            return ans,ans1,ans2,ans3
                
    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=0,y=175,width=698,height=20)
    
    title_3=Label(Frame4,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=2,padx=240,pady=0,sticky="E")
    
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=0,y=195,width=698,height=200)
       
    photo1 = PhotoImage(file=image_path1)
    
    btn1=Button(Frame3,command=Click_Done,text="Run",image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=2,column=0,padx=180,pady=0,sticky="W")

    def Close():
        sys.exit(0)

    photo = PhotoImage(file=image_path)    

    btn2=Button(Frame3,command=Close,text="Close",image=photo,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=2,column=1,padx=60,pady=0,sticky="E")

    def disable_event():
        pass

    txt.focus_set()

    myTip = Hovertip(btn1,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn2,'Click to Exit Process',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)
            
    root.mainloop()

def ping_id():
    global element_1
    root = Tk()
    
    if getattr(sys, 'frozen', False):      
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')

    root.title("Question & Answer Details...")
    root.resizable(False,False)

    w = 900
    h = 200
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)

    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Question & Answer Screen...",font=("Calibri",20,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()

    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=895,height=80)

    title1=Label(Frame2,text="Question :",font=("Calibri",14,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title1.grid(row=0,column=0,padx=5,pady=5,sticky="W")
    
    title2=Label(Frame2,text=element_1,font=("Calibri",12,"bold","italic"),bg="#2c3e50",fg="white",justify="center",width=100)
    title2.grid(row=0,column=1,padx=5,pady=5,sticky="W")
    
    title3=Label(Frame2,text="Answer :",font=("Calibri",14,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title3.grid(row=1,column=0,padx=5,pady=5,sticky="W")
    
    txt=Entry(Frame2,font=("Calibri",12,"bold","italic"),width=100,justify="center")
    txt.grid(row=1,column=1,padx=5,pady=5,sticky="W")

    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=0,y=115,width=895,height=20)

    answer=StringVar()
    answer.set("")

    def Click_Done():        
        global ans5        
        ans5=txt.get()

        if ans5=="" :
           answer.set("Answer Field Empty Is Not Allowed...")        
        else:    
            root.destroy()
            return ans5    

    title_4=Label(Frame3,text=answer.get(),textvariable=answer,font=("Calibri",11,"bold","italic"),bg="#2c3e50",fg="Red",justify="center",width=68)
    title_4.grid(row=0,column=1,columnspan=1,padx=180,pady=0,sticky="E")
    
    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=0,y=140,width=900,height=100)

    photo1 = PhotoImage(file=image_path1)

    btn1=Button(Frame4,command=Click_Done,text="Run",image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=1,column=0,padx=250,pady=0,sticky="W")

    def Close():
        sys.exit(0)

    photo = PhotoImage(file=image_path)

    btn2=Button(Frame4,command=Close,text="Close",image=photo,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=1,column=1,padx=70,pady=0,sticky="E")

    def disable_event():
        pass

    txt.focus_set()

    myTip = Hovertip(btn1,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn2,'Click to Exit Process',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)    

    root.mainloop()

def process():
        global element_1
        path_sam()
        
        fil_path =ans
        fil_name=ans1
        user_name=ans2
        password=ans3
        
        try:
            options = Options()            
            options.add_argument('--ignore-certificate-errors')
            options.add_argument('--ignore-ssl-errors')
            options.add_argument("--disable-popup-blocking")
            driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options)
            driver.maximize_window()
            driver.get('https://promise.dpw.state.pa.us')
        except Exception as e:                        
            try:
                options = Options()            
                options.add_argument('--ignore-certificate-errors')
                options.add_argument('--ignore-ssl-errors')
                options.add_argument("--disable-popup-blocking")
                # response = requests.get('https://chromedriver.storage.googleapis.com/LATEST_RELEASE')
                # latest_version = response.text.strip()
                # chrome_driver_url = f'https://chromedriver.storage.googleapis.com/{latest_version}/chromedriver_win32.zip'
                # response = requests.get(chrome_driver_url)
                # with open('chromedriver_win32.zip', 'wb') as f:
                #     f.write(response.content)
                # with ZipFile('chromedriver_win32.zip', 'r') as zip_ref:
                #     zip_ref.extractall('.')
                driver_path = os.path.abspath('chromedriver.exe')
                driver = webdriver.Chrome(executable_path=driver_path,options=options)
                driver.maximize_window()
                driver.get('https://promise.dpw.state.pa.us')
            except Exception as e:
                messagebox.showinfo("Internet Problem","Pls Check Your Internet Connection")
                sys.exit(0)        
               
        def text_box(xpath,heding,status,key):                
            counter = 0
            while counter < 60:
                try:                          
                    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding, status)
                sys.exit(0)        
                
        def click(xpath,heding,status):
            counter = 0
            while counter < 60:
                try:             
                    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding,status)
                sys.exit(0)                                

        def count(xpath,heding,status):
            global rows
            counter = 0
            while counter < 60:
                try:             
                    rows=len(WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))                    
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding,status)
                sys.exit(0)          
             
        def Alert():
            counter = 0
            while counter < 60:
                try:             
                    WebDriverWait(driver, 10).until (EC.alert_is_present())
                    a=driver.switch_to.alert
                    a.accept()
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo('Alert','Alert Not Present')
                                                          
        xpath= "/html/body/form/div[3]/div/div[2]/div[3]/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td[1]/div[1]/div/div/div/div/div[2]/div[1]/div/div[2]/input"
        heding="User ID"
        status="User ID Field Not Found"
        key= user_name
        text_box(xpath,heding,status,key)
       
        xpath= "/html/body/form/div[3]/div/div[2]/div[3]/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td[1]/div[1]/div/div/div/div/div[2]/div[2]/input"
        heding="Login Button"
        status="Login Button Not Found"
        click(xpath,heding,status)
        
        element_1 = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//label[@id="dnn_ctr661_ChallengeQuestion_ChallengeQuestionHCPTextBox_ControlLabel"]'))).text

        if element_1 !="":
            ping_id()
        else:
            messagebox.showinfo('Question & Answer','Question & Answer Text Not Found')
            sys.exit(0)   
               
        xpath= '//input[@id="dnn_ctr661_ChallengeQuestion_ChallengeResponseHCPTextBox_Control"]'
        heding="Answer Text Field"
        status="Answer Text Field Not Found"
        key=ans5
        text_box(xpath,heding,status,key)

        xpath= '//input[@id="dnn_ctr661_ChallengeQuestion_ContinueHCPButton"]'
        heding="Continue Button"
        status="Continue Button Not Found"
        click(xpath,heding,status)
        
        while True:
            xpath='/html/body/form/div[3]/div/div[2]/div[3]/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td[2]/div/div/div/div/div[2]/div[1]/span[2]'         
            try:
                element_3 = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, xpath))).text
                if element_3=='Your answer was incorrect. Please try again.':         
                    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//input[@id="dnn_ctr661_ChallengeQuestion_ChallengeResponseHCPTextBox_Control"]'))).clear()           
                    messagebox.showinfo('Question & Answer', 'Your Answer Was Incorrect. Please Try Again.')
                    element_1 = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//label[@id="dnn_ctr661_ChallengeQuestion_ChallengeQuestionHCPTextBox_ControlLabel"]'))).text
                    ping_id()    
                    xpath= '//input[@id="dnn_ctr661_ChallengeQuestion_ChallengeResponseHCPTextBox_Control"]'
                    heding="Answer Text Field"
                    status="Answer Text Field Not Found"
                    key=ans5
                    text_box(xpath,heding,status,key)

                    xpath= '//input[@id="dnn_ctr661_ChallengeQuestion_ContinueHCPButton"]'
                    heding="Continue Button"
                    status="Continue Button Not Found"
                    click(xpath,heding,status)
                else:
                    break    
            except Exception as e:                
                    break
        
        xpath= '//input[@id="dnn_ctr662_SiteTokenPassword_PasswordHCPTextBox_Control"]'
        heding="Password Text Field"
        status="Password Text Field Not Found"
        key=password
        text_box(xpath,heding,status,key)
                
        xpath= '//input[@id="dnn_ctr662_SiteTokenPassword_SignInHCPButton"]'
        heding="Password SigIn Button"
        status="Password SigIn Button Not Found"
        click(xpath,heding,status)
        
        xpath= '/html/body/form/div[3]/div/div[2]/div[1]/div/ul/div/li[5]/nobr/a'
        heding="Eligibility Button"
        status="Eligibility Button Not Found"
        click(xpath,heding,status)

        Alert()


        file=pd.read_excel(fil_path + '/' + fil_name + '.xlsx',sheet_name='Data',header=0)
        
        file['Subscriber Recipient ID'] = pd.to_numeric(file['Subscriber Recipient ID'], errors='coerce').astype('float').astype('Int64')
        
        file['Subscriber Recipient ID'] = file['Subscriber Recipient ID'].astype(object).where(file['Subscriber Recipient ID'].notnull(), np.nan)                

        for index, row in file.iterrows():                          
            srid = row[1]              
            dob = row[2]     
            ssn = row[3]
            dob1=row[4]
            sfn=row[5]
            sln=row[6]
            dob2=row[7]
            fdt=row[8]

            if not pd.isna(sfn) and pd.isna(srid) and pd.isna(ssn):
                xpath= '//input[@id="dnn_ctr1732_Eligibility_txtFirstName"]'
                heding="First Name"
                status="First Name Field Not Found"
                key=sfn
                text_box(xpath,heding,status,key)

                xpath= '//input[@id="dnn_ctr1732_Eligibility_txtLastName"]'
                heding="Last Name"
                status="Last Name Field Not Found"
                key=sln
                text_box(xpath,heding,status,key)                                

                date_object = datetime.strptime(str(dob2), "%Y-%m-%d %H:%M:%S")
                dob2 = date_object.strftime("%m/%d/%Y")
               
                WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//input[@id="dnn_ctr1732_Eligibility_txtDob2"]'))).clear() 
            
                xpath= '//input[@id="dnn_ctr1732_Eligibility_txtDob2"]'
                heding="Date of Birth"
                status="Date of Birth Field Not Found"
                key=dob2
                text_box(xpath,heding,status,key)

                date_object = datetime.strptime(str(fdt), "%Y-%m-%d %H:%M:%S")
                fdt = date_object.strftime("%m/%d/%Y")
               
                WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//input[@id="dnn_ctr1732_Eligibility_txtDosFrom"]'))).clear() 
                xpath= '//input[@id="dnn_ctr1732_Eligibility_txtDosFrom"]'
                heding="From DOS"
                status="From DOS Field Not Found"
                key=fdt
                text_box(xpath,heding,status,key)  

                WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//input[@id="dnn_ctr1732_Eligibility_txtDosTo"]'))).clear() 
                xpath= '//input[@id="dnn_ctr1732_Eligibility_txtDosTo"]'
                heding="To DOS"
                status="To DOS Field Not Found"
                key=fdt
                text_box(xpath,heding,status,key)  

                xpath= '//input[@id="dnn_ctr1732_Eligibility_btnSearch"]'
                heding="Search Button"
                status="Search Button Not Found"
                click(xpath,heding,status)

                while True:           
                    try:
                        element_4 = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, '//span[@id="dnn_ctr1732_Eligibility_lblResultsError"]'))).text                                               
                        if element_4=='Subscriber/Insured Not Found' or element_4=='Invalid/Missing Subscriber/Insured ID':   
                            wb1=load_workbook(filename=fil_path + '/' + fil_name + '.xlsx')
                            sheet = wb1['Data']
                            column_letter = 'K'  
                            column_cells = sheet[column_letter]
                            last_row = None
                            for cell in reversed(column_cells):
                                if cell.value:
                                    last_row = cell.row
                                    break                            
                            sheet['K' + str(int(last_row + 1))]='Error'        
                            wb1.save(fil_path + '/' + fil_name + '.xlsx')
                            wb1.close()       
                                
                            xpath= '/html/body/form/div[3]/div/div[2]/div[1]/div/ul/div/li[5]/nobr/a'
                            heding="Eligibility Button"
                            status="Eligibility Button Not Found"
                            click(xpath,heding,status)      
                            break                             
                    except Exception as e:                
                            xpath='//table[@id="dnn_ctr1732_Eligibility_gvRecipient"]/tbody/tr'
                            heding="Recipient Table"
                            status="Recipient Table Not Found"
                            count(xpath,heding,status)                           
                            
                            lst = []                            
                            lstt1 = []
                            j=1
                            while j<rows+1:
                                xpath= '//table[@id="dnn_ctr1732_Eligibility_gvRecipient"]/tbody/tr[{}]/td[1]'
                                cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                if cnm=='Name:':
                                    xpath='//table[@id="dnn_ctr1732_Eligibility_gvRecipient"]/tbody/tr[{}]/td[2]'                                      
                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                    if nm==' ':
                                        nm='N/A' 
                                        lstt1.append(nm)
                                    else:
                                        lstt1.append(nm) 
                                    break
                                j=j+1
                            
                            lt=len(lstt1) 
                            if lt==0:                                    
                                lst.append('N/A')
                            else:
                                vr=lstt1[0]
                                lst.append(vr)

                            lstt2 = []
                            j=1
                            while j<rows+1:
                                xpath= '//table[@id="dnn_ctr1732_Eligibility_gvRecipient"]/tbody/tr[{}]/td[1]'
                                cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                if cnm=='Recipient ID:':
                                    xpath='//table[@id="dnn_ctr1732_Eligibility_gvRecipient"]/tbody/tr[{}]/td[2]'                                      
                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                    if nm==' ':
                                        nm='N/A' 
                                        lstt2.append(nm)
                                    else:
                                        lstt2.append(nm) 
                                    break
                                j=j+1
                            
                            lt=len(lstt2) 
                            if lt==0:                                    
                                lst.append('N/A')
                            else:
                                vr=lstt2[0]
                                lst.append(vr)

                            lstt3 = []
                            j=1
                            while j<rows+1:
                                xpath= '//table[@id="dnn_ctr1732_Eligibility_gvRecipient"]/tbody/tr[{}]/td[1]'
                                cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                if cnm=='Date of Birth:':
                                    xpath='//table[@id="dnn_ctr1732_Eligibility_gvRecipient"]/tbody/tr[{}]/td[2]'                                      
                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                    if nm==' ':
                                        nm='N/A' 
                                        lstt3.append(nm)
                                    else:
                                        lstt3.append(nm) 
                                    break
                                j=j+1
                            
                            lt=len(lstt3) 
                            if lt==0:                                    
                                lst.append('N/A')
                            else:
                                vr=lstt3[0]
                                lst.append(vr)

                            lstt4 = []
                            j=1
                            while j<rows+1:
                                xpath= '//table[@id="dnn_ctr1732_Eligibility_gvRecipient"]/tbody/tr[{}]/td[1]'
                                cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                if cnm=='Gender:':
                                    xpath='//table[@id="dnn_ctr1732_Eligibility_gvRecipient"]/tbody/tr[{}]/td[2]'                                      
                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                    if nm==' ':
                                        nm='N/A' 
                                        lstt4.append(nm)
                                    else:
                                        lstt4.append(nm) 
                                    break
                                j=j+1
                            
                            lt=len(lstt4) 
                            if lt==0:                                    
                                lst.append('N/A')
                            else:
                                vr=lstt4[0]
                                lst.append(vr)
                               
                            wb1=load_workbook(filename=fil_path + '/' + fil_name + '.xlsx')
                            sheet = wb1['Data']
                            column_letter = 'K'  
                            column_cells = sheet[column_letter]
                            last_row = None
                            for cell in reversed(column_cells):
                                if cell.value:
                                    last_row = cell.row
                                    break              
                            start_column = 'L'
                            current_column_index = openpyxl.utils.column_index_from_string(start_column)
                            current_row = last_row + 1

                            for value in lst:
                                current_column = openpyxl.utils.get_column_letter(current_column_index)
                                sheet[current_column + str(current_row)] = value
                                current_column_index += 1
                            
                            wb1.save(filename=fil_path + '/' + fil_name + '.xlsx')
                            wb1.close()
                            
                            try:
                                xpath='//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr'                             
                                counter = 0
                                while counter < 10:
                                    try:
                                        rows1=len(WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))                    
                                        break
                                    except Exception as e:
                                        time.sleep(1)
                                        counter += 1

                                lst1 = []                                                                    
                                lstt1 = []
                                j=1
                                while j<rows+1:
                                    xpath= '//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[1]'
                                    cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                    if cnm=='Status:':
                                        xpath='//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[2]'                                      
                                        nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                        if nm==' ':
                                            nm='N/A' 
                                            lstt1.append(nm)
                                        else:
                                            lstt1.append(nm) 
                                        break
                                    j=j+1
                                
                                lt=len(lstt1) 
                                if lt==0:                                    
                                    lst1.append('N/A')
                                else:
                                    vr=lstt1[0]
                                    lst1.append(vr)

                                lstt2 = []
                                j=1
                                while j<rows+1:
                                    xpath= '//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[1]'
                                    cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                    if cnm=='Service Type:':
                                        xpath='//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[2]'                                      
                                        nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                        if nm==' ':
                                            nm='N/A' 
                                            lstt2.append(nm)
                                        else:
                                            lstt2.append(nm) 
                                        break
                                    j=j+1

                                lt=len(lstt2) 
                                if lt==0:                                    
                                    lst1.append('N/A')
                                else:
                                    vr=lstt2[0]
                                    lst1.append(vr)

                                lstt3 = []
                                j=1
                                while j<rows+1:
                                    xpath= '//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[1]'
                                    cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                    if cnm=='Insurance Type:':
                                        xpath='//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[2]'                                      
                                        nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                        if nm==' ':
                                            nm='N/A' 
                                            lstt3.append(nm)
                                        else:
                                            lstt3.append(nm) 
                                        break
                                    j=j+1
                                
                                lt=len(lstt3) 
                                if lt==0:                                    
                                    lst1.append('N/A')
                                else:
                                    vr=lstt3[0]
                                    lst1.append(vr)

                                lstt4 = []
                                j=1
                                while j<rows+1:
                                    xpath= '//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[1]'
                                    cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                    if cnm=='Coverage Description:':
                                        xpath='//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[2]'                                      
                                        nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                        if nm==' ':
                                            nm='N/A' 
                                            lstt4.append(nm)
                                        else:
                                            lstt4.append(nm) 
                                        break
                                    j=j+1

                                lt=len(lstt4) 
                                if lt==0:                                    
                                    lst1.append('N/A')
                                else:
                                    vr=lstt4[0]
                                    lst1.append(vr)

                                lstt5 = []
                                j=1
                                while j<rows+1:
                                    xpath= '//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[1]'
                                    cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                    if cnm=='Plan':
                                        xpath='//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[2]'                                      
                                        nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                        if nm==' ':
                                            nm='N/A' 
                                            lstt5.append(nm)
                                        else:
                                            lstt5.append(nm) 
                                        break
                                    j=j+1

                                lt=len(lstt5) 
                                if lt==0:                                    
                                    lst1.append('N/A')                                                                              
                                else:
                                    vr=lstt5[0]
                                    lst1.append(vr)
                                     
                                wb1=load_workbook(filename=fil_path + '/' + fil_name + '.xlsx')
                                sheet = wb1['Data']
                                column_letter = 'K'  
                                column_cells = sheet[column_letter]
                                last_row = None
                                for cell in reversed(column_cells):
                                    if cell.value:
                                        last_row = cell.row
                                        break              
                                start_column = 'P'
                                current_column_index = openpyxl.utils.column_index_from_string(start_column)
                                current_row = last_row + 1

                                for value in lst1:
                                    current_column = openpyxl.utils.get_column_letter(current_column_index)
                                    sheet[current_column + str(current_row)] = value
                                    current_column_index += 1                                
                                
                                wb1.save(filename=fil_path + '/' + fil_name + '.xlsx')
                                wb1.close()
                            
                            except Exception as e:   
                                    pass

                            xpath='//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr'
                            heding="Eligibility Summary Table"
                            status="Eligibility Summary Table Not Found"
                            count(xpath,heding,status)      
                            
                            lst2 = []
                            i=1
                            while i<rows+1:        
                                if i > 1:                    
                                    cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,'//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr[{}]/td[1]'.format(i)))).text
                                    if cnm=='Co-Insurance':
                                        xpath='//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr[{}]/td[1]'  
                                        nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(i)))).text
                                        if nm==' ':
                                            nm='N/A' 
                                            lst2.append(nm)
                                        else:
                                            lst2.append(nm)
                                        xpath='//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr[{}]/td[2]'  
                                        nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(i)))).text
                                        if nm==' ':
                                            nm='N/A' 
                                            lst2.append(nm)
                                        else:
                                            lst2.append(nm)
                                        xpath='//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr[{}]/td[3]'  
                                        nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(i)))).text
                                        if nm==' ':
                                            nm='N/A' 
                                            lst2.append(nm)
                                        else:
                                            lst2.append(nm)
                                        xpath='//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr[{}]/td[4]'  
                                        nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(i)))).text
                                        if nm==' ':
                                            nm='N/A' 
                                            lst2.append(nm)
                                        else:
                                            lst2.append(nm)                                                                                
                                        break
                                    else:                                    
                                        xpath='//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr[{}]/td[1]'  
                                        nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(i)))).text
                                        if nm==' ':
                                            nm='N/A' 
                                            lst2.append(nm)
                                        else:
                                            lst2.append(nm)
                                        xpath='//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr[{}]/td[2]'  
                                        nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(i)))).text
                                        if nm==' ':
                                            nm='N/A' 
                                            lst2.append(nm)
                                        else:
                                            lst2.append(nm)
                                        xpath='//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr[{}]/td[3]'  
                                        nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(i)))).text
                                        if nm==' ':
                                            nm='N/A' 
                                            lst2.append(nm)
                                        else:
                                            lst2.append(nm)
                                        xpath='//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr[{}]/td[4]'  
                                        nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(i)))).text
                                        if nm==' ':
                                            nm='N/A' 
                                            lst2.append(nm)
                                        else:
                                            lst2.append(nm)                                                                
                                i=i+1
                            
                            wb1=load_workbook(filename=fil_path + '/' + fil_name + '.xlsx')
                            sheet = wb1['Data']
                            column_letter = 'K'  
                            column_cells = sheet[column_letter]
                            last_row = None
                            for cell in reversed(column_cells):
                                if cell.value:
                                    last_row = cell.row
                                    break              
                            start_column = 'U'
                            current_column_index = openpyxl.utils.column_index_from_string(start_column)
                            current_row = last_row + 1

                            for value in lst2:
                                current_column = openpyxl.utils.get_column_letter(current_column_index)
                                sheet[current_column + str(current_row)] = value
                                current_column_index += 1                                
                            
                            sheet['K' + str(int(last_row + 1))]='Done'

                            wb1.save(filename=fil_path + '/' + fil_name + '.xlsx')
                            wb1.close()                                                

                            xpath= '/html/body/form/div[3]/div/div[2]/div[1]/div/ul/div/li[5]/nobr/a'
                            heding="Eligibility Button"
                            status="Eligibility Button Not Found"
                            click(xpath,heding,status)   
                            break                         
            elif pd.isna(sfn) and pd.isna(srid) and not pd.isna(ssn):     
                xpath= '//input[@id="dnn_ctr1732_Eligibility_txtssn"]'
                heding="SSN ID"
                status="SSN ID Field Not Found"
                key=ssn
                                                   
                counter = 0
                while counter < 60:
                    try:                          
                        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)   
                        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//input[@id="dnn_ctr1732_Eligibility_txtDob1"]'))).click()                     
                        break
                    except Exception as e:
                        time.sleep(1)
                        counter += 1
                else:
                    messagebox.showinfo(heding, status)
                    sys.exit(0)   

                while True:
                    try:    
                        element_5 = WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH, '//span[@id="dnn_ctr1732_Eligibility_regExSSN"]'))).text                    
                        if element_5=='SSN needs to be 9 numeric characters.':
                            wb1=load_workbook(filename=fil_path + '/' + fil_name + '.xlsx')
                            sheet = wb1['Data']
                            column_letter = 'K'  
                            column_cells = sheet[column_letter]
                            last_row = None
                            for cell in reversed(column_cells):
                                if cell.value:
                                    last_row = cell.row
                                    break                            
                            sheet['K' + str(int(last_row + 1))]='Invalid Recipient ID entered'        
                            wb1.save(fil_path + '/' + fil_name + '.xlsx')
                            wb1.close() 

                            xpath= '/html/body/form/div[3]/div/div[2]/div[1]/div/ul/div/li[5]/nobr/a'
                            heding="Eligibility Button"
                            status="Eligibility Button Not Found"
                            click(xpath,heding,status)  

                            break
                    except Exception as e:  

                            date_object = datetime.strptime(str(dob1), "%Y-%m-%d %H:%M:%S")
                            dob2 = date_object.strftime("%m/%d/%Y")
                        
                            WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//input[@id="dnn_ctr1732_Eligibility_txtDob1"]'))).clear() 
                        
                            xpath= '//input[@id="dnn_ctr1732_Eligibility_txtDob1"]'
                            heding="Date of Birth"
                            status="Date of Birth Field Not Found"
                            key=dob2
                            text_box(xpath,heding,status,key)

                            date_object = datetime.strptime(str(fdt), "%Y-%m-%d %H:%M:%S")
                            fdt = date_object.strftime("%m/%d/%Y")
                        
                            WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//input[@id="dnn_ctr1732_Eligibility_txtDosFrom"]'))).clear() 
                            xpath= '//input[@id="dnn_ctr1732_Eligibility_txtDosFrom"]'
                            heding="From DOS"
                            status="From DOS Field Not Found"
                            key=fdt
                            text_box(xpath,heding,status,key)  

                            WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//input[@id="dnn_ctr1732_Eligibility_txtDosTo"]'))).clear() 
                            xpath= '//input[@id="dnn_ctr1732_Eligibility_txtDosTo"]'
                            heding="To DOS"
                            status="To DOS Field Not Found"
                            key=fdt
                            text_box(xpath,heding,status,key)  

                            xpath= '//input[@id="dnn_ctr1732_Eligibility_btnSearch"]'
                            heding="Search Button"
                            status="Search Button Not Found"
                            click(xpath,heding,status)

                            while True:           
                                try:
                                    element_4 = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, '//span[@id="dnn_ctr1732_Eligibility_lblResultsError"]'))).text                                               
                                    if element_4=='Subscriber/Insured Not Found' or element_4=='Invalid/Missing Subscriber/Insured ID':   
                                        wb1=load_workbook(filename=fil_path + '/' + fil_name + '.xlsx')
                                        sheet = wb1['Data']
                                        column_letter = 'K'  
                                        column_cells = sheet[column_letter]
                                        last_row = None
                                        for cell in reversed(column_cells):
                                            if cell.value:
                                                last_row = cell.row
                                                break                            
                                        sheet['K' + str(int(last_row + 1))]='Error'        
                                        wb1.save(fil_path + '/' + fil_name + '.xlsx')
                                        wb1.close()  

                                        xpath= '/html/body/form/div[3]/div/div[2]/div[1]/div/ul/div/li[5]/nobr/a'
                                        heding="Eligibility Button"
                                        status="Eligibility Button Not Found"
                                        click(xpath,heding,status)   
                                        break                      
                                                                        
                                except Exception as e:                
                                        xpath='//table[@id="dnn_ctr1732_Eligibility_gvRecipient"]/tbody/tr'
                                        heding="Recipient Table"
                                        status="Recipient Table Not Found"
                                        count(xpath,heding,status)                           
                                        
                                        lst = []                                                                                    
                                        lstt1 = []                                            
                                        j=1
                                        while j<rows+1:
                                            xpath= '//table[@id="dnn_ctr1732_Eligibility_gvRecipient"]/tbody/tr[{}]/td[1]'
                                            cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                            if cnm=='Name:':
                                                xpath='//table[@id="dnn_ctr1732_Eligibility_gvRecipient"]/tbody/tr[{}]/td[2]'                                      
                                                nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                                if nm==' ':
                                                    nm='N/A' 
                                                    lstt1.append(nm)
                                                else:
                                                    lstt1.append(nm) 
                                                break
                                            j=j+1
                                        
                                        lt=len(lstt1) 
                                        if lt==0:                                    
                                            lst.append('N/A')
                                        else:
                                            vr=lstt1[0]
                                            lst.append(vr)

                                        lstt2 = []                                            
                                        j=1
                                        while j<rows+1:
                                            xpath= '//table[@id="dnn_ctr1732_Eligibility_gvRecipient"]/tbody/tr[{}]/td[1]'
                                            cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                            if cnm=='Recipient ID:':
                                                xpath='//table[@id="dnn_ctr1732_Eligibility_gvRecipient"]/tbody/tr[{}]/td[2]'                                      
                                                nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                                if nm==' ':
                                                    nm='N/A' 
                                                    lstt2.append(nm)
                                                else:
                                                    lstt2.append(nm) 
                                                break
                                            j=j+1
                                        
                                        lt=len(lstt2) 
                                        if lt==0:                                    
                                            lst.append('N/A')
                                        else:
                                            vr=lstt2[0]
                                            lst.append(vr)

                                        lstt3 = []                                            
                                        j=1
                                        while j<rows+1:
                                            xpath= '//table[@id="dnn_ctr1732_Eligibility_gvRecipient"]/tbody/tr[{}]/td[1]'
                                            cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                            if cnm=='Date of Birth:':
                                                xpath='//table[@id="dnn_ctr1732_Eligibility_gvRecipient"]/tbody/tr[{}]/td[2]'                                      
                                                nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                                if nm==' ':
                                                    nm='N/A' 
                                                    lstt3.append(nm)
                                                else:
                                                    lstt3.append(nm) 
                                                break
                                            j=j+1
                                        
                                        lt=len(lstt3) 
                                        if lt==0:                                    
                                            lst.append('N/A')
                                        else:
                                            vr=lstt3[0]
                                            lst.append(vr)

                                        lstt4 = []                                            
                                        j=1
                                        while j<rows+1:
                                            xpath= '//table[@id="dnn_ctr1732_Eligibility_gvRecipient"]/tbody/tr[{}]/td[1]'
                                            cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                            if cnm=='Gender:':
                                                xpath='//table[@id="dnn_ctr1732_Eligibility_gvRecipient"]/tbody/tr[{}]/td[2]'                                      
                                                nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                                if nm==' ':
                                                    nm='N/A' 
                                                    lstt4.append(nm)
                                                else:
                                                    lstt4.append(nm) 
                                                break
                                            j=j+1
                                        
                                        lt=len(lstt4) 
                                        if lt==0:                                    
                                            lst.append('N/A')
                                        else:
                                            vr=lstt4[0]
                                            lst.append(vr)
                                            
                                        wb1=load_workbook(filename=fil_path + '/' + fil_name + '.xlsx')
                                        sheet = wb1['Data']
                                        column_letter = 'K'  
                                        column_cells = sheet[column_letter]
                                        last_row = None
                                        for cell in reversed(column_cells):
                                            if cell.value:
                                                last_row = cell.row
                                                break              
                                        start_column = 'L'
                                        current_column_index = openpyxl.utils.column_index_from_string(start_column)
                                        current_row = last_row + 1

                                        for value in lst:
                                            current_column = openpyxl.utils.get_column_letter(current_column_index)
                                            sheet[current_column + str(current_row)] = value
                                            current_column_index += 1
                                        
                                        wb1.save(filename=fil_path + '/' + fil_name + '.xlsx')
                                        wb1.close()
                                        
                                        try:
                                            xpath='//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr'                             
                                            counter = 0
                                            while counter < 10:
                                                try:                                                       
                                                    rows1=len(WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))                    
                                                    break
                                                except Exception as e:
                                                    time.sleep(1)
                                                    counter += 1                                
                                            
                                            lst1 = []                                            
                                            lstt1 = []
                                            j=1
                                            while j<rows+1:
                                                xpath= '//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[1]'
                                                cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                                if cnm=='Status:':
                                                    xpath='//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[2]'                                      
                                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                                    if nm==' ':
                                                        nm='N/A' 
                                                        lstt1.append(nm)
                                                    else:
                                                        lstt1.append(nm) 
                                                    break
                                                j=j+1
                                            
                                            lt=len(lstt1) 
                                            if lt==0:                                    
                                                lst1.append('N/A')
                                            else:
                                                vr=lstt1[0]
                                                lst1.append(vr)

                                            lstt2 = []
                                            j=1
                                            while j<rows+1:
                                                xpath= '//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[1]'
                                                cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                                if cnm=='Service Type:':
                                                    xpath='//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[2]'                                      
                                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                                    if nm==' ':
                                                        nm='N/A' 
                                                        lstt2.append(nm)
                                                    else:
                                                        lstt2.append(nm) 
                                                    break
                                                j=j+1

                                            lt=len(lstt2) 
                                            if lt==0:                                    
                                                lst1.append('N/A')
                                            else:
                                                vr=lstt2[0]
                                                lst1.append(vr)

                                            lstt3 = []
                                            j=1
                                            while j<rows+1:
                                                xpath= '//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[1]'
                                                cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                                if cnm=='Insurance Type:':
                                                    xpath='//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[2]'                                      
                                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                                    if nm==' ':
                                                        nm='N/A' 
                                                        lstt3.append(nm)
                                                    else:
                                                        lstt3.append(nm) 
                                                    break
                                                j=j+1
                                            
                                            lt=len(lstt3) 
                                            if lt==0:                                    
                                                lst1.append('N/A')
                                            else:
                                                vr=lstt3[0]
                                                lst1.append(vr)

                                            lstt4 = []
                                            j=1
                                            while j<rows+1:
                                                xpath= '//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[1]'
                                                cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                                if cnm=='Coverage Description:':
                                                    xpath='//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[2]'                                      
                                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                                    if nm==' ':
                                                        nm='N/A' 
                                                        lstt4.append(nm)
                                                    else:
                                                        lstt4.append(nm) 
                                                    break
                                                j=j+1

                                            lt=len(lstt4) 
                                            if lt==0:                                    
                                                lst1.append('N/A')
                                            else:
                                                vr=lstt4[0]
                                                lst1.append(vr)

                                            lstt5 = []
                                            j=1
                                            while j<rows+1:
                                                xpath= '//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[1]'
                                                cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                                if cnm=='Plan':
                                                    xpath='//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[2]'                                      
                                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                                    if nm==' ':
                                                        nm='N/A' 
                                                        lstt5.append(nm)
                                                    else:
                                                        lstt5.append(nm) 
                                                    break
                                                j=j+1

                                            lt=len(lstt5) 
                                            if lt==0:                                    
                                                lst1.append('N/A')
                                            else:
                                                vr=lstt5[0]
                                                lst1.append(vr)                                                
                                                    
                                            wb1=load_workbook(filename=fil_path + '/' + fil_name + '.xlsx')
                                            sheet = wb1['Data']
                                            column_letter = 'K'  
                                            column_cells = sheet[column_letter]
                                            last_row = None
                                            for cell in reversed(column_cells):
                                                if cell.value:
                                                    last_row = cell.row
                                                    break              
                                            start_column = 'P'
                                            current_column_index = openpyxl.utils.column_index_from_string(start_column)
                                            current_row = last_row + 1

                                            for value in lst1:
                                                current_column = openpyxl.utils.get_column_letter(current_column_index)
                                                sheet[current_column + str(current_row)] = value
                                                current_column_index += 1                                
                                            
                                            wb1.save(filename=fil_path + '/' + fil_name + '.xlsx')
                                            wb1.close()
                                        
                                        except Exception as e:   
                                                pass

                                        xpath='//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr'
                                        heding="Eligibility Summary Table"
                                        status="Eligibility Summary Table Not Found"
                                        count(xpath,heding,status)      
                                        
                                        lst2 = []
                                        i=1
                                        while i<rows+1:        
                                            if i > 1:                    
                                                cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,'//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr[{}]/td[1]'.format(i)))).text
                                                if cnm=='Co-Insurance':
                                                    xpath='//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr[{}]/td[1]'  
                                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(i)))).text
                                                    if nm==' ':
                                                        nm='N/A' 
                                                        lst2.append(nm)
                                                    else:
                                                        lst2.append(nm)                                        
                                                    xpath='//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr[{}]/td[2]'  
                                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(i)))).text
                                                    if nm==' ':
                                                        nm='N/A' 
                                                        lst2.append(nm)
                                                    else:
                                                        lst2.append(nm)
                                                    xpath='//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr[{}]/td[3]'  
                                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(i)))).text
                                                    if nm==' ':
                                                        nm='N/A' 
                                                        lst2.append(nm)
                                                    else:
                                                        lst2.append(nm)
                                                    xpath='//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr[{}]/td[4]'  
                                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(i)))).text
                                                    if nm==' ':
                                                        nm='N/A' 
                                                        lst2.append(nm)
                                                    else:
                                                        lst2.append(nm)                                          
                                                    break
                                                else:                                    
                                                    xpath='//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr[{}]/td[1]'  
                                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(i)))).text
                                                    if nm==' ':
                                                        nm='N/A' 
                                                        lst2.append(nm)
                                                    else:
                                                        lst2.append(nm)
                                                    xpath='//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr[{}]/td[2]'  
                                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(i)))).text
                                                    if nm==' ':
                                                        nm='N/A' 
                                                        lst2.append(nm)
                                                    else:
                                                        lst2.append(nm)
                                                    xpath='//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr[{}]/td[3]'  
                                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(i)))).text
                                                    if nm==' ':
                                                        nm='N/A' 
                                                        lst2.append(nm)
                                                    else:
                                                        lst2.append(nm)
                                                    xpath='//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr[{}]/td[4]'  
                                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(i)))).text
                                                    if nm==' ':
                                                        nm='N/A' 
                                                        lst2.append(nm)
                                                    else:
                                                        lst2.append(nm)                                                                     
                                            i=i+1

                                        wb1=load_workbook(filename=fil_path + '/' + fil_name + '.xlsx')
                                        sheet = wb1['Data']
                                        column_letter = 'K'  
                                        column_cells = sheet[column_letter]
                                        last_row = None
                                        for cell in reversed(column_cells):
                                            if cell.value:
                                                last_row = cell.row
                                                break              
                                        start_column = 'U'
                                        current_column_index = openpyxl.utils.column_index_from_string(start_column)
                                        current_row = last_row + 1

                                        for value in lst2:
                                            current_column = openpyxl.utils.get_column_letter(current_column_index)
                                            sheet[current_column + str(current_row)] = value
                                            current_column_index += 1                                
                                        
                                        sheet['K' + str(int(last_row + 1))]='Done'
                                        
                                        wb1.save(filename=fil_path + '/' + fil_name + '.xlsx')
                                        wb1.close()                            

                                        xpath= '/html/body/form/div[3]/div/div[2]/div[1]/div/ul/div/li[5]/nobr/a'
                                        heding="Eligibility Button"
                                        status="Eligibility Button Not Found"
                                        click(xpath,heding,status)   
                                        break                    
                    break            
            elif pd.isna(sfn) and not pd.isna(srid) and pd.isna(ssn):              

                xpath= '//input[@id="dnn_ctr1732_Eligibility_txtRecipientID2"]'
                key=srid     
                heding = 'Recipient ID' 
                status = 'Recipient ID Field Not Found'

                counter = 0
                while counter < 60:
                    try:                          
                        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)   
                        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//input[@id="dnn_ctr1732_Eligibility_txtDob3"]'))).click()                     
                        break
                    except Exception as e:
                        time.sleep(1)
                        counter += 1
                else:
                    messagebox.showinfo(heding, status)
                    sys.exit(0)                  
                
                while True:
                    try:    
                        element_5 = WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH, '//span[@id="dnn_ctr1732_Eligibility_cusRecipientID2"]'))).text                    
                        if element_5=='Invalid Recipient ID entered.':
                            key1=str('0' + str(key))                                      
                            WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, xpath))).clear()                 
                            WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key1)
                            WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//input[@id="dnn_ctr1732_Eligibility_txtDob3"]'))).click()
                            break
                    except Exception as e:        
                            break
                
                while True:
                    try:    
                        element_5 = WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH, '//span[@id="dnn_ctr1732_Eligibility_cusRecipientID2"]'))).text                    
                        if element_5=='Invalid Recipient ID entered.':
                            wb1=load_workbook(filename=fil_path + '/' + fil_name + '.xlsx')
                            sheet = wb1['Data']
                            column_letter = 'K'  
                            column_cells = sheet[column_letter]
                            last_row = None
                            for cell in reversed(column_cells):
                                if cell.value:
                                    last_row = cell.row
                                    break                            
                            sheet['K' + str(int(last_row + 1))]='Invalid Recipient ID entered'        
                            wb1.save(fil_path + '/' + fil_name + '.xlsx')
                            wb1.close()  

                            xpath= '/html/body/form/div[3]/div/div[2]/div[1]/div/ul/div/li[5]/nobr/a'
                            heding="Eligibility Button"
                            status="Eligibility Button Not Found"
                            click(xpath,heding,status)  
                            break
                    except Exception as e:                                                                
                            date_object = datetime.strptime(str(dob), "%Y-%m-%d %H:%M:%S")
                            dob2 = date_object.strftime("%m/%d/%Y")
                        
                            WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//input[@id="dnn_ctr1732_Eligibility_txtDob3"]'))).clear() 
                        
                            xpath= '//input[@id="dnn_ctr1732_Eligibility_txtDob3"]'
                            heding="Date of Birth"
                            status="Date of Birth Field Not Found"
                            key=dob2
                            text_box(xpath,heding,status,key)

                            date_object = datetime.strptime(str(fdt), "%Y-%m-%d %H:%M:%S")
                            fdt = date_object.strftime("%m/%d/%Y")
                        
                            WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//input[@id="dnn_ctr1732_Eligibility_txtDosFrom"]'))).clear() 
                            xpath= '//input[@id="dnn_ctr1732_Eligibility_txtDosFrom"]'
                            heding="From DOS"
                            status="From DOS Field Not Found"
                            key=fdt
                            text_box(xpath,heding,status,key)  

                            WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, '//input[@id="dnn_ctr1732_Eligibility_txtDosTo"]'))).clear() 
                            xpath= '//input[@id="dnn_ctr1732_Eligibility_txtDosTo"]'
                            heding="To DOS"
                            status="To DOS Field Not Found"
                            key=fdt
                            text_box(xpath,heding,status,key)  

                            xpath= '//input[@id="dnn_ctr1732_Eligibility_btnSearch"]'
                            heding="Search Button"
                            status="Search Button Not Found"
                            click(xpath,heding,status)

                            while True:           
                                try:
                                    element_4 = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, '//span[@id="dnn_ctr1732_Eligibility_lblResultsError"]'))).text                                               
                                    if element_4=='Subscriber/Insured Not Found' or element_4=='Invalid/Missing Subscriber/Insured ID':   
                                        wb1=load_workbook(filename=fil_path + '/' + fil_name + '.xlsx')
                                        sheet = wb1['Data']
                                        column_letter = 'K'  
                                        column_cells = sheet[column_letter]
                                        last_row = None
                                        for cell in reversed(column_cells):
                                            if cell.value:
                                                last_row = cell.row
                                                break                            
                                        sheet['K' + str(int(last_row + 1))]='Error'        
                                        wb1.save(fil_path + '/' + fil_name + '.xlsx')
                                        wb1.close()       
                                        xpath= '/html/body/form/div[3]/div/div[2]/div[1]/div/ul/div/li[5]/nobr/a'
                                        heding="Eligibility Button"
                                        status="Eligibility Button Not Found"
                                        click(xpath,heding,status)   
                                        break                                                        
                                except Exception as e:                
                                        xpath='//table[@id="dnn_ctr1732_Eligibility_gvRecipient"]/tbody/tr'
                                        heding="Recipient Table"
                                        status="Recipient Table Not Found"
                                        count(xpath,heding,status)                           
                                        
                                        lst = []                                                                                    
                                        lstt1 = []                                            
                                        j=1
                                        while j<rows+1:
                                            xpath= '//table[@id="dnn_ctr1732_Eligibility_gvRecipient"]/tbody/tr[{}]/td[1]'
                                            cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                            if cnm=='Name:':
                                                xpath='//table[@id="dnn_ctr1732_Eligibility_gvRecipient"]/tbody/tr[{}]/td[2]'                                      
                                                nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                                if nm==' ':
                                                    nm='N/A' 
                                                    lstt1.append(nm)
                                                else:
                                                    lstt1.append(nm) 
                                                break
                                            j=j+1
                                        
                                        lt=len(lstt1) 
                                        if lt==0:                                    
                                            lst.append('N/A')
                                        else:
                                            vr=lstt1[0]
                                            lst.append(vr) 

                                        lstt2 = []                                            
                                        j=1
                                        while j<rows+1:
                                            xpath= '//table[@id="dnn_ctr1732_Eligibility_gvRecipient"]/tbody/tr[{}]/td[1]'
                                            cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                            if cnm=='Recipient ID:':
                                                xpath='//table[@id="dnn_ctr1732_Eligibility_gvRecipient"]/tbody/tr[{}]/td[2]'                                      
                                                nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                                if nm==' ':
                                                    nm='N/A' 
                                                    lstt2.append(nm)
                                                else:
                                                    lstt2.append(nm) 
                                                break
                                            j=j+1
                                        
                                        lt=len(lstt2) 
                                        if lt==0:                                    
                                            lst.append('N/A')
                                        else:
                                            vr=lstt2[0]
                                            lst.append(vr) 

                                        lstt3 = []                                            
                                        j=1
                                        while j<rows+1:
                                            xpath= '//table[@id="dnn_ctr1732_Eligibility_gvRecipient"]/tbody/tr[{}]/td[1]'
                                            cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                            if cnm=='Date of Birth:':
                                                xpath='//table[@id="dnn_ctr1732_Eligibility_gvRecipient"]/tbody/tr[{}]/td[2]'                                      
                                                nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                                if nm==' ':
                                                    nm='N/A' 
                                                    lstt3.append(nm)
                                                else:
                                                    lstt3.append(nm) 
                                                break
                                            j=j+1
                                        
                                        lt=len(lstt3) 
                                        if lt==0:                                    
                                            lst.append('N/A')
                                        else:
                                            vr=lstt3[0]
                                            lst.append(vr) 

                                        lstt4 = []                                            
                                        j=1
                                        while j<rows+1:
                                            xpath= '//table[@id="dnn_ctr1732_Eligibility_gvRecipient"]/tbody/tr[{}]/td[1]'
                                            cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                            if cnm=='Gender:':
                                                xpath='//table[@id="dnn_ctr1732_Eligibility_gvRecipient"]/tbody/tr[{}]/td[2]'                                      
                                                nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                                if nm==' ':
                                                    nm='N/A' 
                                                    lstt4.append(nm)
                                                else:
                                                    lstt4.append(nm) 
                                                break
                                            j=j+1
                                        
                                        lt=len(lstt4) 
                                        if lt==0:                                    
                                            lst.append('N/A')
                                        else:
                                            vr=lstt4[0]
                                            lst.append(vr) 
                                            
                                        wb1=load_workbook(filename=fil_path + '/' + fil_name + '.xlsx')
                                        sheet = wb1['Data']
                                        column_letter = 'K'  
                                        column_cells = sheet[column_letter]
                                        last_row = None
                                        for cell in reversed(column_cells):
                                            if cell.value:
                                                last_row = cell.row
                                                break              
                                        start_column = 'L'
                                        current_column_index = openpyxl.utils.column_index_from_string(start_column)
                                        current_row = last_row + 1

                                        for value in lst:
                                            current_column = openpyxl.utils.get_column_letter(current_column_index)
                                            sheet[current_column + str(current_row)] = value
                                            current_column_index += 1
                                        
                                        wb1.save(filename=fil_path + '/' + fil_name + '.xlsx')
                                        wb1.close()
                                        
                                        try:
                                            xpath='//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr'                             
                                            counter = 0
                                            while counter < 10:
                                                try:                                                       
                                                    rows1=len(WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))                    
                                                    break
                                                except Exception as e:
                                                    time.sleep(1)
                                                    counter += 1                                
                                            
                                            lst1 = []                                                                                                                                        
                                            lstt1 = []
                                            j=1
                                            while j<rows+1:
                                                xpath= '//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[1]'
                                                cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                                if cnm=='Status:':
                                                    xpath='//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[2]'                                      
                                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                                    if nm==' ':
                                                        nm='N/A' 
                                                        lstt1.append(nm)
                                                    else:
                                                        lstt1.append(nm) 
                                                    break
                                                j=j+1
                                            
                                            lt=len(lstt1) 
                                            if lt==0:                                    
                                                lst1.append('N/A')
                                            else:
                                                vr=lstt1[0]
                                                lst1.append(vr) 

                                            lstt2 = []
                                            j=1
                                            while j<rows+1:
                                                xpath= '//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[1]'
                                                cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                                if cnm=='Service Type:':
                                                    xpath='//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[2]'                                      
                                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                                    if nm==' ':
                                                        nm='N/A' 
                                                        lstt2.append(nm)
                                                    else:
                                                        lstt2.append(nm) 
                                                    break
                                                j=j+1

                                            lt=len(lstt2) 
                                            if lt==0:                                    
                                                lst1.append('N/A')
                                            else:
                                                vr=lstt2[0]
                                                lst1.append(vr) 

                                            lstt3 = []
                                            j=1
                                            while j<rows+1:
                                                xpath= '//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[1]'
                                                cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                                if cnm=='Insurance Type:':
                                                    xpath='//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[2]'                                      
                                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                                    if nm==' ':
                                                        nm='N/A' 
                                                        lstt3.append(nm)
                                                    else:
                                                        lstt3.append(nm) 
                                                    break
                                                j=j+1
                                            
                                            lt=len(lstt3) 
                                            if lt==0:                                    
                                                lst1.append('N/A')
                                            else:
                                                vr=lstt3[0]
                                                lst1.append(vr) 

                                            lstt4 = []
                                            j=1
                                            while j<rows+1:
                                                xpath= '//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[1]'
                                                cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                                if cnm=='Coverage Description:':
                                                    xpath='//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[2]'                                      
                                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                                    if nm==' ':
                                                        nm='N/A' 
                                                        lstt4.append(nm)
                                                    else:
                                                        lstt4.append(nm) 
                                                    break
                                                j=j+1

                                            lt=len(lstt4) 
                                            if lt==0:                                    
                                                lst1.append('N/A')
                                            else:
                                                vr=lstt4[0]
                                                lst1.append(vr) 

                                            lstt5 = []
                                            j=1
                                            while j<rows+1:
                                                xpath= '//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[1]'
                                                cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                                if cnm=='Plan':
                                                    xpath='//table[@id="dnn_ctr1732_Eligibility_rptDetail_gvDetail_0"]/tbody/tr[{}]/td[2]'                                      
                                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                                    if nm==' ':
                                                        nm='N/A' 
                                                        lstt5.append(nm)
                                                    else:
                                                        lstt5.append(nm) 
                                                    break
                                                j=j+1

                                            lt=len(lstt5) 
                                            if lt==0:                                    
                                                lst1.append('N/A')
                                            else:
                                                vr=lstt5[0]
                                                lst1.append(vr)                                                 
                                                    
                                            wb1=load_workbook(filename=fil_path + '/' + fil_name + '.xlsx')
                                            sheet = wb1['Data']
                                            column_letter = 'K'  
                                            column_cells = sheet[column_letter]
                                            last_row = None
                                            for cell in reversed(column_cells):
                                                if cell.value:
                                                    last_row = cell.row
                                                    break              
                                            start_column = 'P'
                                            current_column_index = openpyxl.utils.column_index_from_string(start_column)
                                            current_row = last_row + 1

                                            for value in lst1:
                                                current_column = openpyxl.utils.get_column_letter(current_column_index)
                                                sheet[current_column + str(current_row)] = value
                                                current_column_index += 1                                
                                            
                                            wb1.save(filename=fil_path + '/' + fil_name + '.xlsx')
                                            wb1.close()
                                        
                                        except Exception as e:   
                                                pass

                                        xpath='//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr'
                                        heding="Eligibility Summary Table"
                                        status="Eligibility Summary Table Not Found"
                                        count(xpath,heding,status)      
                                        
                                        lst2 = []
                                        i=1
                                        while i<rows+1:        
                                            if i > 1:                    
                                                cnm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,'//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr[{}]/td[1]'.format(i)))).text
                                                if cnm=='Co-Insurance':
                                                    xpath='//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr[{}]/td[1]'  
                                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(i)))).text
                                                    if nm==' ':
                                                        nm='N/A' 
                                                        lst2.append(nm)
                                                    else:
                                                        lst2.append(nm)
                                                    xpath='//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr[{}]/td[2]'  
                                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(i)))).text
                                                    if nm==' ':
                                                        nm='N/A' 
                                                        lst2.append(nm)
                                                    else:
                                                        lst2.append(nm)
                                                    xpath='//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr[{}]/td[3]'  
                                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(i)))).text
                                                    if nm==' ':
                                                        nm='N/A' 
                                                        lst2.append(nm)
                                                    else:
                                                        lst2.append(nm)
                                                    xpath='//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr[{}]/td[4]'  
                                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(i)))).text
                                                    if nm==' ':
                                                        nm='N/A' 
                                                        lst2.append(nm)
                                                    else:
                                                        lst2.append(nm)                                        
                                                    break
                                                else:                                    
                                                    xpath='//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr[{}]/td[1]'  
                                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(i)))).text
                                                    if nm==' ':
                                                        nm='N/A' 
                                                        lst2.append(nm)
                                                    else:
                                                        lst2.append(nm)
                                                    xpath='//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr[{}]/td[2]'  
                                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(i)))).text
                                                    if nm==' ':
                                                        nm='N/A' 
                                                        lst2.append(nm)
                                                    else:
                                                        lst2.append(nm)
                                                    xpath='//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr[{}]/td[3]'  
                                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(i)))).text
                                                    if nm==' ':
                                                        nm='N/A' 
                                                        lst2.append(nm)
                                                    else:
                                                        lst2.append(nm)
                                                    xpath='//table[@id="dnn_ctr1732_Eligibility_gvSummary"]/tbody/tr[{}]/td[4]'  
                                                    nm=WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH,xpath.format(i)))).text
                                                    if nm==' ':
                                                        nm='N/A' 
                                                        lst2.append(nm)
                                                    else:
                                                        lst2.append(nm)                                
                                            i=i+1
                                        wb1=load_workbook(filename=fil_path + '/' + fil_name + '.xlsx')
                                        sheet = wb1['Data']
                                        column_letter = 'K'  
                                        column_cells = sheet[column_letter]
                                        last_row = None
                                        for cell in reversed(column_cells):
                                            if cell.value:
                                                last_row = cell.row
                                                break              
                                        start_column = 'U'
                                        current_column_index = openpyxl.utils.column_index_from_string(start_column)
                                        current_row = last_row + 1

                                        for value in lst2:
                                            current_column = openpyxl.utils.get_column_letter(current_column_index)
                                            sheet[current_column + str(current_row)] = value
                                            current_column_index += 1                                
                                        
                                        sheet['K' + str(int(last_row + 1))]='Done'
                                        
                                        wb1.save(filename=fil_path + '/' + fil_name + '.xlsx')
                                        wb1.close()   

                                        xpath= '/html/body/form/div[3]/div/div[2]/div[1]/div/ul/div/li[5]/nobr/a'
                                        heding="Eligibility Button"
                                        status="Eligibility Button Not Found"
                                        click(xpath,heding,status)   
                                        break  
                            break
            else:
                wb1=load_workbook(filename=fil_path + '/' + fil_name + '.xlsx')
                sheet = wb1['Data']
                column_letter = 'K'  
                column_cells = sheet[column_letter]
                last_row = None
                for cell in reversed(column_cells):
                    if cell.value:
                        last_row = cell.row
                        break                                  
                sheet['K' + str(int(last_row + 1))]='Condition - Not Meet Process Scenario'
                wb1.save(filename=fil_path + '/' + fil_name + '.xlsx')
                wb1.close()
        
        driver.close()
        messagebox.showinfo('Process Status', 'Process Completed...')
        sys.exit(0) 
          
if __name__=="__main__":    
    process()
    
