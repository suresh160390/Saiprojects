from tkinter import BOTH, LEFT, TOP, Button, Entry, Frame, Label, PhotoImage, StringVar, Tk,Radiobutton,StringVar,IntVar,filedialog
from idlelib.tooltip import Hovertip
from tkinter import messagebox
import sys
import os
import warnings
from datetime import datetime, date
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, StaleElementReferenceException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import openpyxl
import time
import pandas as pd
from openpyxl import load_workbook
from urllib.parse import urlparse
import warnings
import numpy as np
import requests
from zipfile import ZipFile
from openpyxl.utils import get_column_letter,column_index_from_string
warnings.filterwarnings("ignore")


global rows
global xpath
global heding
global status
global key
global nme
global driver
element_1 = None


def radio():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')

    root.title("Sharepoint - User Login And File Details")
    root.resizable(False,False)

    root.title("PCC Process Report")
    root.resizable(False,False)
   
    w = 600
    h = 230
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Please Select Claims (OR) Eligibility Report",font=("Calibri",18,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
    
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=300)
    
    title1=Label(Frame2,text="Claims Report :",font=("Calibri",17,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title1.grid(row=0,column=0,padx=45,pady=5,sticky="W")

    title2=Label(Frame2,text="Eligibility Report :",font=("Calibri",17,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title2.grid(row=1,column=0,padx=45,pady=5,sticky="W")
    
    answer=StringVar()
    answer.set("")

    title_3=Label(Frame2,text=answer.get(),textvariable=answer,font=("Calibri",12,"bold","italic"),bg="#2c3e50",fg="Red",justify="center",width=68)
    title_3.grid(row=2,column=0,columnspan=2,padx=0,pady=0,sticky="W")

    def Radio(*event):
        answer.set("")

    global var

    var = IntVar()

    R1 = Radiobutton(Frame2,text="Claims     ",variable=var, value=1,font=("Calibri",9,"bold","italic"),bg="Gold",fg="Black",width=20,justify="left",command=lambda: Radio())
    R1.grid(row=0,column=1,padx=0,pady=5,sticky="W")

    R2 = Radiobutton(Frame2,text="Eligibility",variable=var,value=2,font=("Calibri",9,"bold","italic"),bg="Gold",fg="Black",width=20,command=lambda: Radio())
    R2.grid(row=1,column=1,padx=0,pady=5,sticky="W")
        
    def Click_Done():
        selection = str(var.get())

        if selection==str(1):
           answer.set("")
           root.destroy() 
           process()
        elif selection==str(2):
            answer.set("")
            root.destroy()
            eligibility()        
        else:    
            answer.set("Please Select Any One Option...")            
    
    photo1 = PhotoImage(file=image_path1)

    btn=Button(Frame2,command=Click_Done,text="Done",image=photo1,borderwidth=0,bg="#2c3e50")
    btn.grid(row=4,column=0,padx=125,pady=20,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo = PhotoImage(file=image_path)   

    btn1=Button(Frame2,command=Close,text="Exit",image=photo,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=4,column=1,padx=15,pady=20,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn1,'Click to Exit Process',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def file_pick():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
        image_path2 = os.path.join(sys._MEIPASS, 'Static', 'Mapping.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')
        image_path2 = os.path.join(os.getcwd(), 'Static', 'Mapping.png')

    root.title("Process File Picker")
    root.resizable(False,False)
   
    w = 500
    h = 160
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Browse Process - File Picker",font=("Calibri",18,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
          
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=50)
    
    answer=StringVar()
    answer.set("")
       
    def browse_button():
        global filename
        answer.set("")  
        filename = filedialog.askopenfilename()  
        txt.delete(0, 'end') 
        txt.insert(0, filename)

    def Click_Done():
        global ans        
        
        ans=txt.get()        
        
        if ans=="":
           answer.set("File Path Fields Empty Is Not Allowed...")        
        else:      
            root.destroy()          
            return ans
        
    txt=Entry(Frame2,font=("Calibri",12,"bold","italic"),width=50,justify="left")
    txt.grid(row=0,column=0,padx=5,pady=10,sticky="E")    

    photo1 = PhotoImage(file=image_path2)
            
    btn1=Button(Frame2,text="Browse",command=browse_button,image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=0,column=1,padx=3,pady=0,sticky="W")     
                
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=3,y=80,width=490,height=25)
    
    title_3=Label(Frame3,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=1,padx=150,pady=0,sticky="E")

    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=3,y=105,width=500,height=50)
    
    photo2 = PhotoImage(file=image_path1)

    btn2=Button(Frame4,command=Click_Done,text="Done",image=photo2,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=3,column=0,padx=125,pady=0,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo3 = PhotoImage(file=image_path)   

    btn3=Button(Frame4,command=Close,text="Exit",image=photo3,borderwidth=0,bg="#2c3e50")
    btn3.grid(row=3,column=1,padx=15,pady=0,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn2,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn3,'Click to Exit Process',hover_delay=1000)
    myTip2 = Hovertip(btn1,'Click to Pick File',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def file_pick_1():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
        image_path2 = os.path.join(sys._MEIPASS, 'Static', 'Mapping.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')
        image_path2 = os.path.join(os.getcwd(), 'Static', 'Mapping.png')

    root.title("Python File Path")
    root.resizable(False,False)
   
    w = 500
    h = 160
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Python File Path",font=("Calibri",18,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
          
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=50)
    
    answer=StringVar()
    answer.set("")
       
    # def browse_button():
    #     global filename
    #     answer.set("")  
    #     filename = filedialog.askopenfilename()  
    #     txt.delete(0, 'end') 
    #     txt.insert(0, filename)

    def Click_Done():
        global ans4        
        
        ans4=txt.get()        
        
        if ans4=="":
           answer.set("File Path Fields Empty Is Not Allowed...")        
        else:      
            root.destroy()          
            return ans4
        
    txt=Entry(Frame2,font=("Calibri",12,"bold","italic"),width=60,justify="left")
    txt.grid(row=0,column=0,padx=5,pady=10,sticky="E")    

    # photo1 = PhotoImage(file=image_path2)
            
    # btn1=Button(Frame2,text="Browse",command=browse_button,image=photo1,borderwidth=0,bg="#2c3e50")
    # btn1.grid(row=0,column=1,padx=3,pady=0,sticky="W")     
                
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=3,y=80,width=490,height=25)
    
    title_3=Label(Frame3,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=1,padx=150,pady=0,sticky="E")

    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=3,y=105,width=500,height=50)
    
    photo2 = PhotoImage(file=image_path1)

    btn2=Button(Frame4,command=Click_Done,text="Done",image=photo2,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=3,column=0,padx=125,pady=0,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo3 = PhotoImage(file=image_path)   

    btn3=Button(Frame4,command=Close,text="Exit",image=photo3,borderwidth=0,bg="#2c3e50")
    btn3.grid(row=3,column=1,padx=15,pady=0,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn2,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn3,'Click to Exit Process',hover_delay=1000)
    # myTip2 = Hovertip(btn1,'Click to Pick File',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def user_pass():
    global ans1
    global ans2
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')

    root.title("Sharepoint - User Login And File Details")
    root.resizable(False,False)

    w = 600
    h = 180
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)

    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="User Name and Password Details...",font=("Calibri",20,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack() 
 
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=698,height=150)    
    
    title3=Label(Frame2,text="User Name :",font=("Calibri",11,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title3.grid(row=0,column=0,padx=5,pady=5,sticky="W")
    
    txt1=Entry(Frame2,font=("Calibri",11,"bold","italic"),width=50,justify="left")
    txt1.grid(row=0,column=1,padx=30,pady=5,sticky="W")
    
    title4=Label(Frame2,text="Password :",font=("Calibri",11,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title4.grid(row=1,column=0,padx=5,pady=5,sticky="W")
    
    txt2=Entry(Frame2,font=("Calibri",11,"bold","italic"),width=50,justify="left")
    txt2.grid(row=1,column=1,padx=30,pady=5,sticky="W")

    answer=StringVar()
    answer.set("")

    def Click_Done():
        global ans1
        global ans2        
        
        ans1=txt1.get()
        ans2=txt2.get()        

        if ans1=="":
           answer.set("User Name Field Empty Is Not Allowed...")
        elif ans2=="":
            answer.set("Password Field Empty Is Not Allowed...")        
        else:    
            root.destroy()
            return ans1,ans2
                
    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=0,y=105,width=698,height=20)
    
    title_3=Label(Frame4,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=2,padx=200,pady=0,sticky="E")
    
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=0,y=125,width=698,height=200)
       
    photo1 = PhotoImage(file=image_path1)
    
    btn1=Button(Frame3,command=Click_Done,text="Run",image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=2,column=0,padx=140,pady=0,sticky="W")

    def Close():
        sys.exit(0)

    photo = PhotoImage(file=image_path)    

    btn2=Button(Frame3,command=Close,text="Close",image=photo,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=2,column=1,padx=100,pady=0,sticky="E")

    def disable_event():
        pass

    txt1.focus_set()

    myTip = Hovertip(btn1,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn2,'Click to Exit Process',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)
            
    root.mainloop()

def process():        
        global element_1    
        global ans1
        global ans2 
        global fs
        global ans4
        global driver_path
        # ccshssh_sub01
        # Carespan@2023

        user_pass()
        file_pick()
        # file_pick_1()

        user_name =ans1
        password=ans2        
        fil=ans       
        
        try:
            options = Options()            
            options.add_argument('--ignore-certificate-errors')
            options.add_argument('--ignore-ssl-errors')
            options.add_argument("--disable-popup-blocking")
            driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options)
            driver.maximize_window()
            driver.get('https://cms.officeally.com/')
        except Exception as e:                        
            try:
                options =  webdriver.ChromeOptions()        
                options.add_argument('--ignore-certificate-errors')
                options.add_argument('--ignore-ssl-errors')
                options.add_argument("--disable-popup-blocking")
                # response = requests.get('https://chromedriver.storage.googleapis.com/LATEST_RELEASE')
                # latest_version = response.text.strip()
                # chrome_driver_url = f'https://chromedriver.storage.googleapis.com/{latest_version}/chromedriver_win32.zip'
                # response = requests.get(chrome_driver_url)
                # with open('chromedriver_win32.zip', 'wb') as f:
                #     f.write(response.content)
                # with ZipFile('chromedriver_win32.zip', 'r') as zip_ref:
                #     zip_ref.extractall('.')   
                # driver_path=r'C:\Users\sanandrao\Desktop\New folder\chromedriver.exe'
                driver_path = os.path.abspath('chromedriver.exe')                
                driver = webdriver.Chrome(executable_path=driver_path,options=options)             
                driver.maximize_window()
                driver.get('https://cms.officeally.com/')                                         
            except Exception as e:
                messagebox.showinfo("Driver Problem","Pls Check Your Chrome Version Driver")
                sys.exit(0)
        
        
        
        def text_box(xpath,heding,status,key):                
            counter = 0
            while counter < 15:
                try:   
                    WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()                   
                    WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding, status)
                sys.exit(0)        
                
        def click(xpath,heding,status):
            counter = 0
            while counter < 15:
                try:             
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding,status)
                sys.exit(0)                                

        def count(xpath,heding,status):
            global rows
            counter = 0
            while counter < 15:
                try:             
                    rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))                    
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding,status)
                sys.exit(0)                         

        def Alert():
            counter = 0
            while counter < 15:
                try:             
                    WebDriverWait(driver, 0).until (EC.alert_is_present())
                    a=driver.switch_to.alert
                    a.accept()
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo('Alert','Alert Not Present')
                                                          
        xpath= "/html/body/div[8]/div[1]/div[2]/div[1]/div[2]/div[1]/div"
        heding="Login Button"
        status="Login Button Not Found"
        click(xpath,heding,status)
                
        xpath= "/html/body/div[8]/div[1]/div[2]/div[1]/div[2]/div[1]/nav/div/div/a"
        heding="Login Count"
        status="Login Count Field Not Found"        
        count(xpath,heding,status)
              
        xpath= "/html/body/div[8]/div[1]/div[2]/div[1]/div[2]/div[1]/nav/div/div/a[{}]"
        
        j=1
        while j<rows+1:                                
            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
            parts = cnm.split('\n')
            cnm = parts[0]
            if cnm=="Service Center":
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                            
                break
            j=j+1       
                       
        driver.switch_to.window(driver.window_handles[1])        

        xpath= '/html/body/main/section/div/div/div/form/div[2]/div/div[1]/div/input'
        heding="User Name"
        status="User Name Field Not Found"
        key=user_name
        text_box(xpath,heding,status,key)
        
        xpath= '/html/body/main/section/div/div/div/form/div[2]/div/div[2]/div/input'
        heding="Password"
        status="Password Field Not Found"
        key=password
        text_box(xpath,heding,status,key)

        xpath= "/html/body/main/section/div/div/div/form/div[3]/button"
        heding="Continue"
        status="Continue Button Not Found"
        click(xpath,heding,status)
        
        while True:
            xpath='/html/body/main/section/div/div/div/form/div[2]/div/div[2]/span'         
            try:
                element_3 = WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH, xpath))).text
                if element_3.lstrip().rstrip()=='Wrong username or password':                               
                    messagebox.showinfo('Login Status','Please Check - UserName or Password Wrong')                    
                    user_pass()
                    
                    xpath= '/html/body/main/section/div/div/div/form/div[2]/div/div[1]/div/input'
                    heding="User Name"
                    status="User Name Field Not Found"
                    key=ans1
                    text_box(xpath,heding,status,key)
                    
                    xpath= '/html/body/main/section/div/div/div/form/div[2]/div/div[2]/div/input'
                    heding="Password"
                    status="Password Field Not Found"
                    key=ans2
                    text_box(xpath,heding,status,key)

                    xpath= "/html/body/main/section/div/div/div/form/div[3]/button"
                    heding="Continue"
                    status="Continue Button Not Found"
                    click(xpath,heding,status)
                else:
                    break    
            except Exception as e:                
                    break                                 
        
        xpath= "/html/body/app-root/div[1]/oa-layout/section/header/oa-navigation/div/div[3]/button"
        heding="Return to Classic"
        status="Return to Classic Button Not Found"
        click(xpath,heding,status)
        
        xpath= "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td[1]/table/tbody/tr[2]/td[2]/table/tbody/tr"
        heding="Request Count"
        status="Request Count Not Found"        
        count(xpath,heding,status)      
       
        j=1
        while j<rows+1:
            xpath= "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td[1]/table/tbody/tr[2]/td[2]/table/tbody/tr[{}]"                       
            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
            if cnm.lstrip().rstrip()=="Real Time Claim Status":
                xpath= "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td[1]/table/tbody/tr[2]/td[2]/table/tbody/tr[{}]/td/a"
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                            
                break
            j=j+1    
        
        counter = 0
        while counter < 15:
            try: 
                iframe = WebDriverWait(driver, 0).until(EC.presence_of_element_located((By.ID, "Iframe7")))
                driver.switch_to.frame(iframe)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
                messagebox.showinfo('Iframe','Iframe7 Not Present') 
                sys.exit(0)       

        file=pd.read_excel(fil,sheet_name='Claim Status BOT',header=0)

        file['CPT - HC'] = file['CPT - HC'].astype(str)
        file['Modifier'] = file['Modifier'].astype(str)        

        for index, row in file.iterrows():                          
            py_nm = row[9]  
            py_lnm = row[2]            
            py_fnm = row[3]                 
            tx_id=row[6]            
            fdt=row[11]
            tdt=row[12]
                       
            cpt=row[13]
            mod=row[14]

            if pd.isnull(mod) or mod=='nan':
                fcpt='HC:'+str(cpt)
            else:
                fcpt='HC:'+str(cpt)+':'+str(mod)

            xpath= "/html/body/form/div[3]/div[1]/div[1]/div[1]/input"
            heding="Reset All"
            status="Reset All Button Not Found"
            click(xpath,heding,status)   
            
            xpath= '/html/body/form/div[3]/div[1]/div[2]/div/div[2]/input[1]'
            heding="Last Name"
            status="Last Name Field Not Found"
            key=py_lnm
            text_box(xpath,heding,status,key)

            xpath= '/html/body/form/div[3]/div[1]/div[2]/div/div[2]/input[2]'
            heding="First Name"
            status="First Name Field Not Found"
            key=py_fnm
            text_box(xpath,heding,status,key)
            
            xpath= '/html/body/form/div[3]/div[1]/div[2]/div/div[3]/input[2]'
            heding="Tax ID"
            status="Tax ID Field Not Found"
            key=tx_id
            text_box(xpath,heding,status,key)
            
            lst=[]           
            date_object = datetime.strptime(str(fdt), "%Y-%m-%d %H:%M:%S")
            dob2 = date_object.strftime("%#m/%#d/%Y")
            lst=dob2.split('/')
            mn=lst[0]
            dt=lst[1]
            yy=lst[2]
           
            xpath= '/html/body/form/div[3]/div[1]/div[2]/div/div[4]/span[1]/input[1]'
            heding="First Date"
            status="Month Field Not Found"
            key=mn
            text_box(xpath,heding,status,key)
            
            xpath= '/html/body/form/div[3]/div[1]/div[2]/div/div[4]/span[1]/span[1]/input'
            heding="First Date"
            status="Date Field Not Found"
            key=dt
            text_box(xpath,heding,status,key)
            
            xpath= '/html/body/form/div[3]/div[1]/div[2]/div/div[4]/span[1]/input[2]'
            heding="First Date"
            status="Year Field Not Found"
            key=yy
            text_box(xpath,heding,status,key)

            lst=[]           
            date_object = datetime.strptime(str(tdt), "%Y-%m-%d %H:%M:%S")
            dob2 = date_object.strftime("%#m/%#d/%Y")
            lst=dob2.split('/')
            mn=lst[0]
            dt=lst[1]
            yy=lst[2]            
       
            xpath= '/html/body/form/div[3]/div[1]/div[2]/div/div[4]/span[2]/input[1]'
            heding="Last Date"
            status="Month Field Not Found"
            key=mn
            text_box(xpath,heding,status,key)
            
            xpath= '/html/body/form/div[3]/div[1]/div[2]/div/div[4]/span[2]/span[1]/input'
            heding="Last Date"
            status="Date Field Not Found"
            key=dt
            text_box(xpath,heding,status,key)
            
            xpath= '/html/body/form/div[3]/div[1]/div[2]/div/div[4]/span[2]/input[2]'
            heding="Last Date"
            status="Year Field Not Found"
            key=yy
            text_box(xpath,heding,status,key)
            
            xpath= "/html/body/form/div[3]/div[1]/div[2]/div/div[4]/input"
            heding="Claim Search"
            status="Claim Search Button Not Found"
            click(xpath,heding,status)
           
            while True:
                try:
                    element_1 = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[3]/div[1]/div[3]/span'))).text
                    break
                except Exception as e:
                    pass     

            if element_1 =="Search Results - 0 claims found":                
                wb1=load_workbook(filename=fil)
                sheet = wb1['Claim Status BOT']
                column_letter = 'Z'  
                column_cells = sheet[column_letter]
                last_row = None
                for cell in reversed(column_cells):
                    if cell.value:
                        last_row = cell.row
                        break                            
                sheet['Z' + str(int(last_row + 1))]=element_1       
                wb1.save(fil)
                wb1.close()                                                        
            else:
                xpath= "/html/body/form/div[3]/div[1]/div[3]/div[1]/table/tbody/tr[2]/td[11]/a"
                heding="Claim Click"
                status="Claim Click Button Not Found"
                click(xpath,heding,status)                                        
                
                while True:
                    try:
                        element_6 = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[3]/div[2]'))).text
                        break
                    except Exception as e:
                        pass

                if element_6.lstrip().rstrip()=='Backend processing error.':

                    xpath= "/html/body/div[3]/div[1]/button"
                    heding="Close Click"
                    status="Close Click Button Not Found"
                    click(xpath,heding,status)
                    
                    wb1=load_workbook(filename=fil)
                    sheet = wb1['Claim Status BOT']
                    column_letter = 'Z'  
                    column_cells = sheet[column_letter]
                    last_row = None
                    for cell in reversed(column_cells):
                        if cell.value:
                            last_row = cell.row
                            break                            
                    sheet['Z' + str(int(last_row + 1))]='Error'       
                    wb1.save(fil)
                    wb1.close()

                else:                                      
                    try:                                       
                        element_2 = WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[3]/div[2]/dl/dt'))).text
                        if element_2.lstrip().rstrip() =="Message:": 
                            wb1=load_workbook(filename=fil)
                            sheet = wb1['Claim Status BOT']
                            column_letter = 'Z'  
                            column_cells = sheet[column_letter]
                            last_row = None
                            for cell in reversed(column_cells):
                                if cell.value:
                                    last_row = cell.row
                                    break                            
                            sheet['Z' + str(int(last_row + 1))]='No Transaction History'       
                            wb1.save(fil)
                            wb1.close()
                    except Exception as e:  
                                            
                        xpath= "/html/body/div[3]/div[2]/fieldset"
                        heding="Result Count"
                        status="Result Count Not Found"        
                        count(xpath,heding,status) 

                        fs=rows
                        
                        xpath= "/html/body/div[3]/div[2]/fieldset[{}]/fieldset/dl".format(fs)
                        heding="Result Count"
                        status="Result Count Not Found"        
                        count(xpath,heding,status) 

                        lst = []
                        lst1 = []
                        j=1
                        while j<rows+1:                    
                            xpath= "/html/body/div[3]/div[2]/fieldset[{}]/fieldset/dl[{}]/dt"                                
                            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(fs,j)))).text              
                            if 'Payors Claim Number :' in cnm:
                                xpath= "/html/body/div[3]/div[2]/fieldset[{}]/fieldset/dl[{}]/dd"
                                nm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(fs,j)))).text  
                                if nm==' ':
                                    nm='N/A' 
                                    lst1.append(nm)
                                else:
                                    lst1.append(nm)                                        
                                break
                            j=j+1
                        
                        lt=len(lst1)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst1[0]
                            lst.append(vr)

                        lst2=[]
                        j=1
                        while j<rows+1:
                            xpath= "/html/body/div[3]/div[2]/fieldset[{}]/fieldset/dl[{}]/dt"                                
                            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(fs,j)))).text              
                            if cnm.lstrip().rstrip()=="Dependent Account Number :":
                                xpath= "/html/body/div[3]/div[2]/fieldset[{}]/fieldset/dl[{}]/dd"
                                nm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(fs,j)))).text  
                                if nm==' ':
                                    nm='N/A' 
                                    lst2.append(nm)
                                else:
                                    lst2.append(nm)                                                 
                                break
                            j=j+1   
                        
                        lt=len(lst2) 
                        if lt==0:                                    
                            lst.append('N/A')   
                        else:
                            vr=lst2[0]
                            lst.append(vr)

                        lst3=[]
                        j=1
                        while j<rows+1:
                            xpath= "/html/body/div[3]/div[2]/fieldset[{}]/fieldset/dl[{}]/dt"                                
                            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(fs,j)))).text              
                            if cnm.lstrip().rstrip()=="Claim Number :":
                                xpath= "/html/body/div[3]/div[2]/fieldset[{}]/fieldset/dl[{}]/dd"
                                nm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(fs,j)))).text  
                                if nm==' ':
                                    nm='N/A' 
                                    lst3.append(nm)
                                else:
                                    lst3.append(nm)                                                 
                                break
                            j=j+1  

                        lt=len(lst3)
                        if lt==0:
                            lst.append('N/A')
                        else:
                            vr=lst3[0]
                            lst.append(vr)

                        if rows!=0:
                            lst4=[]
                            xpath = "/html/body/div[3]/div[2]/fieldset[{}]/fieldset"
                            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(fs)))).text   
                            lst4=cnm.split('\n')
                            data = {line.split(":")[0].strip(): line.split(":")[1].strip() for line in lst4 if ":" in line}
                            keys_to_remove = ['Payors Claim Number', 'Dependent Account Number', 'Claim Number']

                            for key in keys_to_remove:
                                if key in data:
                                    del data[key]                              
                            
                            cpm = data.get('Claim Payment Amount', '')

                            if cpm =='':
                                lst.append('N/A') 
                            else:
                                lst.append(cpm) 
                            
                            pd1 = data.get('Payment Date', '')

                            if pd1 =='':
                                lst.append('N/A') 
                            else:
                                lst.append(pd1) 
                            
                            ptn = data.get('Payment Trace Number', '')

                            if ptn =='':
                                lst.append('N/A') 
                            else:
                                lst.append(ptn) 
                            
                            tuple_data = tuple(data.items())                       
                            dec=list(data.items())[0]                
                            dec = str(dec)
                            lst.append(dec) 
                            
                            xpath= "/html/body/div[3]/div[2]/div[{}]/fieldset/fieldset/h3/a".format(fs)
                            heding="Claim Click"
                            status="Claim Click Button Not Found"
                            click(xpath,heding,status)
                            
                            lst5=[]   
                            try:
                                xpath = "/html/body/div[3]/div[2]/div[{}]/fieldset/fieldset/div".format(fs)
                                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text                  
                                lst5=cnm.split('\n')
                                
                                if fcpt in lst5:                                
                                    index = lst5.index(fcpt)                               
                                    # if index !=0:
                                    #     del lst5[0:index]
                                    # my_dict = {line.split(":")[0].strip(): line.split(":")[1].strip() for line in lst5 if ":" in line}    
                                    # print(my_dict)
                                    # index = lst5.index(fcpt)
                                    element = lst5[index + 2]
                                    my_dict = {}
                                    key, value = element.split(": ")
                                    my_dict = {key.strip(): value.strip()}

                                    pam = my_dict.get('Paid', '')

                                    if pam =='':
                                        lst.append('N/A') 
                                    else:
                                        lst.append(pam) 

                                    wb1=load_workbook(fil)
                                    sheet = wb1['Claim Status BOT']
                                    column_letter = 'Z'  
                                    column_cells = sheet[column_letter]
                                    last_row = None
                                    for cell in reversed(column_cells):
                                        if cell.value:
                                            last_row = cell.row
                                            break              
                                    start_column = 'R'
                                    current_column_index = openpyxl.utils.column_index_from_string(start_column)
                                    current_row = last_row + 1

                                    for value in lst:
                                        current_column = openpyxl.utils.get_column_letter(current_column_index)
                                        sheet[current_column + str(current_row)] = value
                                        current_column_index += 1                
                                    sheet['Z' + str(int(last_row + 1))]='Done'
                                    wb1.save(filename=fil)
                                    wb1.close()
                                else:
                                    wb1=load_workbook(fil)
                                    sheet = wb1['Claim Status BOT']
                                    column_letter = 'Z'  
                                    column_cells = sheet[column_letter]
                                    last_row = None
                                    for cell in reversed(column_cells):
                                        if cell.value:
                                            last_row = cell.row
                                            break              
                                    start_column = 'R'
                                    current_column_index = openpyxl.utils.column_index_from_string(start_column)
                                    current_row = last_row + 1

                                    for value in lst:
                                        current_column = openpyxl.utils.get_column_letter(current_column_index)
                                        sheet[current_column + str(current_row)] = value
                                        current_column_index += 1                
                                    sheet['Z' + str(int(last_row + 1))]=fcpt + ' - Not Found'
                                    wb1.save(filename=fil)
                                    wb1.close()
                            except Exception as e:
                                wb1=load_workbook(fil)
                                sheet = wb1['Claim Status BOT']
                                column_letter = 'Z'  
                                column_cells = sheet[column_letter]
                                last_row = None
                                for cell in reversed(column_cells):
                                    if cell.value:
                                        last_row = cell.row
                                        break              
                                start_column = 'R'
                                current_column_index = openpyxl.utils.column_index_from_string(start_column)
                                current_row = last_row + 1

                                for value in lst:
                                    current_column = openpyxl.utils.get_column_letter(current_column_index)
                                    sheet[current_column + str(current_row)] = value
                                    current_column_index += 1                
                                sheet['Z' + str(int(last_row + 1))]='List Detailed Items Empty'
                                wb1.save(filename=fil)
                                wb1.close()
                        else:
                            wb1=load_workbook(fil)
                            sheet = wb1['Claim Status BOT']
                            column_letter = 'Z'  
                            column_cells = sheet[column_letter]
                            last_row = None
                            for cell in reversed(column_cells):
                                if cell.value:
                                    last_row = cell.row
                                    break              
                            start_column = 'R'
                            current_column_index = openpyxl.utils.column_index_from_string(start_column)
                            current_row = last_row + 1

                            for value in lst:
                                current_column = openpyxl.utils.get_column_letter(current_column_index)
                                sheet[current_column + str(current_row)] = value
                                current_column_index += 1                
                            sheet['Z' + str(int(last_row + 1))]='Done'
                            wb1.save(filename=fil)
                            wb1.close()

                    xpath= "/html/body/div[3]/div[1]/button/span[1]"
                    heding="Close Click"
                    status="Close Click Button Not Found"
                    click(xpath,heding,status)                

        driver.quit()
        messagebox.showinfo("Process Status", "Process Completed")
        sys.exit(0)                    

def eligibility():        
        global element_1    
        global ans1
        global ans2 
        global driver
        global ans4
        # ccshssh_sub01
        # Carespan@2023

        user_pass()
        file_pick()
        # file_pick_1()
        user_name =ans1
        password=ans2        
        fil=ans       
        
        try:
            options = Options()            
            options.add_argument('--ignore-certificate-errors')
            options.add_argument('--ignore-ssl-errors')
            options.add_argument("--disable-popup-blocking")
            driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options)
            driver.maximize_window()
            driver.get('https://cms.officeally.com/')
        except Exception as e:                        
            try:
                options = Options()            
                options.add_argument('--ignore-certificate-errors')
                options.add_argument('--ignore-ssl-errors')
                options.add_argument("--disable-popup-blocking")
                # response = requests.get('https://chromedriver.storage.googleapis.com/LATEST_RELEASE')
                # latest_version = response.text.strip()
                # chrome_driver_url = f'https://chromedriver.storage.googleapis.com/{latest_version}/chromedriver_win32.zip'
                # response = requests.get(chrome_driver_url)
                # with open('chromedriver_win32.zip', 'wb') as f:
                #     f.write(response.content)
                # with ZipFile('chromedriver_win32.zip', 'r') as zip_ref:
                #     zip_ref.extractall('.')
                
                # driver_path = ans4 + '/' + 'chromedriver.exe'
                # print(driver_path)    
                
                driver_path = os.path.abspath('chromedriver.exe')
                driver = webdriver.Chrome(executable_path=driver_path,options=options)
                driver.maximize_window()
                driver.get('https://cms.officeally.com/')
            except Exception as e:
                messagebox.showinfo("Internet Problem","Pls Check Your Internet Connection")
                sys.exit(0)
                
        def text_box(xpath,heding,status,key):                
            counter = 0
            while counter < 15:
                try:   
                    WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()                   
                    WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding, status)
                sys.exit(0)        
                
        def click(xpath,heding,status):
            counter = 0
            while counter < 15:
                try:             
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding,status)
                sys.exit(0)                                

        def count(xpath,heding,status):
            global rows
            counter = 0
            while counter < 15:
                try:             
                    rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))                    
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding,status)
                sys.exit(0)                         

        def Alert():
            counter = 0
            while counter < 15:
                try:             
                    WebDriverWait(driver, 0).until (EC.alert_is_present())
                    a=driver.switch_to.alert
                    a.accept()
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo('Alert','Alert Not Present')
                                                          
        xpath= "/html/body/div[8]/div[1]/div[2]/div[1]/div[2]/div[1]/div"
        heding="Login Button"
        status="Login Button Not Found"
        click(xpath,heding,status)
                
        xpath= "/html/body/div[8]/div[1]/div[2]/div[1]/div[2]/div[1]/nav/div/div/a"
        heding="Login Count"
        status="Login Count Field Not Found"        
        count(xpath,heding,status)
              
        xpath= "/html/body/div[8]/div[1]/div[2]/div[1]/div[2]/div[1]/nav/div/div/a[{}]"
        
        j=1
        while j<rows+1:                                
            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
            parts = cnm.split('\n')
            cnm = parts[0]
            if cnm=="Service Center":
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                            
                break
            j=j+1       
                       
        driver.switch_to.window(driver.window_handles[1])        

        xpath= '/html/body/main/section/div/div/div/form/div[2]/div/div[1]/div/input'
        heding="User Name"
        status="User Name Field Not Found"
        key=user_name
        text_box(xpath,heding,status,key)
        
        xpath= '/html/body/main/section/div/div/div/form/div[2]/div/div[2]/div/input'
        heding="Password"
        status="Password Field Not Found"
        key=password
        text_box(xpath,heding,status,key)

        xpath= "/html/body/main/section/div/div/div/form/div[3]/button"
        heding="Continue"
        status="Continue Button Not Found"
        click(xpath,heding,status)
        
        while True:
            xpath='/html/body/main/section/div/div/div/form/div[2]/div/div[2]/span'         
            try:
                element_3 = WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH, xpath))).text
                if element_3.lstrip().rstrip()=='Wrong username or password':                               
                    messagebox.showinfo('Login Status','Please Check - UserName or Password Wrong')                    
                    user_pass()
                    
                    xpath= '/html/body/main/section/div/div/div/form/div[2]/div/div[1]/div/input'
                    heding="User Name"
                    status="User Name Field Not Found"
                    key=ans1
                    text_box(xpath,heding,status,key)
                    
                    xpath= '/html/body/main/section/div/div/div/form/div[2]/div/div[2]/div/input'
                    heding="Password"
                    status="Password Field Not Found"
                    key=ans2
                    text_box(xpath,heding,status,key)

                    xpath= "/html/body/main/section/div/div/div/form/div[3]/button"
                    heding="Continue"
                    status="Continue Button Not Found"
                    click(xpath,heding,status)
                else:
                    break    
            except Exception as e:                
                    break
          
        xpath= "/html/body/app-root/div[1]/oa-layout/section/header/oa-navigation/div/div[3]/button"
        heding="Return to Classic"
        status="Return to Classic Button Not Found"
        click(xpath,heding,status)
        
        xpath= "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td[1]/table/tbody/tr[2]/td[2]/table/tbody/tr"
        heding="Request Count"
        status="Request Count Not Found"        
        count(xpath,heding,status)      
       
        j=1
        while j<rows+1:         
            xpath= "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td[1]/table/tbody/tr[2]/td[2]/table/tbody/tr[{}]"                       
            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
            if cnm.lstrip().rstrip()=="Verify Eligibility & Benefits":
                xpath= "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td[1]/table/tbody/tr[2]/td[2]/table/tbody/tr[{}]/td/a"
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                            
                break
            j=j+1                    

        file=pd.read_excel(fil,sheet_name='Eligibility BOT',header=0)
        
        file['NPI'] = file['NPI'].astype(str)

        for index, row in file.iterrows():                      
            py_lnm = row[1]            
            py_fnm = row[2]                                       
            dob=row[3]
            npi=row[4]
            py_nm=row[5]
            mem_id=row[6]
            fdt=row[7]

            py_nm = py_nm.lower()
            py_nm = py_nm.replace(' ', '')

            if pd.isnull(mem_id):
                mem_id=''            
            
            xpath= "/html/body/app-root/oa-find-payer/oa-modal/div/div/div/div/div[2]/oa-combobox/div/div[1]/input"
            heding="Search"
            status="Search Payer Button Not Found"
            click(xpath,heding,status)
            
            # xpath="/html/body/app-root/oa-find-payer/oa-modal/div/div/div/div/div[2]/oa-combobox/div/div[1]/input"
            # heding="Payer"
            # status="Payer Field Not Found"
            # key=py_nm
            # text_box(xpath,heding,status,key)

            xpath= "/html/body/app-root/oa-find-payer/oa-modal/div/div/div/div/div[2]/oa-combobox/div/div[3]/oa-select-option"
            heding="Payer Count"
            status="Payer Count Not Found"        
            count(xpath,heding,status)     
                        
            j=1
            while j<rows+1:         
                xpath= "/html/body/app-root/oa-find-payer/oa-modal/div/div/div/div/div[2]/oa-combobox/div/div[3]/oa-select-option[{}]"                       
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text     
                cnm = cnm.lower()
                cnm = cnm.replace(' ', '')
                
                if cnm.lstrip().rstrip()==py_nm.lstrip().rstrip():                       
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                            
                    break
                j=j+1

            if cnm.lstrip().rstrip()==py_nm.lstrip().rstrip():
                   
                xpath= "/html/body/app-root/div[1]/oa-layout/section/section/main/oa-eligibility-form/div[2]/div/form/div[1]/oa-provider-search/div/div[1]"
                heding="Provider"
                status="Provider Not Found"
                click(xpath,heding,status)                                

                xpath="/html/body/app-root/app-load-eligibility-provider/oa-modal/div/div/div/div/section/div[1]/oa-input/div/div[1]/input"
                heding="NPI"
                status="NPI Field Not Found"
                key=npi
                text_box(xpath,heding,status,key)

                xpath= "/html/body/app-root/app-load-eligibility-provider/oa-modal/div/div/div/div/section/div[2]/ag-grid-angular/div/div[2]/div[2]/div[3]/div[2]/div/div/div"
                heding="Find Provider"
                status="Find Provider Count Not Found"        
                count(xpath,heding,status)

                j=1
                while j<rows+1:                             
                    xpath= "/html/body/app-root/app-load-eligibility-provider/oa-modal/div/div/div/div/section/div[2]/ag-grid-angular/div/div[2]/div[2]/div[3]/div[2]/div/div/div[{}]/div[2]"                                
                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                    if cnm==npi:
                        xpath= "/html/body/app-root/app-load-eligibility-provider/oa-modal/div/div/div/div/section/div[2]/ag-grid-angular/div/div[2]/div[2]/div[3]/div[2]/div/div/div[{}]"
                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                          
                        break
                    j=j+1                                

                xpath= '/html/body/app-root/div[1]/oa-layout/section/section/main/oa-eligibility-form/div[2]/div/form/div[2]/div[2]/oa-input[1]/div/div[1]/input'
                heding="First Name"
                status="First Name Field Not Found"
                key=py_fnm
                text_box(xpath,heding,status,key)
                
                xpath= '/html/body/app-root/div[1]/oa-layout/section/section/main/oa-eligibility-form/div[2]/div/form/div[2]/div[2]/oa-input[2]/div/div[1]/input'
                heding="Last Name"
                status="Last Name Field Not Found"
                key=py_lnm
                text_box(xpath,heding,status,key)

                date_object = datetime.strptime(str(dob), "%Y-%m-%d %H:%M:%S")
                dob2 = date_object.strftime("%m/%d/%Y")

                xpath= '/html/body/app-root/div[1]/oa-layout/section/section/main/oa-eligibility-form/div[2]/div/form/div[2]/div[3]/oa-input[1]/div/div[1]/input'
                heding="Date of Birth"
                status="Date of Birth Field Not Found"
                key=dob2
                text_box(xpath,heding,status,key)

                xpath= '/html/body/app-root/div[1]/oa-layout/section/section/main/oa-eligibility-form/div[2]/div/form/div[2]/div[3]/oa-input[2]/div/div[1]/input'
                heding="Member ID"
                status="Member ID Field Not Found"
                key=mem_id
                text_box(xpath,heding,status,key)
                
                date_object_1 = datetime.strptime(str(fdt), "%Y-%m-%d %H:%M:%S")
                dob3 = date_object_1.strftime("%m/%d/%Y")

                xpath= '/html/body/app-root/div[1]/oa-layout/section/section/main/oa-eligibility-form/div[2]/div/form/div[4]/div[2]/oa-date-input[1]/div/div[1]/input'
                heding="From Date"
                status="From Date Field Not Found"
                key=dob3
                text_box(xpath,heding,status,key)

                xpath= "/html/body/app-root/div[1]/oa-layout/section/section/main/oa-eligibility-form/div[2]/div/div/button"
                heding="Submit"
                status="Submit Button Not Found"
                click(xpath,heding,status)
                
                counter = 0
                while counter < 60:
                    try:                           
                        element_4=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/app-root/div[1]/oa-layout/section/section/main/oa-eligibility-form/oa-searching-loader/div[2]')))                                                                   
                        if element_4:                                                  
                            time.sleep(1)
                            counter += 1
                    except Exception as e:
                        break                                                                                                               
                else:
                    messagebox.showinfo('Time Out', 'Portal Issue, Please try Later')
                    sys.exit(0)

                while True:
                    try:
                        element_1 = WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH, '/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/div[1]/oa-payer-coverage/div/div/div[2]'))).text
                        break
                    except Exception as e:
                        pass
                
                lst = []
                if element_1=='Active Coverage':
                    
                    lst.append(element_1)

                    xpath= "/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/oa-result-summary/div/div[2]/div/div"
                    heding="Payer Response"
                    status="Payer Response Count Not Found"        
                    count(xpath,heding,status)  
                    
                    lst1 = []
                    j=1
                    while j<rows+1:                    
                        xpath= "/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/oa-result-summary/div/div[2]/div/div[{}]"                                
                        cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                        if 'Group Number' in cnm:
                            xpath= "/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/oa-result-summary/div/div[2]/div/div[{}]/div[2]"
                            nm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                            if nm==' ':
                                nm='N/A' 
                                lst1.append(nm)
                            else:
                                lst1.append(nm)                                        
                            break
                        j=j+1
                    
                    lt=len(lst1)
                    if lt==0:                           
                        lst.append('N/A')
                    else:
                        vr=lst1[0]
                        lst.append(vr)

                    lst2 = []
                    j=1
                    while j<rows+1:                    
                        xpath= "/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/oa-result-summary/div/div[2]/div/div[{}]"                                
                        cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                        if 'Medicare Beneficiary ID (MBI)' in cnm:
                            xpath= "/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/oa-result-summary/div/div[2]/div/div[{}]/div[2]"
                            nm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                            if nm==' ':
                                nm='N/A' 
                                lst2.append(nm)
                            else:
                                lst2.append(nm)                                        
                            break
                        j=j+1
                    
                    lt=len(lst2)
                    if lt==0:                           
                        lst.append('N/A')
                    else:
                        vr=lst2[0]
                        lst.append(vr)

                    try:                                                                                                              
                        element_6 = WebDriverWait(driver, 2).until(EC.visibility_of_element_located((By.XPATH, '/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/oa-result-summary/div/div[3]/div/div')))
                        xpath= "/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/oa-result-summary/div/div[3]/div/div"
                        heding="Dependent Information"
                        status="Dependent Information Count Not Found"        
                        count(xpath,heding,status)  

                        lst7 = []
                        j=1
                        while j<rows+1:                    
                            xpath= "/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/oa-result-summary/div/div[3]/div/div[{}]"                                
                            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                            if 'Address' in cnm:                            
                                xpath= "/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/oa-result-summary/div/div[3]/div/div[{}]/div[2]"
                                nm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                if nm==' ':
                                    nm='N/A' 
                                    lst7.append(nm)
                                else:
                                    lst7.append(nm)                                        
                                break
                            j=j+1
                        
                        lt=len(lst7)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst7[0]
                            lst.append(vr)   

                        lst2 = []
                        j=1
                        while j<rows+1:                    
                            xpath= "/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/oa-result-summary/div/div[3]/div/div[{}]"                                
                            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                            if 'Plan Begin Date' in cnm:
                                xpath= "/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/oa-result-summary/div/div[3]/div/div[{}]/div[2]"
                                nm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                if nm==' ':
                                    nm='N/A' 
                                    lst2.append(nm)
                                else:
                                    lst2.append(nm)                                        
                                break
                            j=j+1
                        
                        lt=len(lst2)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst2[0]
                            lst.append(vr)
                        
                        lst3 = []
                        j=1
                        while j<rows+1:                    
                            xpath= "/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/oa-result-summary/div/div[3]/div/div[{}]"                                
                            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                            if 'Plan End Date' in cnm:
                                xpath= "/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/oa-result-summary/div/div[3]/div/div[{}]/div[2]"
                                nm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                if nm==' ':
                                    nm='N/A' 
                                    lst3.append(nm)
                                else:
                                    lst3.append(nm)                                        
                                break
                            j=j+1
                        
                        lt=len(lst3)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst3[0]
                            lst.append(vr)                                                

                    except Exception as e:
                        
                        xpath= "/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/oa-result-summary/div/div[2]/div/div"
                        heding="Payer Response Exception"
                        status="Payer Response Exception Count Not Found"        
                        count(xpath,heding,status)     

                        lst7 = []
                        j=1
                        while j<rows+1:                    
                            xpath= "/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/oa-result-summary/div/div[2]/div/div[{}]"                                
                            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                            if 'Address' in cnm:                            
                                xpath= "/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/oa-result-summary/div/div[2]/div/div[{}]/div[2]"
                                nm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                if nm==' ':
                                    nm='N/A' 
                                    lst7.append(nm)
                                else:
                                    lst7.append(nm)                                        
                                break
                            j=j+1
                        
                        lt=len(lst7)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst7[0]
                            lst.append(vr)                           

                        lst2 = []
                        j=1
                        while j<rows+1:                    
                            xpath= "/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/oa-result-summary/div/div[2]/div/div[{}]"                                
                            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                            if 'Plan Begin Date' in cnm:
                                xpath= "/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/oa-result-summary/div/div[2]/div/div[{}]/div[2]"
                                nm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                if nm==' ':
                                    nm='N/A' 
                                    lst2.append(nm)
                                else:
                                    lst2.append(nm)                                        
                                break
                            j=j+1
                        
                        lt=len(lst2)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst2[0]
                            lst.append(vr)

                        lst3 = []
                        j=1

                        
                        while j<rows+1:                    
                            xpath= "/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/oa-result-summary/div/div[2]/div/div[{}]"                                
                            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                            if 'Plan End Date' in cnm:
                                xpath= "/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/oa-result-summary/div/div[2]/div/div[{}]/div[2]"
                                nm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                                if nm==' ':
                                    nm='N/A' 
                                    lst3.append(nm)
                                else:
                                    lst3.append(nm)                                        
                                break
                            j=j+1
                        
                        lt=len(lst3)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst3[0]
                            lst.append(vr)
                    
                    xpath= "/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/oa-result-plan/div/div[2]/div/div[1]/div[1]/h3"
                    heding="Plan Details"
                    status="Plan Details Count Not Found"        
                    count(xpath,heding,status)  
                    
                    lst4 = []
                    j=1
                    while j<rows+1:                    
                        xpath= "/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/oa-result-plan/div/div[2]/div/div[1]/div[1]/h3[{}]"                                
                        cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                        if 'Plan Name:' in cnm:
                            key, value = cnm.split(':')
                            key = key.strip()
                            value = value.strip()
                            dict_data = {key: value}                            
                            nm = dict_data.get('Plan Name', '')           
                            if nm==' ':
                                nm='N/A'
                                lst4.append(nm)
                            else:
                                lst4.append(nm)                                        
                            break
                        j=j+1
                    
                    lt=len(lst4)
                    if lt==0:                           
                        lst.append('N/A')
                    else:
                        vr=lst4[0]
                        lst.append(vr)

                    lst5 = []
                    j=1
                    while j<rows+1:                    
                        xpath= "/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/oa-result-plan/div/div[2]/div/div[1]/div[1]/h3[{}]"                                
                        cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                        if 'Insurance Type:' in cnm:
                            key, value = cnm.split(':')
                            key = key.strip()
                            value = value.strip()
                            dict_data = {key: value}  
                            nm = dict_data.get('Insurance Type', '')           
                            if nm==' ':
                                nm='N/A' 
                                lst5.append(nm)
                            else:
                                lst5.append(nm)                                        
                            break
                        j=j+1
                    
                    lt=len(lst5)
                    if lt==0:                           
                        lst.append('N/A')
                    else:
                        vr=lst5[0]
                        lst.append(vr)                                        

                    wb1=load_workbook(fil)
                    sheet = wb1['Eligibility BOT']
                    column_letter = 'R'  
                    column_cells = sheet[column_letter]
                    last_row = None
                    for cell in reversed(column_cells):
                        if cell.value:
                            last_row = cell.row
                            break              
                    start_column = 'J'
                    current_column_index = openpyxl.utils.column_index_from_string(start_column)
                    current_row = last_row + 1

                    for value in lst:
                        current_column = openpyxl.utils.get_column_letter(current_column_index)
                        sheet[current_column + str(current_row)] = value
                        current_column_index += 1                
                    sheet['R' + str(int(last_row + 1))]='Done'
                    wb1.save(filename=fil)
                    wb1.close()
                    
                    xpath= "/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/div[1]/oa-eligibility-response-nav/div/div[3]/div/button"
                    heding="New Search"
                    status="New Search Not Found"
                    click(xpath,heding,status)
                    
                    xpath= "/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/div[1]/oa-eligibility-response-nav/div/div[3]/div/div/div[1]/button"
                    heding=" Different Payer"
                    status=" Different Payer Button Not Found"
                    click(xpath,heding,status)                                                               
     
                # elif element_1=='Inactive Coverage':
                #     lst.append(element_1)
                else:                   
                    wb1=load_workbook(filename=fil)
                    sheet = wb1['Eligibility BOT']
                    column_letter = 'R'  
                    column_cells = sheet[column_letter]
                    last_row = None
                    for cell in reversed(column_cells):
                        if cell.value:
                            last_row = cell.row
                            break                            
                    sheet['R' + str(int(last_row + 1))]=element_1      
                    wb1.save(fil)
                    wb1.close()                                         
                    
                    xpath= "/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/div[1]/oa-eligibility-response-nav/div/div[3]/div/button"
                    heding="New Search"
                    status="New Search Not Found"
                    click(xpath,heding,status)
                    
                    xpath= "/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/div[1]/oa-eligibility-response-nav/div/div[3]/div/div/div[1]/button"
                    heding=" Different Payer"
                    status=" Different Payer Button Not Found"
                    click(xpath,heding,status)                                                               
            else:
                wb1=load_workbook(filename=fil)
                sheet = wb1['Eligibility BOT']
                column_letter = 'R'  
                column_cells = sheet[column_letter]
                last_row = None
                for cell in reversed(column_cells):
                    if cell.value:
                        last_row = cell.row
                        break                            
                sheet['R' + str(int(last_row + 1))]='Payer Name Not Found'       
                wb1.save(fil)
                wb1.close()    
                
                xpath= "/html/body/app-root/oa-find-payer/oa-modal/div/div/div/div/div[1]/div/button"
                heding="Payer Cancel"
                status="Payer Cancel Button Not Found"
                click(xpath,heding,status)
                                
                xpath= "/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/div[1]/oa-eligibility-response-nav/div/div[3]/div/button"
                heding="New Search"
                status="New Search Not Found"
                click(xpath,heding,status)

                xpath= "/html/body/app-root/div[1]/oa-layout/section[2]/oa-results/div[1]/oa-eligibility-response-nav/div/div[3]/div/div/div[1]/button"
                heding=" Different Payer"
                status=" Different Payer Button Not Found"
                click(xpath,heding,status)     
                
        driver.quit()
        messagebox.showinfo("Process Status", "Process Completed")
        sys.exit(0)       

if __name__=="__main__":        
    radio()    