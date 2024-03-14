from tkinter import BOTH, LEFT, TOP, Button, Entry, Frame, Label, PhotoImage, StringVar, Tk,Radiobutton,StringVar,IntVar,filedialog
from idlelib.tooltip import Hovertip
from tkinter import messagebox
import sys
import os
import warnings
from datetime import datetime, date
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, StaleElementReferenceException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
import time
import pandas as pd
from openpyxl import load_workbook
from urllib.parse import urlparse
import warnings
import numpy as np
import requests
from zipfile import ZipFile
from openpyxl.utils import get_column_letter,column_index_from_string
warnings.filterwarnings("ignore")


global rows
global rows_1
global xpath
global heding
global status
global key
global nme
global driver
element_1 = None
global j
global i


def radio():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')

    root.title("Psych360 Medicaid")
    root.resizable(False,False)
   
    w = 600
    h = 230
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Please Select Primary (OR) Secondary Claims",font=("Calibri",18,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
    
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=300)
    
    title1=Label(Frame2,text="Primary Claims :",font=("Calibri",17,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title1.grid(row=0,column=0,padx=45,pady=5,sticky="W")

    title2=Label(Frame2,text="Secondary Claims :",font=("Calibri",17,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title2.grid(row=1,column=0,padx=45,pady=5,sticky="W")
    
    answer=StringVar()
    answer.set("")

    title_3=Label(Frame2,text=answer.get(),textvariable=answer,font=("Calibri",12,"bold","italic"),bg="#2c3e50",fg="Red",justify="center",width=68)
    title_3.grid(row=2,column=0,columnspan=2,padx=0,pady=0,sticky="W")

    def Radio(*event):
        answer.set("")

    global var

    var = IntVar()

    R1 = Radiobutton(Frame2,text="Primary     ",variable=var, value=1,font=("Calibri",9,"bold","italic"),bg="Gold",fg="Black",width=20,justify="left",command=lambda: Radio())
    R1.grid(row=0,column=1,padx=0,pady=5,sticky="W")

    R2 = Radiobutton(Frame2,text="Secondary",variable=var,value=2,font=("Calibri",9,"bold","italic"),bg="Gold",fg="Black",width=20,command=lambda: Radio())
    R2.grid(row=1,column=1,padx=0,pady=5,sticky="W")
        
    def Click_Done():
        selection = str(var.get())

        if selection==str(1):
           answer.set("")
           root.destroy() 
           primary()
        elif selection==str(2):
            answer.set("")
            root.destroy()
            secondary()        
        else:    
            answer.set("Please Select Any One Option...")            
    
    photo1 = PhotoImage(file=image_path1)

    btn=Button(Frame2,command=Click_Done,text="Done",image=photo1,borderwidth=0,bg="#2c3e50")
    btn.grid(row=4,column=0,padx=125,pady=20,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo = PhotoImage(file=image_path)   

    btn1=Button(Frame2,command=Close,text="Exit",image=photo,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=4,column=1,padx=15,pady=20,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn1,'Click to Exit Process',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def file_pick():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
        image_path2 = os.path.join(sys._MEIPASS, 'Static', 'Mapping.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')
        image_path2 = os.path.join(os.getcwd(), 'Static', 'Mapping.png')

    root.title("Process File Picker")
    root.resizable(False,False)
   
    w = 500
    h = 160
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Browse Process - File Picker",font=("Calibri",18,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
          
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=50)
    
    answer=StringVar()
    answer.set("")
       
    def browse_button():
        global filename
        answer.set("")  
        filename = filedialog.askopenfilename()  
        txt.delete(0, 'end') 
        txt.insert(0, filename)

    def Click_Done():
        global ans        
        
        ans=txt.get()        
        
        if ans=="":
           answer.set("File Path Fields Empty Is Not Allowed...")        
        else:      
            root.destroy()          
            return ans
        
    txt=Entry(Frame2,font=("Calibri",12,"bold","italic"),width=50,justify="left")
    txt.grid(row=0,column=0,padx=5,pady=10,sticky="E")    

    photo1 = PhotoImage(file=image_path2)
            
    btn1=Button(Frame2,text="Browse",command=browse_button,image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=0,column=1,padx=3,pady=0,sticky="W")     
                
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=3,y=80,width=490,height=25)
    
    title_3=Label(Frame3,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=1,padx=150,pady=0,sticky="E")

    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=3,y=105,width=500,height=50)
    
    photo2 = PhotoImage(file=image_path1)

    btn2=Button(Frame4,command=Click_Done,text="Done",image=photo2,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=3,column=0,padx=125,pady=0,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo3 = PhotoImage(file=image_path)   

    btn3=Button(Frame4,command=Close,text="Exit",image=photo3,borderwidth=0,bg="#2c3e50")
    btn3.grid(row=3,column=1,padx=15,pady=0,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn2,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn3,'Click to Exit Process',hover_delay=1000)
    myTip2 = Hovertip(btn1,'Click to Pick File',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def primary():        
        global element_1    
        global ans1
        global ans2        
        global ans4
        global driver_path
        global j
        global i       
        
        file_pick()
                      
        fil=ans       
        
        try:
            options = Options()            
            options.add_argument('--ignore-certificate-errors')
            options.add_argument('--ignore-ssl-errors')
            options.add_argument("--disable-popup-blocking")
            driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options)
            driver.maximize_window()
            driver.get('https://ohid.ohio.gov/wps/portal/gov/ohid/login')
        except Exception as e:                        
            try:
                options =  webdriver.ChromeOptions()        
                options.add_argument('--ignore-certificate-errors')
                options.add_argument('--ignore-ssl-errors')
                options.add_argument("--disable-popup-blocking")
                # response = requests.get('https://chromedriver.storage.googleapis.com/LATEST_RELEASE')
                # latest_version = response.text.strip()
                # chrome_driver_url = f'https://chromedriver.storage.googleapis.com/{latest_version}/chromedriver_win32.zip'
                # response = requests.get(chrome_driver_url)
                # with open('chromedriver_win32.zip', 'wb') as f:
                #     f.write(response.content)
                # with ZipFile('chromedriver_win32.zip', 'r') as zip_ref:
                #     zip_ref.extractall('.')   
                # driver_path=r'C:\Users\sanandrao\Desktop\New folder\chromedriver.exe'
                driver_path = os.path.abspath('chromedriver.exe')                
                driver = webdriver.Chrome(executable_path=driver_path)                
                driver.maximize_window()
                driver.get('https://ohid.ohio.gov/wps/portal/gov/ohid/login')                                         
            except Exception as e:
                messagebox.showinfo("Internet Problem","Pls Check Your Internet Connection")
                sys.exit(0)
                        
        def text_box(xpath,heding,status,key):                
            counter = 0
            while counter < 15:
                try:   
                    WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()                   
                    WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding, status)
                sys.exit(0)        
        
        def text_box(xpath,heding,status,key):                
            counter = 0
            while counter < 15:
                try:   
                    WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()                   
                    WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding, status)
                sys.exit(0)    
        
        def text_box_key(xpath,heding,status,key):                
            counter = 0
            while counter < 15:
                try:   
                    element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, xpath)))
                    element.send_keys(key)
                    element.send_keys(Keys.TAB)
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding, status)
                sys.exit(0)    


        def click(xpath,heding,status):
            counter = 0
            while counter < 15:
                try:             
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding,status)
                sys.exit(0)                                

        def count(xpath,heding,status):
            global rows
            counter = 0
            while counter < 15:
                try:             
                    rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))                    
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding,status)
                sys.exit(0)                         

        def count_1(xpath,heding,status):
            global rows_1
            counter = 0
            while counter < 15:
                try:             
                    rows_1=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))                    
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding,status)
                sys.exit(0)   

        def Alert():
            counter = 0
            while counter < 15:
                try:             
                    WebDriverWait(driver, 0).until (EC.alert_is_present())
                    a=driver.switch_to.alert
                    a.accept()
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo('Alert','Alert Not Present')
                
        messagebox.showinfo('Waiting','Authentication Waiting')
        
        driver.switch_to.window(driver.window_handles[1])           
                       
        file=pd.read_excel(fil,sheet_name='Primary & Secondary Template',header=0)        

        for index, row in file.iterrows():                          
            bil_num = row[0]
            add_zer=row[1]
            dob = row[2]            
            pat_acn = row[3]                 
            ren_id=row[4]                                    
            dx1=row[5]
            dx2=row[6]
            dx3=row[7]
            dx4=row[8]
            dos=row[9]
            pos=row[10]
            pod_cod=row[11]
            mod1=row[12]
            mod2=row[13]
            amt=row[14]
            
            wait = WebDriverWait(driver, 20)
            wait.until(lambda driver: driver.execute_script("return document.readyState === 'complete';"))
            
            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[1]/td/div[1]/ul/li"
            heding="Claims"
            status="Claims Tab Count Not Found"        
            count(xpath,heding,status)

            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[1]/td/div[1]/ul/li[{}]"

            j=1
            while j<rows+1:                                
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text            
                if cnm.lstrip().rstrip()=="Claims":
                    
                    time.sleep(3)

                    element=WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j))))
                    actions = ActionChains(driver)
                    actions.move_to_element(element).perform()  

                    xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[1]/td/div[1]/ul/li[{}]/ul/li".format(j)
                    heding="Claims Count"
                    status="Claims Count Not Found"        
                    count_1(xpath,heding,status)                
                        
                    i=1
                    while i<rows_1+1:                                       
                        xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[1]/td/div[1]/ul/li[{}]/ul/li[{}]"
                        element_7=WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j,i))))
                        actions = ActionChains(driver)
                        actions.move_to_element(element_7).perform()  
                        element_text = element_7.text                    
                        if element_text.lstrip().rstrip()=="Professional":
                            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j,i)))).click()  
                            break                       
                        i=i+1                                                              
                    break
                j=j+1                            
            
            if pd.isnull(add_zer):
                bil_num = str(bil_num)
            elif add_zer.lower().lstrip().rstrip()=='add':
                bil_num ='0' + str(bil_num)
            else:
                bil_num = str(bil_num)

            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/span/table/tbody/tr/td/table/tbody/tr[6]/td[2]/input'
            heding="Medical Billing Number"
            status="Medical Billing Number Field Not Found"
            key=bil_num
            text_box(xpath,heding,status,key)
            
            date_object_1 = datetime.strptime(str(dob), "%Y-%m-%d %H:%M:%S")
            dob3 = date_object_1.strftime("%m/%d/%Y")
            
            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/span/table/tbody/tr/td/table/tbody/tr[7]/td[2]/input'
            heding="Date of Birth"
            status="Date of Birth Field Not Found"
            key=dob3
            text_box_key(xpath,heding,status,key)
            
            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/span/table/tbody/tr/td/table/tbody/tr[19]/td[2]/select/option"
            heding="Medicare Assignment"
            status="Medicare Assignment Count Not Found"        
            count(xpath,heding,status)   
            
            j=1
            while j<rows+1:                       
                xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/span/table/tbody/tr/td/table/tbody/tr[19]/td[2]/select/option[{}]"
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                if cnm.lstrip().rstrip()=="ASSIGNED":
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()  
                    break                       
                j=j+1    
            
            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/span/table/tbody/tr/td/table/tbody/tr[2]/td[4]/select/option"
            heding="Release of Information"
            status="Release of Information Count Not Found"        
            count(xpath,heding,status)   
            
            j=1
            while j<rows+1:                       
                xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/span/table/tbody/tr/td/table/tbody/tr[2]/td[4]/select/option[{}]"
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                if cnm.lstrip().rstrip()=="SIGNED STMT PERMITTING RELEASE":
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()  
                    break                       
                j=j+1   

            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/span/table/tbody/tr/td/table/tbody/tr[5]/td[4]/select/option"
            heding="Signature Source"
            status="Signature Source Count Not Found"        
            count(xpath,heding,status)   
            
            j=1
            while j<rows+1:                       
                xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/span/table/tbody/tr/td/table/tbody/tr[5]/td[4]/select/option[{}]"
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                if cnm.lstrip().rstrip()=="SIGNED HCFA-1500 ON FILE":
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()  
                    break                       
                j=j+1   

            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/span/table/tbody/tr/td/table/tbody/tr[10]/td[2]/input'
            heding="Patient Account"
            status="Patient Account Field Not Found"
            key=str(pat_acn)
            text_box(xpath,heding,status,key)

            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/span/table/tbody/tr/td/table/tbody/tr[18]/td[2]/input'
            heding="Rendering ID"
            status="Rendering ID Field Not Found"
            key=str(ren_id)
            text_box(xpath,heding,status,key)


            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[2]/td[2]/select/option"
            heding="Sequence"
            status="Sequence Count Not Found"        
            count(xpath,heding,status) 

            if pd.isnull(dx1):
                pass
            else:
                j=1
                while j<rows+1:                       
                    xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[2]/td[2]/select/option[{}]"
                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                    if cnm.lstrip().rstrip()=="01":
                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()  
                        
                        xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[2]/td[4]/span/input'
                        heding="Diagnosis Code - 01"
                        status="Diagnosis Code - 01 Field Not Found"
                        key=str(dx1)
                        text_box(xpath,heding,status,key)

                        break                       
                    j=j+1                                   
                
            if pd.isnull(dx2):
                pass
            else:                
                j=1
                while j<rows+1:                       
                    xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[2]/td[2]/select/option[{}]"
                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                    if cnm.lstrip().rstrip()=="02":
                        xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[1]/td[1]/table/tbody/tr/td[2]/a"
                        heding="Diagnosis add in item"
                        status="Diagnosis add in item Button Not Found"
                        click(xpath,heding,status)

                        xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[2]/td[2]/select/option[{}]"
                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click() 

                        xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[2]/td[4]/span/input'
                        heding="Diagnosis Code - 02"
                        status="Diagnosis Code - 02 Field Not Found"
                        key=str(dx2)
                        text_box(xpath,heding,status,key)

                        break                       
                    j=j+1   
            
            if pd.isnull(dx3):
                pass
            else:                
                j=1
                while j<rows+1:                       
                    xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[2]/td[2]/select/option[{}]"
                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                    if cnm.lstrip().rstrip()=="03":
                        xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[1]/td[1]/table/tbody/tr/td[2]/a"
                        heding="Diagnosis add in item"
                        status="Diagnosis add in item Button Not Found"
                        click(xpath,heding,status)

                        xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[2]/td[2]/select/option[{}]"
                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click() 

                        xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[2]/td[4]/span/input'
                        heding="Diagnosis Code - 03"
                        status="Diagnosis Code - 03 Field Not Found"
                        key=str(dx3)
                        text_box(xpath,heding,status,key)

                        break                       
                    j=j+1   
            
            if pd.isnull(dx4):
                pass
            else:                
                j=1
                while j<rows+1:                       
                    xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[2]/td[2]/select/option[{}]"
                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                    if cnm.lstrip().rstrip()=="04":
                        xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[1]/td[1]/table/tbody/tr/td[2]/a"
                        heding="Diagnosis add in item"
                        status="Diagnosis add in item Button Not Found"
                        click(xpath,heding,status)

                        xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[2]/td[2]/select/option[{}]"
                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click() 

                        xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[2]/td[4]/span/input'
                        heding="Diagnosis Code - 04"
                        status="Diagnosis Code - 04 Field Not Found"
                        key=str(dx4)
                        text_box(xpath,heding,status,key)

                        break                       
                    j=j+1   

            date_object_1 = datetime.strptime(str(dos), "%Y-%m-%d %H:%M:%S")
            dob3 = date_object_1.strftime("%m/%d/%Y")
            
            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[3]/td[2]/input'
            heding="From DOS"
            status="From DOS Field Not Found"
            key=dob3
            text_box(xpath,heding,status,key)

            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[5]/td[2]/input'
            heding="Units"
            status="Units Field Not Found"
            key='1'
            text_box(xpath,heding,status,key)
            
            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[2]/td[4]/span/input'
            heding="Place Of Service"
            status="Place Of Service Field Not Found"
            if len(str(pos))==1:
                pos='0' + str(pos)
            key=str(pos)
            text_box(xpath,heding,status,key)

            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[3]/td[4]/span/input'
            heding="Procedure Code"
            status="Procedure Code Field Not Found"
            key=str(pod_cod)
            text_box(xpath,heding,status,key)

            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[8]/td[2]/input'
            heding="Rendering Provider"
            status="Rendering Provider Field Not Found"
            key=str(ren_id)
            text_box(xpath,heding,status,key)            

            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[7]/td[4]/select[1]/option"
            heding="Diagnosis Code Pointer"
            status="Diagnosis Code Pointer 1 Count Not Found"        
            count(xpath,heding,status) 
            
            if pd.isnull(dx1):
                pass
            else:
                j=1
                while j<rows+1:                       
                    xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[7]/td[4]/select[1]/option[{}]"
                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                    if cnm.lstrip().rstrip()=="01":
                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                                 
                        break                       
                    j=j+1       

            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[7]/td[4]/select[2]/option"
            heding="Diagnosis Code Pointer"
            status="Diagnosis Code Pointer 2 Count Not Found"        
            count(xpath,heding,status) 
            
            if pd.isnull(dx2):
                pass
            else:
                j=1
                while j<rows+1:                       
                    xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[7]/td[4]/select[2]/option[{}]"
                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                    if cnm.lstrip().rstrip()=="02":
                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                                 
                        break                       
                    j=j+1    

            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[7]/td[4]/select[3]/option"
            heding="Diagnosis Code Pointer"
            status="Diagnosis Code Pointer 3 Count Not Found"        
            count(xpath,heding,status) 
            
            if pd.isnull(dx3):
                pass
            else:
                j=1
                while j<rows+1:                       
                    xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[7]/td[4]/select[3]/option[{}]"
                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                    if cnm.lstrip().rstrip()=="03":
                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                                 
                        break                       
                    j=j+1   

            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[7]/td[4]/select[4]/option"
            heding="Diagnosis Code Pointer"
            status="Diagnosis Code Pointer 4 Count Not Found"        
            count(xpath,heding,status) 
            
            if pd.isnull(dx4):
                pass
            else:
                j=1
                while j<rows+1:                       
                    xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[7]/td[4]/select[4]/option[{}]"
                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                    if cnm.lstrip().rstrip()=="04":
                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                                 
                        break                       
                    j=j+1   

            if pd.isnull(mod1):
                pass
            else:
                xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[8]/td[4]/span[1]/input'
                heding="Modifiers1"
                status="Modifiers1 Field Not Found"
                key=str(mod1)
                text_box(xpath,heding,status,key)

            if pd.isnull(mod2):
                pass
            else:
                xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[8]/td[4]/span[3]/input'
                heding="Modifiers2"
                status="Modifiers2 Field Not Found"
                key=str(mod2)
                text_box(xpath,heding,status,key)

            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[6]/td[2]/input'
            heding="Charges"
            status="Charges Field Not Found"
            key=str(amt)
            text_box(xpath,heding,status,key)
                        
            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/span[2]/table/tbody/tr/td[4]/table/tbody/tr/td[1]/a"
            heding="Submit"
            status="Submit Button Not Found"
            click(xpath,heding,status)

            wait = WebDriverWait(driver, 20)
            wait.until(lambda driver: driver.execute_script("return document.readyState === 'complete';"))

            try:                                                            
                xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[6]/span/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[1]/td[2]/input"
                cs=WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH,xpath)))  
                cs = cs.get_attribute("value") 
                
                wb1=load_workbook(filename=fil)
                sheet = wb1['Primary & Secondary Template']
                column_letter = 'AC'
                column_cells = sheet[column_letter]
                last_row = None
                for cell in reversed(column_cells):
                    if cell.value:
                        last_row = cell.row
                        break                            
                sheet['Z' + str(int(last_row + 1))]=cs

                if cs=='DENIED':
                    xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[6]/table/tbody/tr"
                    heding="EOB Information"
                    status="EOB Information Count Not Found"        
                    count(xpath,heding,status) 

                    j=3
                    while j<rows+1:                       
                        xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[6]/table/tbody/tr[{}]/td[2]"
                        cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                        if cnm.lstrip().rstrip()=='DENIED':
                            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[6]/table/tbody/tr[{}]/td[5]"
                            carc=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                                     
                            
                            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[6]/table/tbody/tr[{}]/td[7]"
                            carc_dec=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                                     
                            
                            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[6]/table/tbody/tr[{}]/td[8]"
                            rarc=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                                     

                            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[6]/table/tbody/tr[{}]/td[9]"
                            rarc_dec=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                                     

                            sheet['AD' + str(int(last_row + 1))]=carc
                            sheet['AE' + str(int(last_row + 1))]=carc_dec
                            sheet['AF' + str(int(last_row + 1))]=rarc
                            sheet['AG' + str(int(last_row + 1))]=rarc_dec
                            break                       
                        j=j+1  
                
                xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[6]/span/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[2]/td[2]/input"
                c_icn=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath)))
                c_icn = c_icn.get_attribute("value") 
                sheet['AA' + str(int(last_row + 1))]=c_icn

                xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[6]/span/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[5]/td[2]/input"
                p_amt=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath)))
                p_amt = p_amt.get_attribute("value") 
                sheet['AB' + str(int(last_row + 1))]=p_amt
                                                
                sheet['AC' + str(int(last_row + 1))]='Done'
                wb1.save(fil)
                wb1.close()    
            except Exception as e:
                wb1=load_workbook(filename=fil)
                sheet = wb1['Primary & Secondary Template']
                column_letter = 'AC'  
                column_cells = sheet[column_letter]
                last_row = None
                for cell in reversed(column_cells):
                    if cell.value:
                        last_row = cell.row
                        break                            
                sheet['AC' + str(int(last_row + 1))]='Error'                
                wb1.save(fil)
                wb1.close()   
                
                xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/span[2]/table/tbody/tr/td[4]/table/tbody/tr/td[2]/a"
                heding="Cancel"
                status="Cancel Button Not Found"
                click(xpath,heding,status)
                
                Alert()                    
        
        driver.quit()
        messagebox.showinfo("Process Status", "Process Completed")
        sys.exit(0)                    

def secondary():        
        global element_1    
        global ans1
        global ans2        
        global ans4
        global driver_path
        global j
        global i       
        
        file_pick()
                    
        fil=ans       
        
        try:
            options = Options()            
            options.add_argument('--ignore-certificate-errors')
            options.add_argument('--ignore-ssl-errors')
            options.add_argument("--disable-popup-blocking")
            driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options)
            driver.maximize_window()
            driver.get('https://ohid.ohio.gov/wps/portal/gov/ohid/login')
        except Exception as e:                        
            try:
                options =  webdriver.ChromeOptions()        
                options.add_argument('--ignore-certificate-errors')
                options.add_argument('--ignore-ssl-errors')
                options.add_argument("--disable-popup-blocking")
                # response = requests.get('https://chromedriver.storage.googleapis.com/LATEST_RELEASE')
                # latest_version = response.text.strip()
                # chrome_driver_url = f'https://chromedriver.storage.googleapis.com/{latest_version}/chromedriver_win32.zip'
                # response = requests.get(chrome_driver_url)
                # with open('chromedriver_win32.zip', 'wb') as f:
                #     f.write(response.content)
                # with ZipFile('chromedriver_win32.zip', 'r') as zip_ref:
                #     zip_ref.extractall('.')   
                # driver_path=r'C:\Users\sanandrao\Desktop\New folder\chromedriver.exe'
                driver_path = os.path.abspath('chromedriver.exe')                
                driver = webdriver.Chrome(executable_path=driver_path)                
                driver.maximize_window()
                driver.get('https://ohid.ohio.gov/wps/portal/gov/ohid/login')                                         
            except Exception as e:
                messagebox.showinfo("Driver Problem","Pls Check Your Chrome Version Driver")
                sys.exit(0)
                        
        def text_box(xpath,heding,status,key):                
            counter = 0
            while counter < 15:
                try:   
                    WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()                   
                    WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding, status)
                sys.exit(0)        
        
        def text_box(xpath,heding,status,key):                
            counter = 0
            while counter < 15:
                try:   
                    WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()                   
                    WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                    WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(Keys.TAB)
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding, status)
                sys.exit(0)    
        
        def text_box_key(xpath,heding,status,key):                
            counter = 0
            while counter < 15:
                try:   
                    element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, xpath)))
                    element.send_keys(key)
                    element.send_keys(Keys.TAB)
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding, status)
                sys.exit(0)    


        def click(xpath,heding,status):
            counter = 0
            while counter < 15:
                try:             
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding,status)
                sys.exit(0)                                

        def count(xpath,heding,status):
            global rows
            counter = 0
            while counter < 15:
                try:             
                    rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))                    
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding,status)
                sys.exit(0)                         

        def count_1(xpath,heding,status):
            global rows_1
            counter = 0
            while counter < 15:
                try:             
                    rows_1=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))                    
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding,status)
                sys.exit(0)   

        def Alert():
            counter = 0
            while counter < 15:
                try:             
                    WebDriverWait(driver, 0).until (EC.alert_is_present())
                    a=driver.switch_to.alert
                    a.accept()
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo('Alert','Alert Not Present')
        
        # xpath= '/html/body/div[3]/div/section/div[4]/div/div[2]/div/div[2]/div/div[1]/div/div[1]/div/div[2]/div[2]/div[2]/main/article/section[1]/div[1]/div[2]/div[1]/form/div/div/div/div[5]/div[1]/input'
        # heding="User Name"
        # status="User Name Field Not Found"
        # key=user_name
        # text_box(xpath,heding,status,key)
        
        # xpath= '/html/body/div[3]/div/section/div[4]/div/div[2]/div/div[2]/div/div[1]/div/div[1]/div/div[2]/div[2]/div[2]/main/article/section[1]/div[1]/div[2]/div[1]/form/div/div/div/div[6]/div[1]/input'
        # heding="Password"
        # status="Password Field Not Found"
        # key=password
        # text_box(xpath,heding,status,key)

        # xpath= "/html/body/div[3]/div/section/div[4]/div/div[2]/div/div[2]/div/div[1]/div/div[1]/div/div[2]/div[2]/div[2]/main/article/section[1]/div[1]/div[2]/div[1]/form/div/div/div/div[7]/button"
        # heding="Log in"
        # status="Log in Button Not Found"
        # click(xpath,heding,status)
        
        # while True:
        #     xpath='/html/body/div[3]/div/section/div[4]/div/div[2]/div/div[2]/div/div[1]/div/div[1]/div/div[2]/div[2]/div[2]/main/article/section[1]/div[1]/div[2]/div[1]/form/div/div/div/div[4]/div/p'         
        #     try:
        #         element_3 = WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH, xpath))).text
        #         if element_3.lstrip().rstrip()=="We didn't recognize the username or password you entered. Forgot your":                               
        #             messagebox.showinfo('Login Status','Please Check - UserName or Password Wrong')                    
        #             user_pass()
                    
        #             xpath= '/html/body/div[3]/div/section/div[4]/div/div[2]/div/div[2]/div/div[1]/div/div[1]/div/div[2]/div[2]/div[2]/main/article/section[1]/div[1]/div[2]/div[1]/form/div/div/div/div[5]/div[1]/input'
        #             heding="User Name"
        #             status="User Name Field Not Found"
        #             key=user_name
        #             text_box(xpath,heding,status,key)
                    
        #             xpath= '/html/body/div[3]/div/section/div[4]/div/div[2]/div/div[2]/div/div[1]/div/div[1]/div/div[2]/div[2]/div[2]/main/article/section[1]/div[1]/div[2]/div[1]/form/div/div/div/div[6]/div[1]/input'
        #             heding="Password"
        #             status="Password Field Not Found"
        #             key=password
        #             text_box(xpath,heding,status,key)

        #             xpath= "/html/body/div[3]/div/section/div[4]/div/div[2]/div/div[2]/div/div[1]/div/div[1]/div/div[2]/div[2]/div[2]/main/article/section[1]/div[1]/div[2]/div[1]/form/div/div/div/div[7]/button"
        #             heding="Log in"
        #             status="Log in Button Not Found"
        #             click(xpath,heding,status)
        #         else:
        #             break    
        #     except Exception as e:                
        #             break                    

        messagebox.showinfo('Waiting','Authentication Waiting')
        
        driver.switch_to.window(driver.window_handles[1])           
                       
        file=pd.read_excel(fil,sheet_name='Primary & Secondary Template',header=0)        

        for index, row in file.iterrows():                          
            bil_num = row[0]  
            add_zer=row[1]
            dob = row[2]            
            pat_acn = row[3]                 
            ren_id=row[4]                                    
            dx1=row[5]
            dx2=row[6]
            dx3=row[7]
            dx4=row[8]
            dos=row[9]
            pos=row[10]
            pod_cod=row[11]
            mod1=row[12]
            mod2=row[13]
            amt=row[14]
            cfi=row[15]
            cp_amt=row[16]
            cp_dt=row[17]
            icnm=row[18]
            cas_gc=row[19]
            arc=row[20]
            c_amt=row[21]
            cas_gc_1=row[22]
            arc_1=row[23]
            c_amt1=row[24]

            wait = WebDriverWait(driver, 20)
            wait.until(lambda driver: driver.execute_script("return document.readyState === 'complete';"))

            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[1]/td/div[1]/ul/li"
            heding="Claims"
            status="Claims Tab Count Not Found"        
            count(xpath,heding,status)
            
            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[1]/td/div[1]/ul/li[{}]"

            j=1
            while j<rows+1:                                
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text            
                if cnm.lstrip().rstrip()=="Claims":
                    
                    time.sleep(3)

                    element=WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j))))
                    actions = ActionChains(driver)
                    actions.move_to_element(element).perform()  

                    xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[1]/td/div[1]/ul/li[{}]/ul/li".format(j)
                    heding="Claims Count"
                    status="Claims Count Not Found"        
                    count_1(xpath,heding,status)                
                        
                    i=1
                    while i<rows_1+1:                                       
                        xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[1]/td/div[1]/ul/li[{}]/ul/li[{}]"
                        element_7=WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j,i))))
                        actions = ActionChains(driver)
                        actions.move_to_element(element_7).perform()  
                        element_text = element_7.text                    
                        if element_text.lstrip().rstrip()=="Professional":
                            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j,i)))).click()  
                            break                       
                        i=i+1                                                              
                    break
                j=j+1                            
            
            if pd.isnull(add_zer):
                bil_num = str(bil_num)
            elif add_zer.lower().lstrip().rstrip()=='add':
                bil_num ='0' + str(bil_num)
            else:
                bil_num = str(bil_num)
                
            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/span/table/tbody/tr/td/table/tbody/tr[6]/td[2]/input'
            heding="Medical Billing Number"
            status="Medical Billing Number Field Not Found"
            key=bil_num
            text_box(xpath,heding,status,key)
            
            date_object_1 = datetime.strptime(str(dob), "%Y-%m-%d %H:%M:%S")
            dob3 = date_object_1.strftime("%m/%d/%Y")
            
            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/span/table/tbody/tr/td/table/tbody/tr[7]/td[2]/input'
            heding="Date of Birth"
            status="Date of Birth Field Not Found"
            key=dob3
            text_box_key(xpath,heding,status,key)
            
            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/span/table/tbody/tr/td/table/tbody/tr[19]/td[2]/select/option"
            heding="Medicare Assignment"
            status="Medicare Assignment Count Not Found"        
            count(xpath,heding,status)   
            
            j=1
            while j<rows+1:                       
                xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/span/table/tbody/tr/td/table/tbody/tr[19]/td[2]/select/option[{}]"
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                if cnm.lstrip().rstrip()=="ASSIGNED":
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()  
                    break                       
                j=j+1    
            
            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/span/table/tbody/tr/td/table/tbody/tr[2]/td[4]/select/option"
            heding="Release of Information"
            status="Release of Information Count Not Found"        
            count(xpath,heding,status)   
            
            j=1
            while j<rows+1:                       
                xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/span/table/tbody/tr/td/table/tbody/tr[2]/td[4]/select/option[{}]"
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                if cnm.lstrip().rstrip()=="SIGNED STMT PERMITTING RELEASE":
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()  
                    break                       
                j=j+1   

            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/span/table/tbody/tr/td/table/tbody/tr[5]/td[4]/select/option"
            heding="Signature Source"
            status="Signature Source Count Not Found"        
            count(xpath,heding,status)   
            
            j=1
            while j<rows+1:                       
                xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/span/table/tbody/tr/td/table/tbody/tr[5]/td[4]/select/option[{}]"
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                if cnm.lstrip().rstrip()=="SIGNED HCFA-1500 ON FILE":
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()  
                    break                       
                j=j+1   

            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/span/table/tbody/tr/td/table/tbody/tr[10]/td[2]/input'
            heding="Patient Account"
            status="Patient Account Field Not Found"
            key=str(pat_acn)
            text_box(xpath,heding,status,key)

            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/span/table/tbody/tr/td/table/tbody/tr[18]/td[2]/input'
            heding="Rendering ID"
            status="Rendering ID Field Not Found"
            key=str(ren_id)
            text_box(xpath,heding,status,key)        

            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[2]/td[2]/select/option"
            heding="Sequence"
            status="Sequence Count Not Found"        
            count(xpath,heding,status) 

            if pd.isnull(dx1):
                pass
            else:
                j=1
                while j<rows+1:                       
                    xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[2]/td[2]/select/option[{}]"
                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                    if cnm.lstrip().rstrip()=="01":
                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()  
                        
                        xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[2]/td[4]/span/input'
                        heding="Diagnosis Code - 01"
                        status="Diagnosis Code - 01 Field Not Found"
                        key=str(dx1)
                        text_box(xpath,heding,status,key)

                        break                       
                    j=j+1                                   
                
            if pd.isnull(dx2):
                pass
            else:                
                j=1
                while j<rows+1:                       
                    xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[2]/td[2]/select/option[{}]"
                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                    if cnm.lstrip().rstrip()=="02":
                        xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[1]/td[1]/table/tbody/tr/td[2]/a"
                        heding="Diagnosis add in item"
                        status="Diagnosis add in item Button Not Found"
                        click(xpath,heding,status)

                        xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[2]/td[2]/select/option[{}]"
                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click() 

                        xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[2]/td[4]/span/input'
                        heding="Diagnosis Code - 02"
                        status="Diagnosis Code - 02 Field Not Found"
                        key=str(dx2)
                        text_box(xpath,heding,status,key)

                        break                       
                    j=j+1   
            
            if pd.isnull(dx3):
                pass
            else:                
                j=1
                while j<rows+1:                       
                    xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[2]/td[2]/select/option[{}]"
                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                    if cnm.lstrip().rstrip()=="03":
                        xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[1]/td[1]/table/tbody/tr/td[2]/a"
                        heding="Diagnosis add in item"
                        status="Diagnosis add in item Button Not Found"
                        click(xpath,heding,status)

                        xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[2]/td[2]/select/option[{}]"
                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click() 

                        xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[2]/td[4]/span/input'
                        heding="Diagnosis Code - 03"
                        status="Diagnosis Code - 03 Field Not Found"
                        key=str(dx3)
                        text_box(xpath,heding,status,key)

                        break                       
                    j=j+1   
            
            if pd.isnull(dx4):
                pass
            else:                
                j=1
                while j<rows+1:                       
                    xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[2]/td[2]/select/option[{}]"
                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                    if cnm.lstrip().rstrip()=="04":
                        xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[1]/td[1]/table/tbody/tr/td[2]/a"
                        heding="Diagnosis add in item"
                        status="Diagnosis add in item Button Not Found"
                        click(xpath,heding,status)

                        xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[2]/td[2]/select/option[{}]"
                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click() 

                        xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[2]/td[4]/span/input'
                        heding="Diagnosis Code - 04"
                        status="Diagnosis Code - 04 Field Not Found"
                        key=str(dx4)
                        text_box(xpath,heding,status,key)

                        break                       
                    j=j+1                           

            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[2]/span/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr/td[2]/a"
            heding="Header add in item"
            status="Header add in item Button Not Found"
            click(xpath,heding,status)
            
            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[2]/span/table/tbody/tr/td/table/tbody/tr[2]/td[2]/select/option"
            heding="Header - Other Payer"
            status="Claim Filing Indicator Count Not Found"        
            count(xpath,heding,status)
            
            if pd.isnull(cfi):
                cfi=''
                
            j=1
            while j<rows+1:  
                xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[2]/span/table/tbody/tr/td/table/tbody/tr[2]/td[2]/select/option[{}]"
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                if cnm.lstrip().rstrip()==cfi.lstrip().rstrip():                                        
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click() 
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,"/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[2]/span/table/tbody/tr/td/table/tbody/tr[2]/td[2]/select"))).send_keys(Keys.TAB)
                    break
                j=j+1
                  
            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[2]/span/table/tbody/tr/td/table/tbody/tr[3]/td[2]/select/option"
            heding="Header - Other Payer"
            status="Policy Holder Relationship to Insured Count Not Found"        
            count(xpath,heding,status)

            j=1
            while j<rows+1:                    
                xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[2]/span/table/tbody/tr/td/table/tbody/tr[3]/td[2]/select/option[{}]"
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                if cnm.lstrip().rstrip()=='SELF':                                        
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,"/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[2]/span/table/tbody/tr/td/table/tbody/tr[3]/td[2]/select"))).send_keys(Keys.TAB)                 
                    break                       
                j=j+1   

            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[2]/span/table/tbody/tr/td/table/tbody/tr[2]/td[4]/input'
            heding="Header - Other Payer"
            status="Insurance Carrier Name Field Not Found"
            key=icnm
            text_box(xpath,heding,status,key)

            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[2]/span/table/tbody/tr/td/table/tbody/tr[3]/td[4]/input'
            heding="Header - Other Payer"
            status="Electronic Payer ID Field Not Found"
            key='NONE'
            text_box(xpath,heding,status,key)                        

            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[2]/span/table/tbody/tr/td/table/tbody/tr[8]/td[2]/input'
            heding="Header - Other Payer"
            status="Paid Amount Field Not Found"
            key=cp_amt
            text_box(xpath,heding,status,key)

            date_object_1 = datetime.strptime(str(cp_dt), "%Y-%m-%d %H:%M:%S")
            dob3 = date_object_1.strftime("%m/%d/%Y")

            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[2]/span/table/tbody/tr/td/table/tbody/tr[9]/td[2]/input'
            heding="Header - Other Payer"
            status="Paid Date Field Not Found"
            key=dob3
            text_box(xpath,heding,status,key)

            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[2]/span/table/tbody/tr/td/table/tbody/tr[5]/td[4]/select/option"
            heding="Header - Other Payer"
            status="Payer Sequence Count Not Found"        
            count(xpath,heding,status)

            j=1
            while j<rows+1:                       
                xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[2]/span/table/tbody/tr/td/table/tbody/tr[5]/td[4]/select/option[{}]"
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                if cnm.lstrip().rstrip()=='PRIMARY':                                        
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click() 
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,"/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[2]/span/table/tbody/tr/td/table/tbody/tr[5]/td[4]/select"))).send_keys(Keys.TAB)                      
                    break                       
                j=j+1   

            date_object_1 = datetime.strptime(str(dos), "%Y-%m-%d %H:%M:%S")
            dob3 = date_object_1.strftime("%m/%d/%Y")
            
            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[3]/td[2]/input'
            heding="From DOS 1"
            status="From DOS Field Not Found"
            key=dob3
            text_box(xpath,heding,status,key)

            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[3]/td[2]/input'
            heding="From DOS 2"
            status="From DOS Field Not Found"
            key=dob3
            text_box(xpath,heding,status,key)

            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[5]/td[2]/input'
            heding="Units"
            status="Units Field Not Found"
            key='1'
            text_box(xpath,heding,status,key)
            
            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[2]/td[4]/span/input'
            heding="Place Of Service"
            status="Place Of Service Field Not Found"
            if len(str(pos))==1:
                pos='0' + str(pos)
            key=str(pos)
            text_box(xpath,heding,status,key)

            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[3]/td[4]/span/input'
            heding="Procedure Code"
            status="Procedure Code Field Not Found"
            key=str(pod_cod)
            text_box(xpath,heding,status,key)

            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[8]/td[2]/input'
            heding="Rendering Provider"
            status="Rendering Provider Field Not Found"
            key=str(ren_id)
            text_box(xpath,heding,status,key)            

            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[7]/td[4]/select[1]/option"
            heding="Diagnosis Code Pointer"
            status="Diagnosis Code Pointer 1 Count Not Found"        
            count(xpath,heding,status) 
            
            if pd.isnull(dx1):
                pass
            else:
                j=1
                while j<rows+1:                       
                    xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[7]/td[4]/select[1]/option[{}]"
                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                    if cnm.lstrip().rstrip()=="01":
                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                                 
                        break                       
                    j=j+1       

            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[7]/td[4]/select[2]/option"
            heding="Diagnosis Code Pointer"
            status="Diagnosis Code Pointer 2 Count Not Found"        
            count(xpath,heding,status) 
            
            if pd.isnull(dx2):
                pass
            else:
                j=1
                while j<rows+1:                       
                    xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[7]/td[4]/select[2]/option[{}]"
                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                    if cnm.lstrip().rstrip()=="02":
                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                                 
                        break                       
                    j=j+1    

            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[7]/td[4]/select[3]/option"
            heding="Diagnosis Code Pointer"
            status="Diagnosis Code Pointer 3 Count Not Found"        
            count(xpath,heding,status) 
            
            if pd.isnull(dx3):
                pass
            else:
                j=1
                while j<rows+1:                       
                    xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[7]/td[4]/select[3]/option[{}]"
                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                    if cnm.lstrip().rstrip()=="03":
                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                                 
                        break                       
                    j=j+1   

            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[7]/td[4]/select[4]/option"
            heding="Diagnosis Code Pointer"
            status="Diagnosis Code Pointer 4 Count Not Found"        
            count(xpath,heding,status) 
            
            if pd.isnull(dx4):
                pass
            else:
                j=1
                while j<rows+1:                       
                    xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[7]/td[4]/select[4]/option[{}]"
                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                    if cnm.lstrip().rstrip()=="04":
                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                                 
                        break                       
                    j=j+1   

            if pd.isnull(mod1):
                pass
            else:
                xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[8]/td[4]/span[1]/input'
                heding="Modifiers1"
                status="Modifiers1 Field Not Found"
                key=str(mod1)
                text_box(xpath,heding,status,key)

            if pd.isnull(mod2):
                pass
            else:
                xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[8]/td[4]/span[3]/input'
                heding="Modifiers2"
                status="Modifiers2 Field Not Found"
                key=str(mod2)
                text_box(xpath,heding,status,key)

            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[6]/td[2]/input'
            heding="Charges"
            status="Charges Field Not Found"
            key=str(amt)
            text_box(xpath,heding,status,key)                        
            
            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/div/table/tbody/tr/td[3]/a"
            heding="Detail - Other Payer"
            status="Detail - Other Payer Button Not Found"
            click(xpath,heding,status)
            
            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/div[2]/span/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[2]/a"
            heding="Detail - Other Payer"
            status="Add an item Button Not Found"
            click(xpath,heding,status)

            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/div[2]/span/table/tbody/tr/td/table/tbody/tr[4]/td[2]/select/option"
            heding="Detail - Other Payer"
            status="Detail Item 1 Count Not Found"        
            count(xpath,heding,status) 
            
            j=1
            while j<rows+1:                       
                xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/div[2]/span/table/tbody/tr/td/table/tbody/tr[4]/td[2]/select/option[{}]"
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                if cnm.lstrip().rstrip()=="1":
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                                 
                    break                       
                j=j+1   

            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/div[2]/span/table/tbody/tr/td/table/tbody/tr[5]/td[2]/select/option"
            heding="Detail - Other Payer"
            status="Electronic Payer ID 1 Count Not Found"        
            count(xpath,heding,status) 
            
            j=1
            while j<rows+1:                       
                xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/div[2]/span/table/tbody/tr/td/table/tbody/tr[5]/td[2]/select/option[{}]"
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                if cnm.lstrip().rstrip()=="NONE":
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                                 
                    break                       
                j=j+1  

            date_object_1 = datetime.strptime(str(cp_dt), "%Y-%m-%d %H:%M:%S")
            dob3 = date_object_1.strftime("%m/%d/%Y")

            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/div[2]/span/table/tbody/tr/td/table/tbody/tr[6]/td[2]/input'
            heding="Detail - Other Payer"
            status="Paid Date Field Not Found"
            key=dob3
            text_box(xpath,heding,status,key)

            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/div[2]/span/table/tbody/tr/td/table/tbody/tr[7]/td[2]/input'
            heding="Detail - Other Payer"
            status="Paid Amount Field Not Found"
            key=str(cp_amt)
            text_box(xpath,heding,status,key)          

            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/div[2]/div[2]/table/tbody/tr/td/a"
            heding="Detail - Other Payer Amount"
            status="Detail - Other Payer Amount and Adjustment Reason Codes Button Not Found"
            click(xpath,heding,status)

            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/div[2]/div[3]/span/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr/td[2]/a"
            heding="Detail - Other Payer Amount"
            status="Add An Item Button Not Found"
            click(xpath,heding,status)

            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/div[2]/div[3]/span/table/tbody/tr/td/table/tbody/tr[2]/td[4]/select/option"
            heding="Detail - Other Payer Amount and Adjustment Reason Codes 1"
            status="Detail item/electronic Payer ID Count Not Found"        
            count(xpath,heding,status) 
            
            j=1
            while j<rows+1:                       
                xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/div[2]/div[3]/span/table/tbody/tr/td/table/tbody/tr[2]/td[4]/select/option[{}]"
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                if cnm.lstrip().rstrip()=="1/NONE":
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                                 
                    break                       
                j=j+1  

            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/div[2]/div[3]/span/table/tbody/tr/td/table/tbody/tr[3]/td[4]/select/option"
            heding="Detail - Other Payer Amount and Adjustment Reason Codes 1"
            status="CAS Group Code Count Not Found"        
            count(xpath,heding,status) 
            
            j=1
            while j<rows+1:                       
                xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/div[2]/div[3]/span/table/tbody/tr/td/table/tbody/tr[3]/td[4]/select/option[{}]"
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                if cnm.lstrip().rstrip()==cas_gc.lstrip().rstrip():
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                                 
                    break                       
                j=j+1  

            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/div[2]/div[3]/span/table/tbody/tr/td/table/tbody/tr[4]/td[4]/input'
            heding="Detail - Other Payer Amount and Adjustment Reason Codes 1"
            status="ARC 1 Field Not Found"
            key=str(arc)
            text_box(xpath,heding,status,key)        

            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/div[2]/div[3]/span/table/tbody/tr/td/table/tbody/tr[5]/td[4]/input'
            heding="Detail - Other Payer Amount and Adjustment Reason Codes 1"
            status="Amount Field Not Found"
            key=str(c_amt)
            text_box(xpath,heding,status,key)               

            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/div[2]/div[3]/span/table/tbody/tr/td/table/tbody/tr[1]/td[1]/table/tbody/tr/td[2]/a"
            heding="Detail - Other Payer Amount and Adjustment Reason Codes 2"
            status="Add An Item Button Not Found"
            click(xpath,heding,status)
            
            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/div[2]/div[3]/span/table/tbody/tr/td/table/tbody/tr[2]/td[4]/select/option"
            heding="Detail - Other Payer Amount and Adjustment Reason Codes 2"
            status="Detail item/electronic Payer ID Count Not Found"        
            count(xpath,heding,status) 

            j=1
            while j<rows+1:                       
                xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/div[2]/div[3]/span/table/tbody/tr/td/table/tbody/tr[2]/td[4]/select/option[{}]"
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                if cnm.lstrip().rstrip()=="1/NONE":
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                                 
                    break                       
                j=j+1 

            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/div[2]/div[3]/span/table/tbody/tr/td/table/tbody/tr[3]/td[4]/select/option"
            heding="Detail - Other Payer Amount and Adjustment Reason Codes 2"
            status="CAS Group Code Count Not Found"        
            count(xpath,heding,status) 
            
            j=1
            while j<rows+1:                       
                xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/div[2]/div[3]/span/table/tbody/tr/td/table/tbody/tr[3]/td[4]/select/option[{}]"
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                if cnm.lstrip().rstrip()==cas_gc_1.lstrip().rstrip():
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                                 
                    break                       
                j=j+1  

            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/div[2]/div[3]/span/table/tbody/tr/td/table/tbody/tr[4]/td[4]/input'
            heding="Detail - Other Payer Amount and Adjustment Reason Codes 2"
            status="ARC 1 Field Not Found"
            key=str(arc_1)
            text_box(xpath,heding,status,key)        

            xpath= '/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/div[2]/div[3]/span/table/tbody/tr/td/table/tbody/tr[5]/td[4]/input'
            heding="Detail - Other Payer Amount and Adjustment Reason Codes 2"
            status="Amount Field Not Found"
            key=str( c_amt1)
            text_box(xpath,heding,status,key)              

            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/span[2]/table/tbody/tr/td[4]/table/tbody/tr/td[1]/a"
            heding="Submit"
            status="Submit Button Not Found"
            click(xpath,heding,status)

            wait = WebDriverWait(driver, 20)
            wait.until(lambda driver: driver.execute_script("return document.readyState === 'complete';"))

            try:                 
                xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[6]/span/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[1]/td[2]/input"
                cs=WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH,xpath)))  
                cs = cs.get_attribute("value")                 

                wb1=load_workbook(filename=fil)
                sheet = wb1['Primary & Secondary Template']
                column_letter = 'AC'  
                column_cells = sheet[column_letter]
                last_row = None
                for cell in reversed(column_cells):
                    if cell.value:
                        last_row = cell.row
                        break    
                
                sheet['Z' + str(int(last_row + 1))]=cs

                if cs=='DENIED':
                    xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[6]/table/tbody/tr"
                    heding="EOB Information"
                    status="EOB Information Count Not Found"        
                    count(xpath,heding,status) 

                    j=3
                    while j<rows+1:                       
                        xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[6]/table/tbody/tr[{}]/td[2]"
                        cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                     
                        if cnm.lstrip().rstrip()=='DENIED':
                            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[6]/table/tbody/tr[{}]/td[5]"
                            carc=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                                     
                            
                            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[6]/table/tbody/tr[{}]/td[7]"
                            carc_dec=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                                     
                            
                            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[6]/table/tbody/tr[{}]/td[8]"
                            rarc=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                                     

                            xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[6]/table/tbody/tr[{}]/td[9]"
                            rarc_dec=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                                     

                            sheet['AD' + str(int(last_row + 1))]=carc
                            sheet['AE' + str(int(last_row + 1))]=carc_dec
                            sheet['AF' + str(int(last_row + 1))]=rarc
                            sheet['AG' + str(int(last_row + 1))]=rarc_dec
                            break                       
                        j=j+1  

                xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[6]/span/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[2]/td[2]/input"
                c_icn=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath)))
                c_icn = c_icn.get_attribute("value") 
                sheet['AA' + str(int(last_row + 1))]=c_icn

                xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[6]/span/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[5]/td[2]/input"
                p_amt=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath)))
                p_amt = p_amt.get_attribute("value") 
                sheet['AB' + str(int(last_row + 1))]=p_amt                                                                                        
                
                sheet['AC' + str(int(last_row + 1))]='Done'
                wb1.save(fil)
                wb1.close()    
            except Exception as e:
                wb1=load_workbook(filename=fil)
                sheet = wb1['Primary & Secondary Template']
                column_letter = 'AC'  
                column_cells = sheet[column_letter]
                last_row = None
                for cell in reversed(column_cells):
                    if cell.value:
                        last_row = cell.row
                        break                            
                sheet['AC' + str(int(last_row + 1))]='Error'                
                wb1.save(fil)
                wb1.close()  

                xpath= "/html/body/form/div[3]/table/tbody/tr/td[2]/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/div/div/div[2]/div/div/span[2]/table/tbody/tr/td[4]/table/tbody/tr/td[2]/a"
                heding="Cancel"
                status="Cancel Button Not Found"
                click(xpath,heding,status)
                
                Alert()            

        driver.quit()
        messagebox.showinfo("Process Status", "Process Completed")
        sys.exit(0)                    

if __name__=="__main__":        
    radio()
   