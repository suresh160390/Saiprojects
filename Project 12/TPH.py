from tkinter import BOTH, LEFT, TOP, Button, Entry, Frame, Label, PhotoImage, StringVar, Tk,Radiobutton,StringVar,IntVar,filedialog
from idlelib.tooltip import Hovertip
from tkinter import messagebox
import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter,column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
import os
import shutil
import warnings

import socket
import os
import getpass

warnings.filterwarnings("ignore")


def folder():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
        image_path2 = os.path.join(sys._MEIPASS, 'Static', 'Mapping.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')
        image_path2 = os.path.join(os.getcwd(), 'Static', 'Mapping.png')

    root.title("Process Folder Picker")
    root.resizable(False,False)

    root.title("Process Folder Picker")
    root.resizable(False,False)
   
    w = 500
    h = 160
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Browse Process - Folder Picker",font=("Calibri",18,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
          
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=50)
    
    answer=StringVar()
    answer.set("")
       
    def browse_button():
        global filename
        answer.set("")  
        filename = filedialog.askdirectory()  
        txt.delete(0, 'end') 
        txt.insert(0, filename)

    def Click_Done():
        global ans        
        
        ans=txt.get()        
        
        if ans=="":
           answer.set("Folder Path Fields Empty Is Not Allowed...")        
        else:      
            root.destroy()          
            return ans
        
    txt=Entry(Frame2,font=("Calibri",12,"bold","italic"),width=50,justify="left")
    txt.grid(row=0,column=0,padx=5,pady=10,sticky="E")    

    photo1 = PhotoImage(file=image_path2)
            
    btn1=Button(Frame2,text="Browse",command=browse_button,image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=0,column=1,padx=3,pady=0,sticky="W")     
                
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=3,y=80,width=490,height=25)
    
    title_3=Label(Frame3,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=1,padx=150,pady=0,sticky="E")

    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=3,y=105,width=500,height=50)
    
    photo2 = PhotoImage(file=image_path1)

    btn2=Button(Frame4,command=Click_Done,text="Done",image=photo2,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=3,column=0,padx=125,pady=0,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo3 = PhotoImage(file=image_path)   

    btn3=Button(Frame4,command=Close,text="Exit",image=photo3,borderwidth=0,bg="#2c3e50")
    btn3.grid(row=3,column=1,padx=15,pady=0,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn2,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn3,'Click to Exit Process',hover_delay=1000)
    myTip2 = Hovertip(btn1,'Click to Pick Folder',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def file_pick():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
        image_path2 = os.path.join(sys._MEIPASS, 'Static', 'Mapping.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')
        image_path2 = os.path.join(os.getcwd(), 'Static', 'Mapping.png')

    root.title("Mapping Header - File Picker")
    root.resizable(False,False)
   
    w = 500
    h = 160
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Browse Mapping Header - File Picker",font=("Calibri",18,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
          
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=50)
    
    answer=StringVar()
    answer.set("")
       
    def browse_button():
        global filename
        answer.set("")  
        filename = filedialog.askopenfilename()  
        txt.delete(0, 'end') 
        txt.insert(0, filename)

    def Click_Done():
        global ans1        
        
        ans1=txt.get()        
        
        if ans1=="":
           answer.set("File Path Fields Empty Is Not Allowed...")        
        else:      
            root.destroy()          
            return ans1
        
    txt=Entry(Frame2,font=("Calibri",12,"bold","italic"),width=50,justify="left")
    txt.grid(row=0,column=0,padx=5,pady=10,sticky="E")    

    photo1 = PhotoImage(file=image_path2)
            
    btn1=Button(Frame2,text="Browse",command=browse_button,image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=0,column=1,padx=3,pady=0,sticky="W")     
                
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=3,y=80,width=490,height=25)
    
    title_3=Label(Frame3,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=1,padx=150,pady=0,sticky="E")

    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=3,y=105,width=500,height=50)
    
    photo2 = PhotoImage(file=image_path1)

    btn2=Button(Frame4,command=Click_Done,text="Done",image=photo2,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=3,column=0,padx=125,pady=0,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo3 = PhotoImage(file=image_path)   

    btn3=Button(Frame4,command=Close,text="Exit",image=photo3,borderwidth=0,bg="#2c3e50")
    btn3.grid(row=3,column=1,padx=15,pady=0,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn2,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn3,'Click to Exit Process',hover_delay=1000)
    myTip2 = Hovertip(btn1,'Click to Pick File',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def check():
    folder()
    file_pick()

    folder_pic=ans + '/' 
    fil=ans1 

    files = [item for item in os.listdir(folder_pic) if os.path.isfile(os.path.join(folder_pic, item))]    
    
    mp_fl = pd.read_excel(fil,sheet_name='Data', engine='openpyxl')

    workbook = openpyxl.load_workbook(fil)
    worksheet = workbook["Check"]
    end_row = worksheet.max_row
    worksheet.delete_rows(2, end_row)
    workbook.save(fil)
    workbook.close()

    j=0
    while j < len(files):
        fl=files[j]

        rw_fl = pd.DataFrame()
        
        rw_fl = pd.read_excel(folder_pic + fl,engine='xlrd',header=5)
        lst=list(rw_fl.head())
        
        lst_df = pd.DataFrame(lst, columns=['Name'])       
        
        out = lst_df.merge(mp_fl.drop_duplicates(subset=['PCC Heading']), left_on='Name', right_on='PCC Heading', how='left')

        filt_nan = out[out['PCC Heading'].isna()]              

        filt_nan = filt_nan[['Name']]
        filt_nan['File Name']=fl 

        book = load_workbook(filename=fil)
        
        with pd.ExcelWriter(fil, engine='openpyxl') as writer:
            writer.book = book
            for ws in book.worksheets:
                writer.sheets[ws.title] = writer.book[ws.title]
            worksheet = writer.sheets['Check']
            startrow = worksheet.max_row if worksheet.max_row > 0 else None
            filt_nan.to_excel(writer, sheet_name='Check', index=False, startrow=startrow,header=False)
        
        print(fl + "   -   Done")
        j=j+1
    messagebox.showinfo('Process Status', 'Process Completed...')
    sys.exit(0) 

def console():
    folder()
    # file_pick()

    # map_fil=ans1 
    fil=ans + '/' 
    file_path = fil + 'Console'
    file_con = fil + 'Console' + '/'
    files = [item for item in os.listdir(fil) if os.path.isfile(os.path.join(fil, item))]    

    fld='Console'
        
    if not os.path.isdir(file_path):
        os.mkdir(fil + 'Console')        
    else:
        path1 = os.path.join(fil, fld)
        shutil.rmtree(path1)
        os.mkdir(fil + 'Console')
        
    wb = openpyxl.Workbook()
    ws=wb.active    
    ws['A1']='PCC Name'
    ws['B1']='Link'
    ws['C1']='User Name'
    ws['D1']='Password'
    ws['E1']='Final Status'
    ws['F1']='Previous address'
    ws['G1']='Postal/Zip Code'
    ws['H1']='City'
    ws['I1']='County'
    ws['J1']='Country'
    ws['K1']='Prov/State'
            
    ws.title = 'Data'
    wb.save(filename=file_con +  'Consol.xlsx')
    wb.close()
    
    # mp_fl = pd.read_excel(map_fil,sheet_name='Data', engine='openpyxl')

    # mp_fl = mp_fl[mp_fl['Required Header'] != 'Not Required']
    # mp_fl = pd.DataFrame(mp_fl).reset_index(drop=True)

    # print(mp_fl)

    j=0
           
    while j < len(files):
        fl=files[j]

        ex_fn = pd.DataFrame()
        df = pd.DataFrame()
        
        # ex_fn = pd.read_excel(fil + fl,engine='openpyxl',header=0)

        # ex_fn = pd.read_excel(fil + fl,engine='xlrd',header=0)
        # lst=list(ex_fn.head())
        # filtered_list = [item for item in lst if not item.startswith('Unnamed')]
        # fn=filtered_list[1]
        
        final_lst = []
        
        ex_fil = pd.read_excel(fil + fl,engine='openpyxl',header=1) 

        # ex_fil = pd.read_excel(fil + fl,engine='xlrd',header=5) 
        # Remove columns with all NaN values    
        ex_fil = ex_fil.dropna(axis=1, how='all')
        
        # Write Code ---------------------

        # new_df = pd.DataFrame()
        
        # for index, row in mp_fl.iterrows(): 
        #     org_hed = row[0] 
        #     req_hed = row[1] 

        #     if org_hed in ex_fil.columns:
        #         # Rename the column to 'Suresh-1'
        #         ex_fil.rename(columns={org_hed : req_hed}, inplace=True)
                
        #         new_df[req_hed] = ex_fil[req_hed]

        # print(new_df)
        
        # write code -----------------

        # ex_fil['PCC Name']=fn
        lst1 =list(ex_fil.head())

        con_fil = pd.read_excel(file_con + "Consol.xlsx",sheet_name='Data', engine='openpyxl')
        lst2 =list(con_fil.head())

        for item in lst2:
            if item not in final_lst:
                final_lst.append(item)
        
        for item in lst1:
            if item not in final_lst:
                final_lst.append(item)   
        
        book = load_workbook(filename=file_con + "Consol.xlsx")
        sheet = book['Data']

        for i, value in enumerate(final_lst, start=1):
            sheet.cell(row=1, column=i, value=value)
        
        book.save(filename=file_con + "Consol.xlsx")

        con = pd.read_excel(file_con + "Consol.xlsx",sheet_name='Data', engine='openpyxl')
        con.drop(index=con.index, inplace=True) 
        heading_list = con.columns.tolist()

        df = pd.concat([con, ex_fil], axis=0)
        df = df[heading_list]

        book = load_workbook(filename=file_con + "Consol.xlsx")                

        with pd.ExcelWriter(file_con + "Consol.xlsx", engine='openpyxl') as writer:
            writer.book = book
            for ws in book.worksheets:
                writer.sheets[ws.title] = writer.book[ws.title]
            worksheet = writer.sheets['Data']
            startrow = worksheet.max_row if worksheet.max_row > 0 else None
            df.to_excel(writer, sheet_name='Data', index=False, startrow=startrow,header=False)
        
        print(fl + "   -   Done")
        j=j+1
    
    messagebox.showinfo('Process Status', 'Process Completed...')
    sys.exit(0) 

def test():
    hostname = socket.gethostname()

    ip_address = socket.gethostbyname(hostname)

    print(ip_address)

    username = os.getlogin()

    print(username)

    # username1 = getpass.getuser()

    # print(username1)

if __name__=="__main__":        
    # console()
    test()