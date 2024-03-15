from tkinter import BOTH, LEFT, TOP, Button, Entry, Frame, Label, PhotoImage, StringVar, Tk,Radiobutton,StringVar,IntVar,filedialog
from idlelib.tooltip import Hovertip
import sys
import os
from operator import itemgetter
import time
from tkinter import messagebox
from os import listdir
from os.path import isfile, join
from datetime import datetime, date
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, StaleElementReferenceException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import openpyxl
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from urllib.parse import urlparse
import warnings
import numpy as np
from openpyxl.utils import get_column_letter,column_index_from_string
from datetime import datetime
from pytz import timezone
import requests
from zipfile import ZipFile
import subprocess

warnings.filterwarnings("ignore")

global rows
global xpath
global heding
global status
global key
global nme
element_1 = None
global j


def browse():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
        image_path2 = os.path.join(sys._MEIPASS, 'Static', 'Mapping.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')
        image_path2 = os.path.join(os.getcwd(), 'Static', 'Mapping.png')

    root.title("Process File Picker")
    root.resizable(False,False)

    root.title("Process File Picker")
    root.resizable(False,False)
   
    w = 500
    h = 160
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Browse Process - File Picker",font=("Calibri",18,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
          
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=50)
    
    answer=StringVar()
    answer.set("")
       
    def browse_button():
        global filename
        answer.set("")  
        filename = filedialog.askopenfilename()  
        txt.delete(0, 'end') 
        txt.insert(0, filename)

    def Click_Done():
        global ans        
        
        ans=txt.get()        
        
        if ans=="":
           answer.set("File Path Fields Empty Is Not Allowed...")        
        else:      
            root.destroy()          
            return ans
        
    txt=Entry(Frame2,font=("Calibri",12,"bold","italic"),width=50,justify="left")
    txt.grid(row=0,column=0,padx=5,pady=10,sticky="E")    

    photo1 = PhotoImage(file=image_path2)
            
    btn1=Button(Frame2,text="Browse",command=browse_button,image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=0,column=1,padx=3,pady=0,sticky="W")     
                
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=3,y=80,width=490,height=25)
    
    title_3=Label(Frame3,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=1,padx=150,pady=0,sticky="E")

    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=3,y=105,width=500,height=50)
    
    photo2 = PhotoImage(file=image_path1)

    btn2=Button(Frame4,command=Click_Done,text="Done",image=photo2,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=3,column=0,padx=125,pady=0,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo3 = PhotoImage(file=image_path)   

    btn3=Button(Frame4,command=Close,text="Exit",image=photo3,borderwidth=0,bg="#2c3e50")
    btn3.grid(row=3,column=1,padx=15,pady=0,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn2,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn3,'Click to Exit Process',hover_delay=1000)
    myTip2 = Hovertip(btn1,'Click to Pick File',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def process():        
    global element_1    
    global ans1
    global ans2 
    global driver
    global j
    global dob
    global mid
    global frm_dt
    global to_dt
    global cpt
    global rows
    global rows1
    # https://rajhandicraft.com/collections/bed-bed-sides/products/solid-wood-small-peacock-patra-design-bed-k?variant=41111281926328
    browse()
      
    fil=ans               

    try:
        options = Options()            
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--ignore-ssl-errors')
        options.add_argument("--disable-popup-blocking")
        driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options)
        driver.maximize_window()
        driver.get('https://promise.dpw.state.pa.us/portal/provider/Home/tabid/135/Default.aspx')
    except Exception as e:                        
        try:
            options = Options()
            options.add_argument('--ignore-certificate-errors')
            options.add_argument('--ignore-ssl-errors')
            options.add_argument("--disable-popup-blocking")
            # response = requests.get('https://chromedriver.storage.googleapis.com/LATEST_RELEASE')
            # latest_version = response.text.strip()
            # chrome_driver_url = f'https://chromedriver.storage.googleapis.com/{latest_version}/chromedriver_win32.zip'
            # response = requests.get(chrome_driver_url)
            # with open('chromedriver_win32.zip', 'wb') as f:
            #     f.write(response.content)
            # with ZipFile('chromedriver_win32.zip', 'r') as zip_ref:
            #     zip_ref.extractall('.')
            driver_path = os.path.abspath('chromedriver.exe')
            driver = webdriver.Chrome(executable_path=driver_path,options=options)
            driver.maximize_window()
            # driver.get('https://promise.dpw.state.pa.us/portal/provider/Home/tabid/135/Default.aspx')
        except Exception as e:           
            messagebox.showinfo("Driver Problem","Pls Check Your Chrome Driver Version")
            sys.exit(0)
        
    def text_box(xpath,heding,status,key):                
        counter = 0
        while counter < 5:
            try:   
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()                     
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            print('Moved Exception')
            raise e
            # messagebox.showinfo(heding, status)
            # sys.exit(0)        
    
    def text_box_key(xpath,heding,status,key):                
        counter = 0
        while counter < 5:
            try:                   
                element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, xpath)))
                element.clear()                
                element.send_keys(key)                
                # element.send_keys(Keys.TAB)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            print('Moved Exception')
            raise e
            # messagebox.showinfo(heding, status)
            # sys.exit(0)    

    def text_box_js(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:
                js_code = "document.querySelector('" + xpath + "').setAttribute('value', '" + key + "');"
                WebDriverWait(driver, 0).until(lambda driver: driver.execute_script(js_code))               
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo(heding, status)
            sys.exit(0)       

    def click(xpath,heding,status):
        counter = 0
        while counter < 5:
            try:             
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            print('Moved Exception')
            raise e
            # messagebox.showinfo(heding,status)
            # sys.exit(0)                                

    def count(xpath,heding,status):
        global rows
        counter = 0
        while counter < 5:
            try:             
                rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            print('Moved Exception')
            raise e
            # messagebox.showinfo(heding,status)
            # sys.exit(0)                         

    def count_1(xpath,heding,status):
        global rows1
        counter = 0
        while counter < 5:
            try:             
                rows1=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            print('Moved Exception')
            raise e
            # messagebox.showinfo(heding,status)
            # sys.exit(0)        

    def Alert():
        counter = 0
        while counter < 15:
            try:             
                WebDriverWait(driver, 0).until (EC.alert_is_present())
                a=driver.switch_to.alert
                a.accept()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo('Alert','Alert Not Present')                                                                
    
    # messagebox.showinfo('Waiting','Authentication Waiting')        
    
    # driver.switch_to.window(driver.window_handles[1])      

    file=pd.read_excel(fil,sheet_name='Matrix',header=0)
    
    for index, row in file.iterrows():                                      
        lnk = row[0]
        unm=row[1]
        psw = row[2]                     
        fec_nm = row[4]
        st=row[5]
        lst_nm = row[6]
        fst_nm = row[7]
        
        driver.get(lnk.lstrip().rstrip())
        
        key=unm.lstrip().rstrip()
        counter = 0
        while counter < 5:
            try:                    
                xpath= "/html/body/div/div/main/div/div/div[2]/div/div/form/div[1]/div[2]/div[1]/div[2]/span/input"
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()                     
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)             
                break
            except Exception as e:           
                try:
                    xpath= "/html/body/form/div/div/main/div[1]/div[3]/div[3]/div/input"
                    WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()                     
                    WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1   
        
        key=psw.lstrip().rstrip()
        counter = 0
        while counter < 5:
            try:
                xpath= "/html/body/div/div/main/div/div/div[2]/div/div/form/div[1]/div[2]/div[2]/div[2]/span/input"
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()                     
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)             
                break
            except Exception as e:
                try:
                    xpath= "/html/body/form/div/div/main/div[1]/div[3]/div[4]/div/input"
                    WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()                     
                    WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
       
        counter = 0
        while counter < 5:
            try:
                xpath='/html/body/div/div/main/div/div/div[2]/div/div/form/div[2]/input'      
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()           
                break
            except Exception as e:
                try:
                    xpath='/html/body/form/div/div/main/div[1]/div[3]/div[6]/input'      
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()   
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1

        counter = 0
        while counter < 5:
            try:
                xpath= "/html/body/div/div/main/div/div/div[2]/div/div/form/div[1]/div[1]/div/div/p"
                ck=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text
                
                wb1=load_workbook(filename=fil)
                sheet = wb1['Matrix']
                column_letter = 'W'  
                column_cells = sheet[column_letter]
                last_row = None
                for cell in reversed(column_cells):
                    if cell.value:
                        last_row = cell.row
                        break                            
                sheet['W' + str(int(last_row + 1))]=ck 
                wb1.save(fil)
                wb1.close()                                    

                break 
            except Exception as e:
                try:
                    xpath= "/html/body/form/div/div/main/div[1]/div[3]/div[2]"
                    ck=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text
                    
                    wb1=load_workbook(filename=fil)
                    sheet = wb1['Matrix']
                    column_letter = 'W'  
                    column_cells = sheet[column_letter]
                    last_row = None
                    for cell in reversed(column_cells):
                        if cell.value:
                            last_row = cell.row
                            break                            
                    sheet['W' + str(int(last_row + 1))]=ck 
                    wb1.save(fil)
                    wb1.close()                                    

                    break 
                except Exception as e:
                    try:                                               
                        try:
                            xpath='/html/body/div[1]/nav[1]/div[3]/div/ul[2]/li'
                            ck1=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text
                        except Exception as e:
                            xpath='/html/body/div[1]/nav[1]/div[4]/div/ul[2]/li'
                            ck1=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text
                        
                        if ck1.lstrip().rstrip()=='Log off':
                                  
                            xpath= "/html/body/div[1]/nav[2]/div/div[2]/ul/li"
                            heding="Facility Count"        
                            status="Facility Count Not Found"                    
                            count(xpath,heding,status)
                            
                            j=1
                            while j<rows+1:  
                                sck=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/nav[2]/div/div[2]/ul/li[{}]/a'.format(j)))).text
                                if sck.lstrip().rstrip()=='Facility':
                                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[1]/nav[2]/div/div[2]/ul/li[{}]/a'.format(j)))).click()                                                                         
                                    counter = 0
                                    while counter < 5:
                                        try:           
                                            xpath= "/html/body/div[1]/nav[2]/div/div[2]/ul/li[{}]/ul/div/div/li/ul/li".format(j)  
                                            rows1=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))                                        
                                            break
                                        except Exception as e:
                                            try:
                                                xpath= "/html/body/div[1]/nav[2]/div/div[2]/ul/li[{}]/ul/li".format(j)  
                                                rows1=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))      
                                                break
                                            except Exception as e:                                        
                                                time.sleep(1)
                                                counter += 1
                                                                                                                           
                                    i=1
                                    while i<rows1+1:
                                        try:
                                            sck=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/nav[2]/div/div[2]/ul/li[{}]/ul/div/div/li/ul/li[{}]'.format(j,i)))).text
                                        except Exception as e:
                                            sck=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/nav[2]/div/div[2]/ul/li[{}]/ul/li[{}]'.format(j,i)))).text
                                        if sck.lstrip().rstrip()=='Search Facility':
                                            try:
                                                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[1]/nav[2]/div/div[2]/ul/li[{}]/ul/div/div/li/ul/li[{}]/a'.format(j,i)))).click()
                                                break
                                            except Exception as e:
                                                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[1]/nav[2]/div/div[2]/ul/li[{}]/ul/li[{}]/a'.format(j,i)))).click()
                                                break
                                        i=i+1
                                    break
                                j=j+1
                                  
                            xpath= "/html/body/div[1]/main/form/div[1]/fieldset/div[2]/input"
                            heding="Facility Name"
                            status="Facility Name Field Not Found"
                            key=fec_nm.lstrip().rstrip()
                            text_box(xpath,heding,status,key)

                            xpath= "/html/body/div[1]/main/form/div[1]/fieldset/div[2]/select[2]/option"
                            heding="State Count"        
                            status="State Count Not Found"                    
                            count(xpath,heding,status)
                            
                            j=1
                            while j<rows+1:  
                                sck=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/form/div[1]/fieldset/div[2]/select[2]/option[{}]'.format(j)))).text
                                if sck.lstrip().rstrip()==st.lstrip().rstrip():
                                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[1]/main/form/div[1]/fieldset/div[2]/select[2]/option[{}]'.format(j)))).click()
                                    break
                                j=j+1

                            xpath='/html/body/div[1]/main/form/div[1]/fieldset/div[3]/input'
                            heding="Facility Search"
                            status="Facility Search Button Not Found"        
                            click(xpath,heding,status)

                            time.sleep(2)

                            try:           
                                xpath='/html/body/div[1]/main/form/table/tbody/tr/td[1]/a'
                                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                                
                                try:                
                                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[1]/form/div[1]/a[1]'))).click()
                                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[1]/form/div[2]/input'))).click()

                                    xpath= "/html/body/header/nav[2]/div/ul/li"
                                    heding="Resident Count"
                                    status="Resident Count Not Found"
                                    count(xpath,heding,status)

                                    j=1
                                    while j<rows+1:  
                                        rec=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/header/nav[2]/div/ul/li[{}]'.format(j)))).text
                                        if rec.lstrip().rstrip()=='Resident':
                                            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/header/nav[2]/div/ul/li[{}]/a'.format(j)))).click()
                                            break
                                        j=j+1

                                    xpath= "/html/body/header/nav[2]/div/ul/li[4]/ul/div/div/li/ul/li"
                                    heding="Search Resident Count"        
                                    status="Search Resident Count Not Found"                    
                                    count(xpath,heding,status)

                                    j=1
                                    while j<rows+1:  
                                        rec_ser=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/header/nav[2]/div/ul/li[4]/ul/div/div/li/ul/li[{}]'.format(j)))).text
                                        if rec_ser.lstrip().rstrip()=='Search Resident':
                                            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/header/nav[2]/div/ul/li[4]/ul/div/div/li/ul/li[{}]/a'.format(j)))).click()
                                            break
                                        j=j+1

                                    xpath= "/html/body/div[1]/main/form/fieldset[1]/div[1]/input"
                                    heding="Last Name"
                                    status="Last Name Field Not Found"
                                    key=lst_nm.lstrip().rstrip()
                                    text_box(xpath,heding,status,key)
                                    
                                    xpath= "/html/body/div[1]/main/form/fieldset[1]/div[2]/input"
                                    heding="First Name"
                                    status="First Name Field Not Found"
                                    key=fst_nm.lstrip().rstrip()
                                    text_box(xpath,heding,status,key)

                                    xpath='/html/body/div[1]/main/form/fieldset[7]/section/input'
                                    heding="Search Residents"
                                    status="Search Residents Button Not Found"        
                                    click(xpath,heding,status)
                                    
                                    time.sleep(2)
                                        
                                    xpath='/html/body/div[1]/main/form[2]/table/tbody/tr/td[1]/a'
                                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()

                                    lst=[]

                                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/article[1]/div/div[1]/p[1]/span[2]'))).text

                                    if cnm==' ' or cnm=='':
                                        cnm='N/A'
                                        lst.append(cnm)   
                                    else:
                                        lst.append(cnm) 
                                    
                                    lt=len(lst)
                                    if lt==0:                           
                                        lst.append('N/A') 

                                    lst1=[]
                                    sex=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/article[1]/div/div[1]/p[3]/span[2]'))).text

                                    if sex==' ' or sex=='':
                                        sex='N/A' 
                                        lst1.append(sex)
                                    else:
                                        lst1.append(sex)  

                                    lt=len(lst1)
                                    if lt==0:                           
                                        lst.append('N/A')
                                    else:
                                        vr=lst1[0]
                                        lst.append(vr)
                                
                                    lst2=[]
                                    cdob=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/article[1]/div/div[1]/p[4]/span[2]'))).text

                                    if cdob==' ' or cdob=='':
                                        cdob='N/A' 
                                        lst2.append(cdob)
                                    else:
                                        lst2.append(cdob)  

                                    lt=len(lst2)
                                    if lt==0:                           
                                        lst.append('N/A')
                                    else:
                                        vr=lst2[0]
                                        lst.append(vr)

                                    lst3=[]
                                    sts=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/article[1]/div/div[3]/p[1]/span[2]'))).text

                                    if sts==' ' or sts=='':
                                        sts='N/A' 
                                        lst3.append(sts)
                                    else:
                                        lst3.append(sts)  

                                    lt=len(lst3)
                                    if lt==0:                           
                                        lst.append('N/A')
                                    else:
                                        vr=lst3[0]
                                        lst.append(vr)

                                    lst4=[]
                                    add=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/article[3]/div[1]/div[1]/p[1]/span[2]'))).text

                                    if add==' ' or add=='':
                                        add='N/A' 
                                        lst4.append(add)
                                    else:
                                        lst4.append(add)  

                                    lt=len(lst4)
                                    if lt==0:                           
                                        lst.append('N/A')
                                    else:
                                        vr=lst4[0]
                                        lst.append(vr)
                                    
                                    lst5=[]
                                    ph=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/article[3]/div[1]/div[1]/p[3]/span[2]'))).text

                                    if ph==' ' or ph=='':
                                        ph='N/A' 
                                        lst5.append(ph)
                                    else:
                                        lst5.append(ph)  

                                    lt=len(lst5)
                                    if lt==0:                           
                                        lst.append('N/A')
                                    else:
                                        vr=lst5[0]
                                        lst.append(vr)
                                    
                                    lst6=[]
                                    ssn=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/article[4]/div/div[1]/p[2]/span[2]'))).text

                                    if ssn==' ' or ssn=='':
                                        ssn='N/A' 
                                        lst6.append(ssn)
                                    else:
                                        lst6.append(ssn)  

                                    lt=len(lst6)
                                    if lt==0:                           
                                        lst.append('N/A')
                                    else:
                                        vr=lst6[0]
                                        lst.append(vr)
                                
                                    lst7=[]
                                    med_a=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/article[4]/div/div[2]/p[1]/span[2]'))).text

                                    if med_a==' ' or med_a=='':
                                        med_a='N/A' 
                                        lst7.append(med_a)
                                    else:
                                        lst7.append(med_a)  

                                    lt=len(lst7)
                                    if lt==0:                           
                                        lst.append('N/A')
                                    else:
                                        vr=lst7[0]
                                        lst.append(vr)
                                    
                                    lst8=[]
                                    med_b=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/article[4]/div/div[3]/p[1]/span[2]'))).text

                                    if med_b==' ' or med_b=='':
                                        med_b='N/A' 
                                        lst8.append(med_b)
                                    else:
                                        lst8.append(med_b)  

                                    lt=len(lst8)
                                    if lt==0:                           
                                        lst.append('N/A')
                                    else:
                                        vr=lst8[0]
                                        lst.append(vr)
                                    
                                    lst9=[]
                                    mecd=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/article[4]/div/div[3]/p[2]/span[2]'))).text

                                    if mecd==' ' or mecd=='':
                                        mecd='N/A' 
                                        lst9.append(mecd)
                                    else:
                                        lst9.append(mecd)  

                                    lt=len(lst9)
                                    if lt==0:                           
                                        lst.append('N/A')
                                    else:
                                        vr=lst9[0]
                                        lst.append(vr)
                                    
                                    lst10=[]
                                    ins=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/article[5]/div/table/tbody/tr/td[1]'))).text

                                    if ins==' ' or ins=='':
                                        ins='N/A' 
                                        lst10.append(ins)
                                    else:
                                        lst10.append(ins)  

                                    lt=len(lst10)
                                    if lt==0:                           
                                        lst.append('N/A')
                                    else:
                                        vr=lst10[0]
                                        lst.append(vr)
                                    
                                    lst11=[]
                                    grp_nm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/article[5]/div/table/tbody/tr/td[2]'))).text

                                    if grp_nm==' ' or grp_nm=='':
                                        grp_nm='N/A' 
                                        lst11.append(grp_nm)
                                    else:
                                        lst11.append(grp_nm)  

                                    lt=len(lst11)
                                    if lt==0:                           
                                        lst.append('N/A')
                                    else:
                                        vr=lst11[0]
                                        lst.append(vr)
                                    
                                    lst12=[]
                                    grp_id=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/article[5]/div/table/tbody/tr/td[3]'))).text

                                    if grp_id==' ' or grp_id=='':
                                        grp_id='N/A' 
                                        lst12.append(grp_id)
                                    else:
                                        lst12.append(grp_id)  

                                    lt=len(lst12)
                                    if lt==0:                           
                                        lst.append('N/A')
                                    else:
                                        vr=lst12[0]
                                        lst.append(vr)
                                    
                                    lst13=[]
                                    ins_id=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/article[5]/div/table/tbody/tr/td[4]'))).text

                                    if ins_id==' ' or ins_id=='':
                                        ins_id='N/A' 
                                        lst13.append(ins_id)
                                    else:
                                        lst13.append(ins_id)  

                                    lt=len(lst13)
                                    if lt==0:                           
                                        lst.append('N/A')
                                    else:
                                        vr=lst13[0]
                                        lst.append(vr)

                                    try:                                         
                                        xpath='/html/body/div[1]/nav[1]/div[3]/div/ul[2]/li/a'
                                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                                    except Exception as e:
                                        xpath='/html/body/div[1]/nav[1]/div[4]/div/ul[2]/li/a'
                                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()

                                    wb1=load_workbook(fil)
                                    sheet = wb1['Matrix']
                                    column_letter = 'W'  
                                    column_cells = sheet[column_letter]
                                    last_row = None
                                    for cell in reversed(column_cells):
                                        if cell.value:
                                            last_row = cell.row
                                            break              
                                    start_column = 'I'
                                    current_column_index = openpyxl.utils.column_index_from_string(start_column)
                                    current_row = last_row + 1

                                    for value in lst:
                                        current_column = openpyxl.utils.get_column_letter(current_column_index)
                                        sheet[current_column + str(current_row)] = value
                                        current_column_index += 1                
                                    sheet['W' + str(int(last_row + 1))]='Done'
                                    wb1.save(filename=fil)
                                    wb1.close()    

                                    break
                                except Exception as e:
                                    # Patten 2                                  
                                    try:
                                        xpath= "/html/body/header/nav[2]/div/ul/li"
                                        rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                                    except Exception as e:
                                        xpath= "/html/body/div[1]/nav[2]/div/div[2]/ul/li"
                                        rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))                                

                                    j=1
                                    while j<rows+1: 
                                        try: 
                                            rec=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/header/nav[2]/div/ul/li[{}]'.format(j)))).text
                                        except Exception as e:
                                            rec=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/nav[2]/div/div[2]/ul/li[{}]'.format(j)))).text

                                        if rec.lstrip().rstrip()=='Resident':
                                            try:
                                                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/header/nav[2]/div/ul/li[{}]/a'.format(j)))).click()
                                            except Exception as e:
                                                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[1]/nav[2]/div/div[2]/ul/li[{}]/a'.format(j)))).click()
                                        
                                            try:                                                  
                                                xpath= "/html/body/header/nav[2]/div/ul/li[{}]/ul/div/div/li/ul/li".format(j)
                                                rows1=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                                            except Exception as e:                                                  
                                                xpath= "/html/body/div[1]/nav[2]/div/div[2]/ul/li[{}]/ul/div/div/li/ul/li".format(j)
                                                rows1=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                                            
                                            i=1
                                            while i<rows1+1:  
                                                try:                                                                                                                                
                                                    rec_ser=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/header/nav[2]/div/ul/li[{}]/ul/div/div/li/ul/li[{}]'.format(j,i)))).text
                                                except Exception as e:
                                                    rec_ser=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/nav[2]/div/div[2]/ul/li[{}]/ul/div/div/li/ul/li[{}]'.format(j,i)))).text

                                                if rec_ser.lstrip().rstrip()=='Search Resident':
                                                    try:
                                                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/header/nav[2]/div/ul/li[{}]/ul/div/div/li/ul/li[{}]/a'.format(j,i)))).click()
                                                    except Exception as e:
                                                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[1]/nav[2]/div/div[2]/ul/li[{}]/ul/div/div/li/ul/li[{}]/a'.format(j,i)))).click()
                                                    break
                                                j=j+1                                                                                                            
                                            break
                                        j=j+1
                                                                                                        
                                    xpath= "/html/body/div[1]/main/form/fieldset[1]/div[1]/input"
                                    heding="Last Name"
                                    status="Last Name Field Not Found"
                                    key=lst_nm.lstrip().rstrip()
                                    text_box(xpath,heding,status,key)
                                    
                                    xpath= "/html/body/div[1]/main/form/fieldset[1]/div[2]/input"
                                    heding="First Name"
                                    status="First Name Field Not Found"
                                    key=fst_nm.lstrip().rstrip()
                                    text_box(xpath,heding,status,key)

                                    xpath='/html/body/div[1]/main/form/fieldset[7]/section/input'
                                    heding="Search Residents"
                                    status="Search Residents Button Not Found"        
                                    click(xpath,heding,status)
                                    
                                    time.sleep(2)
                                        
                                    xpath='/html/body/div[1]/main/form[2]/table/tbody/tr/td[1]/a'
                                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()

                                    lst=[]
                                    try:
                                        cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/article[1]/div/div[1]/p[1]/span[2]'))).text
                                    except Exception as e:
                                        cnm='N/A'

                                    if cnm==' ' or cnm=='':
                                        cnm='N/A'
                                        lst.append(cnm)   
                                    else:
                                        lst.append(cnm) 
                                    
                                    lt=len(lst)
                                    if lt==0:                           
                                        lst.append('N/A') 

                                    lst1=[]
                                    try:
                                        sex=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/article[1]/div/div[1]/p[3]/span[2]'))).text
                                    except Exception as e:
                                        sex='N/A'

                                    if sex==' ' or sex=='':
                                        sex='N/A' 
                                        lst1.append(sex)
                                    else:
                                        lst1.append(sex)  

                                    lt=len(lst1)
                                    if lt==0:                           
                                        lst.append('N/A')
                                    else:
                                        vr=lst1[0]
                                        lst.append(vr)
                                
                                    lst2=[]
                                    try:
                                        cdob=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/article[1]/div/div[1]/p[4]/span[2]'))).text
                                    except Exception as e:
                                        cdob='N/A'

                                    if cdob==' ' or cdob=='':
                                        cdob='N/A' 
                                        lst2.append(cdob)
                                    else:
                                        lst2.append(cdob)  

                                    lt=len(lst2)
                                    if lt==0:                           
                                        lst.append('N/A')
                                    else:
                                        vr=lst2[0]
                                        lst.append(vr)

                                    lst3=[]
                                    try:
                                        sts=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/article[1]/div/div[3]/p[1]/span[2]'))).text
                                    except Exception as e:
                                        sts='N/A'

                                    if sts==' ' or sts=='':
                                        sts='N/A' 
                                        lst3.append(sts)
                                    else:
                                        lst3.append(sts)  

                                    lt=len(lst3)
                                    if lt==0:                           
                                        lst.append('N/A')
                                    else:
                                        vr=lst3[0]
                                        lst.append(vr)

                                    lst4=[]
                                    try:
                                        add=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/article[3]/div[1]/div[1]/p[1]/span[2]'))).text
                                    except Exception as e:
                                        add='N/A'

                                    if add==' ' or add=='':
                                        add='N/A' 
                                        lst4.append(add)
                                    else:
                                        lst4.append(add)  

                                    lt=len(lst4)
                                    if lt==0:                           
                                        lst.append('N/A')
                                    else:
                                        vr=lst4[0]
                                        lst.append(vr)
                                    
                                    lst5=[]
                                    try:                                                                            
                                        ph=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/article[3]/div[1]/div[1]/p[3]/span[2]'))).text
                                    except Exception as e:
                                        ph='N/A'

                                    if ph==' ' or ph=='':
                                        ph='N/A' 
                                        lst5.append(ph)
                                    else:
                                        lst5.append(ph)  

                                    lt=len(lst5)
                                    if lt==0:                           
                                        lst.append('N/A')
                                    else:
                                        vr=lst5[0]
                                        lst.append(vr)
                                    
                                    lst6=[]
                                    try:
                                        ssn=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/article[4]/div/div[1]/p[2]/span[2]'))).text
                                    except Exception as e:
                                        ssn='N/A'

                                    if ssn==' ' or ssn=='':
                                        ssn='N/A' 
                                        lst6.append(ssn)
                                    else:
                                        lst6.append(ssn)  

                                    lt=len(lst6)
                                    if lt==0:                           
                                        lst.append('N/A')
                                    else:
                                        vr=lst6[0]
                                        lst.append(vr)
                                
                                    lst7=[]
                                    try:
                                        med_a=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/article[4]/div/div[2]/p[1]/span[2]'))).text
                                    except Exception as e:
                                        med_a='N/A'

                                    if med_a==' ' or med_a=='':
                                        med_a='N/A' 
                                        lst7.append(med_a)
                                    else:
                                        lst7.append(med_a)  

                                    lt=len(lst7)
                                    if lt==0:                           
                                        lst.append('N/A')
                                    else:
                                        vr=lst7[0]
                                        lst.append(vr)
                                    
                                    lst8=[]
                                    try:
                                        med_b=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/article[4]/div/div[3]/p[1]/span[2]'))).text
                                    except Exception as e:
                                        med_b='N/A'

                                    if med_b==' ' or med_b=='':
                                        med_b='N/A' 
                                        lst8.append(med_b)
                                    else:
                                        lst8.append(med_b)  

                                    lt=len(lst8)
                                    if lt==0:                           
                                        lst.append('N/A')
                                    else:
                                        vr=lst8[0]
                                        lst.append(vr)
                                    
                                    lst9=[]
                                    try:
                                        mecd=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/article[4]/div/div[3]/p[2]/span[2]'))).text
                                    except Exception as e:
                                        mecd='N/A'

                                    if mecd==' ' or mecd=='':
                                        mecd='N/A' 
                                        lst9.append(mecd)
                                    else:
                                        lst9.append(mecd)  

                                    lt=len(lst9)
                                    if lt==0:                           
                                        lst.append('N/A')
                                    else:
                                        vr=lst9[0]
                                        lst.append(vr)
                                    
                                    lst10=[]
                                    try:
                                        ins=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/article[5]/div/table/tbody/tr/td[1]'))).text
                                    except Exception as e:
                                        ins='N/A'

                                    if ins==' ' or ins=='':
                                        ins='N/A' 
                                        lst10.append(ins)
                                    else:
                                        lst10.append(ins)  

                                    lt=len(lst10)
                                    if lt==0:                           
                                        lst.append('N/A')
                                    else:
                                        vr=lst10[0]
                                        lst.append(vr)
                                    
                                    lst11=[]
                                    try:
                                        grp_nm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/article[5]/div/table/tbody/tr/td[2]'))).text
                                    except Exception as e:
                                        grp_nm='N/A'

                                    if grp_nm==' ' or grp_nm=='':
                                        grp_nm='N/A' 
                                        lst11.append(grp_nm)
                                    else:
                                        lst11.append(grp_nm)  

                                    lt=len(lst11)
                                    if lt==0:                           
                                        lst.append('N/A')
                                    else:
                                        vr=lst11[0]
                                        lst.append(vr)
                                    
                                    lst12=[]
                                    try:
                                        grp_id=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/article[5]/div/table/tbody/tr/td[3]'))).text
                                    except Exception as e:
                                        grp_id='N/A'

                                    if grp_id==' ' or grp_id=='':
                                        grp_id='N/A' 
                                        lst12.append(grp_id)
                                    else:
                                        lst12.append(grp_id)  

                                    lt=len(lst12)
                                    if lt==0:                           
                                        lst.append('N/A')
                                    else:
                                        vr=lst12[0]
                                        lst.append(vr)
                                    
                                    lst13=[]
                                    try:
                                        ins_id=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/article[5]/div/table/tbody/tr/td[4]'))).text
                                    except Exception as e:
                                        ins_id='N/A'

                                    if ins_id==' ' or ins_id=='':
                                        ins_id='N/A' 
                                        lst13.append(ins_id)
                                    else:
                                        lst13.append(ins_id)  

                                    lt=len(lst13)
                                    if lt==0:                           
                                        lst.append('N/A')
                                    else:
                                        vr=lst13[0]
                                        lst.append(vr)
                                    
                                    try:                                         
                                        xpath='/html/body/div[1]/nav[1]/div[3]/div/ul[2]/li/a'
                                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                                    except Exception as e:
                                        xpath='/html/body/div[1]/nav[1]/div[4]/div/ul[2]/li/a'
                                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                                                                        
                                    wb1=load_workbook(fil)
                                    sheet = wb1['Matrix']
                                    column_letter = 'W'  
                                    column_cells = sheet[column_letter]
                                    last_row = None
                                    for cell in reversed(column_cells):
                                        if cell.value:
                                            last_row = cell.row
                                            break              
                                    start_column = 'I'
                                    current_column_index = openpyxl.utils.column_index_from_string(start_column)
                                    current_row = last_row + 1

                                    for value in lst:
                                        current_column = openpyxl.utils.get_column_letter(current_column_index)
                                        sheet[current_column + str(current_row)] = value
                                        current_column_index += 1                
                                    sheet['W' + str(int(last_row + 1))]='Done'
                                    wb1.save(filename=fil)
                                    wb1.close()                                    

                                    break                                                
                            except Exception as e:
                                wb1=load_workbook(filename=fil)
                                sheet = wb1['Matrix']
                                column_letter = 'W'  
                                column_cells = sheet[column_letter]
                                last_row = None
                                for cell in reversed(column_cells):
                                    if cell.value:
                                        last_row = cell.row
                                        break                            
                                sheet['W' + str(int(last_row + 1))]='Facility Table is Empty' 
                                wb1.save(fil)
                                wb1.close()  

                                try:                                         
                                    xpath='/html/body/div[1]/nav[1]/div[3]/div/ul[2]/li/a'
                                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                                except Exception as e:
                                    xpath='/html/body/div[1]/nav[1]/div[4]/div/ul[2]/li/a'
                                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                                
                                break
                    except Exception as e:
                        time.sleep(1)
                        counter += 1   

    driver.quit()
    messagebox.showinfo("Process Status", "Process Completed")
    sys.exit(0)    

if __name__=="__main__":        
    process()
    