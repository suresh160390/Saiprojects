from tkinter import BOTH, LEFT, TOP, Button, Entry, Frame, Label, PhotoImage, StringVar, Tk,Radiobutton,StringVar,IntVar,filedialog
from idlelib.tooltip import Hovertip
import sys
import os
from operator import itemgetter
import time
from tkinter import messagebox
from os import listdir
from os.path import isfile, join
from datetime import datetime, date
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, StaleElementReferenceException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import openpyxl
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from urllib.parse import urlparse
import warnings
import numpy as np
from openpyxl.utils import get_column_letter,column_index_from_string
from datetime import datetime
from pytz import timezone
import requests
from zipfile import ZipFile
import subprocess

warnings.filterwarnings("ignore")

global rows
global xpath
global heding
global status
global key
global nme
element_1 = None
global j

def user_pass():
    global ans1
    global ans2
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')

    root.title("CCBH - User Login And File Details")
    root.resizable(False,False)

    w = 600
    h = 180
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)

    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="User Name and Password Details...",font=("Calibri",20,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack() 
 
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=698,height=150)    
    
    title3=Label(Frame2,text="User Name :",font=("Calibri",11,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title3.grid(row=0,column=0,padx=5,pady=5,sticky="W")
    
    txt1=Entry(Frame2,font=("Calibri",11,"bold","italic"),width=50,justify="left")
    txt1.grid(row=0,column=1,padx=30,pady=5,sticky="W")
    
    title4=Label(Frame2,text="Password :",font=("Calibri",11,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title4.grid(row=1,column=0,padx=5,pady=5,sticky="W")
    
    txt2=Entry(Frame2,font=("Calibri",11,"bold","italic"),width=50,justify="left")
    txt2.grid(row=1,column=1,padx=30,pady=5,sticky="W")

    answer=StringVar()
    answer.set("")

    def Click_Done():
        global ans1
        global ans2        
        
        ans1=txt1.get()
        ans2=txt2.get()        

        if ans1=="":
           answer.set("User Name Field Empty Is Not Allowed...")
        elif ans2=="":
            answer.set("Password Field Empty Is Not Allowed...")        
        else:    
            root.destroy()
            return ans1,ans2
                
    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=0,y=105,width=698,height=20)
    
    title_3=Label(Frame4,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=2,padx=200,pady=0,sticky="E")
    
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=0,y=125,width=698,height=200)
       
    photo1 = PhotoImage(file=image_path1)
    
    btn1=Button(Frame3,command=Click_Done,text="Run",image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=2,column=0,padx=140,pady=0,sticky="W")

    def Close():
        sys.exit(0)

    photo = PhotoImage(file=image_path)    

    btn2=Button(Frame3,command=Close,text="Close",image=photo,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=2,column=1,padx=100,pady=0,sticky="E")

    def disable_event():
        pass

    txt1.focus_set()

    myTip = Hovertip(btn1,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn2,'Click to Exit Process',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)
            
    root.mainloop()

def browse():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
        image_path2 = os.path.join(sys._MEIPASS, 'Static', 'Mapping.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')
        image_path2 = os.path.join(os.getcwd(), 'Static', 'Mapping.png')

    root.title("Process File Picker")
    root.resizable(False,False)

    root.title("Process File Picker")
    root.resizable(False,False)
   
    w = 500
    h = 160
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Browse Process - File Picker",font=("Calibri",18,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
          
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=50)
    
    answer=StringVar()
    answer.set("")
       
    def browse_button():
        global filename
        answer.set("")  
        filename = filedialog.askopenfilename()  
        txt.delete(0, 'end') 
        txt.insert(0, filename)

    def Click_Done():
        global ans        
        
        ans=txt.get()        
        
        if ans=="":
           answer.set("File Path Fields Empty Is Not Allowed...")        
        else:      
            root.destroy()          
            return ans
        
    txt=Entry(Frame2,font=("Calibri",12,"bold","italic"),width=50,justify="left")
    txt.grid(row=0,column=0,padx=5,pady=10,sticky="E")    

    photo1 = PhotoImage(file=image_path2)
            
    btn1=Button(Frame2,text="Browse",command=browse_button,image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=0,column=1,padx=3,pady=0,sticky="W")     
                
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=3,y=80,width=490,height=25)
    
    title_3=Label(Frame3,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=1,padx=150,pady=0,sticky="E")

    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=3,y=105,width=500,height=50)
    
    photo2 = PhotoImage(file=image_path1)

    btn2=Button(Frame4,command=Click_Done,text="Done",image=photo2,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=3,column=0,padx=125,pady=0,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo3 = PhotoImage(file=image_path)   

    btn3=Button(Frame4,command=Close,text="Exit",image=photo3,borderwidth=0,bg="#2c3e50")
    btn3.grid(row=3,column=1,padx=15,pady=0,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn2,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn3,'Click to Exit Process',hover_delay=1000)
    myTip2 = Hovertip(btn1,'Click to Pick File',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def process():        
    global element_1    
    global ans1
    global ans2 
    global driver
    global j
    global dob
    global mid
    global frm_dt
    global to_dt
    global cpt

    # user_pass()
    browse()

    fil=ans               
    # user = ans1
    # password = ans2

    try:
        options = Options()            
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--ignore-ssl-errors')
        options.add_argument("--disable-popup-blocking")
        driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options)
        driver.maximize_window()
        driver.get('https://secure.ccbh.com/ePortal/login')
    except Exception as e:                        
        try:
            options = Options()
            options.add_argument('--ignore-certificate-errors')
            options.add_argument('--ignore-ssl-errors')
            options.add_argument("--disable-popup-blocking")
            # response = requests.get('https://chromedriver.storage.googleapis.com/LATEST_RELEASE')
            # latest_version = response.text.strip()
            # chrome_driver_url = f'https://chromedriver.storage.googleapis.com/{latest_version}/chromedriver_win32.zip'
            # response = requests.get(chrome_driver_url)
            # with open('chromedriver_win32.zip', 'wb') as f:
            #     f.write(response.content)
            # with ZipFile('chromedriver_win32.zip', 'r') as zip_ref:
            #     zip_ref.extractall('.')
            driver_path = os.path.abspath('chromedriver.exe')
            driver = webdriver.Chrome(executable_path=driver_path,options=options)
            driver.maximize_window()
            driver.get('https://secure.ccbh.com/ePortal/login')
        except Exception as e:           
            messagebox.showinfo("Internet Problem","Pls Check Your Internet Connection")
            sys.exit(0)
        
    def text_box(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:   
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding, status)
            sys.exit(0)        
    
    def text_box_key(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:                   
                input_field =WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath)))
                input_field.send_keys(Keys.BACKSPACE * len(input_field))                
                input_field.send_keys(key)               
                # element.send_keys(Keys.TAB)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding, status)
            sys.exit(0)    

    def text_box_js(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:
                js_code = "document.querySelector('" + xpath + "').setAttribute('value', '" + key + "');"
                WebDriverWait(driver, 0).until(lambda driver: driver.execute_script(js_code))               
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo(heding, status)
            sys.exit(0)       

    def click(xpath,heding,status):
        counter = 0
        while counter < 15:
            try:             
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding,status)
            sys.exit(0)                                

    def count(xpath,heding,status):
        global rows
        counter = 0
        while counter < 15:
            try:             
                rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding,status)
            sys.exit(0)                         

    def Alert():
        counter = 0
        while counter < 15:
            try:             
                WebDriverWait(driver, 0).until (EC.alert_is_present())
                a=driver.switch_to.alert
                a.accept()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo('Alert','Alert Not Present')                                                                
                
    messagebox.showinfo('Waiting','Authentication Waiting')        
    
    driver.switch_to.window(driver.window_handles[1])
    
    file=pd.read_excel(fil,sheet_name='CCBH Claim Status BOT',header=0)
    
    for index, row in file.iterrows():                                             
        mid = row[4]            
        frm_dt=row[5]
        to_dt=row[6]
        cpt = row[7]                 
       
        xpath= "/html/body/form/div[3]/div/div[3]/table/tbody/tr/td[1]/div/table"
        heding="Table Count"
        status="Table Count Not Found"        
        count(xpath,heding,status)  
                            
        j=1
        while j<rows+1:         
            xpath= "/html/body/form/div[3]/div/div[3]/table/tbody/tr/td[1]/div/table[{}]"                       
            cnm=WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
            if cnm.lstrip().rstrip()=="› Claim Inquiry":  
                xpath= "/html/body/form/div[3]/div/div[3]/table/tbody/tr/td[1]/div/table[{}]/tbody/tr/td/a "               
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()
                break
            j=j+1  

        xpath= "/html/body/form/div[3]/div/div[3]/table/tbody/tr/td[2]/div/div[1]/div/div[2]/table[2]/tbody/tr[1]/td[2]/select/option"
        heding="Table Count"
        status="Table Count Not Found"        
        count(xpath,heding,status)

        j=1
        while j<rows+1:
            xpath= "/html/body/form/div[3]/div/div[3]/table/tbody/tr/td[2]/div/div[1]/div/div[2]/table[2]/tbody/tr[1]/td[2]/select/option[{}]"                       
            cnm=WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
            if cnm.lstrip().rstrip()=="VITAL HEALTHCARE SOLUTIONS [VC7897]":                                 
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()
                break
            j=j+1  
        
        xpath= "/html/body/form/div[3]/div/div[3]/table/tbody/tr/td[2]/div/div[1]/div/div[2]/table[2]/tbody/tr[4]/td[2]/input"
        heding="Member ID"
        status="Member ID Field Not Found"
        key=str(mid)
        text_box(xpath,heding,status,key)  
        
        date_object_1 = datetime.strptime(str(frm_dt), "%Y-%m-%d %H:%M:%S")
        dob3 = date_object_1.strftime("%m/%d/%Y")            
        
        xpath= "/html/body/form/div[3]/div/div[3]/table/tbody/tr/td[2]/div/div[1]/div/div[2]/table[2]/tbody/tr[9]/td[2]/input"
        heding="From Date"
        status="From Date Field Not Found"
        key=dob3

        counter = 0
        while counter < 15:
            try:   
                input_field =WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath)))
                input_field.send_keys(Keys.BACKSPACE * len(input_field.get_attribute("value")))
                time.sleep(1)
                input_field.send_keys(key)                  
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo(heding, status)
            sys.exit(0)
        
        date_object_1 = datetime.strptime(str(to_dt), "%Y-%m-%d %H:%M:%S")
        dob3 = date_object_1.strftime("%m/%d/%Y")            
        
        xpath= "/html/body/form/div[3]/div/div[3]/table/tbody/tr/td[2]/div/div[1]/div/div[2]/table[2]/tbody/tr[9]/td[4]/input"
        heding="To Date"
        status="To Date Field Not Found"
        key=dob3

        counter = 0
        while counter < 15:
            try:   
                input_field =WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath)))
                input_field.send_keys(Keys.BACKSPACE * len(input_field.get_attribute("value")))
                time.sleep(1)
                input_field.send_keys(key)   
                input_field.send_keys(Keys.TAB)                 
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo(heding, status)
            sys.exit(0)
                
        xpath= "/html/body/form/div[3]/div/div[3]/table/tbody/tr/td[2]/div/div[1]/div/div[2]/table[2]/tbody/tr[11]/td[1]/input"
        heding="Search"
        status="Search Button Not Found"
        click(xpath,heding,status)
        
        wait = WebDriverWait(driver, 20)
        wait.until(lambda driver: driver.execute_script("return document.readyState === 'complete';"))
        
        try:                                                                                                                  
            element_1 = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[3]/div/div[3]/table/tbody/tr/td[2]/div/div[2]/div[1]/div/table/tbody/tr[2]'))).text                                               
            if element_1.lstrip().rstrip() == 'No claims match this search criteria. Please refine your search.': 
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/form/div[3]/div/div[3]/table/tbody/tr/td[2]/div/div[2]/div[1]/div/table/tbody/tr[3]/td/input'))).click()                
                wb1=load_workbook(filename=fil)
                sheet = wb1['CCBH Claim Status BOT']
                column_letter = 'T'  
                column_cells = sheet[column_letter]
                last_row = None
                for cell in reversed(column_cells):
                    if cell.value:
                        last_row = cell.row
                        break                            
                sheet['T' + str(int(last_row + 1))]=element_1   
                wb1.save(fil)
                wb1.close()  
        except Exception as e:
            
            lst=[]

            xpath= "/html/body/form/div[3]/div/div[3]/table/tbody/tr/td[2]/div/div[1]/div[2]/fieldset/table/tbody/tr/td/div[1]/table/tbody/tr"
            heding="Behavioral Health Claims Table Count"
            status="Behavioral Health Claims Table Count Not Found"        
            count(xpath,heding,status)

            j=2
            while j<rows+1:
                if rows==2:                    
                    st = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[3]/div/div[3]/table/tbody/tr/td[2]/div/div[1]/div[2]/fieldset/table/tbody/tr/td/div[1]/table/tbody/tr[2]/td[5]'))).text     
            
                    if st==' ' or st=='':
                        st='N/A' 
                        lst.append(st)   
                    else:
                        lst.append(st)   

                    lt=len(lst)
                    if lt==0:                           
                        lst.append('N/A')   
                                   
                    xpath= "/html/body/form/div[3]/div/div[3]/table/tbody/tr/td[2]/div/div[1]/div[2]/fieldset/table/tbody/tr/td/div[1]/table/tbody/tr[2]/td[1]/a"
                    heding="Date of Service"
                    status="Date of Service Link Not Found"
                    click(xpath,heding,status)
                    break
                elif j==rows:
                    st = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[3]/div/div[3]/table/tbody/tr/td[2]/div/div[1]/div[2]/fieldset/table/tbody/tr/td/div[1]/table/tbody/tr[{}]/td[5]'.format(j)))).text   
                    
                    if st==' ' or st=='':
                        st='N/A' 
                        lst.append(st)   
                    else:
                        lst.append(st)   

                    lt=len(lst)
                    if lt==0:                           
                        lst.append('N/A')   
                    
                    xpath= "/html/body/form/div[3]/div/div[3]/table/tbody/tr/td[2]/div/div[1]/div[2]/fieldset/table/tbody/tr/td/div[1]/table/tbody/tr[{}]/td[1]/a".format(j)
                    heding="Date of Service"
                    status="Date of Service Link Not Found"
                    click(xpath,heding,status)
                    break
                j=j+1  
                       
            cn = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[3]/div/div[3]/table/tbody/tr/td[2]/div/div[1]/div[3]/div[2]/table/tbody/tr[2]/td/table/tbody/tr[1]/td[1]/span'))).text

            if cn==' ' or cn=='':
                cn='N/A' 
                lst.append(cn)   
            else:
                lst.append(cn)   

            lt=len(lst)
            if lt==0:                           
                lst.append('N/A') 

            rd = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[3]/div/div[3]/table/tbody/tr/td[2]/div/div[1]/div[3]/div[2]/table/tbody/tr[2]/td/table/tbody/tr[1]/td[2]/span'))).text

            if rd==' ' or rd=='':
                rd='N/A' 
                lst.append(rd)   
            else:
                lst.append(rd)   

            lt=len(lst)
            if lt==0:                           
                lst.append('N/A') 

            ald_amt = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[3]/div/div[3]/table/tbody/tr/td[2]/div/div[1]/div[3]/div[2]/table/tbody/tr[2]/td/table/tbody/tr[12]/td/table/tbody/tr[2]/td/table/tbody/tr[2]/td[2]/span'))).text

            if ald_amt==' ' or ald_amt=='':
                ald_amt='N/A' 
                lst.append(ald_amt)   
            else:
                lst.append(ald_amt)   

            lt=len(lst)
            if lt==0:                           
                lst.append('N/A') 

            mem_res = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[3]/div/div[3]/table/tbody/tr/td[2]/div/div[1]/div[3]/div[2]/table/tbody/tr[2]/td/table/tbody/tr[12]/td/table/tbody/tr[2]/td/table/tbody/tr[2]/td[3]/span'))).text

            if mem_res==' ' or mem_res=='':
                mem_res='N/A' 
                lst.append(mem_res)   
            else:
                lst.append(mem_res)   

            lt=len(lst)
            if lt==0:                           
                lst.append('N/A') 

            pid_ccbh = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[3]/div/div[3]/table/tbody/tr/td[2]/div/div[1]/div[3]/div[2]/table/tbody/tr[2]/td/table/tbody/tr[12]/td/table/tbody/tr[2]/td/table/tbody/tr[2]/td[4]/span'))).text

            if pid_ccbh==' ' or pid_ccbh=='':
                pid_ccbh='N/A' 
                lst.append(pid_ccbh)   
            else:
                lst.append(pid_ccbh)   

            lt=len(lst)
            if lt==0:                           
                lst.append('N/A') 
            
            try:                                                                                                                                                                                                #   
                pyt_num = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[3]/div/div[3]/table/tbody/tr/td[2]/div/div[1]/div[3]/div[2]/table/tbody/tr[2]/td/table/tbody/tr[13]/td/table/tbody/tr[2]/td/table/tbody/tr[1]/td[7]/table/tbody/tr[2]/td[1]/table/tbody/tr[2]/td/span'))).text

                if pyt_num==' ' or pyt_num=='':
                    pyt_num='N/A' 
                    lst.append(pyt_num)   
                else:
                    lst.append(pyt_num)   

                lt=len(lst)
                if lt==0:                           
                    lst.append('N/A')
            except Exception as e:
                try:                                                                                    
                    pyt_num = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[3]/div/div[3]/table/tbody/tr/td[2]/div/div[1]/div[3]/div[2]/table/tbody/tr[2]/td/table/tbody/tr[13]/td/table/tbody/tr[2]/td/table/tbody/tr[1]/td[6]/table/tbody/tr[2]/td[1]/table/tbody/tr[2]/td/span'))).text

                    if pyt_num==' ' or pyt_num=='':
                        pyt_num='N/A' 
                        lst.append(pyt_num)   
                    else:
                        lst.append(pyt_num)   

                    lt=len(lst)
                    if lt==0:                           
                        lst.append('N/A')
                except Exception as e:                 
                    lst.append('N/A') 

            try:
                pyt_dt = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[3]/div/div[3]/table/tbody/tr/td[2]/div/div[1]/div[3]/div[2]/table/tbody/tr[2]/td/table/tbody/tr[13]/td/table/tbody/tr[2]/td/table/tbody/tr[1]/td[7]/table/tbody/tr[2]/td[2]/table/tbody/tr[2]/td'))).text

                if pyt_dt==' ' or pyt_dt=='':
                    pyt_dt='N/A' 
                    lst.append(pyt_dt)   
                else:
                    lst.append(pyt_dt)   

                lt=len(lst)
                if lt==0:                           
                    lst.append('N/A')
            except Exception as e:
                try:
                    pyt_dt = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[3]/div/div[3]/table/tbody/tr/td[2]/div/div[1]/div[3]/div[2]/table/tbody/tr[2]/td/table/tbody/tr[13]/td/table/tbody/tr[2]/td/table/tbody/tr[1]/td[6]/table/tbody/tr[2]/td[2]/table/tbody/tr[2]/td'))).text

                    if pyt_dt==' ' or pyt_dt=='':
                        pyt_dt='N/A' 
                        lst.append(pyt_dt)   
                    else:
                        lst.append(pyt_dt)   

                    lt=len(lst)
                    if lt==0:                           
                        lst.append('N/A')
                except Exception as e:
                    lst.append('N/A')  

            # colspan = 3

            dec = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[3]/div/div[3]/table/tbody/tr/td[2]/div/div[1]/div[3]/div[2]/table/tbody/tr[2]/td/table/tbody/tr[13]/td/table/tbody/tr[3]'))).text

            if dec==' ' or dec=='':
                dec='N/A' 
                lst.append(dec)   
            else:
                lst.append(dec)   

            lt=len(lst)
            if lt==0:                           
                lst.append('N/A')        

            wb1=load_workbook(fil)
            sheet = wb1['CCBH Claim Status BOT']
            column_letter = 'T'  
            column_cells = sheet[column_letter]
            last_row = None
            for cell in reversed(column_cells):
                if cell.value:
                    last_row = cell.row
                    break              
            start_column = 'K'
            current_column_index = openpyxl.utils.column_index_from_string(start_column)
            current_row = last_row + 1

            for value in lst:
                current_column = openpyxl.utils.get_column_letter(current_column_index)
                sheet[current_column + str(current_row)] = value
                current_column_index += 1                
            sheet['T' + str(int(last_row + 1))]='Done'
            wb1.save(filename=fil)
            wb1.close()

    driver.quit()
    messagebox.showinfo("Process Status", "Process Completed")
    sys.exit(0)    


if __name__=="__main__":        
    process()
    