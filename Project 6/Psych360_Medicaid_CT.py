from tkinter import BOTH, LEFT, TOP, Button, Entry, Frame, Label, PhotoImage, StringVar, Tk,Radiobutton,StringVar,IntVar,filedialog
from idlelib.tooltip import Hovertip
from tkinter import messagebox
import sys
import os
import warnings
from datetime import datetime, date
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, StaleElementReferenceException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
import time
import pandas as pd
from openpyxl import load_workbook
from urllib.parse import urlparse
import warnings
import numpy as np
import requests
from zipfile import ZipFile
from openpyxl.utils import get_column_letter,column_index_from_string
warnings.filterwarnings("ignore")


global rows
global rows_1
global xpath
global heding
global status
global key
global nme
global driver
element_1 = None
global j
global i


def file_pick():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
        image_path2 = os.path.join(sys._MEIPASS, 'Static', 'Mapping.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')
        image_path2 = os.path.join(os.getcwd(), 'Static', 'Mapping.png')

    root.title("Process File Picker")
    root.resizable(False,False)
   
    w = 500
    h = 160
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Browse Process - File Picker",font=("Calibri",18,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
          
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=50)
    
    answer=StringVar()
    answer.set("")
       
    def browse_button():
        global filename
        answer.set("")  
        filename = filedialog.askopenfilename()  
        txt.delete(0, 'end') 
        txt.insert(0, filename)

    def Click_Done():
        global ans        
        
        ans=txt.get()        
        
        if ans=="":
           answer.set("File Path Fields Empty Is Not Allowed...")        
        else:      
            root.destroy()          
            return ans
        
    txt=Entry(Frame2,font=("Calibri",12,"bold","italic"),width=50,justify="left")
    txt.grid(row=0,column=0,padx=5,pady=10,sticky="E")    

    photo1 = PhotoImage(file=image_path2)
            
    btn1=Button(Frame2,text="Browse",command=browse_button,image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=0,column=1,padx=3,pady=0,sticky="W")     
                
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=3,y=80,width=490,height=25)
    
    title_3=Label(Frame3,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=1,padx=150,pady=0,sticky="E")

    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=3,y=105,width=500,height=50)
    
    photo2 = PhotoImage(file=image_path1)

    btn2=Button(Frame4,command=Click_Done,text="Done",image=photo2,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=3,column=0,padx=125,pady=0,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo3 = PhotoImage(file=image_path)   

    btn3=Button(Frame4,command=Close,text="Exit",image=photo3,borderwidth=0,bg="#2c3e50")
    btn3.grid(row=3,column=1,padx=15,pady=0,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn2,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn3,'Click to Exit Process',hover_delay=1000)
    myTip2 = Hovertip(btn1,'Click to Pick File',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def user_pass():
    global ans1
    global ans2
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')

    root.title("Psych360_Medicaid_CT - User Login And File Details")
    root.resizable(False,False)

    w = 600
    h = 180
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)

    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="User Name and Password Details...",font=("Calibri",20,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack() 
 
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=698,height=150)    
    
    title3=Label(Frame2,text="User Name :",font=("Calibri",11,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title3.grid(row=0,column=0,padx=5,pady=5,sticky="W")
    
    txt1=Entry(Frame2,font=("Calibri",11,"bold","italic"),width=50,justify="left")
    txt1.grid(row=0,column=1,padx=30,pady=5,sticky="W")
    
    title4=Label(Frame2,text="Password :",font=("Calibri",11,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title4.grid(row=1,column=0,padx=5,pady=5,sticky="W")
    
    txt2=Entry(Frame2,font=("Calibri",11,"bold","italic"),width=50,justify="left")
    txt2.grid(row=1,column=1,padx=30,pady=5,sticky="W")

    answer=StringVar()
    answer.set("")

    def Click_Done():
        global ans1
        global ans2        
        
        ans1=txt1.get()
        ans2=txt2.get()        

        if ans1=="":
           answer.set("User Name Field Empty Is Not Allowed...")
        elif ans2=="":
            answer.set("Password Field Empty Is Not Allowed...")        
        else:    
            root.destroy()
            return ans1,ans2
                
    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=0,y=105,width=698,height=20)
    
    title_3=Label(Frame4,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=2,padx=200,pady=0,sticky="E")
    
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=0,y=125,width=698,height=200)
       
    photo1 = PhotoImage(file=image_path1)
    
    btn1=Button(Frame3,command=Click_Done,text="Run",image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=2,column=0,padx=140,pady=0,sticky="W")

    def Close():
        sys.exit(0)

    photo = PhotoImage(file=image_path)    

    btn2=Button(Frame3,command=Close,text="Close",image=photo,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=2,column=1,padx=100,pady=0,sticky="E")

    def disable_event():
        pass

    txt1.focus_set()

    myTip = Hovertip(btn1,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn2,'Click to Exit Process',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)
            
    root.mainloop()

def primary():        
        global element_1    
        global ans1
        global ans2                
        global driver_path
        global j
        global i       
        
        user_pass()
        file_pick()

        user_name =ans1
        password=ans2               
        fil=ans       
        
        try:
            options = Options()            
            options.add_argument('--ignore-certificate-errors')
            options.add_argument('--ignore-ssl-errors')
            options.add_argument("--disable-popup-blocking")
            driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options)
            driver.maximize_window()
            driver.get('https://www.ctdssmap.com/CTPortal/Provider')
        except Exception as e:                        
            try:
                options =  webdriver.ChromeOptions()        
                options.add_argument('--ignore-certificate-errors')
                options.add_argument('--ignore-ssl-errors')
                options.add_argument("--disable-popup-blocking")
                # response = requests.get('https://chromedriver.storage.googleapis.com/LATEST_RELEASE')
                # latest_version = response.text.strip()
                # chrome_driver_url = f'https://chromedriver.storage.googleapis.com/{latest_version}/chromedriver_win32.zip'
                # response = requests.get(chrome_driver_url)
                # with open('chromedriver_win32.zip', 'wb') as f:
                #     f.write(response.content)
                # with ZipFile('chromedriver_win32.zip', 'r') as zip_ref:
                #     zip_ref.extractall('.')   
                # driver_path=r'C:\Users\sanandrao\Desktop\New folder\chromedriver.exe'
                driver_path = os.path.abspath('chromedriver.exe')                
                driver = webdriver.Chrome(executable_path=driver_path)                
                driver.maximize_window()
                driver.get('https://www.ctdssmap.com/CTPortal/Provider')                                         
            except Exception as e:
                messagebox.showinfo("Driver Problem","Pls Check Your Driver Version")
                sys.exit(0)
                        
        def text_box(xpath,heding,status,key):                
            counter = 0
            while counter < 15:
                try:   
                    WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()                   
                    WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding, status)
                sys.exit(0)        
        
        def text_box(xpath,heding,status,key):                
            counter = 0
            while counter < 15:
                try:   
                    WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()                   
                    WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding, status)
                sys.exit(0)    
        
        def text_box_key(xpath,heding,status,key):                
            counter = 0
            while counter < 15:
                try:   
                    element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, xpath)))
                    element.send_keys(key)
                    element.send_keys(Keys.TAB)
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding, status)
                sys.exit(0)    


        def click(xpath,heding,status):
            counter = 0
            while counter < 15:
                try:             
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding,status)
                sys.exit(0)                                

        def count(xpath,heding,status):
            global rows
            counter = 0
            while counter < 15:
                try:             
                    rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))                    
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding,status)
                sys.exit(0)                         

        def count_1(xpath,heding,status):
            global rows_1
            counter = 0
            while counter < 15:
                try:             
                    rows_1=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))                    
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding,status)
                sys.exit(0)   

        def Alert():            
            try:             
                WebDriverWait(driver, 1).until (EC.alert_is_present())
                a=driver.switch_to.alert
                a.accept()                
            except Exception as e:
                pass            
                
        xpath= '/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[3]/div[1]/div/div/div/fieldset/table/tbody/tr[1]/td[2]/input'
        heding="User ID"
        status="User ID Field Not Found"
        key=user_name
        text_box(xpath,heding,status,key)
        
        xpath= '/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[3]/div[1]/div/div/div/fieldset/table/tbody/tr[2]/td[2]/input'
        heding="Password"
        status="Password Field Not Found"
        key=password
        text_box(xpath,heding,status,key)

        xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[3]/div[1]/div/div/div/fieldset/table/tbody/tr[3]/td/a"
        heding="Login"
        status="Login Button Not Found"
        click(xpath,heding,status)
          
        page_title = driver.title

        if page_title.lstrip().rstrip()=='Secure Site':                               
            messagebox.showinfo('Login Status','Please Check - UserName or Password Wrong')                    
            sys.exit(0)                                                      
                       
        file=pd.read_excel(fil,sheet_name='Medicaid Submission',header=0)        

        file['AVRS ID#'] = pd.to_numeric(file['AVRS ID#'], errors='coerce').astype('float').astype('Int64')

        for index, row in file.iterrows():                          
            c_id = row[0]  
            add_zer=row[1]
            p_acn = row[2]            
            dx = row[3]                 
            f_dos=row[4]                                    
            pro=row[5]
            mod=row[6]
            ftc=row[7]
            amt=row[8]
            ren_phy=row[9]
            mp_dt=row[10]
            mc_amt=row[11]
            mp_amt=row[12]
            md_amt=row[13]
            mcins_amt=row[14]
            avrs=row[15]

            avrs_1 = '00' + str(avrs)
            
            wait = WebDriverWait(driver, 20)
            wait.until(lambda driver: driver.execute_script("return document.readyState === 'complete';"))
            
            xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[2]/tbody/tr/td/div[1]/ul/li"
            heding="Account"
            status="Account Tab Count Not Found"        
            count(xpath,heding,status)

            xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[2]/tbody/tr/td/div[1]/ul/li[{}]"

            j=1
            while j<rows+1:                                
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text            
                if cnm.lstrip().rstrip()=="Account":
                    
                    time.sleep(1)

                    element=WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j))))
                    actions = ActionChains(driver)
                    actions.move_to_element(element).perform()  

                    xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[2]/tbody/tr/td/div[1]/ul/li[{}]/ul/li".format(j)
                    heding="Account Count"
                    status="Sub Account Count Not Found"        
                    count_1(xpath,heding,status)                
                        
                    i=1
                    while i<rows_1+1:                                       
                        xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[2]/tbody/tr/td/div[1]/ul/li[{}]/ul/li[{}]"
                        element_7=WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j,i))))
                        actions = ActionChains(driver)
                        actions.move_to_element(element_7).perform()  
                        element_text = element_7.text                    
                        if element_text.lstrip().rstrip()=="Switch Provider":
                            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j,i)))).click()  
                            break                       
                        i=i+1                                                              
                    break
                j=j+1                            
            
            Alert()

            try:
                element_5=WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/table/tbody/tr')))

                xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/table/tbody/tr"
                heding="Switch Provider Table"
                status="Switch Provider Table Count Not Found"        
                count(xpath,heding,status)  
                
                j=2
                while j<rows+1:     
                    xpath="/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/table/tbody/tr[{}]/td[3]"    
                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                    if cnm.lstrip().rstrip()==avrs_1.lstrip().rstrip():
                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()  

                        xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/span[2]/table/tbody/tr/td/table/tbody/tr[1]/td[5]/table/tbody/tr/td/a"
                        heding="Switch"
                        status="Switch Button Not Found"
                        click(xpath,heding,status)

                        Alert()

                        break
                    j=j+1
            except Exception as e:
                pass

            xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[2]/tbody/tr/td/div[1]/ul/li"
            heding="Claims"
            status="Claims Tab Count Not Found"        
            count(xpath,heding,status)

            xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[2]/tbody/tr/td/div[1]/ul/li[{}]"

            j=1            
            while j<rows+1:                                
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text            
                if cnm.lstrip().rstrip()=="Claims":
                    
                    time.sleep(1)

                    element=WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j))))
                    actions = ActionChains(driver)
                    actions.move_to_element(element).perform()  

                    xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[2]/tbody/tr/td/div[1]/ul/li[{}]/ul/li".format(j)
                    heding="Sub Claims Count"
                    status="Sub Claims Count Not Found"        
                    count_1(xpath,heding,status)                
                        
                    i=1
                    while i<rows_1+1:                                       
                        xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[2]/tbody/tr/td/div[1]/ul/li[{}]/ul/li[{}]"
                        element_7=WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j,i))))
                        actions = ActionChains(driver)
                        actions.move_to_element(element_7).perform()  
                        element_text = element_7.text                    
                        if element_text.lstrip().rstrip()=="Professional":
                            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j,i)))).click()  
                            break                       
                        i=i+1                                                              
                    break
                j=j+1       
            
            if pd.isnull(add_zer):
                c_id = str(c_id)
            elif add_zer.lower().lstrip().rstrip()=='add':
                c_id ='00' + str(c_id)
            else:
                c_id = str(c_id)

            Alert()

            xpath= '/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/span/table/tbody/tr[1]/td/table/tbody/tr[4]/td[2]/input'
            heding="Client ID"
            status="Client ID Field Not Found"
            key=c_id
            text_box(xpath,heding,status,key)

            xpath= '/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/span/table/tbody/tr[1]/td/table/tbody/tr[8]/td[2]/input'
            heding="Patient Account"
            status="Patient Account Number Field Not Found"
            key=p_acn
            text_box(xpath,heding,status,key)

            xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/span/table/tbody/tr[1]/td/table/tbody/tr[10]/td[4]/select/option"
            heding="Medicare Crossover"
            status="Medicare Crossover Count Not Found"        
            count(xpath,heding,status)

            j=1
            while j<rows+1:                                       
                xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/span/table/tbody/tr[1]/td/table/tbody/tr[10]/td[4]/select/option[{}]"
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                if cnm.lstrip().rstrip()=='Yes':
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()  
                    break                       
                j=j+1
            
            xpath= '/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[1]/span/table/tbody/tr/td/table/tbody/tr[2]/td[2]/span/input'
            heding="Diagnosis"
            status="Principal Field Not Found"
            key=dx
            text_box(xpath,heding,status,key)

            date_object_1 = datetime.strptime(str(f_dos), "%Y-%m-%d %H:%M:%S")
            dob3 = date_object_1.strftime("%m/%d/%Y")

            xpath= '/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[2]/td[2]/input'
            heding="Detail"
            status="From DOS Field Not Found"
            key=dob3
            text_box(xpath,heding,status,key)

            xpath= '/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[4]/td[2]/span/input'
            heding="Detail"
            status="Procedure Field Not Found"
            key=pro
            text_box(xpath,heding,status,key)

            if pd.isnull(mod):
                mod = str(mod)
            else: 
                xpath= '/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[5]/td[2]/span[1]/input'
                heding="Detail"
                status="Modifiers Field Not Found"
                key=mod
                text_box(xpath,heding,status,key)

            xpath= '/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[8]/td[2]/span/input'
            heding="Detail"
            status="Facility Type Code Field Not Found"
            key=ftc
            text_box(xpath,heding,status,key)
            
            xpath= '/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[9]/td[2]/input'
            heding="Detail"
            status="Charges Field Not Found"
            key=amt
            text_box(xpath,heding,status,key)

            xpath= '/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[10]/td[2]/input[1]'
            heding="Detail"
            status="Rendering Physician Field Not Found"
            key=ren_phy
            text_box(xpath,heding,status,key)
            
            date_object_1 = datetime.strptime(str(mp_dt), "%Y-%m-%d %H:%M:%S")
            dob3 = date_object_1.strftime("%m/%d/%Y")

            xpath= '/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[8]/td[4]/input'
            heding="Detail"
            status="Medicare Paid Date Field Not Found"
            key=dob3
            text_box(xpath,heding,status,key)

            xpath= '/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[9]/td[4]/input'
            heding="Detail"
            status="Medicare Calc Allowed Amt Field Not Found"
            key=mc_amt
            text_box(xpath,heding,status,key)

            xpath= '/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[10]/td[4]/input'
            heding="Detail"
            status="Medicare Paid Amount Field Not Found"
            key=mp_amt
            text_box(xpath,heding,status,key)

            xpath= '/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[11]/td[4]/input'
            heding="Detail"
            status="Medicare Deductible Amount Field Not Found"
            key=md_amt
            text_box(xpath,heding,status,key)

            xpath= '/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[3]/span/table/tbody/tr/td/table/tbody/tr[12]/td[4]/input'
            heding="Detail"
            status="Medicare Coinsurance Amount Field Not Found"
            key=mcins_amt
            text_box(xpath,heding,status,key)
            
            xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/span[2]/table/tbody/tr/td[3]/table/tbody/tr/td[1]/a"
            heding="Submit"
            status="Submit Button Not Found"
            click(xpath,heding,status)

            wait = WebDriverWait(driver, 20)
            wait.until(lambda driver: driver.execute_script("return document.readyState === 'complete';"))

            xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[5]/span/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[1]/td[2]/input"   
            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath)))
            cnm = cnm.get_attribute("value") 
            
            if cnm.lstrip().rstrip()=='Not Submitted yet':                    
                wb1=load_workbook(filename=fil)
                sheet = wb1['Medicaid Submission']
                column_letter = 'AF'  
                column_cells = sheet[column_letter]
                last_row = None
                for cell in reversed(column_cells):
                    if cell.value:
                        last_row = cell.row
                        break                            
                sheet['AF' + str(int(last_row + 1))]='Error'                
                wb1.save(fil)
                wb1.close()   
            else:      
                
                wb1=load_workbook(filename=fil)
                sheet = wb1['Medicaid Submission']
                column_letter = 'AF'  
                column_cells = sheet[column_letter]
                last_row = None
                for cell in reversed(column_cells):
                    if cell.value:
                        last_row = cell.row
                        break         
                
                xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[5]/span/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[1]/td[2]/input"
                cs=WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH,xpath)))  
                cs = cs.get_attribute("value") 
                sheet['S' + str(int(last_row + 1))]=cs

                if cs=='DENIED':
                    try:
                        xpath="/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[5]/table/tbody/tr[3]/td[2]"                                 
                        cd1=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text
                        
                        xpath="/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[5]/table/tbody/tr[3]/td[3]"                                 
                        dec1=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text
                        
                        sheet['V' + str(int(last_row + 1))]=cd1
                        sheet['W' + str(int(last_row + 1))]=dec1
                    except Exception as e:
                        pass
                    
                    try:
                        xpath="/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[5]/table/tbody/tr[4]/td[2]"                                 
                        cd2=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text
                        
                        xpath="/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[5]/table/tbody/tr[4]/td[3]"                                 
                        dec2=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text

                        sheet['X' + str(int(last_row + 1))]=cd2
                        sheet['Y' + str(int(last_row + 1))]=dec2
                    except Exception as e:
                        pass
                    
                    try:
                        xpath="/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[5]/table/tbody/tr[5]/td[2]"                                 
                        cd3=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text
                        
                        xpath="/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[5]/table/tbody/tr[5]/td[3]"                                 
                        dec3=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text

                        sheet['Z' + str(int(last_row + 1))]=cd3
                        sheet['AA' + str(int(last_row + 1))]=dec3
                    except Exception as e:
                        pass
                    
                    try:
                        xpath="/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[5]/table/tbody/tr[6]/td[2]"                                 
                        cd4=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text
                        
                        xpath="/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[5]/table/tbody/tr[6]/td[3]"                                 
                        dec4=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text

                        sheet['AB' + str(int(last_row + 1))]=cd4
                        sheet['AC' + str(int(last_row + 1))]=dec4
                    except Exception as e:
                        pass

                    try:
                        xpath="/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[5]/table/tbody/tr[7]/td[2]"                                 
                        cd5=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text
                        
                        xpath="/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[5]/table/tbody/tr[7]/td[3]"                                 
                        dec5=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text

                        sheet['AD' + str(int(last_row + 1))]=cd5
                        sheet['AE' + str(int(last_row + 1))]=dec5
                    except Exception as e:
                        pass

                xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[5]/span/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[2]/td[2]/input"
                c_icn=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j))))
                c_icn = c_icn.get_attribute("value") 
                sheet['T' + str(int(last_row + 1))]=c_icn

                xpath= "/html/body/form/div[3]/table/tbody/tr[1]/td[2]/table[3]/tbody/tr/td/table/tbody/tr/td[2]/div[2]/div/div/div/div/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td/div/div[5]/span/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[5]/td[2]/input"
                p_amt=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j))))
                p_amt = p_amt.get_attribute("value") 
                sheet['U' + str(int(last_row + 1))]=p_amt
                                                                                                                                                                         
                sheet['AF' + str(int(last_row + 1))]='Done'                
                wb1.save(fil)
                wb1.close()                                             
        
        driver.quit()
        messagebox.showinfo("Process Status", "Process Completed")
        sys.exit(0)                                             

if __name__=="__main__":        
    primary()
   