from tkinter import BOTH, LEFT, TOP, Button, Entry, Frame, Label, PhotoImage, StringVar, Tk,Radiobutton,StringVar,IntVar,filedialog
from idlelib.tooltip import Hovertip
import sys
import os
from operator import itemgetter
import time
from tkinter import messagebox
from os import listdir
from os.path import isfile, join
from datetime import datetime, date
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, StaleElementReferenceException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import openpyxl
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from urllib.parse import urlparse
import warnings
import numpy as np
from openpyxl.utils import get_column_letter,column_index_from_string
from datetime import datetime
from pytz import timezone
import requests
from zipfile import ZipFile
import subprocess

warnings.filterwarnings("ignore")

global rows
global xpath
global heding
global status
global key
global nme
element_1 = None
global j


def browse():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
        image_path2 = os.path.join(sys._MEIPASS, 'Static', 'Mapping.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')
        image_path2 = os.path.join(os.getcwd(), 'Static', 'Mapping.png')

    root.title("Process File Picker")
    root.resizable(False,False)

    root.title("Process File Picker")
    root.resizable(False,False)
   
    w = 500
    h = 160
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Browse Process - File Picker",font=("Calibri",18,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
          
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=50)
    
    answer=StringVar()
    answer.set("")
       
    def browse_button():
        global filename
        answer.set("")  
        filename = filedialog.askopenfilename()  
        txt.delete(0, 'end') 
        txt.insert(0, filename)

    def Click_Done():
        global ans        
        
        ans=txt.get()        
        
        if ans=="":
           answer.set("File Path Fields Empty Is Not Allowed...")        
        else:      
            root.destroy()          
            return ans
        
    txt=Entry(Frame2,font=("Calibri",12,"bold","italic"),width=50,justify="left")
    txt.grid(row=0,column=0,padx=5,pady=10,sticky="E")    

    photo1 = PhotoImage(file=image_path2)
            
    btn1=Button(Frame2,text="Browse",command=browse_button,image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=0,column=1,padx=3,pady=0,sticky="W")     
                
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=3,y=80,width=490,height=25)
    
    title_3=Label(Frame3,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=1,padx=150,pady=0,sticky="E")

    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=3,y=105,width=500,height=50)
    
    photo2 = PhotoImage(file=image_path1)

    btn2=Button(Frame4,command=Click_Done,text="Done",image=photo2,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=3,column=0,padx=125,pady=0,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo3 = PhotoImage(file=image_path)   

    btn3=Button(Frame4,command=Close,text="Exit",image=photo3,borderwidth=0,bg="#2c3e50")
    btn3.grid(row=3,column=1,padx=15,pady=0,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn2,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn3,'Click to Exit Process',hover_delay=1000)
    myTip2 = Hovertip(btn1,'Click to Pick File',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def process():        
    global element_1    
    global ans1
    global ans2 
    global driver
    global j
    global dob
    global mid
    global frm_dt
    global to_dt
    global cpt
    global rows
    global rows1
    
    browse()
    
    fil=ans      

    current_directory = os.getcwd()
    # current_directory += os.path.sep
    check_path=os.path.join(current_directory+'\\Temp')
    folder_create=os.path.join(current_directory+'\\')
    down_path=os.path.join(current_directory+'\\Temp\\')
    fin_file=os.path.join(current_directory+'\\Temp\\Files\\')
   
    fld='Temp'
    
    if not os.path.isdir(check_path):
        os.mkdir(folder_create + 'Temp')
        os.mkdir(down_path + 'Files')
    else:
        path1 = os.path.join(folder_create, fld)
        shutil.rmtree(path1)
        os.mkdir(folder_create + 'Temp')
        os.mkdir(down_path + 'Files')  

    try:
        options = Options()            
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--ignore-ssl-errors')
        # options.add_argument("--disable-popup-blocking")
        # options.add_argument('--disable-crl-sets')
        # options.add_argument('--headless')
        prefs = {"download.default_directory" : down_path,
                    "download.prompt_for_download": False,
                    "download.directory_upgrade": True,
                    "plugins.always_open_pdf_externally": True,
                    "profile.password_manager_enabled": False,
                    "credentials_enable_service": False,
                    # "safebrowsing.enabled": True
                    }
        # options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_experimental_option("prefs",prefs)
        driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options)
        driver.maximize_window()
        driver.get('https://promise.dpw.state.pa.us/portal/provider/Home/tabid/135/Default.aspx')
    except Exception as e:                        
        try:
            options = Options()
            options.add_argument('--ignore-certificate-errors')
            options.add_argument('--ignore-ssl-errors')
            # options.add_argument("--disable-popup-blocking")
            # options.add_argument('--disable-crl-sets')
            # options.add_argument('--headless')
            prefs = {"download.default_directory" : down_path,
                    "download.prompt_for_download": False,
                    "download.directory_upgrade": True,
                    "plugins.always_open_pdf_externally": True,
                    "profile.password_manager_enabled": False,
                    "credentials_enable_service": False,
                    # "safebrowsing.enabled": True
                    }
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_experimental_option("useAutomationExtension", False)
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option("prefs",prefs)            
            driver_path = os.path.abspath('chromedriver.exe')
            driver = webdriver.Chrome(executable_path=driver_path,options=options)
            driver.maximize_window()            
        except Exception as e:           
            messagebox.showinfo("Driver Problem","Pls Check Your Chrome Driver Version")
            sys.exit(0)
        
    def text_box(xpath,heding,status,key):                
        counter = 0
        while counter < 5:
            try:   
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()                     
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            print('Moved Exception')
            raise e
            # messagebox.showinfo(heding, status)
            # sys.exit(0)        
    
    def text_box_key(xpath,heding,status,key):                
        counter = 0
        while counter < 5:
            try:                   
                element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, xpath)))
                element.clear()                
                element.send_keys(key)                
                # element.send_keys(Keys.TAB)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            print('Moved Exception')
            raise e
            # messagebox.showinfo(heding, status)
            # sys.exit(0)    

    def text_box_js(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:
                js_code = "document.querySelector('" + xpath + "').setAttribute('value', '" + key + "');"
                WebDriverWait(driver, 0).until(lambda driver: driver.execute_script(js_code))               
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo(heding, status)
            sys.exit(0)       

    def click(xpath,heding,status):
        counter = 0
        while counter < 5:
            try:             
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            print('Moved Exception')
            raise e
            # messagebox.showinfo(heding,status)
            # sys.exit(0)                                

    def count(xpath,heding,status):
        global rows
        counter = 0
        while counter < 5:
            try:             
                rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            print('Moved Exception')
            raise e
            # messagebox.showinfo(heding,status)
            # sys.exit(0)                         

    def count_1(xpath,heding,status):
        global rows1
        counter = 0
        while counter < 5:
            try:             
                rows1=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            print('Moved Exception')
            raise e
            # messagebox.showinfo(heding,status)
            # sys.exit(0)        

    def Alert():
        counter = 0
        while counter < 15:
            try:             
                WebDriverWait(driver, 0).until (EC.alert_is_present())
                a=driver.switch_to.alert
                a.accept()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo('Alert','Alert Not Present')                                                                
    
    # messagebox.showinfo('Waiting','Authentication Waiting')        
    
    # driver.switch_to.window(driver.window_handles[1])      

    file=pd.read_excel(fil,sheet_name='Matrix',header=0)
    
    for index, row in file.iterrows():                                      
        lnk = row[0]
        unm=row[1]
        psw = row[2]                     
        fec_nm = row[4]
        st=row[5]
        lst_nm = row[6]
        fst_nm = row[7]
        
        driver.get(lnk.lstrip().rstrip())
                      
        xpath= "/html/body/div/div/main/div/div/div[2]/div/div/form/div[1]/div[2]/div[1]/div[2]/span/input"
        heding="User Name"
        status="User Name Field Not Found"
        key=unm.lstrip().rstrip()
        text_box(xpath,heding,status,key)

        xpath= "/html/body/div/div/main/div/div/div[2]/div/div/form/div[1]/div[2]/div[2]/div[2]/span/input"
        heding="Password"
        status="Password Field Not Found"
        key=psw.lstrip().rstrip()
        text_box(xpath,heding,status,key)
        
        xpath='/html/body/div/div/main/div/div/div[2]/div/div/form/div[2]/input'
        heding="Sign In"
        status="Sign In Button Not Found"        
        click(xpath,heding,status)

        counter = 0
        while counter < 5:
            try:
                xpath= "/html/body/div/div/main/div/div/div[2]/div/div/form/div[1]/div[1]/div/div/p"
                ck=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text
                
                wb1=load_workbook(filename=fil)
                sheet = wb1['Matrix']
                column_letter = 'I'  
                column_cells = sheet[column_letter]
                last_row = None
                for cell in reversed(column_cells):
                    if cell.value:
                        last_row = cell.row
                        break                            
                sheet['I' + str(int(last_row + 1))]=ck 
                wb1.save(fil)
                wb1.close()                                    

                break 
            except Exception as e:
                try:                   
                    xpath='/html/body/div[1]/nav[1]/div[3]/div/ul[2]/li'
                    ck1=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text
                    if ck1.lstrip().rstrip()=='Log off':
                        
                        xpath= "/html/body/div[1]/nav[2]/div/div[2]/ul/li"
                        heding="Facility Count"        
                        status="Facility Count Not Found"                    
                        count(xpath,heding,status)
                        
                        j=1
                        while j<rows+1:  
                            sck=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/nav[2]/div/div[2]/ul/li[{}]/a'.format(j)))).text
                            if sck.lstrip().rstrip()=='Facility':
                                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[1]/nav[2]/div/div[2]/ul/li[{}]/a'.format(j)))).click()
                                                                      
                                xpath= "/html/body/div[1]/nav[2]/div/div[2]/ul/li[{}]/ul/div/div/li/ul/li".format(j)
                                heding="Facility Search Count"        
                                status="Facility Search Count Not Found"                    
                                count_1(xpath,heding,status)
                                                
                                i=1
                                while i<rows1+1:  
                                    sck=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/nav[2]/div/div[2]/ul/li[{}]/ul/div/div/li/ul/li[{}]'.format(j,i)))).text
                                    if sck.lstrip().rstrip()=='Search Facility':
                                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[1]/nav[2]/div/div[2]/ul/li[{}]/ul/div/div/li/ul/li[{}]/a'.format(j,i)))).click()
                                        break
                                    i=i+1
                                break
                            j=j+1

                        xpath= "/html/body/div[1]/main/form/div[1]/fieldset/div[2]/input"
                        heding="Facility Name"
                        status="Facility Name Field Not Found"
                        key=fec_nm.lstrip().rstrip()
                        text_box(xpath,heding,status,key)

                        xpath= "/html/body/div[1]/main/form/div[1]/fieldset/div[2]/select[2]/option"
                        heding="State Count"        
                        status="State Count Not Found"                    
                        count(xpath,heding,status)
                        
                        j=1
                        while j<rows+1:  
                            sck=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/form/div[1]/fieldset/div[2]/select[2]/option[{}]'.format(j)))).text
                            if sck.lstrip().rstrip()==st.lstrip().rstrip():
                                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[1]/main/form/div[1]/fieldset/div[2]/select[2]/option[{}]'.format(j)))).click()
                                break
                            j=j+1
                    
                        xpath='/html/body/div[1]/main/form/div[1]/fieldset/div[3]/input'
                        heding="Facility Search"
                        status="Facility Search Button Not Found"        
                        click(xpath,heding,status)

                        time.sleep(2)

                        try:           
                            xpath='/html/body/div[1]/main/form/table/tbody/tr/td[1]/a'
                            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                            
                            try:                
                                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[1]/form/div[1]/a[1]'))).click()
                                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[1]/form/div[2]/input'))).click()

                                xpath= "/html/body/header/nav[2]/div/ul/li"
                                heding="Resident Count"        
                                status="Resident Count Not Found"                    
                                count(xpath,heding,status)

                                j=1
                                while j<rows+1:  
                                    rec=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/header/nav[2]/div/ul/li[{}]'.format(j)))).text
                                    if rec.lstrip().rstrip()=='Resident':
                                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/header/nav[2]/div/ul/li[{}]/a'.format(j)))).click()
                                        break
                                    j=j+1

                                xpath= "/html/body/header/nav[2]/div/ul/li[4]/ul/div/div/li/ul/li"
                                heding="Search Resident Count"        
                                status="Search Resident Count Not Found"                    
                                count(xpath,heding,status)

                                j=1
                                while j<rows+1:  
                                    rec_ser=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/header/nav[2]/div/ul/li[4]/ul/div/div/li/ul/li[{}]'.format(j)))).text
                                    if rec_ser.lstrip().rstrip()=='Search Resident':
                                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/header/nav[2]/div/ul/li[4]/ul/div/div/li/ul/li[{}]/a'.format(j)))).click()
                                        break
                                    j=j+1

                                xpath= "/html/body/div[1]/main/form/fieldset[1]/div[1]/input"
                                heding="Last Name"
                                status="Last Name Field Not Found"
                                key=lst_nm.lstrip().rstrip()
                                text_box(xpath,heding,status,key)
                                
                                xpath= "/html/body/div[1]/main/form/fieldset[1]/div[2]/input"
                                heding="First Name"
                                status="First Name Field Not Found"
                                key=fst_nm.lstrip().rstrip()
                                text_box(xpath,heding,status,key)
                                                                                                        
                                WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/form/fieldset[2]/div[1]/input'))).clear()

                                xpath='/html/body/div[1]/main/form/fieldset[7]/section/input'
                                heding="Search Residents"
                                status="Search Residents Button Not Found"        
                                click(xpath,heding,status)
                                
                                time.sleep(2)
                                    
                                xpath='/html/body/div[1]/main/form[2]/table/tbody/tr/td[1]/a'
                                WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()                                
                                     
                                xpath='/html/body/div[1]/main/div[3]/div[1]/button[1]'
                                heding="File Download"
                                status="File Download Button Not Found"        
                                click(xpath,heding,status)

                                counter=1
                                while counter < 30:                    
                                    file1 = [f for f in listdir(down_path) if isfile(join(down_path, f))]
                                    lol_string1 = ' '.join(file1)
                                    lol_string2=lol_string1.split('.')[-1]                                                                                
                                    if lol_string2!='' and lol_string2!='tmp' and lol_string2 !="crdownload" and lol_string2 !="temp":
                                        time.sleep(1)
                                        file2 = [f for f in os.listdir(down_path) if os.path.isfile(os.path.join(down_path, f))]
                                        if file2:
                                            lol_string3 = os.path.join(down_path, file2[0])  
                                            directory, file_without_extension = os.path.split(lol_string3)
                                            chng_fn = fst_nm + ' ' + lst_nm + '.pdf'
                                            new_file_path = os.path.join(directory, chng_fn)
                                            os.rename(lol_string3,new_file_path)

                                            shutil.move(new_file_path, os.path.join(fin_file, chng_fn))
                                                                                     
                                            wb1=load_workbook(filename=fil)
                                            sheet = wb1['Matrix']
                                            column_letter = 'I'  
                                            column_cells = sheet[column_letter]
                                            last_row = None
                                            for cell in reversed(column_cells):
                                                if cell.value:
                                                    last_row = cell.row
                                                    break                            
                                            sheet['I' + str(int(last_row + 1))]='Done'      
                                            wb1.save(fil)
                                            wb1.close()                                                
                                            break
                                    else:
                                        time.sleep(1)    
                                        counter += 1
                                else:
                                    files = os.listdir(check_path)
                                    for file_name in files:
                                            file_path = os.path.join(check_path, file_name)
                                            try:
                                                if os.path.isfile(file_path):
                                                    os.remove(file_path)                                   
                                            except OSError as e:
                                                pass    
                                    wb1=load_workbook(filename=ans)
                                    sheet = wb1['Matrix']
                                    column_letter = 'I'  
                                    column_cells = sheet[column_letter]
                                    last_row = None
                                    for cell in reversed(column_cells):
                                        if cell.value:
                                            last_row = cell.row
                                            break                            
                                    sheet['I' + str(int(last_row + 1))]='File Error'        
                                    wb1.save(fil)
                                    wb1.close()                                       

                                counter = 0
                                while counter < 8:
                                    try:                                                     
                                        main_window_handle = driver.window_handles[0]
                                        popup_window_handle = None

                                        for handle in driver.window_handles:
                                            if handle != main_window_handle:
                                                popup_window_handle = handle
                                        driver.switch_to.window(popup_window_handle)                                                                                                                    
                                        driver.close()
                                        driver.switch_to.window(main_window_handle)
                                        break
                                    except Exception as e:
                                        time.sleep(1)
                                        counter += 1                                    
                                else:
                                    driver.switch_to.window(main_window_handle) 
                                                                                                                                 
                                xpath='/html/body/div[1]/nav[1]/div[3]/div/ul[2]/li/a'
                                heding="Log Off"
                                status="Log Off Button Not Found"        
                                click(xpath,heding,status)

                                break                                  
                            except Exception as e:
                                # Patten 2                                  
                                try:
                                    xpath= "/html/body/header/nav[2]/div/ul/li"
                                    rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                                except Exception as e:
                                    xpath= "/html/body/div[1]/nav[2]/div/div[2]/ul/li"
                                    rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))                                

                                j=1
                                while j<rows+1: 
                                    try: 
                                        rec=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/header/nav[2]/div/ul/li[{}]'.format(j)))).text
                                    except Exception as e:
                                        rec=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/nav[2]/div/div[2]/ul/li[{}]'.format(j)))).text

                                    if rec.lstrip().rstrip()=='Resident':
                                        try:
                                            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/header/nav[2]/div/ul/li[{}]/a'.format(j)))).click()
                                        except Exception as e:
                                            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[1]/nav[2]/div/div[2]/ul/li[{}]/a'.format(j)))).click()
                                    
                                        try:
                                            xpath= "/html/body/header/nav[2]/div/ul/li[4]/ul/div/div/li/ul/li"
                                            rows1=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                                        except Exception as e:
                                            xpath= "/html/body/div[1]/nav[2]/div/div[2]/ul/li[5]/ul/div/div/li/ul/li"
                                            rows1=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                                        
                                        i=1
                                        while i<rows1+1:  
                                            try:                                                                                                                                
                                                rec_ser=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/header/nav[2]/div/ul/li[{}]/ul/div/div/li/ul/li[{}]'.format(j,i)))).text
                                            except Exception as e:
                                                rec_ser=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/nav[2]/div/div[2]/ul/li[{}]/ul/div/div/li/ul/li[{}]'.format(j,i)))).text

                                            if rec_ser.lstrip().rstrip()=='Search Resident':
                                                try:
                                                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/header/nav[2]/div/ul/li[{}]/ul/div/div/li/ul/li[{}]/a'.format(j,i)))).click()
                                                except Exception as e:
                                                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[1]/nav[2]/div/div[2]/ul/li[{}]/ul/div/div/li/ul/li[{}]/a'.format(j,i)))).click()
                                                break
                                            j=j+1                                                                                                            
                                        break
                                    j=j+1                                                            
                                    
                                xpath= "/html/body/div[1]/main/form/fieldset[1]/div[1]/input"
                                heding="Last Name"
                                status="Last Name Field Not Found"
                                key=lst_nm.lstrip().rstrip()
                                text_box(xpath,heding,status,key)
                                
                                xpath= "/html/body/div[1]/main/form/fieldset[1]/div[2]/input"
                                heding="First Name"
                                status="First Name Field Not Found"
                                key=fst_nm.lstrip().rstrip()
                                text_box(xpath,heding,status,key)

                                WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[1]/main/form/fieldset[2]/div[1]/input'))).clear()

                                xpath='/html/body/div[1]/main/form/fieldset[7]/section/input'
                                heding="Search Residents"
                                status="Search Residents Button Not Found"        
                                click(xpath,heding,status)
                                
                                time.sleep(2)
                                    
                                xpath='/html/body/div[1]/main/form[2]/table/tbody/tr/td[1]/a'
                                WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()

                                xpath='/html/body/div[1]/main/div[3]/div[1]/button[1]'
                                heding="File Download"
                                status="File Download Button Not Found"        
                                click(xpath,heding,status)

                                counter=1
                                while counter < 15:                    
                                    file1 = [f for f in listdir(down_path) if isfile(join(down_path, f))]
                                    lol_string1 = ' '.join(file1)
                                    lol_string2=lol_string1.split('.')[-1]                                                                                
                                    if lol_string2!='' and lol_string2!='tmp' and lol_string2 !="crdownload" and lol_string2 !="temp":
                                        time.sleep(1)
                                        file2 = [f for f in os.listdir(down_path) if os.path.isfile(os.path.join(down_path, f))]
                                        if file2:
                                            lol_string3 = os.path.join(down_path, file2[0])  
                                            directory, file_without_extension = os.path.split(lol_string3)
                                            chng_fn = fst_nm + ' ' + lst_nm + '.pdf'
                                            new_file_path = os.path.join(directory, chng_fn)
                                            os.rename(lol_string3,new_file_path)

                                            shutil.move(new_file_path, os.path.join(fin_file, chng_fn))
                                                                                     
                                            wb1=load_workbook(filename=fil)
                                            sheet = wb1['Matrix']
                                            column_letter = 'I'  
                                            column_cells = sheet[column_letter]
                                            last_row = None
                                            for cell in reversed(column_cells):
                                                if cell.value:
                                                    last_row = cell.row
                                                    break                            
                                            sheet['I' + str(int(last_row + 1))]='Done'      
                                            wb1.save(fil)
                                            wb1.close()                                                
                                            break
                                    else:
                                        time.sleep(1)    
                                        counter += 1
                                else:
                                    files = os.listdir(check_path)
                                    for file_name in files:
                                            file_path = os.path.join(check_path, file_name)
                                            try:
                                                if os.path.isfile(file_path):
                                                    os.remove(file_path)                                   
                                            except OSError as e:
                                                pass    
                                    wb1=load_workbook(filename=fil)
                                    sheet = wb1['Matrix']
                                    column_letter = 'I'  
                                    column_cells = sheet[column_letter]
                                    last_row = None
                                    for cell in reversed(column_cells):
                                        if cell.value:
                                            last_row = cell.row
                                            break                            
                                    sheet['I' + str(int(last_row + 1))]='File Error'        
                                    wb1.save(fil)
                                    wb1.close()                                       

                                counter = 0
                                while counter < 8:
                                    try:                                                     
                                        main_window_handle = driver.window_handles[0]
                                        popup_window_handle = None

                                        for handle in driver.window_handles:
                                            if handle != main_window_handle:
                                                popup_window_handle = handle
                                        driver.switch_to.window(popup_window_handle)                                                                                                                    
                                        driver.close()
                                        driver.switch_to.window(main_window_handle)
                                        break
                                    except Exception as e:
                                        time.sleep(1)
                                        counter += 1                                    
                                else:
                                    driver.switch_to.window(main_window_handle) 
                                                                                                                                 
                                xpath='/html/body/div[1]/nav[1]/div[3]/div/ul[2]/li/a'
                                heding="Log Off"
                                status="Log Off Button Not Found"        
                                click(xpath,heding,status)

                                break                                                
                        except Exception as e:
                            wb1=load_workbook(filename=fil)
                            sheet = wb1['Matrix']
                            column_letter = 'I'  
                            column_cells = sheet[column_letter]
                            last_row = None
                            for cell in reversed(column_cells):
                                if cell.value:
                                    last_row = cell.row
                                    break                            
                            sheet['I' + str(int(last_row + 1))]='Facility Table is Empty' 
                            wb1.save(fil)
                            wb1.close()  

                            xpath='/html/body/div[1]/nav[1]/div[3]/div/ul[2]/li/a'
                            heding="Log Off"
                            status="Log Off Button Not Found"        
                            click(xpath,heding,status)
                            
                            break
                except Exception as e:
                    time.sleep(1)
                    counter += 1   

    driver.quit()
    messagebox.showinfo("Process Status", "Process Completed")
    sys.exit(0)    

if __name__=="__main__":        
    process()
    