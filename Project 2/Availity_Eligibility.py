from tkinter import BOTH, LEFT, TOP, Button, Entry, Frame, Label, PhotoImage, StringVar, Tk,Radiobutton,StringVar,IntVar,filedialog
from idlelib.tooltip import Hovertip
from tkinter import messagebox
import sys
import os
import warnings
from datetime import datetime, date
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, StaleElementReferenceException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import openpyxl
import time
import pandas as pd
from openpyxl import load_workbook
from urllib.parse import urlparse
import warnings
import numpy as np
import requests
from zipfile import ZipFile
from openpyxl.utils import get_column_letter,column_index_from_string
warnings.filterwarnings("ignore")


global rows
global xpath
global heding
global status
global key
global nme
element_1 = None
global driver

def file_pick():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
        image_path2 = os.path.join(sys._MEIPASS, 'Static', 'Mapping.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')
        image_path2 = os.path.join(os.getcwd(), 'Static', 'Mapping.png')

    root.title("Process File Picker")
    root.resizable(False,False)
   
    w = 500
    h = 160
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Browse Process - File Picker",font=("Calibri",18,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
          
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=50)
    
    answer=StringVar()
    answer.set("")
       
    def browse_button():
        global filename
        answer.set("")  
        filename = filedialog.askopenfilename()  
        txt.delete(0, 'end') 
        txt.insert(0, filename)

    def Click_Done():
        global ans        
        
        ans=txt.get()        
        
        if ans=="":
           answer.set("File Path Fields Empty Is Not Allowed...")        
        else:      
            root.destroy()          
            return ans
        
    txt=Entry(Frame2,font=("Calibri",12,"bold","italic"),width=50,justify="left")
    txt.grid(row=0,column=0,padx=5,pady=10,sticky="E")    

    photo1 = PhotoImage(file=image_path2)
            
    btn1=Button(Frame2,text="Browse",command=browse_button,image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=0,column=1,padx=3,pady=0,sticky="W")     
                
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=3,y=80,width=490,height=25)
    
    title_3=Label(Frame3,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=1,padx=150,pady=0,sticky="E")

    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=3,y=105,width=500,height=50)
    
    photo2 = PhotoImage(file=image_path1)

    btn2=Button(Frame4,command=Click_Done,text="Done",image=photo2,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=3,column=0,padx=125,pady=0,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo3 = PhotoImage(file=image_path)   

    btn3=Button(Frame4,command=Close,text="Exit",image=photo3,borderwidth=0,bg="#2c3e50")
    btn3.grid(row=3,column=1,padx=15,pady=0,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn2,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn3,'Click to Exit Process',hover_delay=1000)
    myTip2 = Hovertip(btn1,'Click to Pick File',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def user_pass():
    global ans1
    global ans2
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')

    root.title("Sharepoint - User Login And File Details")
    root.resizable(False,False)

    w = 600
    h = 180
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)

    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="User Name and Password Details...",font=("Calibri",20,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack() 
 
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=698,height=150)    
    
    title3=Label(Frame2,text="User Name :",font=("Calibri",11,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title3.grid(row=0,column=0,padx=5,pady=5,sticky="W")
    
    txt1=Entry(Frame2,font=("Calibri",11,"bold","italic"),width=50,justify="left")
    txt1.grid(row=0,column=1,padx=30,pady=5,sticky="W")
    
    title4=Label(Frame2,text="Password :",font=("Calibri",11,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title4.grid(row=1,column=0,padx=5,pady=5,sticky="W")
    
    txt2=Entry(Frame2,font=("Calibri",11,"bold","italic"),width=50,justify="left")
    txt2.grid(row=1,column=1,padx=30,pady=5,sticky="W")

    answer=StringVar()
    answer.set("")

    def Click_Done():
        global ans1
        global ans2        
        
        ans1=txt1.get()
        ans2=txt2.get()        

        if ans1=="":
           answer.set("User Name Field Empty Is Not Allowed...")
        elif ans2=="":
            answer.set("Password Field Empty Is Not Allowed...")        
        else:    
            root.destroy()
            return ans1,ans2
                
    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=0,y=105,width=698,height=20)
    
    title_3=Label(Frame4,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=2,padx=200,pady=0,sticky="E")
    
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=0,y=125,width=698,height=200)
       
    photo1 = PhotoImage(file=image_path1)
    
    btn1=Button(Frame3,command=Click_Done,text="Run",image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=2,column=0,padx=140,pady=0,sticky="W")

    def Close():
        sys.exit(0)

    photo = PhotoImage(file=image_path)    

    btn2=Button(Frame3,command=Close,text="Close",image=photo,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=2,column=1,padx=100,pady=0,sticky="E")

    def disable_event():
        pass

    txt1.focus_set()

    myTip = Hovertip(btn1,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn2,'Click to Exit Process',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)
            
    root.mainloop()

def process():        
    global element_1    
    global ans1
    global ans2 
    global driver
    # sugan2022
    # Vital@2023%

    # user_pass()
    file_pick()

    # user_name =ans1
    # password=ans2        
    fil=ans       
    
    try:
        options = Options()            
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--ignore-ssl-errors')
        options.add_argument("--disable-popup-blocking")
        driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options)
        driver.maximize_window()
        driver.get('https://apps.availity.com/')
    except Exception as e:                        
        try:
            options = Options()            
            options.add_argument('--ignore-certificate-errors')
            options.add_argument('--ignore-ssl-errors')
            options.add_argument("--disable-popup-blocking")
            # response = requests.get('https://chromedriver.storage.googleapis.com/LATEST_RELEASE')
            # latest_version = response.text.strip()
            # chrome_driver_url = f'https://chromedriver.storage.googleapis.com/{latest_version}/chromedriver_win32.zip'
            # response = requests.get(chrome_driver_url)
            # with open('chromedriver_win32.zip', 'wb') as f:
            #     f.write(response.content)
            # with ZipFile('chromedriver_win32.zip', 'r') as zip_ref:
            #     zip_ref.extractall('.')
            driver_path = os.path.abspath('chromedriver.exe')
            driver = webdriver.Chrome(executable_path=driver_path,options=options)
            driver.maximize_window()
            driver.get('https://apps.availity.com/')
        except Exception as e:
            messagebox.showinfo("Internet Problem","Pls Check Your Internet Connection")
            sys.exit(0)
            
    def text_box(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:   
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()                     
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo(heding, status)
            sys.exit(0)        
    
    def text_box_key(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:   
                element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, xpath)))
                element.send_keys(key)
                time.sleep(1)
                element.send_keys(Keys.TAB)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo(heding, status)
            sys.exit(0)    

    def text_box_js(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:
                js_code = "document.querySelector('" + xpath + "').setAttribute('value', '" + key + "');"
                WebDriverWait(driver, 0).until(lambda driver: driver.execute_script(js_code))               
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo(heding, status)
            sys.exit(0)       

    def click(xpath,heding,status):
        counter = 0
        while counter < 15:
            try:             
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo(heding,status)
            sys.exit(0)                                

    def count(xpath,heding,status):
        global rows
        counter = 0
        while counter < 15:
            try:             
                rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))                    
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo(heding,status)
            sys.exit(0)                         

    def Alert():
        counter = 0
        while counter < 15:
            try:             
                WebDriverWait(driver, 0).until (EC.alert_is_present())
                a=driver.switch_to.alert
                a.accept()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo('Alert','Alert Not Present')                                                                
    
    messagebox.showinfo('Waiting','Authentication Waiting')
    
    xpath= "/html/body/navigation/div/bottom-nav/nav/div/div[2]/ul[1]/li[1]/a"
    heding="Patient Registration"
    status="Patient Registration Click Not Found"
    click(xpath,heding,status)

    xpath= "/html/body/navigation/div/bottom-nav/nav/div/div[2]/ul[1]/li[1]/ul/li/div[1]/div/div/div/div"
    heding="Patient Registration"
    status="Patient Registration Count Not Found"        
    count(xpath,heding,status)   
    
    j=1
    while j<rows+1:         
        xpath= "/html/body/navigation/div/bottom-nav/nav/div/div[2]/ul[1]/li[1]/ul/li/div[1]/div/div/div/div[{}]"                       
        cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
        if cnm.lstrip().rstrip()=="EB Eligibility and Benefits Inquiry":                
            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                            
            break
        j=j+1

    counter = 0
    while counter < 15:
        try: 
            iframe = WebDriverWait(driver, 0).until(EC.presence_of_element_located((By.ID, "newBodyFrame")))
            driver.switch_to.frame(iframe)
            break
        except Exception as e:
            time.sleep(1)
            counter += 1
    else:
            messagebox.showinfo('Iframe','Iframe-newBodyFrame Not Present') 
            sys.exit(0)   
    
    file=pd.read_excel(fil,sheet_name='Eligibility BOT',header=0)

    file['NPI'] = file['NPI'].astype(str)
    file['Insurance ID #'] = file['Insurance ID #'].astype(str)        

    for index, row in file.iterrows():                                      
        py_lnm = row[1]            
        py_fnm = row[2]                 
        dob=row[3]            
        npi=row[4]
        st=row[5]                       
        py_nm=row[6]
        in_id=row[7]
        frm_dt=row[8]           
        
        # xpath///*[@id="react-select-4-option-1-19"]
        
        xpath= "/html/body/div[1]/div/div/main/div/div/div/form/div/div[2]/div[2]/div/div/div/div/div[1]/div[2]/input"
        heding="Payer Name"
        status="Payer Name Field Not Found"
        key=py_nm.lstrip().rstrip()
        text_box_key(xpath,heding,status,key)

        xpath= "/html/body/div[1]/div/div/main/div/div/div/form/div/div[2]/div[2]/div/div/div/div/div[1]/div[1]"
        pn=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text

        if pn.lstrip().rstrip()==py_nm:
            pass
        else:
            messagebox.showinfo("Payer Name","Payer Name is Wrong - Please Check")
            sys.exit(0) 

        xpath="/html/body/div[1]/div/div/main/div/div/div/div[2]/form/div[1]/div[2]/div/div/div/div/div/div[1]/div[2]/input"
        heding="Provider Infromation"
        status="Provider Infromation Field Not Found"
        key=str(npi)
        text_box_key(xpath,heding,status,key)

               #/html/body/div[1]/div/div/main/div/div/div/div[2]/form/div[2]/div[4]/div[1]/div/input   
        xpath= "/html/body/div[1]/div/div/main/div/div/div/div[2]/form/div[2]/div[4]/div[1]/div/input"
        heding="Patient ID"
        status="Patient ID Field Not Found"
        key=str(in_id)
        text_box(xpath,heding,status,key)
        
        date_object_1 = datetime.strptime(str(dob), "%Y-%m-%d %H:%M:%S")
        dob3 = date_object_1.strftime("%m/%d/%Y")
            #  /html/body/div[1]/div/div/main/div/div/div/div[2]/form/div[2]/div[5]/div/div[2]/div[2]/div/div/div/div/div/div[1]/input
        xpath= "/html/body/div[1]/div/div/main/div/div/div/div[2]/form/div[2]/div[4]/div[2]/div/div/div/div/div/div/input"
        heding="Date of Birth"
        status="Date of Birth Field Not Found"
        key=dob3
        text_box(xpath,heding,status,key)

        date_object_1 = datetime.strptime(str(frm_dt), "%Y-%m-%d %H:%M:%S")
        dob3 = date_object_1.strftime("%m/%d/%Y")

        xpath= "/html/body/div[1]/div/div/main/div/div/div/div[2]/form/div[3]/div[1]/div[1]/div/div/div/div/div/div[1]/input"
        heding="From Date"
        status="As of Date Field Not Found"
        key=dob3           
        
        counter = 0
        while counter < 15:
            try:   
                input_field =WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath)))
                input_field.send_keys(Keys.BACKSPACE * len(input_field.get_attribute("value")))
                input_field.send_keys(key)                  
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo(heding, status)
            sys.exit(0)    

        xpath="/html/body/div[1]/div/div/main/div/div/div/div[2]/form/div[3]/div[2]/div/div/div/div/div/div[1]/div[2]/input"
        heding="Benefit/Service Type"
        status="Benefit/Service Type Field Not Found"
        key='Health Benefit Plan Coverage'
        text_box_key(xpath,heding,status,key)
                            
        xpath= "/html/body/div[1]/div/div/main/div/div/div/div[2]/form/div[4]/div/button"                         
        heding="Submit"
        status="Submit Button Not Found"
        click(xpath,heding,status)
        
        wait = WebDriverWait(driver, 20)
        wait.until(lambda driver: driver.execute_script("return document.readyState === 'complete';"))
        
        try:                
            element_2 = WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/main/div/div/div/div[1]/div/div[2]'))).text                
            wb1=load_workbook(filename=fil)
            sheet = wb1['Eligibility BOT']
            column_letter = 'R'  
            column_cells = sheet[column_letter]
            last_row = None
            for cell in reversed(column_cells):
                if cell.value:
                    last_row = cell.row
                    break                            
            sheet['R' + str(int(last_row + 1))]=element_2       
            wb1.save(fil)
            wb1.close()   

            xpath= "/html/body/div[1]/nav/div[3]/button"                         
            heding="New Request"
            status="New Request Button Not Found"
            click(xpath,heding,status)
            
        except Exception as e:
            xpath= "/html/body/div[1]/div/div/main/div/div/div/div/header/div/div[2]/div[1]/div/div[3]/div"
            heding="Patient Registration"
            status="Patient Registration Count Not Found"        
            count(xpath,heding,status)  

            j=1
            while j<rows+1:         
                xpath= "/html/body/div[1]/div/div/main/div/div/div/div/header/div/div[2]/div[1]/div/div[3]/div[{}]/p"                       
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                if cnm.lstrip().rstrip()=="Member Status":         
                    xpath= "/html/body/div[1]/div/div/main/div/div/div/div/header/div/div[2]/div[1]/div/div[3]/div[{}]/span"
                    ms=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                                                   
                    break
                j=j+1  

            if ms.lstrip().rstrip()=='Active Coverage':                    
                xpath= "/html/body/div[1]/div/div/main/div/div/div/div/header/div/div[3]/div/div[1]/ul[1]/li"
                heding="Member Status"
                status="Member Status Count Not Found"        
                count(xpath,heding,status) 

                lst = []
                lst.append(ms)

                lst1 = []
                j=1
                while j<rows+1:         
                    xpath= "/html/body/div[1]/div/div/main/div/div/div/div/header/div/div[3]/div/div[1]/ul[1]/li[{}]/span[1]"                       
                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                    if cnm.lstrip().rstrip()=="Group Number:":         
                        xpath= "/html/body/div[1]/div/div/main/div/div/div/div/header/div/div[3]/div/div[1]/ul[1]/li[{}]/span[2]"
                        nm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                        if nm==' ' or nm=='':
                            nm='N/A' 
                            lst1.append(nm)
                        else:
                            lst1.append(nm)                                        
                        break                                                                             
                    j=j+1  

                lt=len(lst1)
                if lt==0:                           
                    lst.append('N/A')
                else:
                    vr=lst1[0]
                    lst.append(vr)

                xpath= "/html/body/div[1]/div/div/main/div/div/div/div/header/div/div[3]/div/div[1]/ul[2]/li"
                heding="Member Status - Date"
                status="Member Status - Date Count Not Found"        
                count(xpath,heding,status)                    

                lst2 = []
                j=1
                while j<rows+1:         
                    xpath= "/html/body/div[1]/div/div/main/div/div/div/div/header/div/div[3]/div/div[1]/ul[2]/li[{}]/span[1]"                       
                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                    if cnm.lstrip().rstrip()=="Plan Begin Date:":         
                        xpath= "/html/body/div[1]/div/div/main/div/div/div/div/header/div/div[3]/div/div[1]/ul[2]/li[{}]/span[2]" 
                        nm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                        if nm==' ' or nm=='':
                            nm='N/A' 
                            lst2.append(nm)
                        else:
                            lst2.append(nm)                                        
                        break                                                                             
                    j=j+1  

                lt=len(lst2)
                if lt==0:                           
                    lst.append('N/A')
                else:
                    vr=lst2[0]
                    lst.append(vr)

                lst3 = []
                j=1
                while j<rows+1:         
                    xpath= "/html/body/div[1]/div/div/main/div/div/div/div/header/div/div[3]/div/div[1]/ul[2]/li[{}]/span[1]"                       
                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                    if cnm.lstrip().rstrip()=="Plan End Date:":         
                        xpath= "/html/body/div[1]/div/div/main/div/div/div/div/header/div/div[3]/div/div[1]/ul[2]/li[{}]/span[2]" 
                        nm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                        if nm==' ' or nm=='':
                            nm='N/A' 
                            lst3.append(nm)
                        else:
                            lst3.append(nm)                                        
                        break                                                                             
                    j=j+1  

                lt=len(lst3)
                if lt==0:                           
                    lst.append('N/A')
                else:
                    vr=lst3[0]
                    lst.append(vr)
                
                xpath= "/html/body/div[1]/div/div/main/div/div/div/div/header/div/div[3]/div/div[2]/div"
                heding="Additional Information"
                status="Additional Information Count Not Found"        
                count(xpath,heding,status)

                lst4 = []
                j=1
                while j<rows+1:         
                    xpath= "/html/body/div[1]/div/div/main/div/div/div/div/header/div/div[3]/div/div[2]/div[{}]/p"                       
                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                    if cnm.lstrip().rstrip()=="Other or Additional Payer Information":         
                        xpath= "/html/body/div[1]/div/div/main/div/div/div/div/header/div/div[3]/div/div[2]/div[{}]/span" 
                        nm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                        if nm==' ' or nm=='':
                            nm='N/A' 
                            lst4.append(nm)
                        else:
                            lst4.append(nm)                                        
                        break                                                                             
                    j=j+1  

                lt=len(lst4)
                if lt==0:                           
                    lst.append('N/A')
                else:
                    vr=lst4[0]
                    lst.append(vr)

                xpath= "/html/body/div[1]/div/div/main/div/div/div/div/main[2]/div/div[3]/div/div/div[1]/div/div[1]/p"
                heding="Plan & Insurance"
                status="Plan & Insurance Count Not Found"        
                count(xpath,heding,status)

                lst5 = []
                j=1
                while j<rows+1:                                 
                    xpath= "/html/body/div[1]/div/div/main/div/div/div/div/main[2]/div/div[3]/div/div/div[1]/div/div[1]/p[{}]/strong"                       
                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                    if cnm.lstrip().rstrip()=="Insurance Type:":         
                        xpath= "/html/body/div[1]/div/div/main/div/div/div/div/main[2]/div/div[3]/div/div/div[1]/div/div[1]/p[{}]/span" 
                        nm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                        if nm==' ' or nm=='':
                            nm='N/A' 
                            lst5.append(nm)
                        else:
                            lst5.append(nm)                                        
                        break                                                                             
                    j=j+1  

                lt=len(lst5)
                if lt==0:                           
                    lst.append('N/A')
                else:
                    vr=lst5[0]
                    lst.append(vr)

                lst6 = []
                j=1
                while j<rows+1:                                 
                    xpath= "/html/body/div[1]/div/div/main/div/div/div/div/main[2]/div/div[3]/div/div/div[1]/div/div[1]/p[{}]/strong"                       
                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
                    if cnm.lstrip().rstrip()=="Plan / Product:":         
                        xpath= "/html/body/div[1]/div/div/main/div/div/div/div/main[2]/div/div[3]/div/div/div[1]/div/div[1]/p[{}]/span" 
                        nm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text  
                        if nm==' ' or nm=='':
                            nm='N/A' 
                            lst6.append(nm)
                        else:
                            lst6.append(nm)                                        
                        break                                                                             
                    j=j+1  

                lt=len(lst6)
                if lt==0:                           
                    lst.append('N/A')
                else:
                    vr=lst6[0]
                    lst.append(vr)
                
                wb1=load_workbook(fil)
                sheet = wb1['Eligibility BOT']
                column_letter = 'R'  
                column_cells = sheet[column_letter]
                last_row = None
                for cell in reversed(column_cells):
                    if cell.value:
                        last_row = cell.row
                        break              
                start_column = 'K'
                current_column_index = openpyxl.utils.column_index_from_string(start_column)
                current_row = last_row + 1

                for value in lst:
                    current_column = openpyxl.utils.get_column_letter(current_column_index)
                    sheet[current_column + str(current_row)] = value
                    current_column_index += 1                
                sheet['R' + str(int(last_row + 1))]='Done'
                wb1.save(filename=fil)
                wb1.close()

                xpath= "/html/body/div[1]/nav/div[3]/button"                         
                heding="New Request"
                status="New Request Button Not Found"
                click(xpath,heding,status)

            else:
                wb1=load_workbook(filename=fil)
                sheet = wb1['Eligibility BOT']
                column_letter = 'R'  
                column_cells = sheet[column_letter]
                last_row = None
                for cell in reversed(column_cells):
                    if cell.value:
                        last_row = cell.row
                        break                            
                sheet['R' + str(int(last_row + 1))]=ms      
                wb1.save(fil)
                wb1.close()   

                xpath= "/html/body/div[1]/nav/div[3]/button"                         
                heding="New Request"
                status="New Request Button Not Found"
                click(xpath,heding,status)
    
    driver.quit()
    messagebox.showinfo("Process Status", "Process Completed")
    sys.exit(0)    

if __name__=="__main__":        
    process()
    