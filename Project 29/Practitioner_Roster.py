from tkinter import BOTH, LEFT, TOP, Button, Entry, Frame, Label, PhotoImage, StringVar, Tk,Radiobutton,StringVar,IntVar,filedialog
from idlelib.tooltip import Hovertip
import sys
import os
from operator import itemgetter
import time
from tkinter import messagebox
from os import listdir
from os.path import isfile, join
from datetime import datetime, date
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, StaleElementReferenceException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver import ActionChains
import openpyxl
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from urllib.parse import urlparse
import warnings
import numpy as np
from openpyxl.utils import get_column_letter,column_index_from_string
from datetime import datetime
from pytz import timezone
import requests
from zipfile import ZipFile
import subprocess


warnings.filterwarnings("ignore")

global rows
global xpath
global heding
global status
global key
global nme
element_1 = None
global j


def browse():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
        image_path2 = os.path.join(sys._MEIPASS, 'Static', 'Mapping.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')
        image_path2 = os.path.join(os.getcwd(), 'Static', 'Mapping.png')

    root.title("Process File Picker")
    root.resizable(False,False)

    root.title("Process File Picker")
    root.resizable(False,False)
   
    w = 500
    h = 160
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Browse Process - File Picker",font=("Calibri",18,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
          
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=50)
    
    answer=StringVar()
    answer.set("")
       
    def browse_button():
        global filename
        answer.set("")  
        filename = filedialog.askopenfilename()  
        txt.delete(0, 'end') 
        txt.insert(0, filename)

    def Click_Done():
        global ans        
        
        ans=txt.get()        
        
        if ans=="":
           answer.set("File Path Fields Empty Is Not Allowed...")        
        else:      
            root.destroy()          
            return ans
        
    txt=Entry(Frame2,font=("Calibri",12,"bold","italic"),width=50,justify="left")
    txt.grid(row=0,column=0,padx=5,pady=10,sticky="E")    

    photo1 = PhotoImage(file=image_path2)
            
    btn1=Button(Frame2,text="Browse",command=browse_button,image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=0,column=1,padx=3,pady=0,sticky="W")     
                
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=3,y=80,width=490,height=25)
    
    title_3=Label(Frame3,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=1,padx=150,pady=0,sticky="E")

    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=3,y=105,width=500,height=50)
    
    photo2 = PhotoImage(file=image_path1)

    btn2=Button(Frame4,command=Click_Done,text="Done",image=photo2,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=3,column=0,padx=125,pady=0,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo3 = PhotoImage(file=image_path)   

    btn3=Button(Frame4,command=Close,text="Exit",image=photo3,borderwidth=0,bg="#2c3e50")
    btn3.grid(row=3,column=1,padx=15,pady=0,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn2,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn3,'Click to Exit Process',hover_delay=1000)
    myTip2 = Hovertip(btn1,'Click to Pick File',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def process():        
    global element_1    
    global ans1
    global ans2 
    global driver
    global j
    global dob
    global mid
    global frm_dt
    global to_dt
    global cpt
    global rows
    global rows1
    # https://rajhandicraft.com/collections/bed-bed-sides/products/solid-wood-small-peacock-patra-design-bed-k?variant=41111281926328
    browse()
      
    fil=ans      

    # current_directory = os.getcwd()
    # # current_directory += os.path.sep
    # check_path=os.path.join(current_directory+'\\Temp')
    # folder_create=os.path.join(current_directory+'\\')
    # down_path=os.path.join(current_directory+'\\Temp\\')
    # fin_file=os.path.join(current_directory+'\\Temp\\Files\\')
   
    # fld='Temp'
    
    # if not os.path.isdir(check_path):
    #     os.mkdir(folder_create + 'Temp')
    #     os.mkdir(down_path + 'Files')
    # else:
    #     path1 = os.path.join(folder_create, fld)
    #     shutil.rmtree(path1)
    #     os.mkdir(folder_create + 'Temp')
    #     os.mkdir(down_path + 'Files')  

    try:
        options = Options()            
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--ignore-ssl-errors')
        # options.add_argument("--disable-popup-blocking")
        # options.add_argument('--disable-crl-sets')
        # options.add_argument('--headless')
        prefs = {"download.prompt_for_download": False,
                    "download.directory_upgrade": True,
                    "plugins.always_open_pdf_externally": True,
                    "profile.password_manager_enabled": False,
                    "credentials_enable_service": False,
                    # "safebrowsing.enabled": True
                    }
        # options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_experimental_option("prefs",prefs)
        driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options)
        driver.maximize_window()
        # driver.get('https://promise.dpw.state.pa.us/portal/provider/Home/tabid/135/Default.aspx')
    except Exception as e:                        
        try:
            options = Options()
            options.add_argument('--ignore-certificate-errors')
            options.add_argument('--ignore-ssl-errors')
            # options.add_argument("--disable-popup-blocking")
            # options.add_argument('--disable-crl-sets')
            # options.add_argument('--headless')
            prefs = {"download.prompt_for_download": False,
                    "download.directory_upgrade": True,
                    "plugins.always_open_pdf_externally": True,
                    "profile.password_manager_enabled": False,
                    "credentials_enable_service": False,
                    # "safebrowsing.enabled": True
                    }
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_experimental_option("useAutomationExtension", False)
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option("prefs",prefs)            
            driver_path = os.path.abspath('chromedriver.exe')
            driver = webdriver.Chrome(executable_path=driver_path,options=options)
            driver.maximize_window()            
        except Exception as e:           
            messagebox.showinfo("Driver Problem","Pls Check Your Chrome Driver Version")
            sys.exit(0)        
        
    def text_box(xpath,heding,status,key):                
        counter = 0
        while counter < 20:
            try:   
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()                     
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding, status)
            sys.exit(0)        
    
    def text_box_key(xpath,heding,status,key):                
        counter = 0
        while counter < 5:
            try:                   
                element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, xpath)))
                element.clear()                
                element.send_keys(key)                
                # element.send_keys(Keys.TAB)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding, status)
            sys.exit(0)    

    def text_box_js(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:
                js_code = "document.querySelector('" + xpath + "').setAttribute('value', '" + key + "');"
                WebDriverWait(driver, 0).until(lambda driver: driver.execute_script(js_code))               
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo(heding, status)
            sys.exit(0)       

    def click(xpath,heding,status):
        counter = 0
        while counter < 20:
            try:             
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding,status)
            sys.exit(0)                                

    def count(xpath,heding,status):
        global rows
        counter = 0
        while counter < 20:
            try:             
                rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding,status)
            sys.exit(0)                         

    def count_1(xpath,heding,status):
        global rows1
        counter = 0
        while counter < 5:
            try:             
                rows1=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding,status)
            sys.exit(0)        

    def Alert(msg):
        counter = 0
        while counter < 5:
            try:             
                WebDriverWait(driver, 0).until (EC.alert_is_present())
                a=driver.switch_to.alert
                a.accept()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            print('Alert - ',msg)                                                                     

    file=pd.read_excel(fil,sheet_name='CAQH',header=0)
    
    for index, row in file.iterrows():                                      
        pvd_nm = row[0]
        lnk=row[1]
        usr_nm = row[2]                     
        pwd = row[3]
                
        driver.get(lnk.lstrip().rstrip())

        try:
            xpath= "/html/body/div[2]/main/div/div[2]/div[1]/div[2]/div/form/div[1]/input"
            heding="User Name"
            status="User Name Field Not Found"
            key=usr_nm.lstrip().rstrip()
            text_box(xpath,heding,status,key)

            xpath= "/html/body/div[2]/main/div/div[2]/div[1]/div[2]/div/form/div[2]/input"
            heding="Password"
            status="Password Field Not Found"
            key=pwd.lstrip().rstrip()
            text_box(xpath,heding,status,key)

            xpath='/html/body/div[2]/main/div/div[2]/div[1]/div[2]/div/form/button'
            heding="Sign In"
            status="Sign In Button Not Found"        
            click(xpath,heding,status)          

            counter = 0
            while counter < 5:
                try:
                    ck=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[2]/div[1]/div[2]/div/form/div[1]/ul/li'))).text         

                    wb1=load_workbook(filename=fil)
                    sheet = wb1['CAQH']
                    column_letter = 'NY'  
                    column_cells = sheet[column_letter]
                    last_row = None
                    for cell in reversed(column_cells):
                        if cell.value:
                            last_row = cell.row
                            break                            
                    sheet['NY' + str(int(last_row + 1))]=ck 
                    wb1.save(fil)
                    wb1.close()  

                    break
                except Exception as e:
                    try:
                        xpath= "/html/body/div[2]/header/div[2]/div/ul/li"
                        heding="Home Screen"        
                        status="Home Screen Count Not Found"                    
                        count(xpath,heding,status)

                        j=1
                        while j<rows+1:  
                            hck=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/header/div[2]/div/ul/li[{}]'.format(j)))).text
                            if hck.lstrip().rstrip()=='Profile Data':
                                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[2]/header/div[2]/div/ul/li[{}]'.format(j)))).click()
                                break
                            j=j+1

                        lst=[]
                        counter = 0
                        while counter < 5:
                            try:
                                hck=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[2]/div[3]/div/span[2]/span[1]/span/span[1]'))).text                                
                                if hck==' ' or hck=='':
                                    hck='N/A'
                                    lst.append(hck)   
                                else:
                                    lst.append(hck)                             
                                break
                            except Exception as e:
                                time.sleep(1)
                                counter += 1
                        
                        lt=len(lst)
                        if lt==0:                           
                            lst.append('N/A') 
                        
                        lst1=[]
                        pt=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[2]/div[4]/div/span[1]/span[1]/span/span[1]'))).text

                        if pt==' ' or pt=='':
                            pt='N/A'
                            lst1.append(pt)   
                        else:
                            lst1.append(pt) 

                        lt=len(lst1)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst1[0]
                            lst.append(vr)

                        lst2=[]
                        ps=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[2]/div[5]/div/span[2]/span[1]/span'))).text

                        if ps==' ' or ps=='':
                            ps='N/A'
                            lst2.append(ps)   
                        else:
                            lst2.append(ps) 

                        lt=len(lst2)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst2[0]
                            lst.append(vr)

                        lst3=[]
                        pps=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[2]/div[6]/div/div/div[1]/span[2]/span[1]/span/span[1]'))).text

                        if pps==' ' or pps=='':
                            pps='N/A'
                            lst3.append(pps)   
                        else:
                            lst3.append(pps) 

                        lt=len(lst3)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst3[0]
                            lst.append(vr)

                        lst4=[]
                        aps=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[2]/div[7]/div/span/span[1]/span'))).text

                        if aps==' ' or aps=='':
                            aps='N/A'
                            lst4.append(aps)   
                        else:
                            lst4.append(aps) 

                        lt=len(lst4)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst4[0]
                            lst.append(vr)

                        xpath='/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[3]/div[2]/div/div/div[2]/a'
                        heding="Name Edit"
                        status="Name Edit Button Not Found"        
                        click(xpath,heding,status) 
                        
                        lst5=[]
                        counter = 0
                        while counter < 5:
                            try:
                                nm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[2]/div/div/div/div/div[2]/div[1]/div[1]/input'))).text                                
                                if nm==' ' or nm=='':
                                    nm='N/A'
                                    lst5.append(nm)   
                                else:
                                    lst5.append(nm)                             
                                break
                            except Exception as e:
                                time.sleep(1)
                                counter += 1

                        lt=len(lst5)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst5[0]
                            lst.append(vr)
                        
                        lst6=[]
                        mn=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[2]/div/div/div/div/div[2]/div[1]/div[2]/input'))).text
                                                                                                    
                        if mn==' ' or mn=='':
                            mn='N/A'
                            lst6.append(mn)   
                        else:
                            lst6.append(mn) 

                        lt=len(lst6)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst6[0]
                            lst.append(vr)
                       
                        lst7=[]
                        ln=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[2]/div/div/div/div/div[2]/div[2]/div[1]/input'))).text
                                                                                                    
                        if ln==' ' or ln=='':
                            ln='N/A'
                            lst7.append(ln)   
                        else:
                            lst7.append(ln) 

                        lt=len(lst7)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst7[0]
                            lst.append(vr)

                        lst8=[]
                        suf=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[2]/div/div/div/div/div[2]/div[2]/div[2]/span[1]/span[1]/span/span[1]'))).text
                                                                                                    
                        if suf==' ' or suf=='':
                            suf='N/A'
                            lst8.append(suf)   
                        else:
                            lst8.append(suf) 

                        lt=len(lst8)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst8[0]
                            lst.append(vr)

                        xpath='/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[2]/div/div/div/div/div[1]/span'
                        heding="Name Edit Close"
                        status="Name Edit Close Button Not Found"        
                        click(xpath,heding,status)                        
                        
                        lst9=[]
                        counter = 0
                        while counter < 5:
                            try:
                                otnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[4]/div/div[1]/div[2]/div[2]/div[1]/input'))).text                                
                                if otnm==' ' or otnm=='':
                                    otnm='N/A'
                                    lst9.append(otnm)   
                                else:
                                    lst9.append(otnm)                             
                                break
                            except Exception as e:
                                time.sleep(1)
                                counter += 1

                        lt=len(lst9)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst9[0]
                            lst.append(vr)
                       
                        lst10=[]
                        try:
                            otmn=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[4]/div/div[1]/div[2]/div[2]/div[2]/input'))).text
                                                                                                        
                            if otmn==' ' or otmn=='':
                                otmn='N/A'
                                lst10.append(otmn)   
                            else:
                                lst10.append(otmn) 
                        except Exception as e:
                            otmn='N/A'
                            lst10.append(otmn) 
                        
                        lt=len(lst10)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst10[0]
                            lst.append(vr)

                        lst11=[]
                        try:
                            otln=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[4]/div/div[1]/div[2]/div[3]/div[1]/input'))).text
                                                                                                        
                            if otln==' ' or otln=='':
                                otln='N/A'
                                lst11.append(otln)   
                            else:
                                lst11.append(otln) 
                        except Exception as e:
                            otln='N/A'
                            lst11.append(otln)  

                        lt=len(lst11)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst11[0]
                            lst.append(vr)

                        lst12=[]
                        try:
                            otsuf=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[4]/div/div[1]/div[2]/div[3]/div[2]/span[1]/span[1]/span/span[1]'))).text
                                                                                                        
                            if otsuf==' ' or otsuf=='':
                                otsuf='N/A'
                                lst12.append(otsuf)   
                            else:
                                lst12.append(otsuf) 
                        except Exception as e:
                            otsuf='N/A'
                            lst12.append(otsuf)  

                        lt=len(lst12)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst12[0]
                            lst.append(vr)

                        lst13=[]
                        try:
                            hadd_st1=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[5]/div[4]/div[2]/div/input'))).text
                                                                                                        
                            if hadd_st1==' ' or hadd_st1=='':
                                hadd_st1='N/A'
                                lst13.append(hadd_st1)   
                            else:
                                lst13.append(hadd_st1) 
                        except Exception as e:
                            hadd_st1='N/A'
                            lst13.append(hadd_st1)  

                        lt=len(lst13)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst13[0]
                            lst.append(vr)
                        
                        lst14=[]
                        try:
                            hadd_st2=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[5]/div[4]/div[3]/div/input'))).text
                                                                                                        
                            if hadd_st2==' ' or hadd_st2=='':
                                hadd_st2='N/A'
                                lst14.append(hadd_st2)   
                            else:
                                lst14.append(hadd_st2) 
                        except Exception as e:
                            hadd_st2='N/A'
                            lst14.append(hadd_st2)  

                        lt=len(lst14)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst14[0]
                            lst.append(vr)
                      
                        lst15=[]
                        try:
                            hadd_cty=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[5]/div[4]/div[4]/div[1]/input'))).text
                                                                                                        
                            if hadd_cty==' ' or hadd_cty=='':
                                hadd_cty='N/A'
                                lst15.append(hadd_cty)   
                            else:
                                lst15.append(hadd_cty) 
                        except Exception as e:
                            hadd_cty='N/A'
                            lst15.append(hadd_cty)  

                        lt=len(lst15)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst15[0]
                            lst.append(vr)

                        lst16=[]
                        try:
                            hadd_st=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[5]/div[4]/div[4]/div[2]/span[1]/span[1]/span/span[1]'))).text
                                                                                                        
                            if hadd_st==' ' or hadd_st=='':
                                hadd_st='N/A'
                                lst16.append(hadd_st)   
                            else:
                                lst16.append(hadd_st) 
                        except Exception as e:
                            hadd_st='N/A'
                            lst16.append(hadd_st)  

                        lt=len(lst16)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst16[0]
                            lst.append(vr)
                       
                        lst17=[]
                        try:
                            hadd_zpc=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[5]/div[4]/div[4]/div[3]/input'))).text
                                                                                                        
                            if hadd_zpc==' ' or hadd_zpc=='':
                                hadd_zpc='N/A'
                                lst17.append(hadd_zpc)   
                            else:
                                lst17.append(hadd_zpc) 
                        except Exception as e:
                            hadd_zpc='N/A'
                            lst17.append(hadd_zpc)  

                        lt=len(lst17)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst17[0]
                            lst.append(vr)

                        lst18=[]
                        try:
                            hadd_cntry=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[5]/div[4]/div[5]/div[1]/span[1]/span[1]/span/span[1]'))).text
                                                                                                        
                            if hadd_cntry==' ' or hadd_cntry=='':
                                hadd_cntry='N/A'
                                lst18.append(hadd_cntry)   
                            else:
                                lst18.append(hadd_cntry) 
                        except Exception as e:
                            hadd_cntry='N/A'
                            lst18.append(hadd_cntry)  

                        lt=len(lst18)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst18[0]
                            lst.append(vr)

                        lst19=[]
                        try:
                            hadd_cnty=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[5]/div[4]/div[5]/div[2]/span[1]/span[1]/span/span[1]'))).text
                                                                                                        
                            if hadd_cnty==' ' or hadd_cnty=='':
                                hadd_cnty='N/A'
                                lst19.append(hadd_cnty)   
                            else:
                                lst19.append(hadd_cnty) 
                        except Exception as e:
                            hadd_cnty='N/A'
                            lst19.append(hadd_cnty)  

                        lt=len(lst19)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst19[0]
                            lst.append(vr)

                        lst20=[]
                        try:
                            madd_st1=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[5]/div[6]/div[3]/div[1]/div/input'))).text
                                                                                                        
                            if madd_st1==' ' or madd_st1=='':
                                madd_st1='N/A'
                                lst20.append(madd_st1)   
                            else:
                                lst20.append(madd_st1) 
                        except Exception as e:
                            madd_st1='N/A'
                            lst20.append(madd_st1)  

                        lt=len(lst20)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst20[0]
                            lst.append(vr)

                        lst21=[]
                        try:
                            madd_st2=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[5]/div[6]/div[3]/div[2]/div/input'))).text
                                                                                                        
                            if madd_st2==' ' or madd_st2=='':
                                madd_st2='N/A'
                                lst21.append(madd_st2)   
                            else:
                                lst21.append(madd_st2) 
                        except Exception as e:
                            madd_st2='N/A'
                            lst21.append(madd_st2)  

                        lt=len(lst21)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst21[0]
                            lst.append(vr)

                        lst22=[]
                        try:
                            madd_cty=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[5]/div[6]/div[3]/div[3]/div[1]/input'))).text
                                                                                                        
                            if madd_cty==' ' or madd_cty=='':
                                madd_cty='N/A'
                                lst22.append(madd_cty)   
                            else:
                                lst22.append(madd_cty) 
                        except Exception as e:
                            madd_cty='N/A'
                            lst22.append(madd_cty)  

                        lt=len(lst22)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst22[0]
                            lst.append(vr)

                        lst23=[]
                        try:
                            madd_st=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[5]/div[6]/div[3]/div[3]/div[2]/span[1]/span[1]/span/span[1]'))).text
                                                                                                        
                            if madd_st==' ' or madd_st=='':
                                madd_st='N/A'
                                lst23.append(madd_st)   
                            else:
                                lst23.append(madd_st) 
                        except Exception as e:
                            madd_st='N/A'
                            lst23.append(madd_st)  

                        lt=len(lst23)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst23[0]
                            lst.append(vr)

                        lst24=[]
                        try:
                            madd_zip=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[5]/div[6]/div[3]/div[3]/div[3]/input'))).text
                                                                                                        
                            if madd_zip==' ' or madd_zip=='':
                                madd_zip='N/A'
                                lst24.append(madd_zip)   
                            else:
                                lst24.append(madd_zip) 
                        except Exception as e:
                            madd_zip='N/A'
                            lst24.append(madd_zip)  

                        lt=len(lst24)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst24[0]
                            lst.append(vr)

                        lst25=[]
                        try:
                            madd_conty=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[5]/div[6]/div[3]/div[4]/div[1]/span[1]/span[1]/span/span[1]'))).text
                                                                                                        
                            if madd_conty==' ' or madd_conty=='':
                                madd_conty='N/A'
                                lst25.append(madd_conty)   
                            else:
                                lst25.append(madd_conty) 
                        except Exception as e:
                            madd_conty='N/A'
                            lst25.append(madd_conty)  

                        lt=len(lst25)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst25[0]
                            lst.append(vr)

                        lst26=[]
                        try:
                            madd_cnty=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[5]/div[6]/div[3]/div[4]/div[2]/span[1]/span[1]/span'))).text
                                                                                                        
                            if madd_cnty==' ' or madd_cnty=='':
                                madd_cnty='N/A'
                                lst26.append(madd_cnty)   
                            else:
                                lst26.append(madd_cnty) 
                        except Exception as e:
                            madd_cnty='N/A'
                            lst26.append(madd_cnty)  

                        lt=len(lst26)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst26[0]
                            lst.append(vr)
                        
                        lst27=[]
                        try:
                            pe_add=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[6]/div[3]/div[2]/div/div/div[1]/p'))).text
                                                                                                        
                            if pe_add==' ' or pe_add=='':
                                pe_add='N/A'
                                lst27.append(pe_add)   
                            else:
                                lst27.append(pe_add) 
                        except Exception as e:
                            pe_add='N/A'
                            lst27.append(pe_add)  

                        lt=len(lst27)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst27[0]
                            lst.append(vr)
                       
                        lst28=[]
                        try:
                            ae_add1=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[6]/div[5]/div[1]/div/input'))).text
                                                                                                        
                            if ae_add1==' ' or ae_add1=='':
                                ae_add1='N/A'
                                lst28.append(ae_add1)   
                            else:
                                lst28.append(ae_add1) 
                        except Exception as e:
                            ae_add1='N/A'
                            lst28.append(ae_add1)  

                        lt=len(lst28)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst28[0]
                            lst.append(vr)

                        lst29=[]
                        try:
                            ae_add2=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[6]/div[5]/div[2]/div/input'))).text
                                                                                                        
                            if ae_add2==' ' or ae_add2=='':
                                ae_add2='N/A'
                                lst29.append(ae_add2)   
                            else:
                                lst29.append(ae_add2) 
                        except Exception as e:
                            ae_add2='N/A'
                            lst29.append(ae_add2)  

                        lt=len(lst29)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst29[0]
                            lst.append(vr)

                        lst30=[]
                        try:
                            ae_add3=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[6]/div[5]/div[3]/div/input'))).text
                                                                                                        
                            if ae_add3==' ' or ae_add3=='':
                                ae_add3='N/A'
                                lst30.append(ae_add3)   
                            else:
                                lst30.append(ae_add3) 
                        except Exception as e:
                            ae_add3='N/A'
                            lst30.append(ae_add3)  

                        lt=len(lst30)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst30[0]
                            lst.append(vr)

                        lst31=[]
                        try:
                            ppn=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[6]/div[7]/div[2]/div[1]/input'))).text
                                                                                                        
                            if ppn==' ' or ppn=='':
                                ppn='N/A'
                                lst31.append(ppn)   
                            else:
                                lst31.append(ppn) 
                        except Exception as e:
                            ppn='N/A'
                            lst31.append(ppn)  

                        lt=len(lst31)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst31[0]
                            lst.append(vr)

                        lst32=[]
                        try:
                            pcn=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[6]/div[7]/div[3]/div[1]/input'))).text
                                                                                                        
                            if pcn==' ' or pcn=='':
                                pcn='N/A'
                                lst32.append(pcn)   
                            else:
                                lst32.append(pcn) 
                        except Exception as e:
                            pcn='N/A'
                            lst32.append(pcn)  

                        lt=len(lst32)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst32[0]
                            lst.append(vr)

                        lst33=[]
                        try:
                            pfn=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[6]/div[7]/div[3]/div[2]/input'))).text
                                                                                                        
                            if pfn==' ' or pfn=='':
                                pfn='N/A'
                                lst33.append(pfn)   
                            else:
                                lst33.append(pfn) 
                        except Exception as e:
                            pfn='N/A'
                            lst33.append(pfn)  

                        lt=len(lst33)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst33[0]
                            lst.append(vr)

                        lst34=[]
                        try:
                            ssn=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[7]/div[2]/div/input[1]'))).text
                                                                                                        
                            if ssn==' ' or ssn=='':
                                ssn='N/A'
                                lst34.append(ssn)   
                            else:
                                lst34.append(ssn) 
                        except Exception as e:
                            ssn='N/A'
                            lst34.append(ssn)  

                        lt=len(lst34)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst34[0]
                            lst.append(vr)
                        
                        lst35=[]
                        try:
                            ind_npi=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[7]/div[4]/div[2]/input'))).text
                                                                                                        
                            if ind_npi==' ' or ind_npi=='':
                                ind_npi='N/A'
                                lst35.append(ind_npi)   
                            else:
                                lst35.append(ind_npi) 
                        except Exception as e:
                            ind_npi='N/A'
                            lst35.append(ind_npi)  

                        lt=len(lst35)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst35[0]
                            lst.append(vr)
                        
                        lst36=[]
                        try:
                            fnin=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[7]/div[6]/div/div[3]/div[1]/input'))).text
                                                                                                        
                            if fnin==' ' or fnin=='':
                                fnin='N/A'
                                lst36.append(fnin)   
                            else:
                                lst36.append(fnin) 
                        except Exception as e:
                            fnin='N/A'
                            lst36.append(fnin)  

                        lt=len(lst36)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst36[0]
                            lst.append(vr)

                        lst37=[]
                        try:
                            fnin_conty=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[7]/div[6]/div/div[3]/div[2]/span[1]/span[1]/span/span[1]'))).text
                                                                                                        
                            if fnin_conty==' ' or fnin_conty=='':
                                fnin_conty='N/A'
                                lst37.append(fnin_conty)   
                            else:
                                lst37.append(fnin_conty) 
                        except Exception as e:
                            fnin_conty='N/A'
                            lst37.append(fnin_conty)  

                        lt=len(lst37)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst37[0]
                            lst.append(vr)
                        
                        lst38=[]
                        try:
                            upin=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[7]/div[7]/div[2]/div[2]/div/input'))).text
                                                                                                        
                            if upin==' ' or upin=='':
                                upin='N/A'
                                lst38.append(upin)   
                            else:
                                lst38.append(upin) 
                        except Exception as e:
                            upin='N/A'
                            lst38.append(upin)  

                        lt=len(lst38)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst38[0]
                            lst.append(vr)

                        lst39=[]
                        try:
                            gen=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[8]/div[2]/div[1]/div[1]/span/span[1]/span/span[1]'))).text
                                                                                                        
                            if gen==' ' or gen=='':
                                gen='N/A'
                                lst39.append(gen)   
                            else:
                                lst39.append(gen) 
                        except Exception as e:
                            gen='N/A'
                            lst39.append(gen)  

                        lt=len(lst39)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst39[0]
                            lst.append(vr)

                        lst40=[]
                        try:
                            bd=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[8]/div[4]/div[2]/span[1]/span[1]/span/span[1]'))).text
                                                                                                        
                            if bd==' ' or bd=='':
                                bd='N/A'
                                lst40.append(bd)   
                            else:
                                lst40.append(bd) 
                        except Exception as e:
                            bd='N/A'
                            lst40.append(bd)  

                        lt=len(lst40)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst40[0]
                            lst.append(vr)
                        
                        lst41=[]
                        try:
                            bc=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[8]/div[4]/div[1]/input'))).text
                                                                                                        
                            if bc==' ' or bc=='':
                                bc='N/A'
                                lst41.append(bc)   
                            else:
                                lst41.append(bc) 
                        except Exception as e:
                            bc='N/A'
                            lst41.append(bc)  

                        lt=len(lst41)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst41[0]
                            lst.append(vr)
                       
                        lst42=[]
                        try:
                            bs=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[8]/div[4]/div[2]/span[1]/span[1]/span/span[1]'))).text
                                                                                                        
                            if bs==' ' or bs=='':
                                bs='N/A'
                                lst42.append(bs)   
                            else:
                                lst42.append(bs) 
                        except Exception as e:
                            bs='N/A'
                            lst42.append(bs)  

                        lt=len(lst42)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst42[0]
                            lst.append(vr)

                        lst43=[]
                        try:
                            bcon=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[8]/div[5]/div/span[1]/span[1]/span/span[1]'))).text
                                                                                                        
                            if bcon==' ' or bcon=='':
                                bcon='N/A'
                                lst43.append(bcon)   
                            else:
                                lst43.append(bcon) 
                        except Exception as e:
                            bcon='N/A'
                            lst43.append(bcon)  

                        lt=len(lst43)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst43[0]
                            lst.append(vr)                        

                        xpath= "/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[8]/div[6]/div/div/div"                              
                        heding="Race/Ethnicity"        
                        status="Race/Ethnicity Count Not Found"                    
                        count(xpath,heding,status)

                        lst44=[]
                        j=1
                        while j<rows+1:  
                            checkbox_xpath ='/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[8]/div[6]/div/div/div[{}]/div/span'.format(j)
                            checkbox = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, checkbox_xpath)))
                            if checkbox.is_selected():
                                re=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[8]/div[6]/div/div/div[{}]'.format(j)))).text

                                if re==' ' or re=='':
                                    re='N/A'
                                    lst44.append(re)   
                                else:
                                    lst44.append(re)                                     
                                break
                            j=j+1
                        
                        lt=len(lst44)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst44[0]
                            lst.append(vr)  
                                                   
                        lst45=[]
                        try:
                            lan=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[2]/main/div/div[5]/div/div[1]/form/div[1]/div[10]/div[2]/div/span[1]/span[1]/span'))).text
                                                                                                        
                            if lan==' ' or lan=='':
                                lan='N/A'
                                lst45.append(lan)   
                            else:
                                lst45.append(lan) 
                        except Exception as e:
                            lan='N/A'
                            lst45.append(lan)  

                        lt=len(lst45)
                        if lt==0:                           
                            lst.append('N/A')
                        else:
                            vr=lst45[0]
                            lst.append(vr)

                        # Professional IDs


                        break
                    except Exception as e:
                        time.sleep(1)
                        counter += 1


        except Exception as e:
            pass

    driver.quit()
    messagebox.showinfo("Process Status", "Process Completed")
    sys.exit(0)    

if __name__=="__main__":        
    process()
    