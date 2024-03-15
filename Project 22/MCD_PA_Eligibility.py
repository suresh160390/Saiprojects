from tkinter import BOTH, LEFT, TOP, Button, Entry, Frame, Label, PhotoImage, StringVar, Tk,Radiobutton,StringVar,IntVar,filedialog
from idlelib.tooltip import Hovertip
import sys
import os
from operator import itemgetter
import time
from tkinter import messagebox
from os import listdir
from os.path import isfile, join
from datetime import datetime, date
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, StaleElementReferenceException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import openpyxl
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from urllib.parse import urlparse
import warnings
import numpy as np
from openpyxl.utils import get_column_letter,column_index_from_string
from datetime import datetime
from pytz import timezone
import requests
from zipfile import ZipFile
import subprocess

warnings.filterwarnings("ignore")

global rows
global xpath
global heding
global status
global key
global nme
element_1 = None
global j


def browse():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
        image_path2 = os.path.join(sys._MEIPASS, 'Static', 'Mapping.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')
        image_path2 = os.path.join(os.getcwd(), 'Static', 'Mapping.png')

    root.title("Process File Picker")
    root.resizable(False,False)

    root.title("Process File Picker")
    root.resizable(False,False)
   
    w = 500
    h = 160
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Browse Process - File Picker",font=("Calibri",18,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
          
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=50)
    
    answer=StringVar()
    answer.set("")
       
    def browse_button():
        global filename
        answer.set("")  
        filename = filedialog.askopenfilename()  
        txt.delete(0, 'end') 
        txt.insert(0, filename)

    def Click_Done():
        global ans        
        
        ans=txt.get()        
        
        if ans=="":
           answer.set("File Path Fields Empty Is Not Allowed...")        
        else:      
            root.destroy()          
            return ans
        
    txt=Entry(Frame2,font=("Calibri",12,"bold","italic"),width=50,justify="left")
    txt.grid(row=0,column=0,padx=5,pady=10,sticky="E")    

    photo1 = PhotoImage(file=image_path2)
            
    btn1=Button(Frame2,text="Browse",command=browse_button,image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=0,column=1,padx=3,pady=0,sticky="W")     
                
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=3,y=80,width=490,height=25)
    
    title_3=Label(Frame3,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=1,padx=150,pady=0,sticky="E")

    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=3,y=105,width=500,height=50)
    
    photo2 = PhotoImage(file=image_path1)

    btn2=Button(Frame4,command=Click_Done,text="Done",image=photo2,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=3,column=0,padx=125,pady=0,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo3 = PhotoImage(file=image_path)   

    btn3=Button(Frame4,command=Close,text="Exit",image=photo3,borderwidth=0,bg="#2c3e50")
    btn3.grid(row=3,column=1,padx=15,pady=0,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn2,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn3,'Click to Exit Process',hover_delay=1000)
    myTip2 = Hovertip(btn1,'Click to Pick File',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def process():        
    global element_1    
    global ans1
    global ans2 
    global driver
    global j
    global dob
    global mid
    global frm_dt
    global to_dt
    global cpt
    global rows
    global rows1
    # https://rajhandicraft.com/collections/bed-bed-sides/products/solid-wood-small-peacock-patra-design-bed-k?variant=41111281926328
    browse()
      
    fil=ans               

    try:
        options = Options()            
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--ignore-ssl-errors')
        options.add_argument("--disable-popup-blocking")
        driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options)
        driver.maximize_window()
        driver.get('https://promise.dpw.state.pa.us/portal/provider/Home/tabid/135/Default.aspx')
    except Exception as e:                        
        try:
            options = Options()
            options.add_argument('--ignore-certificate-errors')
            options.add_argument('--ignore-ssl-errors')
            options.add_argument("--disable-popup-blocking")
            # response = requests.get('https://chromedriver.storage.googleapis.com/LATEST_RELEASE')
            # latest_version = response.text.strip()
            # chrome_driver_url = f'https://chromedriver.storage.googleapis.com/{latest_version}/chromedriver_win32.zip'
            # response = requests.get(chrome_driver_url)
            # with open('chromedriver_win32.zip', 'wb') as f:
            #     f.write(response.content)
            # with ZipFile('chromedriver_win32.zip', 'r') as zip_ref:
            #     zip_ref.extractall('.')
            driver_path = os.path.abspath('chromedriver.exe')
            driver = webdriver.Chrome(executable_path=driver_path,options=options)
            driver.maximize_window()
            driver.get('https://promise.dpw.state.pa.us/portal/provider/Home/tabid/135/Default.aspx')
        except Exception as e:           
            messagebox.showinfo("Driver Problem","Pls Check Your Chrome Driver Version")
            sys.exit(0)
        
    def text_box(xpath,heding,status,key):                
        counter = 0
        while counter < 5:
            try:   
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()                     
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            print('Moved Exception')
            raise e
            # messagebox.showinfo(heding, status)
            # sys.exit(0)        
    
    def text_box_key(xpath,heding,status,key):                
        counter = 0
        while counter < 5:
            try:                   
                element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, xpath)))
                element.clear()                
                element.send_keys(key)                
                # element.send_keys(Keys.TAB)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            print('Moved Exception')
            raise e
            # messagebox.showinfo(heding, status)
            # sys.exit(0)    

    def text_box_js(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:
                js_code = "document.querySelector('" + xpath + "').setAttribute('value', '" + key + "');"
                WebDriverWait(driver, 0).until(lambda driver: driver.execute_script(js_code))               
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo(heding, status)
            sys.exit(0)       

    def click(xpath,heding,status):
        counter = 0
        while counter < 5:
            try:             
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            print('Moved Exception')
            raise e
            # messagebox.showinfo(heding,status)
            # sys.exit(0)                                

    def count(xpath,heding,status):
        global rows
        counter = 0
        while counter < 5:
            try:             
                rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            print('Moved Exception')
            raise e
            # messagebox.showinfo(heding,status)
            # sys.exit(0)                         

    def count_1(xpath,heding,status):
        global rows1
        counter = 0
        while counter < 5:
            try:             
                rows1=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            print('Moved Exception')
            raise e
            # messagebox.showinfo(heding,status)
            # sys.exit(0)        

    def Alert():
        counter = 0
        while counter < 15:
            try:             
                WebDriverWait(driver, 0).until (EC.alert_is_present())
                a=driver.switch_to.alert
                a.accept()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo('Alert','Alert Not Present')                                                                
    
    messagebox.showinfo('Waiting','Authentication Waiting')        
    
    # driver.switch_to.window(driver.window_handles[1])      

    file=pd.read_excel(fil,sheet_name='MCD PA Eligibility BOT',header=0)
    
    for index, row in file.iterrows():                                      
        fnm = row[2]  
        lnm=row[3]
        dob = row[4]                     
        frm_dos = row[5]
        to_dos = row[6]
        
        xpath= "/html/body/form/div[3]/div/div[2]/div[3]/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td[2]/div/div/div/div/fieldset/table/tbody/tr[10]/td[3]/input[1]"
        heding="First Name"
        status=" First Name Not Found"
        key=fnm.lstrip().rstrip()
        text_box(xpath,heding,status,key)

        xpath= "/html/body/form/div[3]/div/div[2]/div[3]/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td[2]/div/div/div/div/fieldset/table/tbody/tr[10]/td[3]/input[3]"
        heding="Last Name"
        status=" Last Name Not Found"
        key=lnm.lstrip().rstrip()
        text_box(xpath,heding,status,key)

        date_object_1 = datetime.strptime(str(dob), "%Y-%m-%d %H:%M:%S")
        dob1 = date_object_1.strftime("%m/%d/%Y")  

        xpath= "/html/body/form/div[3]/div/div[2]/div[3]/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td[2]/div/div/div/div/fieldset/table/tbody/tr[11]/td[3]/div/div[1]/input"
        heding="Birth Date"        
        status="Birth Date Field Not Found"
        key=dob1
        text_box(xpath,heding,status,key)

        date_object_2 = datetime.strptime(str(frm_dos), "%Y-%m-%d %H:%M:%S")
        dob2 = date_object_2.strftime("%m/%d/%Y")

        xpath= "/html/body/form/div[3]/div/div[2]/div[3]/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td[2]/div/div/div/div/fieldset/table/tbody/tr[15]/td[3]/div/div[1]/input"
        heding="Date of Service From"        
        status="Date of Service From Field Not Found"
        key=dob2
        text_box(xpath,heding,status,key)
        
        date_object_3 = datetime.strptime(str(to_dos), "%Y-%m-%d %H:%M:%S")
        dob3 = date_object_3.strftime("%m/%d/%Y")  

        xpath= "/html/body/form/div[3]/div/div[2]/div[3]/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td[2]/div/div/div/div/fieldset/table/tbody/tr[15]/td[3]/div/div[4]/input"
        heding="Date of Service To"        
        status="Date of Service To Field Not Found"
        key=dob3
        text_box(xpath,heding,status,key)

        xpath='/html/body/form/div[3]/div/div[2]/div[3]/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td[2]/div/div/div/div/fieldset/table/tbody/tr[27]/td/input[1]'
        heding="Search"
        status="Search Button Not Found"        
        click(xpath,heding,status)
        
        counter = 0
        while counter < 5:
            try:            
                ck=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/form/div[3]/div/div[2]/div[3]/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td[2]/div/div/div/fieldset/legend/span'))).text

                err=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/form/div[3]/div/div[2]/div[3]/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td[2]/div/div/div/fieldset/table/tbody/tr/td/span'))).text
                
                wb1=load_workbook(filename=fil)
                sheet = wb1['MCD PA Eligibility BOT']
                column_letter = 'J'  
                column_cells = sheet[column_letter]
                last_row = None
                for cell in reversed(column_cells):
                    if cell.value:
                        last_row = cell.row
                        break                            
                sheet['J' + str(int(last_row + 1))]=err 
                wb1.save(fil)
                wb1.close()               

                xpath='/html/body/form/div[3]/div/div[2]/div[3]/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td[2]/div/div/div/div/fieldset/table/tbody/tr[27]/td/input[2]'
                heding="Clear"
                status="Clear Button Not Found"        
                click(xpath,heding,status)
                break
            except Exception as e:            
                try:
                    xpath= "/html/body/form/div[3]/div/div[2]/div[3]/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td[2]/div/div/div/table[1]/tbody/tr/td/fieldset/table/tbody/tr/td/div/table/tbody/tr"
                    heding="Recipient Tabel Count"        
                    status="Recipient Tabel Count Not Found"                    
                    count(xpath,heding,status)
                    
                    lst=[]
                    j=1
                    while j<rows+1:   
                        ck=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/form/div[3]/div/div[2]/div[3]/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td[2]/div/div/div/table[1]/tbody/tr/td/fieldset/table/tbody/tr/td/div/table/tbody/tr[{}]/td[1]'.format(j)))).text
                        if ck.lstrip().rstrip()=='Recipient ID:':
                            rec=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/form/div[3]/div/div[2]/div[3]/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td[2]/div/div/div/table[1]/tbody/tr/td/fieldset/table/tbody/tr/td/div/table/tbody/tr[{}]/td[2]'.format(j)))).text
                            if rec==' ' or rec=='':
                                rec='N/A'
                                lst.append(rec)   
                            else:
                                lst.append(rec) 
                            break
                        j=j+1

                    lt=len(lst)
                    if lt==0:                           
                        lst.append('N/A') 

                    xpath= "/html/body/form/div[3]/div/div[2]/div[3]/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td[2]/div/div/div/table[2]/tbody/tr/td/fieldset/table/tbody/tr/td/div/table/tbody/tr"
                    heding="Eligibility Summary Tabel Count"        
                    status="Eligibility Summary Tabel Count Not Found"                    
                    count(xpath,heding,status)

                    lst1 = []
                    lst2=[]
                    j=2
                    while j<rows+1:                         
                        ck=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/form/div[3]/div/div[2]/div[3]/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td[2]/div/div/div/table[2]/tbody/tr/td/fieldset/table/tbody/tr/td/div/table/tbody/tr[{}]/td[1]'.format(j)))).text
                        if ck.lstrip().rstrip()=='Managed Care':                            
                            nm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/form/div[3]/div/div[2]/div[3]/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td[2]/div/div/div/table[2]/tbody/tr/td/fieldset/table/tbody/tr/td/div/table/tbody/tr[{}]/td[2]'.format(j)))).text

                            st=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/form/div[3]/div/div[2]/div[3]/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td[2]/div/div/div/table[2]/tbody/tr/td/fieldset/table/tbody/tr/td/div/table/tbody/tr[{}]/td[3]'.format(j)))).text

                            ed=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/form/div[3]/div/div[2]/div[3]/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td[2]/div/div/div/table[2]/tbody/tr/td/fieldset/table/tbody/tr/td/div/table/tbody/tr[{}]/td[4]'.format(j)))).text

                            if nm==' ' or nm=='':
                                  nm='N/A'      
                            if st==' ' or st=='':
                                  st='N/A'  
                            if ed==' ' or ed=='':
                                  ed='N/A'
                            
                            fin_mda = f"{ck}:{nm}:{st}:{ed};"
                            
                            lst1.append(fin_mda)
                            
                            result_string = lst1[0]

                            lst2.append(result_string)                           

                            lst3 = ' '.join(lst2)

                            lst2=[]

                            lst2.append(lst3)                                                 

                            lst1=[]                        
                        j=j+1
                    
                    lt=len(lst2)
                    if lt==0:                           
                        lst.append('N/A')
                    else:                        
                        lst = lst + lst2
                                            
                    wb1=load_workbook(fil)
                    sheet = wb1['MCD PA Eligibility BOT']
                    column_letter = 'J'  
                    column_cells = sheet[column_letter]
                    last_row = None
                    for cell in reversed(column_cells):
                        if cell.value:
                            last_row = cell.row
                            break
                    start_column = 'H'
                    current_column_index = openpyxl.utils.column_index_from_string(start_column)
                    current_row = last_row + 1

                    for value in lst:
                        current_column = openpyxl.utils.get_column_letter(current_column_index)
                        sheet[current_column + str(current_row)] = value
                        current_column_index += 1                
                    sheet['J' + str(int(last_row + 1))]='Done'
                    wb1.save(filename=fil)
                    wb1.close() 
                     
                    xpath='/html/body/form/div[3]/div/div[2]/div[3]/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td[2]/div/div/div/div/fieldset/table/tbody/tr[27]/td/input[2]'
                    heding="Clear"
                    status="Clear Button Not Found"        
                    click(xpath,heding,status)
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1             
        else:
            wb1=load_workbook(filename=fil)
            sheet = wb1['MCD PA Eligibility BOT']
            column_letter = 'J'  
            column_cells = sheet[column_letter]
            last_row = None
            for cell in reversed(column_cells):
                if cell.value:
                    last_row = cell.row
                    break                            
            sheet['J' + str(int(last_row + 1))]='Error'   
            wb1.save(fil)
            wb1.close()    

            xpath='/html/body/form/div[3]/div/div[2]/div[3]/div[2]/table/tbody/tr/td/table/tbody/tr[2]/td[2]/div/div/div/div/fieldset/table/tbody/tr[27]/td/input[2]'
            heding="Clear"
            status="Clear Button Not Found"        
            click(xpath,heding,status)        
    
    driver.quit()
    messagebox.showinfo("Process Status", "Process Completed")
    sys.exit(0)    

if __name__=="__main__":        
    process()
    