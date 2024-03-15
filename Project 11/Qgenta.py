from tkinter import BOTH, LEFT, TOP, Button, Entry, Frame, Label, PhotoImage, StringVar, Tk,Radiobutton,StringVar,IntVar,filedialog
from idlelib.tooltip import Hovertip
import sys
import os
from operator import itemgetter
import time
from tkinter import messagebox
from os import listdir
from os.path import isfile, join
from datetime import datetime, date
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, StaleElementReferenceException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import openpyxl
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from urllib.parse import urlparse
import warnings
import numpy as np
from openpyxl.utils import get_column_letter,column_index_from_string
from datetime import datetime
from pytz import timezone
import requests
from zipfile import ZipFile

warnings.filterwarnings("ignore")

global rows
global xpath
global heding
global status
global key
global nme
element_1 = None


def active():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')

    root.title("Qgenta - Process Report")
    root.resizable(False,False)
   
    w = 530
    h = 230
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Please Select Qgenta Active Report (OR) Qgenta All Report",font=("Calibri",12,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
    
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=300)
    
    title1=Label(Frame2,text="Qgenta Active Report :",font=("Calibri",17,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title1.grid(row=0,column=0,padx=45,pady=5,sticky="W")

    title2=Label(Frame2,text="Qgenta All Report :",font=("Calibri",17,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title2.grid(row=1,column=0,padx=45,pady=5,sticky="W")    

    answer=StringVar()
    answer.set("")

    title_3=Label(Frame2,text=answer.get(),textvariable=answer,font=("Calibri",12,"bold","italic"),bg="#2c3e50",fg="Red",justify="center",width=68)
    title_3.grid(row=2,column=0,columnspan=2,padx=0,pady=0,sticky="W")

    def Radio(*event):
         answer.set("")

    global var

    var = IntVar()

    R1 = Radiobutton(Frame2,text="Qgenta Active     ",variable=var, value=1,font=("Calibri",9,"bold","italic"),bg="Gold",fg="Black",width=25,justify="left",command=lambda: Radio())
    R1.grid(row=0,column=1,padx=0,pady=5,sticky="W")

    R2 = Radiobutton(Frame2,text="Qgenta All           ",variable=var,value=2,font=("Calibri",9,"bold","italic"),bg="Gold",fg="Black",width=25,command=lambda: Radio())
    R2.grid(row=1,column=1,padx=0,pady=5,sticky="W")  
    
    def Click_Done():
        global slt
        selection = str(var.get())

        if selection==str(1):
           answer.set("")
           slt=1
           root.destroy()            
           process()       
        elif selection==str(2):
            answer.set("")
            slt=2
            root.destroy()
            process()
        else:    
            answer.set("Please Select Any One Option...")            
    
    photo1 = PhotoImage(file=image_path1)

    btn=Button(Frame2,command=Click_Done,text="Done",image=photo1,borderwidth=0,bg="#2c3e50")
    btn.grid(row=3,column=0,padx=125,pady=20,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo = PhotoImage(file=image_path)   

    btn1=Button(Frame2,command=Close,text="Exit",image=photo,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=3,column=1,padx=15,pady=20,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn1,'Click to Exit Process',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def radio():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')

    root.title("Qgenta Process Report")
    root.resizable(False,False)
   
    w = 530
    h = 230
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Please Select Qgenta Files Download Report (OR) Qgenta Console Report",font=("Calibri",12,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
    
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=300)
    
    title1=Label(Frame2,text="Qgenta Download Report :",font=("Calibri",17,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title1.grid(row=0,column=0,padx=45,pady=5,sticky="W")

    title2=Label(Frame2,text="Qgenta Console Report :",font=("Calibri",17,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title2.grid(row=1,column=0,padx=45,pady=5,sticky="W")    

    answer=StringVar()
    answer.set("")

    title_3=Label(Frame2,text=answer.get(),textvariable=answer,font=("Calibri",12,"bold","italic"),bg="#2c3e50",fg="Red",justify="center",width=68)
    title_3.grid(row=2,column=0,columnspan=2,padx=0,pady=0,sticky="W")

    def Radio(*event):
         answer.set("")

    global var

    var = IntVar()

    R1 = Radiobutton(Frame2,text="Qgenta Download     ",variable=var, value=1,font=("Calibri",9,"bold","italic"),bg="Gold",fg="Black",width=25,justify="left",command=lambda: Radio())
    R1.grid(row=0,column=1,padx=0,pady=5,sticky="W")

    R2 = Radiobutton(Frame2,text="Qgenta Consol           ",variable=var,value=2,font=("Calibri",9,"bold","italic"),bg="Gold",fg="Black",width=25,command=lambda: Radio())
    R2.grid(row=1,column=1,padx=0,pady=5,sticky="W")  
    
    def Click_Done():
        selection = str(var.get())

        if selection==str(1):
           answer.set("")
           root.destroy()            
           active()       
        elif selection==str(2):
            answer.set("")
            root.destroy()
            console()
        else:    
            answer.set("Please Select Any One Option...")            
    
    photo1 = PhotoImage(file=image_path1)

    btn=Button(Frame2,command=Click_Done,text="Done",image=photo1,borderwidth=0,bg="#2c3e50")
    btn.grid(row=3,column=0,padx=125,pady=20,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo = PhotoImage(file=image_path)   

    btn1=Button(Frame2,command=Close,text="Exit",image=photo,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=3,column=1,padx=15,pady=20,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn1,'Click to Exit Process',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def browse():
    global ans
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
        image_path2 = os.path.join(sys._MEIPASS, 'Static', 'Mapping.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')
        image_path2 = os.path.join(os.getcwd(), 'Static', 'Mapping.png')    

    root.title("Process - File Picker")
    root.resizable(False,False)
   
    w = 500
    h = 160
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Browse Process - File Picker",font=("Calibri",18,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
          
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=50)
    
    answer=StringVar()
    answer.set("")
       
    def browse_button():
        global filename
        answer.set("")  
        filename = filedialog.askopenfilename()  
        txt.delete(0, 'end') 
        txt.insert(0, filename)

    def Click_Done():
        global ans        
        
        ans=txt.get()        
        
        if ans=="":
           answer.set("File Path Fields Empty Is Not Allowed...")        
        else:      
            root.destroy()          
            return ans
        
    txt=Entry(Frame2,font=("Calibri",12,"bold","italic"),width=50,justify="left")
    txt.grid(row=0,column=0,padx=5,pady=10,sticky="E")    

    photo1 = PhotoImage(file=image_path2)
            
    btn1=Button(Frame2,text="Browse",command=browse_button,image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=0,column=1,padx=3,pady=0,sticky="W")     
                
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=3,y=80,width=490,height=25)
    
    title_3=Label(Frame3,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=1,padx=150,pady=0,sticky="E")

    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=3,y=105,width=500,height=50)
    
    photo2 = PhotoImage(file=image_path1)

    btn2=Button(Frame4,command=Click_Done,text="Done",image=photo2,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=3,column=0,padx=125,pady=0,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo3 = PhotoImage(file=image_path)   

    btn3=Button(Frame4,command=Close,text="Exit",image=photo3,borderwidth=0,bg="#2c3e50")
    btn3.grid(row=3,column=1,padx=15,pady=0,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn2,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn3,'Click to Exit Process',hover_delay=1000)
    myTip2 = Hovertip(btn1,'Click to Pick File',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def folder():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
        image_path2 = os.path.join(sys._MEIPASS, 'Static', 'Mapping.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')
        image_path2 = os.path.join(os.getcwd(), 'Static', 'Mapping.png')

    root.title("Process Folder Picker")
    root.resizable(False,False)

    root.title("Process Folder Picker")
    root.resizable(False,False)
   
    w = 500
    h = 160
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Browse Process - Folder Picker",font=("Calibri",18,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
          
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=50)
    
    answer=StringVar()
    answer.set("")
       
    def browse_button():
        global filename
        answer.set("")  
        filename = filedialog.askdirectory()  
        txt.delete(0, 'end') 
        txt.insert(0, filename)

    def Click_Done():
        global ans        
        
        ans=txt.get()        
        
        if ans=="":
           answer.set("Folder Path Fields Empty Is Not Allowed...")        
        else:      
            root.destroy()          
            return ans
        
    txt=Entry(Frame2,font=("Calibri",12,"bold","italic"),width=50,justify="left")
    txt.grid(row=0,column=0,padx=5,pady=10,sticky="E")    

    photo1 = PhotoImage(file=image_path2)
            
    btn1=Button(Frame2,text="Browse",command=browse_button,image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=0,column=1,padx=3,pady=0,sticky="W")     
                
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=3,y=80,width=490,height=25)
    
    title_3=Label(Frame3,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=1,padx=150,pady=0,sticky="E")

    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=3,y=105,width=500,height=50)
    
    photo2 = PhotoImage(file=image_path1)

    btn2=Button(Frame4,command=Click_Done,text="Done",image=photo2,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=3,column=0,padx=125,pady=0,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo3 = PhotoImage(file=image_path)   

    btn3=Button(Frame4,command=Close,text="Exit",image=photo3,borderwidth=0,bg="#2c3e50")
    btn3.grid(row=3,column=1,padx=15,pady=0,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn2,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn3,'Click to Exit Process',hover_delay=1000)
    myTip2 = Hovertip(btn1,'Click to Pick Folder',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def user_pass():
    global ans1
    global ans2
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')

    root.title("QGenda - User Login Details")
    root.resizable(False,False)

    w = 600
    h = 180
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)

    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="User Name and Password Details...",font=("Calibri",20,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack() 
 
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=698,height=150)    
    
    title3=Label(Frame2,text="User Name :",font=("Calibri",11,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title3.grid(row=0,column=0,padx=5,pady=5,sticky="W")
    
    txt1=Entry(Frame2,font=("Calibri",11,"bold","italic"),width=50,justify="left")
    txt1.grid(row=0,column=1,padx=30,pady=5,sticky="W")
    
    title4=Label(Frame2,text="Password :",font=("Calibri",11,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title4.grid(row=1,column=0,padx=5,pady=5,sticky="W")
    
    txt2=Entry(Frame2,font=("Calibri",11,"bold","italic"),width=50,justify="left")
    txt2.grid(row=1,column=1,padx=30,pady=5,sticky="W")

    answer=StringVar()
    answer.set("")

    def Click_Done():
        global ans1
        global ans2        
        
        ans1=txt1.get()
        ans2=txt2.get()        

        if ans1=="":
           answer.set("User Name Field Empty Is Not Allowed...")
        elif ans2=="":
            answer.set("Password Field Empty Is Not Allowed...")        
        else:    
            root.destroy()
            return ans1,ans2
                
    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=0,y=105,width=698,height=20)
    
    title_3=Label(Frame4,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=2,padx=200,pady=0,sticky="E")
    
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=0,y=125,width=698,height=200)
       
    photo1 = PhotoImage(file=image_path1)
    
    btn1=Button(Frame3,command=Click_Done,text="Run",image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=2,column=0,padx=140,pady=0,sticky="W")

    def Close():
        sys.exit(0)

    photo = PhotoImage(file=image_path)    

    btn2=Button(Frame3,command=Close,text="Close",image=photo,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=2,column=1,padx=100,pady=0,sticky="E")

    def disable_event():
        pass

    txt1.focus_set()

    myTip = Hovertip(btn1,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn2,'Click to Exit Process',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)
            
    root.mainloop()

def process():
        global element_1
        global down_path
        global ans1
        global ans2 
        global ans

        user_pass()
        
        browse()       
                                
        folder_create = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Downloads/')
        check_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Downloads/Temp')
        down_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Downloads\\Temp\\')
        file = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Downloads\\Temp\\Files\\')
            
        fld='Temp'
        
        if not os.path.isdir(check_path):
            os.mkdir(folder_create + 'Temp')
            os.mkdir(down_path + 'Files')
        else:
            path1 = os.path.join(folder_create, fld)
            shutil.rmtree(path1)
            os.mkdir(folder_create + 'Temp')
            os.mkdir(down_path + 'Files')                 

        try:
            options = Options()            
            options.add_argument('--ignore-certificate-errors')
            options.add_argument('--ignore-ssl-errors')
            prefs = {"download.default_directory" : down_path,
                    "download.prompt_for_download": False,
                    "download.directory_upgrade": True,
                    "safebrowsing.enabled": True}
            options.add_experimental_option("prefs",prefs)
            driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options)
            driver.maximize_window()
            driver.get('https://app.credentialgenie.com/users/sign_in')            
        except Exception as e:
            try:
                options =  webdriver.ChromeOptions()        
                options.add_argument('--ignore-certificate-errors')
                options.add_argument('--ignore-ssl-errors')
                options.add_argument("--disable-popup-blocking")
                prefs = {"download.default_directory" : down_path,
                        "download.prompt_for_download": False,
                        "download.directory_upgrade": True,
                        "safebrowsing.enabled": True}
                options.add_experimental_option("prefs",prefs)
                # response = requests.get('https://chromedriver.storage.googleapis.com/LATEST_RELEASE')
                # latest_version = response.text.strip()
                # chrome_driver_url = f'https://chromedriver.storage.googleapis.com/{latest_version}/chromedriver_win32.zip'
                # response = requests.get(chrome_driver_url)
                # with open('chromedriver_win32.zip', 'wb') as f:
                #     f.write(response.content)
                # with ZipFile('chromedriver_win32.zip', 'r') as zip_ref:
                #     zip_ref.extractall('.')   
                # driver_path=r'C:\Users\sanandrao\Desktop\New folder\chromedriver.exe'
                driver_path = os.path.abspath('chromedriver.exe')                
                driver = webdriver.Chrome(executable_path=driver_path,options=options)                
                driver.maximize_window()
                driver.get('https://app.credentialgenie.com/users/sign_in')                                         
            except Exception as e:
                messagebox.showinfo("Internet Problem","Pls Check Your Internet Connection")
                sys.exit(0)
             
        def text_box(xpath,heding,status,key):                
            counter = 0
            while counter < 5:
                try:                          
                    WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding, status)
                sys.exit(0)        
                
        def click(xpath,heding,status):
            counter = 0
            while counter < 5:
                try:             
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding,status)
                sys.exit(0)                                

        def count(xpath,heding,status):
            global rows
            counter = 0
            while counter < 5:
                try:             
                    rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))                    
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding,status)
                sys.exit(0)          
             
        def Alert():
            counter = 0
            while counter < 5:
                try:             
                    WebDriverWait(driver, 0).until (EC.alert_is_present())
                    a=driver.switch_to.alert
                    a.accept()
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo('Alert','Alert Not Present')

        xpath= '/html/body/div/div/form/div[1]/input'
        heding="User ID"
        status="User ID Field Not Found"
        key= ans1
        text_box(xpath,heding,status,key)
        
        xpath= '/html/body/div/div/form/div[2]/input'
        heding="Password"
        status="Password Field Not Found"
        key= ans2
        text_box(xpath,heding,status,key)

        xpath= '/html/body/div/div/form/div[3]/input'
        heding="Log in Button"
        status="Log in Button Not Found"
        click(xpath,heding,status)
        
        df=pd.read_excel(ans,sheet_name='Data',header=0)                         
        
        for index, row in df.iterrows():  
            pyr_nm = row[0]                                          
           
            pyr_nm = pyr_nm.lower()
            pyr_nm = pyr_nm.replace(' ', '')
            
            xpath='/html/body/header/div[1]/ul/li'
            heding="Reports Count"
            status="Reports Count Not Found"
            count(xpath,heding,status)
            
            j=1
            while j<rows+1:
                xpath= '/html/body/header/div[1]/ul/li[{}]'                      
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                if cnm.lstrip().rstrip()=='Reports': 
                    xpath= '/html/body/header/div[1]/ul/li[{}]/a'                                
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()
                    break
                j=j+1
                    
            xpath='/html/body/div[1]/div/div/div[2]/div[1]/ul/li'
            heding="Payor Count"
            status="Payor Count Not Found"
            count(xpath,heding,status)

            j=1
            while j<rows+1:
                xpath= '/html/body/div[1]/div/div/div[2]/div[1]/ul/li[{}]'                      
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                if cnm.lstrip().rstrip()=='Payor Enrollments': 
                    xpath= '/html/body/div[1]/div/div/div[2]/div[1]/ul/li[{}]/a'                                
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()
                    break
                j=j+1           

            if slt == 1:
                xpath='/html/body/div[1]/div/div/div[2]/div[2]/div[4]/form/div[1]/div[1]/div/select/option'
                heding="Provider Status"
                status="Provider Status Count Not Found"
                count(xpath,heding,status)

                j=1
                while j<rows+1:
                    xpath= '/html/body/div[1]/div/div/div[2]/div[2]/div[4]/form/div[1]/div[1]/div/select/option[{}]'                      
                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                    
                    if cnm.lstrip().rstrip()=='Active':                                                                         
                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()
                        break
                    j=j+1           

            xpath= '/html/body/div[1]/div/div/div[2]/div[2]/div[4]/form/div[3]/div[1]/div/div/div[1]/input'
            heding="Location"
            status="Location Field Not Found"
            click(xpath,heding,status)

            xpath='/html/body/div[1]/div/div/div[2]/div[2]/div[4]/form/div[3]/div[1]/div/div/div[2]/div/div'
            heding="Location Count"
            status="Location Count Not Found"
            count(xpath,heding,status)

            j=1
            while j<rows+1:
                xpath= '/html/body/div[1]/div/div/div[2]/div[2]/div[4]/form/div[3]/div[1]/div/div/div[2]/div/div[{}]'                      
                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                cnm = cnm.lower()
                cnm = cnm.replace(' ', '')
                if cnm.lstrip().rstrip()==pyr_nm:                          
                    xpath= '/html/body/div[1]/div/div/div[2]/div[2]/div[4]/form/div[3]/div[1]/div/div/div[2]/div/div[{}]'                             
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()
                    break
                j=j+1
            
            xpath= '/html/body/div[1]/div/div/div[2]/div[2]/div[4]/form/div[6]/div/input[1]'
            heding="Download CSV"
            status="Download CSV Button Not Found"
            click(xpath,heding,status)          
            
            counter=1
            while counter < 5:                    
                file1 = [f for f in listdir(down_path) if isfile(join(down_path, f))]
                lol_string1 = ' '.join(file1)
                lol_string2=lol_string1.split('.')[-1]                                                                                
                if lol_string2!='' and lol_string2!='tmp' and lol_string2 !="crdownload" and lol_string2 !="temp":
                    time.sleep(1)
                    file2 = [f for f in os.listdir(down_path) if os.path.isfile(os.path.join(down_path, f))]
                    if file2:
                        lol_string3 = os.path.join(down_path, file2[0])  
                        directory, file_without_extension = os.path.split(lol_string3)
                        file_name = os.path.splitext(file_without_extension)[0]

                        xlsx_file_path = os.path.join(file, f'{file_name}.xlsx')

                        dfcolumns1 = pd.read_csv(lol_string3,nrows = 1)                      
                        df = pd.read_csv(lol_string3,sep=',',usecols = list(range(len(dfcolumns1.columns))))

                        # with open(lol_string3, 'r') as file_1:
                        #     first_line = file_1.readline().strip()
                        #     header_count = len(first_line.split(','))                        
                        # try:
                        #     df = pd.read_csv(lol_string3,delimiter=',',header='infer')
                        # except Exception as e:
                        #     df = pd.read_csv(lol_string3,delimiter=',',usecols=range(header_count),header='infer')
                        
                        df.to_excel(xlsx_file_path, index=False)
                        
                        os.remove(lol_string3)

                        # filename = os.path.basename(lol_string3)
                        # new_destination_path = os.path.join(file, filename)                                                                            
                        
                        # shutil.move(lol_string3, new_destination_path)
                        # # os.rename(lol_string3, new_destination_path)                                                                        
                        wb1=load_workbook(filename=ans)
                        sheet = wb1['Data']
                        column_letter = 'B'  
                        column_cells = sheet[column_letter]
                        last_row = None
                        for cell in reversed(column_cells):
                            if cell.value:
                                last_row = cell.row
                                break                            
                        sheet['B' + str(int(last_row + 1))]='Done'      
                        wb1.save(ans)
                        wb1.close()                                                
                        break
                else:
                    time.sleep(1)    
                    counter += 1
            else:
                files = os.listdir(check_path)
                for file_name in files:
                        file_path = os.path.join(check_path, file_name)
                        try:
                            if os.path.isfile(file_path):
                                os.remove(file_path)                                   
                        except OSError as e:
                            pass    
                wb1=load_workbook(filename=ans)
                sheet = wb1['Data']
                column_letter = 'B'  
                column_cells = sheet[column_letter]
                last_row = None
                for cell in reversed(column_cells):
                    if cell.value:
                        last_row = cell.row
                        break                            
                sheet['B' + str(int(last_row + 1))]='File Error'        
                wb1.save(ans)
                wb1.close()   
                

        driver.close()
        messagebox.showinfo('Process Status', 'Process Completed...')
        sys.exit(0) 

def console():
    folder()
    
    fil=ans + '/' 
    file_path = fil + 'Console'
    file_con = fil + 'Console' + '/'
    files = [item for item in os.listdir(fil) if os.path.isfile(os.path.join(fil, item))]    

    fld='Console'
        
    if not os.path.isdir(file_path):
        os.mkdir(fil + 'Console')        
    else:
        path1 = os.path.join(fil, fld)
        shutil.rmtree(path1)
        os.mkdir(fil + 'Console')
        
    wb = openpyxl.Workbook()
    ws=wb.active    
    ws['A1']='First Name'
    ws['B1']='Last Name'
    ws['C1']='NPI'
    ws['D1']='CAQH ID'
    ws['E1']='Department'
    ws['F1']='Location'
    ws['G1']='Position'
    ws['H1']='Insurance Name'
    ws['I1']='Status'
    ws['J1']='Effective Date'
    ws['K1']='Revalidation Date'
    ws['L1']='Provider ID'
    ws['M1']='File Name'
    

    ws.title = 'Status'
    wb.save(filename=file_con +  'Consol.xlsx')
    wb.close()
    
    j=0
           
    while j < len(files):
        fl=files[j]

        ex_fn = pd.DataFrame()
        df = pd.DataFrame()
        
        ex_fn = pd.read_excel(fil + fl,engine='openpyxl',header=0)
        lst=len(list(ex_fn.head()))
        
        cum = ex_fn.iloc[:, :7]         

        tol = ex_fn.iloc[:, 7: + lst]             

        tol_lst=len(list(tol.head()))  

        k=0
        i=4
        while i < tol_lst + 4 :
            set = tol.iloc[:, k: + i]            
            hd_lst=list(set.head()) 
            hd=hd_lst[0]
            # fin_hd = hd.split('(')[1].split(')')[0]          
            set['Insurance Name']=hd
            set['File Name']=fl   
            set.replace(np.nan, '', inplace=True)
            man_lst=['Status','Effective Date','Revalidation Date','Provider ID','Insurance Name','File Name']            
            set.columns = man_lst
            
            df = cum.join(set[man_lst])
         
            book = load_workbook(filename=file_con + "Consol.xlsx")

            with pd.ExcelWriter(file_con + "Consol.xlsx", engine='openpyxl') as writer:
                writer.book = book
                for ws in book.worksheets:
                    writer.sheets[ws.title] = writer.book[ws.title]
                worksheet = writer.sheets['Status']
                startrow = worksheet.max_row if worksheet.max_row > 0 else None
                df.to_excel(writer, sheet_name='Status', index=False, startrow=startrow,header=False)
                       
            set=pd.DataFrame()
            k=i
            i=i+4        

        print(fl + "   -   Done")
        j=j+1
    
    messagebox.showinfo('Process Status', 'Process Completed...')
    sys.exit(0) 

if __name__=="__main__":        
    radio()
    