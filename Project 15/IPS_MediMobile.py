from tkinter import BOTH, LEFT, TOP, Button, Entry, Frame, Label, PhotoImage, StringVar, Tk,Radiobutton,StringVar,IntVar,filedialog
from idlelib.tooltip import Hovertip
import sys
import os
from operator import itemgetter
import time
from tkinter import messagebox
from os import listdir
from os.path import isfile, join
from datetime import datetime, date
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, StaleElementReferenceException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from urllib.parse import urlparse
import warnings
import numpy as np
from openpyxl.utils import get_column_letter,column_index_from_string
from datetime import datetime
from pytz import timezone
import requests
from zipfile import ZipFile
import subprocess

warnings.filterwarnings("ignore")

global rows
global xpath
global heding
global status
global key
global nme
global element_1  
global j

def browse():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
        image_path2 = os.path.join(sys._MEIPASS, 'Static', 'Mapping.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')
        image_path2 = os.path.join(os.getcwd(), 'Static', 'Mapping.png')

    root.title("Process File Picker")
    root.resizable(False,False)

    root.title("Process File Picker")
    root.resizable(False,False)
   
    w = 500
    h = 160
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Browse Process - File Picker",font=("Calibri",18,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
          
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=50)
    
    answer=StringVar()
    answer.set("")
       
    def browse_button():
        global filename
        answer.set("")  
        filename = filedialog.askopenfilename()  
        txt.delete(0, 'end') 
        txt.insert(0, filename)

    def Click_Done():
        global ans        
        
        ans=txt.get()        
        
        if ans=="":
           answer.set("File Path Fields Empty Is Not Allowed...")        
        else:      
            root.destroy()          
            return ans
        
    txt=Entry(Frame2,font=("Calibri",12,"bold","italic"),width=50,justify="left")
    txt.grid(row=0,column=0,padx=5,pady=10,sticky="E")    

    photo1 = PhotoImage(file=image_path2)
            
    btn1=Button(Frame2,text="Browse",command=browse_button,image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=0,column=1,padx=3,pady=0,sticky="W")     
                
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=3,y=80,width=490,height=25)
    
    title_3=Label(Frame3,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=1,padx=150,pady=0,sticky="E")

    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=3,y=105,width=500,height=50)
    
    photo2 = PhotoImage(file=image_path1)

    btn2=Button(Frame4,command=Click_Done,text="Done",image=photo2,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=3,column=0,padx=125,pady=0,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo3 = PhotoImage(file=image_path)   

    btn3=Button(Frame4,command=Close,text="Exit",image=photo3,borderwidth=0,bg="#2c3e50")
    btn3.grid(row=3,column=1,padx=15,pady=0,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn2,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn3,'Click to Exit Process',hover_delay=1000)
    myTip2 = Hovertip(btn1,'Click to Pick File',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def user_pass():
    global ans1
    global ans2
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')

    root.title("Sharepoint - User Login And File Details")
    root.resizable(False,False)

    w = 600
    h = 180
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)

    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="User Name and Password Details...",font=("Calibri",20,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack() 
 
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=698,height=150)    
    
    title3=Label(Frame2,text="User Name :",font=("Calibri",11,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title3.grid(row=0,column=0,padx=5,pady=5,sticky="W")
    
    txt1=Entry(Frame2,font=("Calibri",11,"bold","italic"),width=50,justify="left")
    txt1.grid(row=0,column=1,padx=30,pady=5,sticky="W")
    
    title4=Label(Frame2,text="Password :",font=("Calibri",11,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title4.grid(row=1,column=0,padx=5,pady=5,sticky="W")
    
    txt2=Entry(Frame2,font=("Calibri",11,"bold","italic"),width=50,justify="left")
    txt2.grid(row=1,column=1,padx=30,pady=5,sticky="W")

    answer=StringVar()
    answer.set("")

    def Click_Done():
        global ans1
        global ans2        
        
        ans1=txt1.get()
        ans2=txt2.get()        

        if ans1=="":
           answer.set("User Name Field Empty Is Not Allowed...")
        elif ans2=="":
            answer.set("Password Field Empty Is Not Allowed...")        
        else:    
            root.destroy()
            return ans1,ans2
                
    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=0,y=105,width=698,height=20)
    
    title_3=Label(Frame4,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=2,padx=200,pady=0,sticky="E")
    
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=0,y=125,width=698,height=200)
       
    photo1 = PhotoImage(file=image_path1)
    
    btn1=Button(Frame3,command=Click_Done,text="Run",image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=2,column=0,padx=140,pady=0,sticky="W")

    def Close():
        sys.exit(0)

    photo = PhotoImage(file=image_path)    

    btn2=Button(Frame3,command=Close,text="Close",image=photo,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=2,column=1,padx=100,pady=0,sticky="E")

    def disable_event():
        pass

    txt1.focus_set()

    myTip = Hovertip(btn1,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn2,'Click to Exit Process',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)
            
    root.mainloop()

def process():        
    global element_1    
    global ans1
    global ans2 
    global driver
    global j
    global dob
    global mid
    global frm_dt
    global to_dt
    global cpt

    user_pass()
    browse()

    fil=ans               
    user = ans1
    password = ans2

    try:
        options = Options()            
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--ignore-ssl-errors')
        options.add_argument("--disable-popup-blocking")
        driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options)
        driver.maximize_window()
        driver.get('https://login.medimobile.com/')
    except Exception as e:                        
        try:
            options = Options()
            options.add_argument('--ignore-certificate-errors')
            options.add_argument('--ignore-ssl-errors')
            options.add_argument("--disable-popup-blocking")
            # response = requests.get('https://chromedriver.storage.googleapis.com/LATEST_RELEASE')
            # latest_version = response.text.strip()
            # chrome_driver_url = f'https://chromedriver.storage.googleapis.com/{latest_version}/chromedriver_win32.zip'
            # response = requests.get(chrome_driver_url)
            # with open('chromedriver_win32.zip', 'wb') as f:
            #     f.write(response.content)
            # with ZipFile('chromedriver_win32.zip', 'r') as zip_ref:
            #     zip_ref.extractall('.')
            driver_path = os.path.abspath('chromedriver.exe')
            driver = webdriver.Chrome(executable_path=driver_path,options=options)
            driver.maximize_window()
            driver.get('https://login.medimobile.com/')
        except Exception as e:           
            messagebox.showinfo("Driver Version Problem","Pls Check Your Chrome Driver Version")
            sys.exit(0)
        
    def text_box(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:   
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding, status)
            sys.exit(0)        
    
    def text_box_key(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:                   
                input_field =WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath)))
                input_field.send_keys(Keys.CONTROL + "a") 
                input_field.send_keys(Keys.BACKSPACE)                 
                input_field.send_keys(key)               
                input_field.send_keys(Keys.TAB)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding, status)
            sys.exit(0)     

    def text_box_js(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:
                js_code = "document.querySelector('" + xpath + "').setAttribute('value', '" + key + "');"
                WebDriverWait(driver, 0).until(lambda driver: driver.execute_script(js_code))               
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo(heding, status)
            sys.exit(0)       

    def click(xpath,heding,status):
        counter = 0
        while counter < 15:
            try:             
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding,status)
            sys.exit(0)                                

    def count(xpath,heding,status):
        global rows
        counter = 0
        while counter < 15:
            try:             
                rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding,status)
            sys.exit(0)                         
    
    def count_1(xpath,heding,status):
        global rows_1
        counter = 0
        while counter < 15:
            try:             
                rows_1=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))                    
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo(heding,status)
            sys.exit(0) 

    def Alert():
        counter = 0
        while counter < 15:
            try:             
                WebDriverWait(driver, 0).until (EC.alert_is_present())
                a=driver.switch_to.alert
                a.accept()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo('Alert','Alert Not Present')                                                                

    xpath= "/html/body/div/div[2]/form/fieldset/div/div[2]/input"
    heding="User Name"
    status="User Name Field Not Found"
    key=user
    text_box(xpath,heding,status,key)
                
    xpath= "/html/body/div/div[2]/form/fieldset/div/div[4]/input"
    heding="Password"
    status="Password Field Not Found"
    key=password
    text_box(xpath,heding,status,key)
    
    xpath= "/html/body/div/div[2]/form/fieldset/div/div[5]/div[2]/input"
    heding="Log In"
    status="Log In Button Not Found"
    click(xpath,heding,status)   
    
    while True:
            xpath='/html/body/div/div[2]/form/div[1]/ul/li'         
            try:
                element_3 = WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH, xpath))).text
                if element_3.lstrip().rstrip()=='Username or password is incorrect.':                               
                    messagebox.showinfo('Login Status','Please Check - UserName or Password Wrong')                    
                    user_pass()
                    
                    xpath= '/html/body/div/div[2]/form/fieldset/div/div[2]/input'
                    heding="User Name"
                    status="User Name Field Not Found"
                    key=ans1
                    text_box_key(xpath,heding,status,key)
                    
                    xpath= '/html/body/div/div[2]/form/fieldset/div/div[4]/input'
                    heding="Password"
                    status="Password Field Not Found"
                    key=ans2
                    text_box(xpath,heding,status,key)
                    
                    xpath= "/html/body/div/div[2]/form/fieldset/div/div[5]/div[2]/input"
                    heding="Log In"
                    status="Log In Button Not Found"
                    click(xpath,heding,status)
                else:
                    break    
            except Exception as e:                
                    break    
    
    xpath= "/html/body/div/div[2]/form/div[1]/a"
    heding="Select Location"
    status="Select Location Table Count Not Found"        
    count(xpath,heding,status) 
    
    j=1
    while j<rows+1: 
        xpath= "/html/body/div/div[2]/form/div[1]/a[{}]/h3"                                
        cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text     
        if cnm.lstrip().rstrip()=='IPS - GA':
            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div/div[2]/form/div[1]/a[{}]'.format(j)))).click()   
            break
        j=j+1
    
    xpath= "/html/body/form/div[6]/div/div/div[3]/div/div/ul/li"
    heding="Charge Capture"
    status="Charge Capture Table Count Not Found"        
    count(xpath,heding,status) 
    
    j=1
    while j<rows+1: 
        xpath= "/html/body/form/div[6]/div/div/div[3]/div/div/ul/li[{}]/span"                                
        cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text     
        if cnm.lstrip().rstrip()=='Charge Capture':
            time.sleep(3)
            element=WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j))))
            actions = ActionChains(driver)
            actions.move_to_element(element).perform()  
            
            xpath= "/html/body/form/div[6]/div/div/div[3]/div/div/ul/li[{}]/ul/li".format(j)
            heding="Charge Capture"
            status="Charge Capture Menu Count Not Found"        
            count_1(xpath,heding,status) 
            
            i=1
            while i<rows_1+1:                                       
                xpath= "/html/body/form/div[6]/div/div/div[3]/div/div/ul/li[{}]/ul/li[{}]/a"
                element_7=WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j,i))))
                actions = ActionChains(driver)
                actions.move_to_element(element_7).perform()  
                element_text = element_7.text                    
                if element_text.lstrip().rstrip()=="Charge Review":
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j,i)))).click()  
                    break                       
                i=i+1                                                              
            break                                
        j=j+1

    file=pd.read_excel(fil,sheet_name='Medimobile',header=0)
    
    for index, row in file.iterrows():                                             
        lnm = row[0]            
        fnm=row[1]
        dob=row[2]
        st = row[3]                 
        frm_dt = row[4]    
        to_dt = row[5]
        
        wait = WebDriverWait(driver, 20)
        wait.until(lambda driver: driver.execute_script("return document.readyState === 'complete';"))

        xpath= "/html/body/form/div[3]/div/table/tbody/tr/td/div/div/table/tbody/tr/td/table[1]/tbody/tr/td/div/table/tbody/tr[2]/td/div[3]/div[1]/select/option"
        heding="Practice Count"
        status="Practice Count Not Found"        
        count(xpath,heding,status) 

        j=1
        while j<rows+1: 
            xpath= "/html/body/form/div[3]/div/table/tbody/tr/td/div/div/table/tbody/tr/td/table[1]/tbody/tr/td/div/table/tbody/tr[2]/td/div[3]/div[1]/select/option[{}]"                                
            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text     
            if cnm.lstrip().rstrip()==st:
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()   
                break
            j=j+1
        
        wait = WebDriverWait(driver, 20)
        wait.until(lambda driver: driver.execute_script("return document.readyState === 'complete';"))

        xpath= "/html/body/form/div[3]/div/table/tbody/tr/td/div/div/table/tbody/tr/td/table[1]/tbody/tr/td/div/table/tbody/tr[2]/td/div[3]/div[2]/select/option"
        heding="Status Count"
        status="Status Count Not Found"        
        count(xpath,heding,status) 

        j=1
        while j<rows+1: 
            xpath= "/html/body/form/div[3]/div/table/tbody/tr/td/div/div/table/tbody/tr/td/table[1]/tbody/tr/td/div/table/tbody/tr[2]/td/div[3]/div[2]/select/option[{}]"                                
            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text     
            if cnm.lstrip().rstrip()=='-All Status-':
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()   
                break
            j=j+1
        
        wait = WebDriverWait(driver, 20)
        wait.until(lambda driver: driver.execute_script("return document.readyState === 'complete';"))

        xpath= "/html/body/form/div[3]/div/table/tbody/tr/td/div/div/table/tbody/tr/td/table[1]/tbody/tr/td/div/table/tbody/tr[2]/td/div[4]/table/tbody/tr[1]/td[1]/input"
        heding="Last Name"
        status="Last Name Field Not Found"
        key=lnm.lstrip().rstrip()
        text_box(xpath,heding,status,key)  
        
        xpath= "/html/body/form/div[3]/div/table/tbody/tr/td/div/div/table/tbody/tr/td/table[1]/tbody/tr/td/div/table/tbody/tr[2]/td/div[4]/table/tbody/tr[1]/td[2]/input"
        heding="First Name"
        status="First Name Field Not Found"
        key=fnm.lstrip().rstrip()
        text_box(xpath,heding,status,key)  

        date_object_1 = datetime.strptime(str(dob), "%Y-%m-%d %H:%M:%S")
        dob3 = date_object_1.strftime("%m/%d/%Y")            
        
        xpath= "/html/body/form/div[3]/div/table/tbody/tr/td/div/div/table/tbody/tr/td/table[1]/tbody/tr/td/div/table/tbody/tr[2]/td/div[4]/table/tbody/tr[1]/td[4]/span/input"
        heding="Date of Birth"
        status="DOB Field Not Found"
        key=dob3
        text_box_key(xpath,heding,status,key)
        
        date_object_1 = datetime.strptime(str(frm_dt), "%Y-%m-%d %H:%M:%S")
        dob3 = date_object_1.strftime("%m/%d/%Y")            
        
        xpath= "/html/body/form/div[3]/div/table/tbody/tr/td/div/div/table/tbody/tr/td/table[1]/tbody/tr/td/div/table/tbody/tr[2]/td/div[4]/table/tbody/tr[2]/td[3]/span/input"
        heding="From Date"
        status="From Date Field Not Found"
        key=dob3
        text_box_key(xpath,heding,status,key)
        
        date_object_1 = datetime.strptime(str(to_dt), "%Y-%m-%d %H:%M:%S")
        dob3 = date_object_1.strftime("%m/%d/%Y")            
        
        xpath= "/html/body/form/div[3]/div/table/tbody/tr/td/div/div/table/tbody/tr/td/table[1]/tbody/tr/td/div/table/tbody/tr[2]/td/div[4]/table/tbody/tr[2]/td[4]/span/input"
        heding="To Date"
        status="To Date Field Not Found"
        key=dob3
        text_box_key(xpath,heding,status,key)
        
        xpath= "/html/body/form/div[3]/div/table/tbody/tr/td/div/div/table/tbody/tr/td/table[1]/tbody/tr/td/div/table/tbody/tr[2]/td/div[3]/div[5]/input[1]"
        heding="Claims Click"
        status="Claims Click Button Not Found"
        click(xpath,heding,status) 

        wait = WebDriverWait(driver, 20)
        wait.until(lambda driver: driver.execute_script("return document.readyState === 'complete';"))
        
        time.sleep(2)

        # desired_zoom = 0.75  
        # driver.execute_script(f"document.body.style.zoom = '{desired_zoom}';")
        
        lst=[]

        # try:                
        #     xpath="/html/body/form/div[3]/div/table/tbody/tr/td/div/div/table/tbody/tr/td/table[2]/tbody/tr/td/div/table/tbody/tr[2]/td/div[2]/div[3]/table/tbody/tr[2]"
        #     WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath))).click()
        #     actions = ActionChains(driver)
        #     actions.move_to_element(element).perform()
        #     print('Yes')
        # except Exception as e:
        #     print(e)



        # counter = 0
        # while counter < 5:
                 # /html/body/form/div[3]/div/table/tbody/tr/td/div/div/table/tbody/tr/td/table[2]/tbody/tr/td/div/table/tbody/tr[2]/td/div[2]/div[2]/table/tbody/tr/td[1]/span
        try:                
            xpath="/html/body/form/div[3]/div/table/tbody/tr/td/div/div/table/tbody/tr/td/table[2]/tbody/tr/td/div/table/tbody/tr[2]/td/div[2]/div[3]/table/tbody/tr[2]"
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath))).click()
            # element = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, xpath)))
            # actions = ActionChains(driver)
            # actions.move_to_element(element).perform()
            # element.click()
            
            # WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, xpath))).click()
            # driver.execute_script("arguments[0].scrollIntoView(true);", element)
            # element.click()            
            
            xpath= "/html/body/form/div[6]/div/div[5]/div/div[1]/div[1]/div/div[1]/div[1]/div[2]/a[3]"
            heding="Search Medimobile Click"
            status="Search Medimobile Click Button Not Found"
            click(xpath,heding,status)  
                                                                    
            try:  
                no_pit = WebDriverWait(driver, 2).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[6]/div/div[5]/div/div[1]/div[10]/div/div[1]/div[3]/span/b'))).text
                
                wb1=load_workbook(filename=fil)
                sheet = wb1['Medimobile']
                column_letter = 'L'  
                column_cells = sheet[column_letter]
                last_row = None
                for cell in reversed(column_cells):
                    if cell.value:
                        last_row = cell.row
                        break                            
                sheet['L' + str(int(last_row + 1))]=no_pit 
                wb1.save(fil)
                wb1.close()  

                xpath= "/html/body/form/div[6]/div/div[5]/div/div[1]/div[10]/div/div[1]/div[2]/div[5]/input[2]"
                close_button = driver.find_element(By.XPATH, xpath)
                driver.execute_script("arguments[0].click();", close_button)

                desired_zoom = 0.75  
                driver.execute_script(f"document.body.style.zoom = '{desired_zoom}';")

                xpath= "/html/body/form/div[6]/div/div[5]/div/div[2]/div[2]/div/div[5]/input[2]"
                cancel_button = driver.find_element(By.XPATH, xpath)
                driver.execute_script("arguments[0].click();", cancel_button)

                xpath= "/html/body/form/div[3]/div/div/div[1]/div[1]/ul/li[1]/a"
                home_button = driver.find_element(By.XPATH, xpath)
                driver.execute_script("arguments[0].click();", home_button)                                 

                counter = 0
                while counter < 10:
                    try:
                        xpath= "/html/body/header/div[3]/div/ul/li"                                                    
                        element=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath)))
                        actions = ActionChains(driver)
                        actions.move_to_element(element).perform()  
                        
                        xpath= "/html/body/header/div[3]/div/ul/li/ul/li"
                        heding="Charge Review"
                        status="Charge Review Home Count Not Found"        
                        count_1(xpath,heding,status) 
                        
                        i=1
                        while i<rows_1+1:                                       
                            xpath= "/html/body/header/div[3]/div/ul/li/ul/li[{}]/a"
                            element_7=WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(i))))
                            actions = ActionChains(driver)
                            actions.move_to_element(element_7).perform()  
                            element_text = element_7.text                    
                            if element_text.lstrip().rstrip()=="Charge Review":
                                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(i)))).click()  
                                break                       
                            i=i+1                                                              
                        break     
                    except Exception as e:
                        time.sleep(1)
                        counter += 1                   
                else:
                    messagebox.showinfo("Mouse perform", "Mouse Not Move In Home")
                    sys.exit(0)  
                    
            except Exception as e:
                xpath= "/html/body/form/div[6]/div/div[5]/div/div[1]/div[10]/div/div[1]/div[3]/table/tbody/tr/td[7]/a"
                heding="Copy Data Click"
                status="Copy Data Click Button Not Found"
                click(xpath,heding,status)  

                desired_zoom = 0.75  
                driver.execute_script(f"document.body.style.zoom = '{desired_zoom}';")                            

                xpath = '/html/body/form/div[6]/div/div[5]/div/div[1]/div[10]/div/div[2]/div[2]/input[1]'
                scroll_script = f"var element = document.evaluate('{xpath}', document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue; element.scrollIntoView();"
                driver.execute_script(scroll_script)                                        
                
                apply_button = driver.find_element(By.XPATH, xpath)
                driver.execute_script("arguments[0].click();", apply_button)

                xpath= '/html/body/div[6]/div/div[2]/input[1]'
                ok_button = driver.find_element(By.XPATH, xpath)
                driver.execute_script("arguments[0].click();", ok_button)
                                                
                xpath= "/html/body/form/div[6]/div/div[5]/div/div[1]/div[10]/div/div[1]/div[2]/div[5]/input[2]"
                close_button = driver.find_element(By.XPATH, xpath)
                driver.execute_script("arguments[0].click();", close_button)
        
                lst1=[]
                                                                                                # /html/body/form/div[6]/div/div[5]/div/div[1]/div[1]/div/div[1]/table/tbody/tr[7]/td[4]/input[1]
                ssn = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[6]/div/div[5]/div/div[1]/div[1]/div/div[1]/table/tbody/tr[7]/td[4]/input[1]')))
                ssn = ssn.get_attribute('value')  
                if ssn==' ' or ssn=='':
                    ssn='N/A' 
                    lst1.append(ssn)   
                else:
                    lst1.append(ssn)

                lt=len(lst1)
                if lt==0:                           
                    lst.append('N/A') 
                else:
                    vr=lst1[0]
                    lst.append(vr)

                xpath= "/html/body/form/div[6]/div/div[4]/div/div/div/ul/li[2]/a"
                payer_tab = driver.find_element(By.XPATH, xpath)
                driver.execute_script("arguments[0].click();", payer_tab)

                lst2=[]
                hsp_ins = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[6]/div/div[5]/div/div[1]/div[2]/div/div[1]/table[1]/tbody/tr[7]/td[2]/input')))
                hsp_ins = hsp_ins.get_attribute('value') 

                if hsp_ins==' ' or hsp_ins=='':
                    hsp_ins='N/A' 
                    lst2.append(hsp_ins)   
                else:
                    lst2.append(hsp_ins)   

                lt=len(lst2)
                if lt==0:                           
                    lst.append('N/A') 
                else:
                    vr=lst2[0]
                    lst.append(vr)
                
                lst3=[]                                                                              
                pln_id = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[6]/div/div[5]/div/div[1]/div[2]/div/div[1]/table[1]/tbody/tr[8]/td[2]/input')))
                pln_id = pln_id.get_attribute('value') 

                if pln_id==' ' or pln_id=='':
                    pln_id='N/A' 
                    lst3.append(pln_id)   
                else:
                    lst3.append(pln_id)   

                lt=len(lst3)
                if lt==0:                           
                    lst.append('N/A') 
                else:
                    vr=lst3[0]
                    lst.append(vr)
                
                lst4=[]                                                                                                   
                pol_num = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[6]/div/div[5]/div/div[1]/div[2]/div/div[1]/table[1]/tbody/tr[23]/td[2]/input')))
                pol_num = pol_num.get_attribute('value') 

                if pol_num==' ' or pol_num=='':
                    pol_num='N/A' 
                    lst4.append(pol_num)   
                else:
                    lst4.append(pol_num)   

                lt=len(lst4)
                if lt==0:                           
                    lst.append('N/A') 
                else:
                    vr=lst4[0]
                    lst.append(vr)

                lst5=[]                                                                                                  
                aut_num = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div[6]/div/div[5]/div/div[1]/div[2]/div/div[1]/table[1]/tbody/tr[25]/td[2]/input')))
                aut_num = aut_num.get_attribute('value') 

                if aut_num==' ' or aut_num=='':
                    aut_num='N/A' 
                    lst5.append(aut_num)   
                else:
                    lst5.append(aut_num)   

                lt=len(lst5)
                if lt==0:                           
                    lst.append('N/A') 
                else:
                    vr=lst5[0]
                    lst.append(vr)
                
                xpath= "/html/body/form/div[6]/div/div[5]/div/div[2]/div[2]/div/div[5]/input[2]"
                cancel_button = driver.find_element(By.XPATH, xpath)
                driver.execute_script("arguments[0].click();", cancel_button)                             

                wb1=load_workbook(fil)
                sheet = wb1['Medimobile']
                column_letter = 'L'  
                column_cells = sheet[column_letter]
                last_row = None
                for cell in reversed(column_cells):
                    if cell.value:
                        last_row = cell.row
                        break              
                start_column = 'G'
                current_column_index = openpyxl.utils.column_index_from_string(start_column)
                current_row = last_row + 1

                for value in lst:
                    current_column = openpyxl.utils.get_column_letter(current_column_index)
                    sheet[current_column + str(current_row)] = value
                    current_column_index += 1                
                sheet['L' + str(int(last_row + 1))]='Done'
                wb1.save(filename=fil)
                wb1.close()     
                        
                xpath= "/html/body/form/div[3]/div/div/div[1]/div[1]/ul/li[1]/a"
                home_button = driver.find_element(By.XPATH, xpath)
                driver.execute_script("arguments[0].click();", home_button)                 

                counter = 0
                while counter < 10:
                    try:
                        xpath= "/html/body/header/div[3]/div/ul/li"                                                    
                        element=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath)))
                        actions = ActionChains(driver)
                        actions.move_to_element(element).perform()  
                        
                        xpath= "/html/body/header/div[3]/div/ul/li/ul/li"
                        heding="Charge Review"
                        status="Charge Review Home Count Not Found"        
                        count_1(xpath,heding,status) 
                        
                        i=1
                        while i<rows_1+1:                                       
                            xpath= "/html/body/header/div[3]/div/ul/li/ul/li[{}]/a"
                            element_7=WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(i))))
                            actions = ActionChains(driver)
                            actions.move_to_element(element_7).perform()  
                            element_text = element_7.text                    
                            if element_text.lstrip().rstrip()=="Charge Review":
                                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(i)))).click()  
                                break                       
                            i=i+1                                                              
                        break                             
                    except Exception as e:
                        time.sleep(1)
                        counter += 1                  
                else:
                    messagebox.showinfo("Mouse perform", "Mouse Not Move In Home")
                    sys.exit(0)                            

        except Exception as e:                    
            try:                                   
                xpath="/html/body/form/div[3]/div/table/tbody/tr/td/div/div/table/tbody/tr/td/table[2]/tbody/tr/td/div/table/tbody/tr[2]/td/div[2]/div[2]/table/tbody/tr/td[1]/span"
                st = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, xpath))).text    
                
                wb1=load_workbook(filename=fil)
                sheet = wb1['Medimobile']
                column_letter = 'L'  
                column_cells = sheet[column_letter]
                last_row = None
                for cell in reversed(column_cells):
                    if cell.value:
                        last_row = cell.row
                        break                            
                sheet['L' + str(int(last_row + 1))]=st 
                wb1.save(fil)
                wb1.close()  
                
                xpath= "/html/body/form/div[3]/div/div/div[1]/div[1]/ul/li[1]/a"
                home_button = driver.find_element(By.XPATH, xpath)
                driver.execute_script("arguments[0].click();", home_button)  

                counter = 0
                while counter < 10:
                    try:
                        xpath= "/html/body/header/div[3]/div/ul/li"                                                    
                        element=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath)))
                        actions = ActionChains(driver)
                        actions.move_to_element(element).perform()  
                        
                        xpath= "/html/body/header/div[3]/div/ul/li/ul/li"
                        heding="Charge Review"
                        status="Charge Review Home Count Not Found"        
                        count_1(xpath,heding,status) 
                        
                        i=1
                        while i<rows_1+1:                                       
                            xpath= "/html/body/header/div[3]/div/ul/li/ul/li[{}]/a"
                            element_7=WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(i))))
                            actions = ActionChains(driver)
                            actions.move_to_element(element_7).perform()  
                            element_text = element_7.text                    
                            if element_text.lstrip().rstrip()=="Charge Review":
                                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(i)))).click()  
                                break                       
                            i=i+1                                                              
                        break     
                    except Exception as e:
                        time.sleep(1)
                        counter += 1                   
                else:
                    messagebox.showinfo("Mouse perform", "Mouse Not Move In Home")
                    sys.exit(0)  
                
            except Exception as e:

                wb1=load_workbook(filename=fil)
                sheet = wb1['Medimobile']
                column_letter = 'L'  
                column_cells = sheet[column_letter]
                last_row = None
                for cell in reversed(column_cells):
                    if cell.value:
                        last_row = cell.row
                        break                            
                sheet['L' + str(int(last_row + 1))]='Error' 
                wb1.save(fil)
                wb1.close()                                   

                xpath= "/html/body/form/div[3]/div/div/div[1]/div[1]/ul/li[1]/a"
                home_button = driver.find_element(By.XPATH, xpath)
                driver.execute_script("arguments[0].click();", home_button)  

                counter = 0
                while counter < 10:
                    try:
                        xpath= "/html/body/header/div[3]/div/ul/li"                                                    
                        element=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath)))
                        actions = ActionChains(driver)
                        actions.move_to_element(element).perform()  
                        
                        xpath= "/html/body/header/div[3]/div/ul/li/ul/li"
                        heding="Charge Review"
                        status="Charge Review Home Count Not Found"        
                        count_1(xpath,heding,status) 
                        
                        i=1
                        while i<rows_1+1:                                       
                            xpath= "/html/body/header/div[3]/div/ul/li/ul/li[{}]/a"
                            element_7=WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(i))))
                            actions = ActionChains(driver)
                            actions.move_to_element(element_7).perform()  
                            element_text = element_7.text                    
                            if element_text.lstrip().rstrip()=="Charge Review":
                                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(i)))).click()  
                                break                       
                            i=i+1                                                              
                        break     
                    except Exception as e:
                        time.sleep(1)
                        counter += 1                   
                else:
                    messagebox.showinfo("Mouse perform", "Mouse Not Move In Home")
                    sys.exit(0)         
    driver.quit()
    messagebox.showinfo("Process Status", "Process Completed")
    sys.exit(0)    

if __name__=="__main__":        
    process()