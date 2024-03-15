from tkinter import BOTH, LEFT, TOP, Button, Entry, Frame, Label, PhotoImage, StringVar, Tk,Radiobutton,StringVar,IntVar,filedialog
from idlelib.tooltip import Hovertip
import sys
import os
from operator import itemgetter
import time
from tkinter import messagebox
from os import listdir
from os.path import isfile, join
from datetime import datetime, date
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, StaleElementReferenceException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import openpyxl
import shutil
import xlrd
import pyxlsb
import pandas as pd
import win32com.client as win32
from win32com.client import constants
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from urllib.parse import urlparse
import warnings
import numpy as np
from openpyxl.utils import get_column_letter,column_index_from_string
from datetime import datetime
import pywintypes
from pytz import timezone
import pythoncom
import requests
from zipfile import ZipFile

warnings.filterwarnings("ignore")

global rows
global xpath
global heding
global status
global key
global nme
element_1 = None


def radio():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')

    root.title("Sharepoint - User Login And File Details")
    root.resizable(False,False)

    root.title("PCC Process Report")
    root.resizable(False,False)
   
    w = 600
    h = 280
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Please Select PCC Files Download Report (OR) PCC Data Extract Report (OR) PCC Console Report",font=("Calibri",11,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
    
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=300)
    
    title1=Label(Frame2,text="PCC Download Report :",font=("Calibri",17,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title1.grid(row=0,column=0,padx=45,pady=5,sticky="W")

    title2=Label(Frame2,text="PCC Data Extract Report :",font=("Calibri",17,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title2.grid(row=1,column=0,padx=45,pady=5,sticky="W")

    title3=Label(Frame2,text="PCC Console Report :",font=("Calibri",17,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title3.grid(row=2,column=0,padx=45,pady=5,sticky="W")

    answer=StringVar()
    answer.set("")

    title_3=Label(Frame2,text=answer.get(),textvariable=answer,font=("Calibri",12,"bold","italic"),bg="#2c3e50",fg="Red",justify="center",width=68)
    title_3.grid(row=3,column=0,columnspan=2,padx=0,pady=0,sticky="W")

    def Radio(*event):
         answer.set("")

    global var

    var = IntVar()

    R1 = Radiobutton(Frame2,text="PCC Download",variable=var, value=1,font=("Calibri",9,"bold","italic"),bg="Gold",fg="Black",width=20,justify="left",command=lambda: Radio())
    R1.grid(row=0,column=1,padx=0,pady=5,sticky="W")

    R2 = Radiobutton(Frame2,text="PCC Data",variable=var,value=2,font=("Calibri",9,"bold","italic"),bg="Gold",fg="Black",width=20,command=lambda: Radio())
    R2.grid(row=1,column=1,padx=0,pady=5,sticky="W")

    R3 = Radiobutton(Frame2,text="PCC Consol",variable=var,value=3,font=("Calibri",9,"bold","italic"),bg="Gold",fg="Black",width=20,command=lambda: Radio())
    R3.grid(row=2,column=1,padx=0,pady=5,sticky="W")
    
    
    def Click_Done():
        selection = str(var.get())

        if selection==str(1):
           answer.set("")
           root.destroy() 
           process()
        elif selection==str(2):
            answer.set("")
            root.destroy()
            upload()
        elif selection==str(3):
            answer.set("")
            root.destroy()
            console()
        else:    
            answer.set("Please Select Any One Option...")            
    
    photo1 = PhotoImage(file=image_path1)

    btn=Button(Frame2,command=Click_Done,text="Done",image=photo1,borderwidth=0,bg="#2c3e50")
    btn.grid(row=4,column=0,padx=125,pady=20,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo = PhotoImage(file=image_path)   

    btn1=Button(Frame2,command=Close,text="Exit",image=photo,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=4,column=1,padx=15,pady=20,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn1,'Click to Exit Process',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def browse():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
        image_path2 = os.path.join(sys._MEIPASS, 'Static', 'Mapping.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')
        image_path2 = os.path.join(os.getcwd(), 'Static', 'Mapping.png')

    root.title("Process File Picker")
    root.resizable(False,False)

    root.title("Process File Picker")
    root.resizable(False,False)
   
    w = 500
    h = 160
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Browse Process - File Picker",font=("Calibri",18,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
          
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=50)
    
    answer=StringVar()
    answer.set("")
       
    def browse_button():
        global filename
        answer.set("")  
        filename = filedialog.askopenfilename()  
        txt.delete(0, 'end') 
        txt.insert(0, filename)

    def Click_Done():
        global ans        
        
        ans=txt.get()        
        
        if ans=="":
           answer.set("File Path Fields Empty Is Not Allowed...")        
        else:      
            root.destroy()          
            return ans
        
    txt=Entry(Frame2,font=("Calibri",12,"bold","italic"),width=50,justify="left")
    txt.grid(row=0,column=0,padx=5,pady=10,sticky="E")    

    photo1 = PhotoImage(file=image_path2)
            
    btn1=Button(Frame2,text="Browse",command=browse_button,image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=0,column=1,padx=3,pady=0,sticky="W")     
                
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=3,y=80,width=490,height=25)
    
    title_3=Label(Frame3,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=1,padx=150,pady=0,sticky="E")

    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=3,y=105,width=500,height=50)
    
    photo2 = PhotoImage(file=image_path1)

    btn2=Button(Frame4,command=Click_Done,text="Done",image=photo2,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=3,column=0,padx=125,pady=0,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo3 = PhotoImage(file=image_path)   

    btn3=Button(Frame4,command=Close,text="Exit",image=photo3,borderwidth=0,bg="#2c3e50")
    btn3.grid(row=3,column=1,padx=15,pady=0,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn2,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn3,'Click to Exit Process',hover_delay=1000)
    myTip2 = Hovertip(btn1,'Click to Pick File',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def folder():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
        image_path2 = os.path.join(sys._MEIPASS, 'Static', 'Mapping.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')
        image_path2 = os.path.join(os.getcwd(), 'Static', 'Mapping.png')

    root.title("Process Folder Picker")
    root.resizable(False,False)

    root.title("Process Folder Picker")
    root.resizable(False,False)
   
    w = 500
    h = 160
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Browse Process - Folder Picker",font=("Calibri",18,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
          
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=50)
    
    answer=StringVar()
    answer.set("")
       
    def browse_button():
        global filename
        answer.set("")  
        filename = filedialog.askdirectory()  
        txt.delete(0, 'end') 
        txt.insert(0, filename)

    def Click_Done():
        global ans        
        
        ans=txt.get()        
        
        if ans=="":
           answer.set("Folder Path Fields Empty Is Not Allowed...")        
        else:      
            root.destroy()          
            return ans
        
    txt=Entry(Frame2,font=("Calibri",12,"bold","italic"),width=50,justify="left")
    txt.grid(row=0,column=0,padx=5,pady=10,sticky="E")    

    photo1 = PhotoImage(file=image_path2)
            
    btn1=Button(Frame2,text="Browse",command=browse_button,image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=0,column=1,padx=3,pady=0,sticky="W")     
                
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=3,y=80,width=490,height=25)
    
    title_3=Label(Frame3,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=1,padx=150,pady=0,sticky="E")

    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=3,y=105,width=500,height=50)
    
    photo2 = PhotoImage(file=image_path1)

    btn2=Button(Frame4,command=Click_Done,text="Done",image=photo2,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=3,column=0,padx=125,pady=0,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo3 = PhotoImage(file=image_path)   

    btn3=Button(Frame4,command=Close,text="Exit",image=photo3,borderwidth=0,bg="#2c3e50")
    btn3.grid(row=3,column=1,padx=15,pady=0,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn2,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn3,'Click to Exit Process',hover_delay=1000)
    myTip2 = Hovertip(btn1,'Click to Pick Folder',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def process():
        global element_1
        global down_path
        browse()
        
        fil_path =ans
                        
        fil=fil_path 
                                
        folder_create = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Downloads/')
        check_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Downloads/Temp')
        down_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Downloads\\Temp\\')
        file = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Downloads\\Temp\\Files\\')
            
        fld='Temp'
        
        if not os.path.isdir(check_path):
            os.mkdir(folder_create + 'Temp')
            os.mkdir(down_path + 'Files')
        else:
            path1 = os.path.join(folder_create, fld)
            shutil.rmtree(path1)
            os.mkdir(folder_create + 'Temp')
            os.mkdir(down_path + 'Files')
        
        # wb = openpyxl.Workbook()
        # ws=wb.active    
        # ws['A1']='PCC Name'
        # ws['B1']='Link'
        # ws['C1']='User Name'
        # ws['D1']='Password'
        # ws['E1']='Final Status'
        # ws['F1']='Previous address'
        # ws['G1']='Postal/Zip Code'
        # ws['H1']='City'
        # ws['I1']='County'
        # ws['J1']='Country'
        # ws['K1']='Prov/State'
                
        # ws.title = 'Data'
        # wb.save(filename=file + 'Consol.xlsx')
        # wb.close()
        

        # wb1=load_workbook(filename=fil)
        # sheet = wb1['Data']
        # sheet.delete_cols(14) 
        # sheet['N1']='Status'
        # wb1.save(fil)
        # wb1.close()   

        try:
            options = Options()            
            options.add_argument('--ignore-certificate-errors')
            options.add_argument('--ignore-ssl-errors')
            prefs = {"download.default_directory" : down_path,
                            "download.prompt_for_download": False,
                            "download.directory_upgrade": True,
                            "plugins.always_open_pdf_externally": True,
                            "profile.password_manager_enabled": False,
                            "credentials_enable_service": False,                            
                            }
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_experimental_option("useAutomationExtension", False)
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option("prefs",prefs)
            driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options)
            driver.maximize_window()            
        except Exception as e:
            try:
                options =  webdriver.ChromeOptions()        
                options.add_argument('--ignore-certificate-errors')
                options.add_argument('--ignore-ssl-errors')                
                prefs = {"download.default_directory" : down_path,
                            "download.prompt_for_download": False,
                            "download.directory_upgrade": True,
                            "plugins.always_open_pdf_externally": True,
                            "profile.password_manager_enabled": False,
                            "credentials_enable_service": False,                            
                            }
                options.add_argument("--disable-blink-features=AutomationControlled")
                options.add_experimental_option("useAutomationExtension", False)
                options.add_experimental_option("excludeSwitches", ["enable-automation"])
                options.add_experimental_option("prefs",prefs)                
                driver_path = os.path.abspath('chromedriver.exe')                
                driver = webdriver.Chrome(executable_path=driver_path,options=options)                
                driver.maximize_window()                                                         
            except Exception as e:
                messagebox.showinfo("Driver Problem","Pls Check Your Driver Version VS Chrome Version")
                sys.exit(0)
             
        def text_box(xpath,heding,status,key):                
            counter = 0
            while counter < 15:
                try:                          
                    WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding, status)
                sys.exit(0)        
                
        def click(xpath,heding,status):
            counter = 0
            while counter < 15:
                try:             
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding,status)
                sys.exit(0)                                

        def count(xpath,heding,status):
            global rows
            counter = 0
            while counter < 15:
                try:             
                    rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))                    
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding,status)
                sys.exit(0)          
        
        def count1(xpath,heding,status):
            global rows1
            counter = 0
            while counter < 15:
                try:             
                    rows1=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))                    
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding,status)
                sys.exit(0)   

        def Alert():
            counter = 0
            while counter < 5:
                try:             
                    WebDriverWait(driver, 0).until (EC.alert_is_present())
                    a=driver.switch_to.alert
                    a.accept()
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo('Alert','Alert Not Present')

        df=pd.read_excel(fil,sheet_name='Data',header=0)                         
        m=1
        for index, row in df.iterrows():  
            nme = row[1]                                          
            lnk = row[2]              
            umn = row[3]     
            pas = row[4]
            # dt_in=row[11]            
            # dt_out=row[12]            

            driver.get(lnk.lstrip().rstrip())
            tit = driver.title

            time.sleep(2)

            if tit=='PointClickCare Login':                    
                # xpath= '//input[@id="un"]'                
                heding="User ID"
                status="User ID Field Not Found"
                key= umn
                # text_box(xpath,heding,status,key)

                counter = 0
                while counter < 15:
                    try:                          
                        WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'//input[@id="un"]'))).send_keys(key)
                        break
                    except Exception as e:
                        try:
                            WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'//input[@id="username"]'))).send_keys(key)
                            break
                        except Exception as e:
                            time.sleep(1)
                            counter += 1
                else:
                    messagebox.showinfo(heding, status)
                    sys.exit(0)        
              
                xpath= '//button[@id="id-next"]'               
                heding="Next Button"
                status="Next Button Not Found"
                click(xpath,heding,status)
                
                xpath= '//input[@id="password"]'
                heding="Password"
                status="Password Field Not Found"
                key= pas
                text_box(xpath,heding,status,key)
                
                xpath= '//button[@id="id-submit"]'
                heding="SIGN IN Button"
                status="SIGN IN Button Not Found"
                click(xpath,heding,status)

                wait = WebDriverWait(driver, 20)
                wait.until(lambda driver: driver.execute_script("return document.readyState === 'complete';"))

                time.sleep(2)
                try:                    
                    element_1 = WebDriverWait(driver, 2).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="app"]/div/main/div/div[1]/nav[1]/div[2]/div/div/div[1]/div[3]/div[1]/div/div'))).text
                    if element_1.lstrip().rstrip() =="There was an error logging you in. Contact your System Administrator to check your account. PointClickCare cannot reset passwords.":
                        wb1=load_workbook(filename=fil)
                        sheet = wb1['Data']
                        column_letter = 'N'  
                        column_cells = sheet[column_letter]
                        last_row = None
                        for cell in reversed(column_cells):
                            if cell.value:
                                last_row = cell.row
                                break                            
                        sheet['N' + str(int(last_row + 1))]='Login Error'        
                        wb1.save(fil)
                        wb1.close()         

                except Exception as e:           
                
                    try:
                        element_1 = WebDriverWait(driver, 2).until(EC.visibility_of_element_located((By.XPATH, '//div[@id="loginerror"]'))).text
                        if element_1 !="":
                            wb1=load_workbook(filename=fil)
                            sheet = wb1['Data']
                            column_letter = 'N'  
                            column_cells = sheet[column_letter]
                            last_row = None
                            for cell in reversed(column_cells):
                                if cell.value:
                                    last_row = cell.row
                                    break                            
                            sheet['N' + str(int(last_row + 1))]='Error'        
                            wb1.save(fil)
                            wb1.close()                                             
                    except Exception as e:                                   
                        xpath= '//li[@id="QTF_reportingTab"]'
                        heding="Report Button"
                        status="Report Button Not Found"
                        click(xpath,heding,status)                                        
                        
                        xpath= '//a[@id="pccFacLink"]'
                        heding="Facilities Button"
                        status="Facilities Button Not Found"
                        click(xpath,heding,status)

                        xpath='//ul[@id="optionList"]/li'
                        heding="Facilities Count"
                        status="Facilities Count Not Found"
                        count(xpath,heding,status) 
                                            
                        j=1
                        while j<rows+1:
                            xpath= '//ul[@id="optionList"]/li[{}]'                      
                            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                            if cnm==nme: 
                                xpath= '//ul[@id="optionList"]/li[{}]/a'                                
                                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()
                                break
                            j=j+1
                        
                        xpath= '/html/body/div[2]/div/div[1]/div/div/input'
                        heding="Search Report"
                        status="Search Report Field Not Found"
                        key= 'Resident List Report *NEW*'

                        counter = 0
                        while counter < 5:
                            try:                          
                                textbox=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath)))
                                textbox.send_keys(key)
                                textbox.send_keys(Keys.TAB)                            
                                break
                            except Exception as e:
                                time.sleep(1)
                                counter += 1
                        else:
                            messagebox.showinfo(heding, status)
                            sys.exit(0)                                          

                        xpath= '/html/body/div[2]/div/div[2]/div/div/div/a[2]/span'
                        heding="All Button"
                        status="All Button Not Found"
                        click(xpath,heding,status)
                        
                        xpath='/html/body/div[2]/div/div[3]/div[2]/div[2]/li'
                        heding="Report Name List"
                        status="Report Name List Not Found"
                        count(xpath,heding,status)                                                                   
                        
                        if rows != 0:
                            while counter < 5:
                                try: 
                                    j=1
                                    while j<rows+1:
                                        xpath= '/html/body/div[2]/div/div[3]/div[2]/div[2]/li[{}]'                      
                                        cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                        if cnm.startswith('Resident List Report *NEW*'): 
                                            xpath= '/html/body/div[2]/div/div[3]/div[2]/div[2]/li[{}]/a'                                
                                            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()
                                            break
                                        j=j+1
                                    break
                                except Exception as e:
                                    time.sleep(1)
                                    counter += 1
                            else:
                                messagebox.showinfo(heding, status)
                                sys.exit(0) 

                            xpath='//table[@id="filteroptionstable"]/tbody/tr/td/table[1]/tbody/tr'
                            heding="Filteroptionstable Table"
                            status="Filteroptionstable Table Not Found"
                            count(xpath,heding,status)                                                                                                                                                           
                            
                            j=1
                            while j<rows+1:                        
                                try:                            
                                    xpath= '//table[@id="filteroptionstable"]/tbody/tr/td/table[1]/tbody/tr[{}]/td[2]//div/label'                 
                                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text                            
                                    if cnm=='Include Outpatients':
                                        xpath= '//table[@id="filteroptionstable"]/tbody/tr/td/table[1]/tbody/tr[{}]/td[2]//div/input'
                                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click() 
                                    elif cnm=='Gender':                        
                                        xpath= '//table[@id="filteroptionstable"]/tbody/tr/td/table[1]/tbody/tr[{}]/td[2]//div/input'
                                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()
                                        break                                
                                    # elif cnm=='Admission Date Range':  
                                    #     xpath= '//table[@id="filteroptionstable"]/tbody/tr/td/table[1]/tbody/tr[{}]/td[2]/div/input'
                                    #     WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()
                                        
                                    #     date_object = datetime.strptime(str(dt_in), "%Y-%m-%d %H:%M:%S")
                                    #     fdt = date_object.strftime("%m/%d/%Y")                                    
                                        
                                    #     key= fdt  
                                                                            
                                    #     # xpath= '//table[@id="filteroptionstable"]/tbody/tr/td/table[1]/tbody/tr[{}]/td[3]/div/input[2]'                                    
                                    #     xpath= '//*[@id="daterange_6_1_dummy"]'

                                    #     WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                                                        
                                    #     WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).send_keys(key)
                                        
                                    #     date_object = datetime.strptime(str(dt_out), "%Y-%m-%d %H:%M:%S")
                                    #     ldt = date_object.strftime("%m/%d/%Y")                                                                        
                                        
                                    #     key= ldt                                                                       
                                        
                                    #     # xpath= '//table[@id="filteroptionstable"]/tbody/tr/td/table[1]/tbody/tr[{}]/td[3]/div/input[4]'
                                    #     xpath= '//*[@id="daterange_6_2_dummy"]'

                                    #     WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()
                                    #     WebDriverWait(driver, 1).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).send_keys(key)                                
                                except Exception as e:   
                                    pass
                                j=j+1                                                
                            
                            xpath='//input[@class="resident_field_rl_home_phone"]'
                            heding="Home Phone"
                            status="Home Phone Not Found"
                            click(xpath,heding,status) 
                            
                            xpath='//input[@class="resident_field_rl_marital_status"]'
                            heding="Marital Status"
                            status="Marital Status Not Found"
                            click(xpath,heding,status) 

                            xpath='//input[@class="resident_field_rl_opt_genger"]'
                            heding="Gender"
                            status="Gender Not Found"
                            click(xpath,heding,status) 
                           
                            xpath='//input[@class="resident_field_rl_birth_date"]'
                            heding="Birth Date"
                            status="Birth Date Not Found"
                            click(xpath,heding,status)                                                 
                             
                            xpath='//div[@id="Identifiers"]/table/tbody/tr'
                            heding="Identifiers Table"
                            status="Identifiers Table Not Found"
                            count(xpath,heding,status)   
                        
                            j=2
                            while j<rows+1:                                  
                                xpath='//div[@id="Identifiers"]/table/tbody/tr[{}]/td'.format(j)
                                heding="Identifiers Count"
                                status="Identifiers Count Not Found"
                                count1(xpath,heding,status)    
                                                          
                                i=1
                                while i<rows1+1:                                  
                                    try:                                    
                                            xpath= '//div[@id="Identifiers"]/table/tbody/tr[{}]/td[{}]/label/input'
                                            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j,i)))).click()                                                                    
                                    except Exception as e:   
                                        pass
                                    i=i+1

                                j=j+1

                            # xpath='//input[@name="ESOLreportcolumnid_5" and contains(following-sibling::text()[1], "Veteran")]'                                               
                            # try:             
                            #     WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()                            
                            # except Exception as e:
                            #     pass                        
                            # //*[@id="reportFormatType"]
                                 # /html/body/table[2]/tbody/tr/td/form[2]/table[7]/tbody/tr[2]/td[2]/select/option[1]
                            # xpath='/html/body/table[3]/tbody/tr/td/form[2]/table[7]/tbody/tr[2]/td[2]/select/option'
                            heding="Report Output Format"
                            status="Report Output Format Not Found"
                            # count(xpath,heding,status)  
                            
                            counter = 0
                            while counter < 15:                                
                                try:
                                    xpath='/html/body/table[3]/tbody/tr/td/form[2]/table[7]/tbody/tr[2]/td[2]/select/option'
                                    rows2=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))    
                                    break
                                except Exception as e:
                                    try:
                                        xpath='/html/body/table[2]/tbody/tr/td/form[2]/table[7]/tbody/tr[2]/td[2]/select/option'
                                        rows2=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))    
                                        break
                                    except Exception as e:
                                        time.sleep(1)
                                        counter += 1
                            else:
                                messagebox.showinfo(heding,status)
                                sys.exit(0)    

                            j=1
                            while j<rows2+1:                                                
                                try:
                                    xpath= '/html/body/table[3]/tbody/tr/td/form[2]/table[7]/tbody/tr[2]/td[2]/select/option[{}]'                 
                                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                    if cnm.strip().lstrip().rstrip()=='EXCEL':
                                        xpath= '/html/body/table[3]/tbody/tr/td/form[2]/table[7]/tbody/tr[2]/td[2]/select/option[{}]'
                                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                                                    
                                        break
                                except Exception as e:   
                                    try:
                                        xpath= '/html/body/table[2]/tbody/tr/td/form[2]/table[7]/tbody/tr[2]/td[2]/select/option[{}]'                 
                                        cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                        if cnm.strip().lstrip().rstrip()=='EXCEL':
                                            xpath= '/html/body/table[2]/tbody/tr/td/form[2]/table[7]/tbody/tr[2]/td[2]/select/option[{}]'
                                            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                                                    
                                            break
                                    except Exception as e:                                    
                                        pass
                                j=j+1   
                            
                            xpath= '//input[@id="runButton"]'
                            heding="Run Report Button"
                            status="Run Report Button Not Found"
                            click(xpath,heding,status)
                                                    
                            counter=1
                            while True:
                                try: 
                                    while True:
                                        file1 = [f for f in listdir(down_path) if isfile(join(down_path, f))]
                                        lol_string1 = ' '.join(file1)
                                        lol_string2=lol_string1.split('.')[-1]                                                                
                                        if lol_string2!='' and lol_string2!='tmp' and lol_string2 !="crdownload" and lol_string2 !="temp":
                                            time.sleep(1)
                                            file2 = [f for f in os.listdir(down_path) if os.path.isfile(os.path.join(down_path, f))]
                                            if file2:
                                                lol_string3 = os.path.join(down_path, file2[0])
                                                new_destination_path = os.path.join(file, str(m) + '.xls')                                            
                                                os.rename(lol_string3, new_destination_path)                                              
                                                break
                                        else:
                                            pass
                                    break
                                except Exception as e:
                                        time.sleep(1)
                                        if counter < 10:                                 
                                            counter = counter+1
                                        else:
                                            files = os.listdir(check_path)
                                            for file_name in files:
                                                    file_path = os.path.join(check_path, file_name)
                                                    try:
                                                        if os.path.isfile(file_path):
                                                            os.remove(file_path)                                   
                                                    except OSError as e:
                                                        pass    
                                            wb1=load_workbook(filename=fil)
                                            sheet = wb1['Data']
                                            column_letter = 'N'  
                                            column_cells = sheet[column_letter]
                                            last_row = None
                                            for cell in reversed(column_cells):
                                                if cell.value:
                                                    last_row = cell.row
                                                    break                            
                                            sheet['N' + str(int(last_row + 1))]='File Error'        
                                            wb1.save(fil)
                                            wb1.close()   
                                            break     
                            
                            wb1=load_workbook(filename=fil)
                            sheet = wb1['Data']
                            column_letter = 'N'  
                            column_cells = sheet[column_letter]
                            last_row = None
                            for cell in reversed(column_cells):
                                if cell.value:
                                    last_row = cell.row
                                    break                            
                            sheet['N' + str(int(last_row + 1))]='Done'        
                            wb1.save(fil)
                            wb1.close()   

                            driver.execute_script("window.scrollTo(0, -document.body.scrollHeight);")
                                                        
                            try:
                                xpath= '/html/body/header/table/tbody/tr/td[2]/span[3]/a'
                                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                            except Exception as e:    
                                try:                            
                                    xpath= '/html/body/header/table/tbody/tr/td[2]/span[4]/a'
                                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                                except Exception as e:
                                    try:
                                        xpath= '/html/body/div[1]/header/table/tbody/tr/td[2]/span[3]/a'
                                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                                    except Exception as e:
                                        xpath= '/html/body/div[1]/header/table/tbody/tr/td[2]/span[4]/a'
                                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                                                                                            
                            # xpath='/html/body/header/table/tbody/tr/td[2]/div[2]/ul/li'
                            heding="Sign Out Count"
                            status="Sign Out Count Not Found"
                            # count(xpath,heding,status)  
                            
                            counter = 0
                            while counter < 15:                                
                                try:
                                    xpath='/html/body/header/table/tbody/tr/td[2]/div[2]/ul/li'
                                    rows2=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))    
                                    break
                                except Exception as e:
                                    try:
                                        xpath='/html/body/div[1]/header/table/tbody/tr/td[2]/div[2]/ul/li'
                                        rows2=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))    
                                        break
                                    except Exception as e:
                                        time.sleep(1)
                                        counter += 1
                            else:
                                messagebox.showinfo(heding,status)
                                sys.exit(0)    

                            j=1
                            while j<rows2+1:                                                
                                try:
                                    xpath= '/html/body/header/table/tbody/tr/td[2]/div[2]/ul/li[{}]'                 
                                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                    if cnm.strip().lstrip().rstrip()=='Sign Out':
                                        xpath= '/html/body/header/table/tbody/tr/td[2]/div[2]/ul/li[{}]/a'
                                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                                                    
                                        break
                                except Exception as e:   
                                    try:
                                        xpath='/html/body/div[1]/header/table/tbody/tr/td[2]/div[2]/ul/li[{}]'                 
                                        cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                        if cnm.strip().lstrip().rstrip()=='Sign Out':
                                            xpath= '/html/body/div[1]/header/table/tbody/tr/td[2]/div[2]/ul/li[{}]/a'
                                            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                                                    
                                            break
                                    except Exception as e:
                                        pass
                                j=j+1                         
                        else:
                            wb1=load_workbook(filename=fil)
                            sheet = wb1['Data']
                            column_letter = 'N'  
                            column_cells = sheet[column_letter]
                            last_row = None
                            for cell in reversed(column_cells):
                                if cell.value:
                                    last_row = cell.row
                                    break                            
                            sheet['N' + str(int(last_row + 1))]='Resident List Report *NEW* - Not Found'        
                            wb1.save(fil)
                            wb1.close()   

                            driver.execute_script("window.scrollTo(0, -document.body.scrollHeight);")

                            try:
                                xpath= '/html/body/header/table/tbody/tr/td[2]/span[3]/a'
                                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                            except Exception as e:    
                                try:                            
                                    xpath= '/html/body/header/table/tbody/tr/td[2]/span[4]/a'
                                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                                except Exception as e:
                                    try:
                                        xpath= '/html/body/div[1]/header/table/tbody/tr/td[2]/span[3]/a'
                                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                                    except Exception as e:
                                        xpath= '/html/body/div[1]/header/table/tbody/tr/td[2]/span[4]/a'
                                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                                        
                            heding="Sign Out Count"
                            status="Sign Out Count Not Found"                            

                            counter = 0
                            while counter < 15:                                
                                try:
                                    xpath='/html/body/header/table/tbody/tr/td[2]/div[2]/ul/li'
                                    rows2=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))    
                                    break
                                except Exception as e:
                                    try:
                                        xpath='/html/body/div[1]/header/table/tbody/tr/td[2]/div[2]/ul/li'
                                        rows2=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))    
                                        break
                                    except Exception as e:
                                        time.sleep(1)
                                        counter += 1
                            else:
                                messagebox.showinfo(heding,status)
                                sys.exit(0)    

                            j=1
                            while j<rows2+1:                                                
                                try:
                                    xpath= '/html/body/header/table/tbody/tr/td[2]/div[2]/ul/li[{}]'                 
                                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                    if cnm.strip().lstrip().rstrip()=='Sign Out':
                                        xpath= '/html/body/header/table/tbody/tr/td[2]/div[2]/ul/li[{}]/a'
                                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                                                    
                                        break
                                except Exception as e:   
                                    try:
                                        xpath='/html/body/div[1]/header/table/tbody/tr/td[2]/div[2]/ul/li[{}]'                 
                                        cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                        if cnm.strip().lstrip().rstrip()=='Sign Out':
                                            xpath= '/html/body/div[1]/header/table/tbody/tr/td[2]/div[2]/ul/li[{}]/a'
                                            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                                                    
                                            break
                                    except Exception as e:
                                        pass
                                j=j+1           
                        m=m+1
            else:
                wb1=load_workbook(filename=fil)
                sheet = wb1['Data']
                column_letter = 'N'  
                column_cells = sheet[column_letter]
                last_row = None
                for cell in reversed(column_cells):
                    if cell.value:
                        last_row = cell.row
                        break                            
                sheet['N' + str(int(last_row + 1))]='Link Error'        
                wb1.save(fil)
                wb1.close()     

        driver.close()
        messagebox.showinfo('Process Status', 'Process Completed...')
        sys.exit(0) 

# def Test():
#         global element_1
#         global down_path
#         browse()
        
#         fil_path =ans
                        
#         fil=fil_path 
                                
#         folder_create = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Downloads/')
#         check_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Downloads/Temp')
#         down_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Downloads\\Temp\\')
#         file = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Downloads\\Temp\\Files\\')
            
#         # fld='Temp'
        
#         # if not os.path.isdir(check_path):
#         #     os.mkdir(folder_create + 'Temp')
#         #     os.mkdir(down_path + 'Files')
#         # else:
#         #     path1 = os.path.join(folder_create, fld)
#         #     shutil.rmtree(path1)
#         #     os.mkdir(folder_create + 'Temp')
#         #     os.mkdir(down_path + 'Files')
        
#         # wb = openpyxl.Workbook()
#         # ws=wb.active    
#         # ws['A1']='PCC Name'
#         # ws['B1']='Link'
#         # ws['C1']='User Name'
#         # ws['D1']='Password'
#         # ws['E1']='Final Status'
#         # ws['F1']='Previous address'
#         # ws['G1']='Postal/Zip Code'
#         # ws['H1']='City'
#         # ws['I1']='County'
#         # ws['J1']='Country'
#         # ws['K1']='Prov/State'
                
#         # ws.title = 'Data'
#         # wb.save(filename=file + 'Consol.xlsx')
#         # wb.close()
        

#         # wb1=load_workbook(filename=fil)
#         # sheet = wb1['Data']
#         # sheet.delete_cols(14) 
#         # sheet['N1']='Status'
#         # wb1.save(fil)
#         # wb1.close()   

#         try:
#             options = Options()            
#             options.add_argument('--ignore-certificate-errors')
#             options.add_argument('--ignore-ssl-errors')
#             prefs = {"download.default_directory" : down_path,
#                     "download.prompt_for_download": False,
#                     "download.directory_upgrade": True,
#                     "safebrowsing.enabled": True}
#             options.add_experimental_option("prefs",prefs)
#             driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options)
#             driver.maximize_window()            
#         except Exception as e:
#             try:
#                 options =  webdriver.ChromeOptions()        
#                 options.add_argument('--ignore-certificate-errors')
#                 options.add_argument('--ignore-ssl-errors')
#                 options.add_argument("--disable-popup-blocking")
#                 prefs = {"download.default_directory" : down_path,
#                         "download.prompt_for_download": False,
#                         "download.directory_upgrade": True,
#                         "safebrowsing.enabled": True}
#                 options.add_experimental_option("prefs",prefs)
#                 # response = requests.get('https://chromedriver.storage.googleapis.com/LATEST_RELEASE')
#                 # latest_version = response.text.strip()
#                 # chrome_driver_url = f'https://chromedriver.storage.googleapis.com/{latest_version}/chromedriver_win32.zip'
#                 # response = requests.get(chrome_driver_url)
#                 # with open('chromedriver_win32.zip', 'wb') as f:
#                 #     f.write(response.content)
#                 # with ZipFile('chromedriver_win32.zip', 'r') as zip_ref:
#                 #     zip_ref.extractall('.')   
#                 # driver_path=r'C:\Users\sanandrao\Desktop\New folder\chromedriver.exe'
#                 driver_path = os.path.abspath('chromedriver.exe')                
#                 driver = webdriver.Chrome(executable_path=driver_path,options=options)                
#                 driver.maximize_window()
#                 driver.get('https://www.google.com')                                         
#             except Exception as e:
#                 messagebox.showinfo("Driver Problem","Pls Check Your Driver Version")
#                 sys.exit(0)
             
#         def text_box(xpath,heding,status,key):                
#             counter = 0
#             while counter < 15:
#                 try:                          
#                     WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
#                     break
#                 except Exception as e:
#                     time.sleep(1)
#                     counter += 1
#             else:
#                 messagebox.showinfo(heding, status)
#                 sys.exit(0)        
                
#         def click(xpath,heding,status):
#             counter = 0
#             while counter < 15:
#                 try:             
#                     WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
#                     break
#                 except Exception as e:
#                     time.sleep(1)
#                     counter += 1
#             else:
#                 messagebox.showinfo(heding,status)
#                 sys.exit(0)                                

#         def count(xpath,heding,status):
#             global rows
#             counter = 0
#             while counter < 15:
#                 try:             
#                     rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))                    
#                     break
#                 except Exception as e:
#                     time.sleep(1)
#                     counter += 1
#             else:
#                 messagebox.showinfo(heding,status)
#                 sys.exit(0)          
             
#         def Alert():
#             counter = 0
#             while counter < 5:
#                 try:             
#                     WebDriverWait(driver, 0).until (EC.alert_is_present())
#                     a=driver.switch_to.alert
#                     a.accept()
#                     break
#                 except Exception as e:
#                     time.sleep(1)
#                     counter += 1
#             else:
#                 messagebox.showinfo('Alert','Alert Not Present')

#         df=pd.read_excel(fil,sheet_name='Data',header=0)                         
#         m=1
#         for index, row in df.iterrows():  
#             # nme = row[1]                                          
#             lnk = row[2]              
#             umn = row[3]     
#             pas = row[4]
#             # dt_in=row[11]            
#             # dt_out=row[12]            

#             driver.get(lnk.lstrip().rstrip())
#             tit = driver.title

#             if tit=='PointClickCare Login':                
#                 xpath= '//input[@id="username"]'
#                 heding="User ID"
#                 status="User ID Field Not Found"
#                 key= umn
#                 text_box(xpath,heding,status,key)
            
#                 xpath= '//button[@id="id-next"]'
#                 heding="Next Button"
#                 status="Next Button Not Found"
#                 click(xpath,heding,status)

#                 xpath= '//input[@id="password"]'
#                 heding="Password"
#                 status="Password Field Not Found"
#                 key= pas
#                 text_box(xpath,heding,status,key)

#                 xpath= '//button[@id="id-submit"]'
#                 heding="SIGN IN Button"
#                 status="SIGN IN Button Not Found"
#                 click(xpath,heding,status)

#                 wait = WebDriverWait(driver, 20)
#                 wait.until(lambda driver: driver.execute_script("return document.readyState === 'complete';"))

#                 time.sleep(2)
#                 try:                    
#                     element_1 = WebDriverWait(driver, 2).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="app"]/div/main/div/div[1]/nav[1]/div[2]/div/div/div[1]/div[3]/div[1]/div/div'))).text
#                     if element_1.lstrip().rstrip() =="There was an error logging you in. Contact your System Administrator to check your account. PointClickCare cannot reset passwords.":
#                         wb1=load_workbook(filename=fil)
#                         sheet = wb1['Data']
#                         column_letter = 'N'  
#                         column_cells = sheet[column_letter]
#                         last_row = None
#                         for cell in reversed(column_cells):
#                             if cell.value:
#                                 last_row = cell.row
#                                 break                            
#                         sheet['N' + str(int(last_row + 1))]='Login Error'        
#                         wb1.save(fil)
#                         wb1.close()         

#                 except Exception as e:           
                
#                     try:
#                         element_1 = WebDriverWait(driver, 2).until(EC.visibility_of_element_located((By.XPATH, '//div[@id="loginerror"]'))).text
#                         if element_1 !="":
#                             wb1=load_workbook(filename=fil)
#                             sheet = wb1['Data']
#                             column_letter = 'N'  
#                             column_cells = sheet[column_letter]
#                             last_row = None
#                             for cell in reversed(column_cells):
#                                 if cell.value:
#                                     last_row = cell.row
#                                     break                            
#                             sheet['N' + str(int(last_row + 1))]='Error'        
#                             wb1.save(fil)
#                             wb1.close()                                             
#                     except Exception as e:                                   
#                         xpath= '//li[@id="QTF_reportingTab"]'
#                         heding="Report Button"
#                         status="Report Button Not Found"
#                         click(xpath,heding,status)                                        
                        
#                         xpath= '//a[@id="pccFacLink"]'
#                         heding="Facilities Button"
#                         status="Facilities Button Not Found"
#                         click(xpath,heding,status)

#                         xpath='//ul[@id="optionList"]/li'
#                         heding="Facilities Count"
#                         status="Facilities Count Not Found"
#                         count(xpath,heding,status) 

#                         try:                    
#                             j=1
#                             while j<rows+1:
#                                 xpath= '//ul[@id="optionList"]/li[{}]'                      
#                                 cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
#                                 wb1=load_workbook(filename='D:\\Project 7\\Output.xlsx')
#                                 sheet = wb1['Data']
#                                 column_letter = 'A'  
#                                 column_cells = sheet[column_letter]
#                                 last_row = None
#                                 for cell in reversed(column_cells):
#                                     if cell.value:
#                                         last_row = cell.row
#                                         break                            
#                                 sheet['A' + str(int(last_row + 1))]=cnm   
#                                 sheet['B' + str(int(last_row + 1))]=lnk   
#                                 sheet['C' + str(int(last_row + 1))]=umn   
#                                 sheet['D' + str(int(last_row + 1))]=pas  
#                                 sheet['E' + str(int(last_row + 1))]='Done'   

#                                 wb1.save('D:\\Project 7\\Output.xlsx')
#                                 wb1.close()  
                                
#                                 j=j+1

#                             wb1=load_workbook(filename=fil)
#                             sheet = wb1['Data']
#                             column_letter = 'N'  
#                             column_cells = sheet[column_letter]
#                             last_row = None
#                             for cell in reversed(column_cells):
#                                 if cell.value:
#                                     last_row = cell.row
#                                     break                            
#                             sheet['N' + str(int(last_row + 1))]='Done'        
#                             wb1.save(fil)
#                             wb1.close()    
#                         except Exception as e:
#                             wb1=load_workbook(filename=fil)
#                             sheet = wb1['Data']
#                             column_letter = 'N'  
#                             column_cells = sheet[column_letter]
#                             last_row = None
#                             for cell in reversed(column_cells):
#                                 if cell.value:
#                                     last_row = cell.row
#                                     break                            
#                             sheet['N' + str(int(last_row + 1))]='Error'        
#                             wb1.save(fil)
#                             wb1.close()    

#                         driver.execute_script("window.scrollTo(0, -document.body.scrollHeight);")
                            
#                         try:
#                             xpath= '/html/body/header/table/tbody/tr/td[2]/span[3]/a'
#                             WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
#                         except Exception as e:
#                             try:
#                                 xpath= '/html/body/header/table/tbody/tr/td[2]/span[4]/a'
#                                 WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
#                             except Exception as e:
#                                 xpath= '/html/body/header/table/tbody/tr/td[2]/span[2]/a/span'
#                                 WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                            
                        
#                         xpath='/html/body/header/table/tbody/tr/td[2]/div[2]/ul/li'
#                         heding="Sign Out Count"
#                         status="Sign Out Count Not Found"
#                         count(xpath,heding,status)  

#                         j=1
#                         while j<rows+1:                                                
#                             try:
#                                 xpath= '/html/body/header/table/tbody/tr/td[2]/div[2]/ul/li[{}]'                 
#                                 cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
#                                 if cnm.strip().lstrip().rstrip()=='Sign Out':
#                                     xpath= '/html/body/header/table/tbody/tr/td[2]/div[2]/ul/li[{}]/a'
#                                     WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                                                    
#                                     break
#                             except Exception as e:   
#                                 pass
#                             j=j+1    

#                         else:
#                             wb1=load_workbook(filename=fil)
#                             sheet = wb1['Data']
#                             column_letter = 'N'  
#                             column_cells = sheet[column_letter]
#                             last_row = None
#                             for cell in reversed(column_cells):
#                                 if cell.value:
#                                     last_row = cell.row
#                                     break                            
#                             sheet['N' + str(int(last_row + 1))]='Resident List Report *NEW* - Not Found'        
#                             wb1.save(fil)
#                             wb1.close()   

#                             driver.execute_script("window.scrollTo(0, -document.body.scrollHeight);")

#                             try:
#                                 xpath= '/html/body/header/table/tbody/tr/td[2]/span[3]/a'
#                                 WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
#                             except Exception as e:
#                                 xpath= '/html/body/header/table/tbody/tr/td[2]/span[4]/a'
#                                 WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()                        

#                             xpath='/html/body/header/table/tbody/tr/td[2]/div[2]/ul/li'
#                             heding="Sign Out Count"
#                             status="Sign Out Count Not Found"
#                             count(xpath,heding,status)  

#                             j=1
#                             while j<rows+1:                                                
#                                 try:
#                                     xpath= '/html/body/header/table/tbody/tr/td[2]/div[2]/ul/li[{}]'                 
#                                     cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
#                                     if cnm.strip().lstrip().rstrip()=='Sign Out':
#                                         xpath= '/html/body/header/table/tbody/tr/td[2]/div[2]/ul/li[{}]/a'
#                                         WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                                                                    
#                                         break
#                                 except Exception as e:   
#                                     pass
#                                 j=j+1         
#                         m=m+1
#             else:
#                 wb1=load_workbook(filename=fil)
#                 sheet = wb1['Data']
#                 column_letter = 'N'  
#                 column_cells = sheet[column_letter]
#                 last_row = None
#                 for cell in reversed(column_cells):
#                     if cell.value:
#                         last_row = cell.row
#                         break                            
#                 sheet['N' + str(int(last_row + 1))]='Link Error'        
#                 wb1.save(fil)
#                 wb1.close()     

#         driver.close()
#         messagebox.showinfo('Process Status', 'Process Completed...')
#         sys.exit(0) 


def upload():
    global element_1
    browse()
    
    fil_path =ans    
                                        
    fil=fil_path 
    
    try:
            options = Options()                        
            driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options)
            driver.maximize_window()            
    except Exception as e:
            try:
                options =  webdriver.ChromeOptions()        
                options.add_argument('--ignore-certificate-errors')
                options.add_argument('--ignore-ssl-errors')
                options.add_argument("--disable-popup-blocking")
                # response = requests.get('https://chromedriver.storage.googleapis.com/LATEST_RELEASE')
                # latest_version = response.text.strip()
                # chrome_driver_url = f'https://chromedriver.storage.googleapis.com/{latest_version}/chromedriver_win32.zip'
                # response = requests.get(chrome_driver_url)
                # with open('chromedriver_win32.zip', 'wb') as f:
                #     f.write(response.content)
                # with ZipFile('chromedriver_win32.zip', 'r') as zip_ref:
                #     zip_ref.extractall('.')   
                # driver_path=r'C:\Users\sanandrao\Desktop\New folder\chromedriver.exe'
                driver_path = os.path.abspath('chromedriver.exe')                
                driver = webdriver.Chrome(executable_path=driver_path)                
                driver.maximize_window()                                                  
            except Exception as e:
                messagebox.showinfo("Internet Problem","Pls Check Your Internet Connection")
                sys.exit(0)

    def text_box(xpath,heding,status,key):                
            counter = 0
            while counter < 5:
                try:                          
                    WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding, status)
                sys.exit(0)        
                
    def click(xpath,heding,status):
        counter = 0
        while counter < 5:
            try:             
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo(heding,status)
            sys.exit(0)                                

    def count(xpath,heding,status):
        global rows
        counter = 0
        while counter < 5:
            try:             
                rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))                    
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo(heding,status)
            sys.exit(0)      

    df=pd.read_excel(fil,sheet_name='Data',header=0)  

    df1=df[['PCC Name','Link','User Name','Password']].drop_duplicates().reset_index(drop=True)                      
    df['Resident Id'] = df['Resident Id'].astype(str)
        
    lst=[]
    for index, row in df.iterrows():
        nme = row[0]
        lnk = row[1]
        umn = row[2]
        pas = row[3]
        res_id=row[13]
        
        tmp=nme+lnk+umn+pas
        
        lst.append(tmp)

        if len(lst)==1:

            driver.get(lnk)
            
            tit = driver.title

            if tit=='PointClickCare Login':                
                xpath= '//input[@id="username"]'
                heding="User ID"
                status="User ID Field Not Found"
                key= umn
                # text_box(xpath,heding,status,key)

                counter = 0
                while counter < 5:
                    try:                          
                        WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'//input[@id="un"]'))).send_keys(key)
                        break
                    except Exception as e:
                        try:
                            WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'//input[@id="username"]'))).send_keys(key)
                            break
                        except Exception as e:
                            time.sleep(1)
                            counter += 1
                else:
                    messagebox.showinfo(heding, status)
                    sys.exit(0)   

                xpath= '//button[@id="id-next"]'
                heding="Next Button"
                status="Next Button Not Found"
                click(xpath,heding,status)

                xpath= '//input[@id="password"]'
                heding="Password"
                status="Password Field Not Found"
                key= pas
                text_box(xpath,heding,status,key)

                xpath= '//button[@id="id-submit"]'
                heding="SIGN IN Button"
                status="SIGN IN Button Not Found"
                click(xpath,heding,status)

                try:
                    element_1 = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, '//div[@id="loginerror"]'))).text
                    if element_1 !="":
                        wb1=load_workbook(filename=fil)
                        sheet = wb1['Data']
                        column_letter = 'E'  
                        column_cells = sheet[column_letter]
                        last_row = None
                        for cell in reversed(column_cells):
                            if cell.value:
                                last_row = cell.row
                                break                            
                        sheet['E' + str(int(last_row + 1))]='SIGN Error'        
                        wb1.save(fil)
                        wb1.close()   
                except Exception as e:                                             
                    xpath= '//a[@id="pccFacLink"]'
                    heding="Facilities Button"
                    status="Facilities Button Not Found"
                    click(xpath,heding,status)

                    xpath='//ul[@id="optionList"]/li'
                    heding="Facilities Count"
                    status="Facilities Count Not Found"
                    count(xpath,heding,status) 
                                        
                    j=1
                    while j<rows+1:
                        xpath= '//ul[@id="optionList"]/li[{}]'                      
                        cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                        if cnm==nme: 
                            xpath= '//ul[@id="optionList"]/li[{}]/a'                                
                            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()
                            break
                        j=j+1
                    
                    time.sleep(2)

                    xpath= '//table[@id="mainPageNavigation"]/tbody/tr/td/div[2]/form/table/tbody/tr/td[2]/input'
                    heding="Search"
                    status="Search Field Not Found"
                    key= res_id
                    text_box(xpath,heding,status,key)                                

                    xpath= '//table[@id="mainPageNavigation"]/tbody/tr/td/div[2]/form/table/tbody/tr/td[3]/input'
                    heding="Search Button"
                    status="Search Button Not Found"                
                    click(xpath,heding,status)
                    
                    try:                                                                                                                                                   
                        element_1 = WebDriverWait(driver, 2).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div/table/tbody/tr[2]/td'))).text                        
                        if element_1 =="No records found.":
                            wb1=load_workbook(filename=fil)
                            sheet = wb1['Data']
                            column_letter = 'E'  
                            column_cells = sheet[column_letter]
                            last_row = None
                            for cell in reversed(column_cells):
                                if cell.value:
                                    last_row = cell.row
                                    break                            
                            sheet['E' + str(int(last_row + 1))]='No Records'        
                            wb1.save(fil)
                            wb1.close() 

                            xpath= '//a[@id="pccFacLink"]'
                            heding="Facilities Button"
                            status="Facilities Button Not Found"
                            click(xpath,heding,status)

                            xpath='//ul[@id="optionList"]/li'
                            heding="Facilities Count"
                            status="Facilities Count Not Found"
                            count(xpath,heding,status) 

                            j=1
                            while j<rows+1:
                                xpath= '//ul[@id="optionList"]/li[{}]'                      
                                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                if cnm==nme: 
                                    xpath= '//ul[@id="optionList"]/li[{}]/a'                                
                                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()
                                    break
                                j=j+1
                        
                        else:                           
                            wb1=load_workbook(filename=fil)
                            sheet = wb1['Data']
                            column_letter = 'E'  
                            column_cells = sheet[column_letter]
                            last_row = None
                            for cell in reversed(column_cells):
                                if cell.value:
                                    last_row = cell.row
                                    break                                                        
                            sheet['E' + str(int(last_row + 1))]='Link Found' 
                            wb1.save(fil)
                            wb1.close()  

                            xpath= '//a[@id="pccFacLink"]'
                            heding="Facilities Button"
                            status="Facilities Button Not Found"
                            click(xpath,heding,status)

                            xpath='//ul[@id="optionList"]/li'
                            heding="Facilities Count"
                            status="Facilities Count Not Found"
                            count(xpath,heding,status) 

                            j=1
                            while j<rows+1:
                                xpath= '//ul[@id="optionList"]/li[{}]'                      
                                cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                if cnm==nme: 
                                    xpath= '//ul[@id="optionList"]/li[{}]/a'                                
                                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()
                                    break
                                j=j+1
                            
                    except Exception as e:  
                        xpath= '/html/body/table[3]/tbody/tr/td[2]/table[2]/tbody/tr/td[1]/table/tbody/tr/td[2]/div/div[1]/a/span'
                        heding="Edit Button"
                        status="Edit Button Not Found"
                        click(xpath,heding,status) 

                        xpath= '/html/body/table[3]/tbody/tr/td[2]/table[2]/tbody/tr/td[1]/table/tbody/tr/td[2]/div/div[2]/ul'
                        heding="Edit Button"
                        status="Edit Button Not Found"
                        count(xpath,heding,status) 
                        
                        j=1
                        while j<rows+1:
                            xpath= '/html/body/table[3]/tbody/tr/td[2]/table[2]/tbody/tr/td[1]/table/tbody/tr/td[2]/div/div[2]/ul/li[{}]'                      
                            cnm=WebDriverWait(driver, 2).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                            if cnm=='Demographics': 
                                xpath= '/html/body/table[3]/tbody/tr/td[2]/table[2]/tbody/tr/td[1]/table/tbody/tr/td[2]/div/div[2]/ul/li[{}]/a'                                
                                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()
                                break
                            j=j+1
                        
                        prv_add=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'//input[@name="address1"]')))
                        prv_add = prv_add.get_attribute('value')

                        prv_add1=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'//input[@name="address2"]')))
                        prv_add1 = prv_add1.get_attribute('value')
                        prv_fin=prv_add+' ' + prv_add1

                        pos_cod=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'//input[@name="postal_zip_code"]')))
                        pos_cod = pos_cod.get_attribute('value')
                        
                        cty=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'//input[@name="city"]')))
                        cty = cty.get_attribute('value')

                        cnt=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'//input[@name="county"]')))
                        cnt = cnt.get_attribute('value')
                        
                        cnty=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'//input[@name="country"]')))
                        cnty = cnty.get_attribute('value')

                        st=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'//input[@name="provstate"]')))
                        st = st.get_attribute('value')

                        wb1=load_workbook(filename=fil)
                        sheet = wb1['Data']
                        column_letter = 'E'  
                        column_cells = sheet[column_letter]
                        last_row = None
                        for cell in reversed(column_cells):
                            if cell.value:
                                last_row = cell.row
                                break                            
                        sheet['F' + str(int(last_row + 1))]=prv_fin  
                        sheet['G' + str(int(last_row + 1))]=pos_cod  
                        sheet['H' + str(int(last_row + 1))]=cty
                        sheet['I' + str(int(last_row + 1))]=cnt
                        sheet['J' + str(int(last_row + 1))]=cnty
                        sheet['K' + str(int(last_row + 1))]=st
                        sheet['E' + str(int(last_row + 1))]='Done' 
                        wb1.save(fil)
                        wb1.close()                                                     
                
                        xpath= '/html/body/table[4]/tbody/tr[1]/td/form/table/tbody/tr[3]/td/div[1]/input[2]'
                        heding="Cancel Button"
                        status="Cancel Button Not Found"
                        click(xpath,heding,status)                                                                                                                    
            else:
                    wb1=load_workbook(filename=fil)
                    sheet = wb1['Data']
                    column_letter = 'E'  
                    column_cells = sheet[column_letter]
                    last_row = None
                    for cell in reversed(column_cells):
                        if cell.value:
                            last_row = cell.row
                            break                            
                    sheet['E' + str(int(last_row + 1))]='Link Error'        
                    wb1.save(fil)
                    wb1.close()     
        else:
            lst1 = lst[0]
            lst2 = lst[1]
            if lst1==lst2:
                xpath= '//table[@id="mainPageNavigation"]/tbody/tr/td/div[2]/form/table/tbody/tr/td[2]/input'
                heding="Search"
                status="Search Field Not Found"
                key= res_id
                text_box(xpath,heding,status,key)                                

                xpath= '//table[@id="mainPageNavigation"]/tbody/tr/td/div[2]/form/table/tbody/tr/td[3]/input'
                heding="Search Button"
                status="Search Button Not Found"                
                click(xpath,heding,status)
                
                try:
                    element_1 = WebDriverWait(driver, 2).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div/table/tbody/tr[2]/td'))).text
                    if element_1 =="No records found.":
                        wb1=load_workbook(filename=fil)
                        sheet = wb1['Data']
                        column_letter = 'E'  
                        column_cells = sheet[column_letter]
                        last_row = None
                        for cell in reversed(column_cells):
                            if cell.value:
                                last_row = cell.row
                                break                            
                        sheet['E' + str(int(last_row + 1))]='No Records'        
                        wb1.save(fil)
                        wb1.close() 
                        
                        xpath= '//a[@id="pccFacLink"]'
                        heding="Facilities Button"
                        status="Facilities Button Not Found"
                        click(xpath,heding,status)
                        
                        xpath='//ul[@id="optionList"]/li'
                        heding="Facilities Count"
                        status="Facilities Count Not Found"
                        count(xpath,heding,status) 

                        j=1
                        while j<rows+1:
                            xpath= '//ul[@id="optionList"]/li[{}]'                      
                            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                            if cnm==nme: 
                                xpath= '//ul[@id="optionList"]/li[{}]/a'                                
                                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()
                                break
                            j=j+1

                    else:                           
                        wb1=load_workbook(filename=fil)
                        sheet = wb1['Data']
                        column_letter = 'E'  
                        column_cells = sheet[column_letter]
                        last_row = None
                        for cell in reversed(column_cells):
                            if cell.value:
                                last_row = cell.row
                                break                                                        
                        sheet['E' + str(int(last_row + 1))]='Link Found' 
                        wb1.save(fil)
                        wb1.close()  

                        xpath= '//a[@id="pccFacLink"]'
                        heding="Facilities Button"
                        status="Facilities Button Not Found"
                        click(xpath,heding,status)

                        xpath='//ul[@id="optionList"]/li'
                        heding="Facilities Count"
                        status="Facilities Count Not Found"
                        count(xpath,heding,status) 

                        j=1
                        while j<rows+1:
                            xpath= '//ul[@id="optionList"]/li[{}]'                      
                            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                            if cnm==nme: 
                                xpath= '//ul[@id="optionList"]/li[{}]/a'                                
                                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()
                                break
                            j=j+1                  
                except Exception as e:  
                    xpath= '/html/body/table[3]/tbody/tr/td[2]/table[2]/tbody/tr/td[1]/table/tbody/tr/td[2]/div/div[1]/a/span'
                    heding="Edit Button"
                    status="Edit Button Not Found"
                    click(xpath,heding,status) 

                    xpath= '/html/body/table[3]/tbody/tr/td[2]/table[2]/tbody/tr/td[1]/table/tbody/tr/td[2]/div/div[2]/ul'
                    heding="Edit Button"
                    status="Edit Button Not Found"
                    count(xpath,heding,status) 
                    
                    j=1
                    while j<rows+1:
                        xpath= '/html/body/table[3]/tbody/tr/td[2]/table[2]/tbody/tr/td[1]/table/tbody/tr/td[2]/div/div[2]/ul/li[{}]'                      
                        cnm=WebDriverWait(driver, 2).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                        if cnm=='Demographics': 
                            xpath= '/html/body/table[3]/tbody/tr/td[2]/table[2]/tbody/tr/td[1]/table/tbody/tr/td[2]/div/div[2]/ul/li[{}]/a'                                
                            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()
                            break
                        j=j+1
                    
                    prv_add=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'//input[@name="address1"]')))
                    prv_add = prv_add.get_attribute('value')

                    prv_add1=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'//input[@name="address2"]')))
                    prv_add1 = prv_add1.get_attribute('value')
                    prv_fin=prv_add+' ' + prv_add1

                    pos_cod=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'//input[@name="postal_zip_code"]')))
                    pos_cod = pos_cod.get_attribute('value')
                    
                    cty=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'//input[@name="city"]')))
                    cty = cty.get_attribute('value')

                    cnt=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'//input[@name="county"]')))
                    cnt = cnt.get_attribute('value')
                    
                    cnty=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'//input[@name="country"]')))
                    cnty = cnty.get_attribute('value')

                    st=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'//input[@name="provstate"]')))
                    st = st.get_attribute('value')

                    wb1=load_workbook(filename=fil)
                    sheet = wb1['Data']
                    column_letter = 'E'  
                    column_cells = sheet[column_letter]
                    last_row = None
                    for cell in reversed(column_cells):
                        if cell.value:
                            last_row = cell.row
                            break                            
                    sheet['F' + str(int(last_row + 1))]=prv_fin  
                    sheet['G' + str(int(last_row + 1))]=pos_cod  
                    sheet['H' + str(int(last_row + 1))]=cty
                    sheet['I' + str(int(last_row + 1))]=cnt
                    sheet['J' + str(int(last_row + 1))]=cnty
                    sheet['K' + str(int(last_row + 1))]=st
                    sheet['E' + str(int(last_row + 1))]='Done' 
                    wb1.save(fil)
                    wb1.close()                                                     
            
                    xpath= '/html/body/table[4]/tbody/tr[1]/td/form/table/tbody/tr[3]/td/div[1]/input[2]'
                    heding="Cancel Button"
                    status="Cancel Button Not Found"
                    click(xpath,heding,status)
                del lst[0]   
            else:
                driver.get(lnk)
            
                tit = driver.title

                if tit=='PointClickCare Login':                
                    xpath= '//input[@id="username"]'
                    heding="User ID"
                    status="User ID Field Not Found"
                    key= umn
                    text_box(xpath,heding,status,key)
                
                    xpath= '//button[@id="id-next"]'
                    heding="Next Button"
                    status="Next Button Not Found"
                    click(xpath,heding,status)

                    xpath= '//input[@id="password"]'
                    heding="Password"
                    status="Password Field Not Found"
                    key= pas
                    text_box(xpath,heding,status,key)

                    xpath= '//button[@id="id-submit"]'
                    heding="SIGN IN Button"
                    status="SIGN IN Button Not Found"
                    click(xpath,heding,status)

                    try:
                        element_1 = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, '//div[@id="loginerror"]'))).text
                        if element_1 !="":
                            wb1=load_workbook(filename=fil)
                            sheet = wb1['Data']
                            column_letter = 'E'  
                            column_cells = sheet[column_letter]
                            last_row = None
                            for cell in reversed(column_cells):
                                if cell.value:
                                    last_row = cell.row
                                    break                            
                            sheet['E' + str(int(last_row + 1))]='Error'        
                            wb1.save(fil)
                            wb1.close()   
                    except Exception as e:                                             
                        xpath= '//a[@id="pccFacLink"]'
                        heding="Facilities Button"
                        status="Facilities Button Not Found"
                        click(xpath,heding,status)

                        xpath='//ul[@id="optionList"]/li'
                        heding="Facilities Count"
                        status="Facilities Count Not Found"
                        count(xpath,heding,status) 
                                            
                        j=1
                        while j<rows+1:
                            xpath= '//ul[@id="optionList"]/li[{}]'                      
                            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                            if cnm==nme: 
                                xpath= '//ul[@id="optionList"]/li[{}]/a'                                
                                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()
                                break
                            j=j+1
                        
                        time.sleep(2)

                        xpath= '//table[@id="mainPageNavigation"]/tbody/tr/td/div[2]/form/table/tbody/tr/td[2]/input'
                        heding="Search"
                        status="Search Field Not Found"
                        key= res_id
                        text_box(xpath,heding,status,key)                                

                        xpath= '//table[@id="mainPageNavigation"]/tbody/tr/td/div[2]/form/table/tbody/tr/td[3]/input'
                        heding="Search Button"
                        status="Search Button Not Found"                
                        click(xpath,heding,status)
                        
                        try:
                            element_1 = WebDriverWait(driver, 2).until(EC.visibility_of_element_located((By.XPATH, '/html/body/form/div/table/tbody/tr[2]/td'))).text
                            if element_1 =="No records found.":
                                wb1=load_workbook(filename=fil)
                                sheet = wb1['Data']
                                column_letter = 'E'  
                                column_cells = sheet[column_letter]
                                last_row = None
                                for cell in reversed(column_cells):
                                    if cell.value:
                                        last_row = cell.row
                                        break                            
                                sheet['E' + str(int(last_row + 1))]='No Records'        
                                wb1.save(fil)
                                wb1.close() 

                                xpath= '//a[@id="pccFacLink"]'
                                heding="Facilities Button"
                                status="Facilities Button Not Found"
                                click(xpath,heding,status)

                                xpath='//ul[@id="optionList"]/li'
                                heding="Facilities Count"
                                status="Facilities Count Not Found"
                                count(xpath,heding,status) 

                                j=1
                                while j<rows+1:
                                    xpath= '//ul[@id="optionList"]/li[{}]'                      
                                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                    if cnm==nme: 
                                        xpath= '//ul[@id="optionList"]/li[{}]/a'                                
                                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()
                                        break
                                    j=j+1

                            else:                           
                                wb1=load_workbook(filename=fil)
                                sheet = wb1['Data']
                                column_letter = 'E'  
                                column_cells = sheet[column_letter]
                                last_row = None
                                for cell in reversed(column_cells):
                                    if cell.value:
                                        last_row = cell.row
                                        break                                                        
                                sheet['E' + str(int(last_row + 1))]='Link Found' 
                                wb1.save(fil)
                                wb1.close()    

                                xpath= '//a[@id="pccFacLink"]'
                                heding="Facilities Button"
                                status="Facilities Button Not Found"
                                click(xpath,heding,status)

                                xpath='//ul[@id="optionList"]/li'
                                heding="Facilities Count"
                                status="Facilities Count Not Found"
                                count(xpath,heding,status) 

                                j=1
                                while j<rows+1:
                                    xpath= '//ul[@id="optionList"]/li[{}]'                      
                                    cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                    if cnm==nme: 
                                        xpath= '//ul[@id="optionList"]/li[{}]/a'                                
                                        WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()
                                        break
                                    j=j+1

                        except Exception as e:  
                            xpath= '/html/body/table[3]/tbody/tr/td[2]/table[2]/tbody/tr/td[1]/table/tbody/tr/td[2]/div/div[1]/a/span'
                            heding="Edit Button"
                            status="Edit Button Not Found"
                            click(xpath,heding,status) 

                            xpath= '/html/body/table[3]/tbody/tr/td[2]/table[2]/tbody/tr/td[1]/table/tbody/tr/td[2]/div/div[2]/ul'
                            heding="Edit Button"
                            status="Edit Button Not Found"
                            count(xpath,heding,status) 
                            
                            j=1
                            while j<rows+1:
                                xpath= '/html/body/table[3]/tbody/tr/td[2]/table[2]/tbody/tr/td[1]/table/tbody/tr/td[2]/div/div[2]/ul/li[{}]'                      
                                cnm=WebDriverWait(driver, 2).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
                                if cnm=='Demographics': 
                                    xpath= '/html/body/table[3]/tbody/tr/td[2]/table[2]/tbody/tr/td[1]/table/tbody/tr/td[2]/div/div[2]/ul/li[{}]/a'                                
                                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()
                                    break
                                j=j+1
                            
                            prv_add=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'//input[@name="address1"]')))
                            prv_add = prv_add.get_attribute('value')

                            prv_add1=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'//input[@name="address2"]')))
                            prv_add1 = prv_add1.get_attribute('value')
                            prv_fin=prv_add+' ' + prv_add1

                            pos_cod=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'//input[@name="postal_zip_code"]')))
                            pos_cod = pos_cod.get_attribute('value')
                            
                            cty=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'//input[@name="city"]')))
                            cty = cty.get_attribute('value')

                            cnt=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'//input[@name="county"]')))
                            cnt = cnt.get_attribute('value')
                            
                            cnty=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'//input[@name="country"]')))
                            cnty = cnty.get_attribute('value')

                            st=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'//input[@name="provstate"]')))
                            st = st.get_attribute('value')

                            wb1=load_workbook(filename=fil)
                            sheet = wb1['Data']
                            column_letter = 'E'  
                            column_cells = sheet[column_letter]
                            last_row = None
                            for cell in reversed(column_cells):
                                if cell.value:
                                    last_row = cell.row
                                    break                            
                            sheet['F' + str(int(last_row + 1))]=prv_fin  
                            sheet['G' + str(int(last_row + 1))]=pos_cod  
                            sheet['H' + str(int(last_row + 1))]=cty
                            sheet['I' + str(int(last_row + 1))]=cnt
                            sheet['J' + str(int(last_row + 1))]=cnty
                            sheet['K' + str(int(last_row + 1))]=st
                            sheet['E' + str(int(last_row + 1))]='Done' 
                            wb1.save(fil)
                            wb1.close()                                                     
                    
                            xpath= '/html/body/table[4]/tbody/tr[1]/td/form/table/tbody/tr[3]/td/div[1]/input[2]'
                            heding="Cancel Button"
                            status="Cancel Button Not Found"
                            click(xpath,heding,status)                                                                                                                    
                else:
                        wb1=load_workbook(filename=fil)
                        sheet = wb1['Data']
                        column_letter = 'E'  
                        column_cells = sheet[column_letter]
                        last_row = None
                        for cell in reversed(column_cells):
                            if cell.value:
                                last_row = cell.row
                                break                            
                        sheet['E' + str(int(last_row + 1))]='Link Error'        
                        wb1.save(fil)
                        wb1.close()     
                del lst[0]

    driver.close()
    messagebox.showinfo('Process Status', 'Process Completed...')
    sys.exit(0) 

def console():
    folder()
    
    fil=ans + '/' 
    file_path = fil + 'Console'
    file_con = fil + 'Console' + '/'
    files = [item for item in os.listdir(fil) if os.path.isfile(os.path.join(fil, item))]    

    fld='Console'
        
    if not os.path.isdir(file_path):
        os.mkdir(fil + 'Console')        
    else:
        path1 = os.path.join(fil, fld)
        shutil.rmtree(path1)
        os.mkdir(fil + 'Console')
        
    wb = openpyxl.Workbook()
    ws=wb.active    
    ws['A1']='PCC Name'
    # ws['B1']='Link'
    # ws['C1']='User Name'
    # ws['D1']='Password'
    # ws['E1']='Final Status'
    # ws['F1']='Previous address'
    # ws['G1']='Postal/Zip Code'
    # ws['H1']='City'
    # ws['I1']='County'
    # ws['J1']='Country'
    # ws['K1']='Prov/State'
    ws['B1']='Resident Last Name'
    ws['C1']='Resident First Name'
    ws['D1']='Resident Id'
    ws['E1']='Birth Date'
    ws['F1']='Home Phone'
    ws['G1']='Gender'

    ws.title = 'Data'
    wb.save(filename=file_con +  'Consol.xlsx')
    wb.close()
    
    j=0
           
    while j < len(files):
        fl=files[j]

        ex_fn = pd.DataFrame()
        df = pd.DataFrame()
        
        ex_fn = pd.read_excel(fil + fl,engine='xlrd',header=0)
        lst=list(ex_fn.head())
        filtered_list = [item for item in lst if not item.startswith('Unnamed')]
        fn=filtered_list[1]
        
        final_lst = []
        ex_fil = pd.read_excel(fil + fl,engine='xlrd',header=5)        
        ex_fil['PCC Name']=fn
        
        # ex_fil=ex_fil[['PCC Name','Resident Last Name','Resident First Name','Resident Id','Birth Date','Home Phone','Gender']]
        
        lst1 =list(ex_fil.head())
        
        # lst1 =list(['PCC Name','Resident Last Name','Resident First Name','Resident Id','Gender'])

        con_fil = pd.read_excel(file_con + "Consol.xlsx",sheet_name='Data', engine='openpyxl')
        lst2 =list(con_fil.head())

        for item in lst2:
            if item not in final_lst:
                final_lst.append(item)
        
        for item in lst1:
            if item not in final_lst:
                final_lst.append(item)   
        
        book = load_workbook(filename=file_con + "Consol.xlsx")
        sheet = book['Data']

        for i, value in enumerate(final_lst, start=1):
            sheet.cell(row=1, column=i, value=value)
        
        book.save(filename=file_con + "Consol.xlsx")

        con = pd.read_excel(file_con + "Consol.xlsx",sheet_name='Data', engine='openpyxl')
        con.drop(index=con.index, inplace=True) 
        heading_list = con.columns.tolist()

        df = pd.concat([con, ex_fil], axis=0)
        df = df[heading_list]

        book = load_workbook(filename=file_con + "Consol.xlsx")

        with pd.ExcelWriter(file_con + "Consol.xlsx", engine='openpyxl') as writer:
            writer.book = book
            for ws in book.worksheets:
                writer.sheets[ws.title] = writer.book[ws.title]
            worksheet = writer.sheets['Data']
            startrow = worksheet.max_row if worksheet.max_row > 0 else None
            df.to_excel(writer, sheet_name='Data', index=False, startrow=startrow,header=False)
        
        print(fl + "   -   Done")
        j=j+1
    
    messagebox.showinfo('Process Status', 'Process Completed...')
    sys.exit(0) 

if __name__=="__main__":        
    radio()
    # Test()
    