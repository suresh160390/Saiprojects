from tkinter import BOTH, LEFT, TOP, Button, Entry, Frame, Label, PhotoImage, StringVar, Tk,Radiobutton,StringVar,IntVar,filedialog
from idlelib.tooltip import Hovertip
import sys
import os
from operator import itemgetter
import time
from tkinter import messagebox
from os import listdir
from os.path import isfile, join
from datetime import datetime, date
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, StaleElementReferenceException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver import ActionChains
import openpyxl
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from urllib.parse import urlparse
import warnings
import numpy as np
from openpyxl.utils import get_column_letter,column_index_from_string
from datetime import datetime
from pytz import timezone
import requests
from zipfile import ZipFile
import subprocess
import pyperclip

warnings.filterwarnings("ignore")

global rows
global xpath
global heding
global status
global key
global nme
element_1 = None
global j


def browse():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
        image_path2 = os.path.join(sys._MEIPASS, 'Static', 'Mapping.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')
        image_path2 = os.path.join(os.getcwd(), 'Static', 'Mapping.png')

    root.title("Process File Picker")
    root.resizable(False,False)

    root.title("Process File Picker")
    root.resizable(False,False)
   
    w = 500
    h = 160
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Browse Process - File Picker",font=("Calibri",18,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
          
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=50)
    
    answer=StringVar()
    answer.set("")
       
    def browse_button():
        global filename
        answer.set("")  
        filename = filedialog.askopenfilename()  
        txt.delete(0, 'end') 
        txt.insert(0, filename)

    def Click_Done():
        global ans        
        
        ans=txt.get()        
        
        if ans=="":
           answer.set("File Path Fields Empty Is Not Allowed...")        
        else:      
            root.destroy()          
            return ans
        
    txt=Entry(Frame2,font=("Calibri",12,"bold","italic"),width=50,justify="left")
    txt.grid(row=0,column=0,padx=5,pady=10,sticky="E")    

    photo1 = PhotoImage(file=image_path2)
            
    btn1=Button(Frame2,text="Browse",command=browse_button,image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=0,column=1,padx=3,pady=0,sticky="W")     
                
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=3,y=80,width=490,height=25)
    
    title_3=Label(Frame3,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=1,padx=150,pady=0,sticky="E")

    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=3,y=105,width=500,height=50)
    
    photo2 = PhotoImage(file=image_path1)

    btn2=Button(Frame4,command=Click_Done,text="Done",image=photo2,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=3,column=0,padx=125,pady=0,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo3 = PhotoImage(file=image_path)   

    btn3=Button(Frame4,command=Close,text="Exit",image=photo3,borderwidth=0,bg="#2c3e50")
    btn3.grid(row=3,column=1,padx=15,pady=0,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn2,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn3,'Click to Exit Process',hover_delay=1000)
    myTip2 = Hovertip(btn1,'Click to Pick File',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def process():        
    global element_1    
    global ans1
    global ans2 
    global driver
    global j
    global dob
    global mid
    global frm_dt
    global to_dt
    global cpt
    global rows
    global rows1
    # https://rajhandicraft.com/collections/bed-bed-sides/products/solid-wood-small-peacock-patra-design-bed-k?variant=41111281926328
    browse()
      
    fil=ans      

    current_directory = os.getcwd()
    # current_directory += os.path.sep
    check_path=os.path.join(current_directory+'\\Temp')
    folder_create=os.path.join(current_directory+'\\')
    down_path=os.path.join(current_directory+'\\Temp\\')
    fin_file=os.path.join(current_directory+'\\Temp\\Files\\')
   
    fld='Temp'
    
    if not os.path.isdir(check_path):
        os.mkdir(folder_create + 'Temp')
        os.mkdir(down_path + 'Files')
    else:
        path1 = os.path.join(folder_create, fld)
        shutil.rmtree(path1)
        os.mkdir(folder_create + 'Temp')
        os.mkdir(down_path + 'Files')  

    try:
        options = Options()            
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--ignore-ssl-errors')
        # options.add_argument("--disable-popup-blocking")
        # options.add_argument('--disable-crl-sets')
        # options.add_argument('--headless')
        prefs = {"download.default_directory" : down_path,
                    "download.prompt_for_download": False,
                    "download.directory_upgrade": True,
                    "plugins.always_open_pdf_externally": True,
                    "profile.password_manager_enabled": False,
                    "credentials_enable_service": False,
                    # "safebrowsing.enabled": True
                    }
        # options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_experimental_option("prefs",prefs)
        driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options)
        driver.maximize_window()
        driver.get('https://promise.dpw.state.pa.us/portal/provider/Home/tabid/135/Default.aspx')
    except Exception as e:                        
        try:
            options = Options()
            options.add_argument('--ignore-certificate-errors')
            options.add_argument('--ignore-ssl-errors')
            # options.add_argument("--disable-popup-blocking")
            # options.add_argument('--disable-crl-sets')
            # options.add_argument('--headless')
            prefs = {"download.default_directory" : down_path,
                    "download.prompt_for_download": False,
                    "download.directory_upgrade": True,
                    "plugins.always_open_pdf_externally": True,
                    "profile.password_manager_enabled": False,
                    "credentials_enable_service": False,
                    # "safebrowsing.enabled": True
                    }
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_experimental_option("useAutomationExtension", False)
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option("prefs",prefs)            
            driver_path = os.path.abspath('chromedriver.exe')
            driver = webdriver.Chrome(executable_path=driver_path,options=options)
            driver.maximize_window()            
        except Exception as e:           
            messagebox.showinfo("Driver Problem","Pls Check Your Chrome Driver Version")
            sys.exit(0)        
        
    def text_box(xpath,heding,status,key):                
        counter = 0
        while counter < 20:
            try:   
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()                     
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            print('Moved Exception')
            raise e
            # messagebox.showinfo(heding, status)
            # sys.exit(0)        
    
    def text_box_key(xpath,heding,status,key):                
        counter = 0
        while counter < 5:
            try:                   
                element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, xpath)))
                element.clear()                
                element.send_keys(key)                
                # element.send_keys(Keys.TAB)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            print('Moved Exception')
            raise e
            # messagebox.showinfo(heding, status)
            # sys.exit(0)    

    def text_box_js(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:
                js_code = "document.querySelector('" + xpath + "').setAttribute('value', '" + key + "');"
                WebDriverWait(driver, 0).until(lambda driver: driver.execute_script(js_code))               
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo(heding, status)
            sys.exit(0)       

    def click(xpath,heding,status):
        counter = 0
        while counter < 20:
            try:             
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            print('Moved Exception')
            raise e
            # messagebox.showinfo(heding,status)
            # sys.exit(0)                                

    def count(xpath,heding,status):
        global rows
        counter = 0
        while counter < 20:
            try:             
                rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            print('Moved Exception')
            raise e
            # messagebox.showinfo(heding,status)
            # sys.exit(0)                         

    def count_1(xpath,heding,status):
        global rows1
        counter = 0
        while counter < 5:
            try:             
                rows1=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            print('Moved Exception')
            raise e
            # messagebox.showinfo(heding,status)
            # sys.exit(0)        

    def Alert(msg):
        counter = 0
        while counter < 5:
            try:             
                WebDriverWait(driver, 0).until (EC.alert_is_present())
                a=driver.switch_to.alert
                a.accept()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            print('Alert - ',msg)                                                                     

    file=pd.read_excel(fil,sheet_name='Sigma',header=0)
    
    for index, row in file.iterrows():                                      
        lnk = row[0]
        Acc=row[1]
        usr_nm = row[2]                     
        pwd = row[3]
        pt_nm=row[4]
        
        driver.get(lnk.lstrip().rstrip())

        xpath= "/html/body/main/div[3]/div[2]/div/form[1]/div[2]/input"
        heding="Account"
        status="Account Field Not Found"
        key=Acc.lstrip().rstrip()
        text_box(xpath,heding,status,key)

        xpath= "/html/body/main/div[3]/div[2]/div/form[1]/div[3]/input"
        heding="User Name"
        status="User Name Field Not Found"
        key=usr_nm.lstrip().rstrip()
        text_box(xpath,heding,status,key)
        
        xpath= "/html/body/main/div[3]/div[2]/div/form[1]/div[4]/input"
        heding="Password"
        status="Password Field Not Found"
        key=pwd.lstrip().rstrip()
        text_box(xpath,heding,status,key)

        xpath='/html/body/main/div[3]/div[2]/div/form[1]/input'
        heding="Sign In"
        status="Sign In Button Not Found"        
        click(xpath,heding,status)

        counter = 0
        while counter < 5:
            try:
                xpath= "/html/body/main/div[3]/div[2]/div/form[1]/div[6]"
                ck=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).text                
                if 'Login using a User Name and Password is not allowed.' in ck:
                    wb1=load_workbook(filename=fil)
                    sheet = wb1['Sigma']
                    column_letter = 'F'  
                    column_cells = sheet[column_letter]
                    last_row = None
                    for cell in reversed(column_cells):
                        if cell.value:
                            last_row = cell.row
                            break                            
                    sheet['F' + str(int(last_row + 1))]=ck 
                    wb1.save(fil)
                    wb1.close()                                    

                    break 
                else:
                    counter = 0
                    while counter < 5:
                        try:             
                            try:
                                WebDriverWait(driver, 0).until (EC.alert_is_present())
                                a=driver.switch_to.alert
                                a.accept()
                            except Exception as e:
                                pass
                            main_window_handle = driver.window_handles[0]
                            popup_window_handle = None

                            for handle in driver.window_handles:
                                if handle != main_window_handle:
                                    popup_window_handle = handle
                            driver.switch_to.window(popup_window_handle)
                            driver.close()
                            driver.switch_to.window(main_window_handle)
                            break
                        except Exception as e:
                            time.sleep(1)
                            counter += 1
                    else:
                        driver.switch_to.window(main_window_handle)   

                    counter = 0
                    while counter < 5:
                        try:                                                                                                  
                            WebDriverWait(driver, 3).until(EC.frame_to_be_available_and_switch_to_it((By.NAME, "Main")))
                            break
                        except Exception as e:
                            time.sleep(1)
                            counter += 1

                    xpath= '/html/body/div[1]/div/div[2]/table/tbody/tr[2]/td[1]/span/span/span/table[2]/tbody/tr[2]/td/table/tbody/tr/td[4]/div/span/table/tbody/tr/td[1]/input'
                    heding="Patient Name"
                    status="Patient Name Field Not Found"
                    key=pt_nm.lstrip().rstrip()
                    text_box(xpath,heding,status,key)

                    xpath='/html/body/div[1]/div/div[2]/table/tbody/tr[2]/td[1]/span/span/span/table[2]/tbody/tr[2]/td/table/tbody/tr/td[4]/div/span/table/tbody/tr/td[2]/span/a'
                    heding="Go Button"
                    status="Go Button Button Not Found"        
                    click(xpath,heding,status)

                    xpath='/html/body/div[1]/div/div[2]/table/tbody/tr[3]/td/div/table[1]/tbody/tr[2]/td/div/table/tbody/tr[5]/td[6]/table/tbody/tr/td[2]/img'
                    heding="Drop Down"
                    status="Drop Down Button Click Not Found"        
                    click(xpath,heding,status)
                    
                    xpath= "/html/body/div[4]/div/table/tbody/tr"
                    heding="Status Dropdown"        
                    status="Status Dropdown Not Found"                    
                    count(xpath,heding,status)
                    
                    j=2
                    while j<rows+1:  
                        dck=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div[4]/div/table/tbody/tr[{}]'.format(j)))).text
                        if 'All' in dck:
                            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[4]/div/table/tbody/tr[{}]'.format(j)))).click()
                            break
                        j=j+1
                    
                    xpath='/html/body/div[1]/div/div[2]/table/tbody/tr[3]/td/div/table[1]/tbody/tr[2]/td/div/table/tbody/tr[7]/td[2]/table/tbody/tr/td[2]/a[2]/img'
                    heding="Patient Search"
                    status="Patient Search Button Not Found"        
                    click(xpath,heding,status)                

                    time.sleep(2)

                    try:                                                                                                            
                        counter = 0
                        while counter < 20:
                            try:                                                                                                  
                                WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[1]/div/div[2]/table/tbody/tr[3]/td/div/table[2]/tbody/tr[2]/td/div/table[2]/tbody/tr[2]/td/div/table/tbody/tr[3]/td[4]/a'))).click()
                                break
                            except Exception as e:
                                time.sleep(1)
                                counter += 1
                                                
                        xpath='/html/body/div[1]/div/div[2]/table/tbody/tr[3]/td/div/table[1]/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/a[1]'
                        heding="Print FaceSheet Link"
                        status="Print FaceSheet Link Not Found"        
                        click(xpath,heding,status)
                                            
                        counter=0
                        while counter < 3:                                     
                            try:                                
                                xpath='/html/body/div[4]/span/span/table/tbody/tr/td/span/table/tbody/tr[2]/td/span/table/tbody/tr[3]/td/span/table/tbody/tr/td[3]/table/tbody/tr/td/span/a/img'
                                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()                                                                
                                
                                xpath='/html/body/div[1]/div/div[2]/table/tbody/tr[3]/td/div/table[1]/tbody/tr/td[1]/table/tbody/tr[2]/td/table/tbody/tr[2]/td/table/tbody/tr/td[2]/a[1]'
                                heding="Print FaceSheet Link"
                                status="Print FaceSheet Link Not Found"        
                                click(xpath,heding,status)

                                break
                            except Exception as e:
                                time.sleep(1)
                                counter += 1                           

                        flag=False
                        
                        counter=0
                        while counter < 3:                                     
                            try:
                                WebDriverWait(driver, 0).until (EC.alert_is_present())
                                a=driver.switch_to.alert
                                a.accept()
                                flag=True
                                break
                            except Exception as e:
                                time.sleep(1)
                                counter += 1                                                
                        if flag:
                            xpath='/html/body/div[1]/div/div[2]/table/tbody/tr[2]/td[1]/span/span/span/table[2]/tbody/tr[1]/td/table/tbody/tr[1]/td[6]/a'
                            heding="Log Out"
                            status="Log Out Button Not Found"        
                            click(xpath,heding,status)

                            wb1=load_workbook(filename=fil)
                            sheet = wb1['Sigma']
                            column_letter = 'F'  
                            column_cells = sheet[column_letter]
                            last_row = None
                            for cell in reversed(column_cells):
                                if cell.value:
                                    last_row = cell.row
                                    break                            
                            sheet['F' + str(int(last_row + 1))]='Error'        
                            wb1.save(fil)
                            wb1.close()   

                            break    
                                                 
                        counter=1
                        while counter < 15:                    
                            file1 = [f for f in listdir(down_path) if isfile(join(down_path, f))]
                            lol_string1 = ' '.join(file1)
                            lol_string2=lol_string1.split('.')[-1]                                                                                
                            if lol_string2!='' and lol_string2!='tmp' and lol_string2 !="crdownload" and lol_string2 !="temp":
                                time.sleep(1)
                                file2 = [f for f in os.listdir(down_path) if os.path.isfile(os.path.join(down_path, f))]
                                if file2:
                                    lol_string3 = os.path.join(down_path, file2[0])  
                                    directory, file_without_extension = os.path.split(lol_string3)
                                    chng_fn = pt_nm +'.pdf'
                                    new_file_path = os.path.join(directory, chng_fn)
                                    os.rename(lol_string3,new_file_path)

                                    shutil.move(new_file_path, os.path.join(fin_file, chng_fn))
                                                                                
                                    wb1=load_workbook(filename=fil)
                                    sheet = wb1['Sigma']
                                    column_letter = 'F'  
                                    column_cells = sheet[column_letter]
                                    last_row = None
                                    for cell in reversed(column_cells):
                                        if cell.value:
                                            last_row = cell.row
                                            break                            
                                    sheet['F' + str(int(last_row + 1))]='Done'      
                                    wb1.save(fil)
                                    wb1.close()                                                
                                    break
                            else:
                                time.sleep(1)    
                                counter += 1
                        else:
                            files = os.listdir(check_path)
                            for file_name in files:
                                    file_path = os.path.join(check_path, file_name)
                                    try:
                                        if os.path.isfile(file_path):
                                            os.remove(file_path)                                   
                                    except OSError as e:
                                        pass    
                            wb1=load_workbook(filename=fil)
                            sheet = wb1['Sigma']
                            column_letter = 'F'  
                            column_cells = sheet[column_letter]
                            last_row = None
                            for cell in reversed(column_cells):
                                if cell.value:
                                    last_row = cell.row
                                    break                            
                            sheet['F' + str(int(last_row + 1))]='File Error'        
                            wb1.save(fil)
                            wb1.close()                                       

                        counter = 0
                        while counter < 8:
                            try:                                                     
                                main_window_handle = driver.window_handles[0]
                                popup_window_handle = None

                                for handle in driver.window_handles:
                                    if handle != main_window_handle:
                                        popup_window_handle = handle
                                driver.switch_to.window(popup_window_handle)                                                                                                                    
                                driver.close()
                                driver.switch_to.window(main_window_handle)
                                break
                            except Exception as e:
                                time.sleep(1)
                                counter += 1                                    
                        else:
                            driver.switch_to.window(main_window_handle) 
                        
                        counter = 0
                        while counter < 10:
                            try:                                                                                                  
                                WebDriverWait(driver, 0).until(EC.frame_to_be_available_and_switch_to_it((By.NAME, "Main")))
                                break
                            except Exception as e:
                                time.sleep(1)
                                counter += 1

                        xpath='/html/body/div[1]/div/div[2]/table/tbody/tr[2]/td[1]/span/span/span/table[2]/tbody/tr[1]/td/table/tbody/tr[1]/td[6]/a'
                        heding="Log Out"
                        status="Log Out Button Not Found"        
                        click(xpath,heding,status)

                        break                      
                    except Exception as e:
                        wb1=load_workbook(filename=fil)
                        sheet = wb1['Sigma']
                        column_letter = 'F'  
                        column_cells = sheet[column_letter]
                        last_row = None
                        for cell in reversed(column_cells):
                            if cell.value:
                                last_row = cell.row
                                break                            
                        sheet['F' + str(int(last_row + 1))]='Patient Resident Link Not Found'        
                        wb1.save(fil)
                        wb1.close()  
                        
                        xpath='/html/body/div[1]/div/div[2]/table/tbody/tr[2]/td[1]/span/span/span/table[2]/tbody/tr[1]/td/table/tbody/tr[1]/td[6]/a'
                        heding="Log Out"
                        status="Log Out Button Not Found"        
                        click(xpath,heding,status)

                        break      

            except Exception as e:                              
                time.sleep(1)
                counter += 1
        else:
            wb1=load_workbook(filename=fil)
            sheet = wb1['Sigma']
            column_letter = 'F'  
            column_cells = sheet[column_letter]
            last_row = None
            for cell in reversed(column_cells):
                if cell.value:
                    last_row = cell.row
                    break                            
            sheet['F' + str(int(last_row + 1))]='Error'        
            wb1.save(fil)
            wb1.close()     

    driver.quit()
    messagebox.showinfo("Process Status", "Process Completed")
    sys.exit(0)    

if __name__=="__main__":        
    process()
    