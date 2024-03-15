from tkinter import BOTH, LEFT, TOP, Button, Entry, Frame, Label, PhotoImage, StringVar, Tk,Radiobutton,StringVar,IntVar,filedialog
from idlelib.tooltip import Hovertip
import sys
import os
from operator import itemgetter
import time
from tkinter import messagebox
from os import listdir
from os.path import isfile, join
from datetime import datetime, date
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, StaleElementReferenceException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver import ActionChains
import openpyxl
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from urllib.parse import urlparse
import warnings
import numpy as np
from openpyxl.utils import get_column_letter,column_index_from_string
from datetime import datetime
from pytz import timezone
import requests
from zipfile import ZipFile
import subprocess


warnings.filterwarnings("ignore")

global rows
global xpath
global heding
global status
global key
global nme
element_1 = None
global j


def browse():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
        image_path2 = os.path.join(sys._MEIPASS, 'Static', 'Mapping.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')
        image_path2 = os.path.join(os.getcwd(), 'Static', 'Mapping.png')

    root.title("Process File Picker")
    root.resizable(False,False)

    root.title("Process File Picker")
    root.resizable(False,False)
   
    w = 500
    h = 160
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Browse Process - File Picker",font=("Calibri",18,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
          
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=50)
    
    answer=StringVar()
    answer.set("")
       
    def browse_button():
        global filename
        answer.set("")  
        filename = filedialog.askopenfilename()  
        txt.delete(0, 'end') 
        txt.insert(0, filename)

    def Click_Done():
        global ans        
        
        ans=txt.get()        
        
        if ans=="":
           answer.set("File Path Fields Empty Is Not Allowed...")        
        else:      
            root.destroy()          
            return ans
        
    txt=Entry(Frame2,font=("Calibri",12,"bold","italic"),width=50,justify="left")
    txt.grid(row=0,column=0,padx=5,pady=10,sticky="E")    

    photo1 = PhotoImage(file=image_path2)
            
    btn1=Button(Frame2,text="Browse",command=browse_button,image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=0,column=1,padx=3,pady=0,sticky="W")     
                
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=3,y=80,width=490,height=25)
    
    title_3=Label(Frame3,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=1,padx=150,pady=0,sticky="E")

    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=3,y=105,width=500,height=50)
    
    photo2 = PhotoImage(file=image_path1)

    btn2=Button(Frame4,command=Click_Done,text="Done",image=photo2,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=3,column=0,padx=125,pady=0,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo3 = PhotoImage(file=image_path)   

    btn3=Button(Frame4,command=Close,text="Exit",image=photo3,borderwidth=0,bg="#2c3e50")
    btn3.grid(row=3,column=1,padx=15,pady=0,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn2,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn3,'Click to Exit Process',hover_delay=1000)
    myTip2 = Hovertip(btn1,'Click to Pick File',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def process():        
    global element_1    
    global ans1
    global ans2 
    global driver
    global j
    global dob
    global mid
    global frm_dt
    global to_dt
    global cpt
    global rows
    global rows1
    # https://rajhandicraft.com/collections/bed-bed-sides/products/solid-wood-small-peacock-patra-design-bed-k?variant=41111281926328
    browse()
      
    fil=ans      

    # current_directory = os.getcwd()
    # # current_directory += os.path.sep
    # check_path=os.path.join(current_directory+'\\Temp')
    # folder_create=os.path.join(current_directory+'\\')
    # down_path=os.path.join(current_directory+'\\Temp\\')
    # fin_file=os.path.join(current_directory+'\\Temp\\Files\\')
   
    # fld='Temp'
    
    # if not os.path.isdir(check_path):
    #     os.mkdir(folder_create + 'Temp')
    #     os.mkdir(down_path + 'Files')
    # else:
    #     path1 = os.path.join(folder_create, fld)
    #     shutil.rmtree(path1)
    #     os.mkdir(folder_create + 'Temp')
    #     os.mkdir(down_path + 'Files')  

    try:
        options = Options()            
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--ignore-ssl-errors')
        # options.add_argument("--disable-popup-blocking")
        # options.add_argument('--disable-crl-sets')
        # options.add_argument('--headless')
        prefs = {"download.prompt_for_download": False,
                    "download.directory_upgrade": True,
                    "plugins.always_open_pdf_externally": True,
                    "profile.password_manager_enabled": False,
                    "credentials_enable_service": False,
                    # "safebrowsing.enabled": True
                    }
        # options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_experimental_option("prefs",prefs)
        driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options)
        driver.maximize_window()
        driver.get('https://data.cms.gov/tools/medicare-revalidation-list')
    except Exception as e:                        
        try:
            options = Options()
            options.add_argument('--ignore-certificate-errors')
            options.add_argument('--ignore-ssl-errors')
            # options.add_argument("--disable-popup-blocking")
            # options.add_argument('--disable-crl-sets')
            # options.add_argument('--headless')
            prefs = {"download.prompt_for_download": False,
                    "download.directory_upgrade": True,
                    "plugins.always_open_pdf_externally": True,
                    "profile.password_manager_enabled": False,
                    "credentials_enable_service": False,
                    # "safebrowsing.enabled": True
                    }
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_experimental_option("useAutomationExtension", False)
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option("prefs",prefs)            
            driver_path = os.path.abspath('chromedriver.exe')
            driver = webdriver.Chrome(executable_path=driver_path,options=options)
            driver.maximize_window()            
            driver.get('https://data.cms.gov/tools/medicare-revalidation-list')
        except Exception as e:           
            messagebox.showinfo("Driver Problem","Pls Check Your Chrome Driver Version")
            sys.exit(0)        
        
    def text_box(xpath,heding,status,key):                
        counter = 0
        while counter < 20:
            try:   
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()                     
                WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding, status)
            sys.exit(0)        
    
    def text_box_key(xpath,heding,status,key):                
        counter = 0
        while counter < 5:
            try:                   
                element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, xpath)))
                element.clear()                
                element.send_keys(key)                
                # element.send_keys(Keys.TAB)
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding, status)
            sys.exit(0)    

    def text_box_js(xpath,heding,status,key):                
        counter = 0
        while counter < 15:
            try:
                js_code = "document.querySelector('" + xpath + "').setAttribute('value', '" + key + "');"
                WebDriverWait(driver, 0).until(lambda driver: driver.execute_script(js_code))               
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            messagebox.showinfo(heding, status)
            sys.exit(0)       

    def click(xpath,heding,status):
        counter = 0
        while counter < 20:
            try:             
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding,status)
            sys.exit(0)                                

    def count(xpath,heding,status):
        global rows
        counter = 0
        while counter < 20:
            try:             
                rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding,status)
            sys.exit(0)                         

    def count_1(xpath,heding,status):
        global rows1
        counter = 0
        while counter < 5:
            try:             
                rows1=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            # print('Moved Exception')
            # raise e
            messagebox.showinfo(heding,status)
            sys.exit(0)        

    def Alert(msg):
        counter = 0
        while counter < 5:
            try:             
                WebDriverWait(driver, 0).until (EC.alert_is_present())
                a=driver.switch_to.alert
                a.accept()
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:
            print('Alert - ',msg)                                                                     

    file=pd.read_excel(fil,sheet_name='Medicare Revalidation',header=0)
    
    for index, row in file.iterrows():                                      
        npi = row[0]
        st=row[1]
        
        try:
            xpath='/html/body/div/div/div/div/div/div/div/div[3]/div[2]/div/div[2]/form/div[2]/div[3]/div/div[1]/button'
            WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
        except Exception as e:
            pass
                      
        xpath= "/html/body/div/div/div/div/div/div/div/div[3]/div[2]/div/div[2]/form/div[1]/div[1]/input"
        heding="Search by NPI"
        status="Search by NPI Field Not Found"
        key=str(npi)
        text_box(xpath,heding,status,key)
            
        # input_xpath = "//input[@id='react-select-5-input']"
        input_xpath = '/html/body/div/div/div/div/div/div/div/div[3]/div[2]/div/div[2]/form/div[2]/div[1]/div/div/div[1]/div[2]/div/input'
        heding="Location"
        status="Location Field Not Found"

        counter = 0
        while counter < 10:
            try:                                  
                input_field = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, input_xpath)))
                input_field.clear()
                input_field.send_keys(st)   
                # input_field.send_keys(Keys.TAB)     
                input_field.send_keys(Keys.ENTER) 
                break
            except Exception as e:
                time.sleep(1)
                counter += 1
        else:            
            messagebox.showinfo(heding, status)
            sys.exit(0) 
        
        xpath='/html/body/div/div/div/div/div/div/div/div[3]/div[2]/div/div[2]/form/div[2]/div[3]/div/div[2]/button'
        heding="Find Provider"
        status="Find Provider Button Not Found"        
        click(xpath,heding,status)          

        counter = 0
        while counter < 5:
            try:                              
                ck=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div/div/div/div/div/div/div/div[5]/div[1]/div/div/div[1]/div[1]'))).text                         
                if ck=='Displaying 0 - 0 of 0 records':
                    wb1=load_workbook(filename=fil)
                    sheet = wb1['Medicare Revalidation']
                    column_letter = 'J'  
                    column_cells = sheet[column_letter]
                    last_row = None
                    for cell in reversed(column_cells):
                        if cell.value:
                            last_row = cell.row
                            break                            
                    sheet['J' + str(int(last_row + 1))]=ck 
                    wb1.save(filename=fil)
                    wb1.close()  

                    break
                else:
                    lst=[]
                  
                    ind=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div/div/div/div/div/div/div/div[5]/div[1]/div/div/div[2]/div[2]/div[1]/a/h4'))).text                                
                    if ind==' ' or ind=='':
                        ind='N/A'
                        lst.append(ind)
                    else:
                        lst.append(ind)
                    
                    lt=len(lst)
                    if lt==0:
                        lst.append('N/A')

                    lst1=[]
                    np=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div/div/div/div/div/div/div/div[5]/div[1]/div/div/div[2]/div[2]/div[1]/div[2]'))).text

                    if np==' ' or np=='':
                        np='N/A'
                        lst1.append(np)
                    else:
                        lst1.append(np) 

                    lt=len(lst1)
                    if lt==0:
                        lst.append('N/A')
                    else:
                        vr=lst1[0]
                        lst.append(vr)

                    lst2=[]
                    dt=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div/div/div/div/div/div/div/div[5]/div[1]/div/div/div[2]/div[2]/div[2]/div[2]'))).text

                    if dt==' ' or dt=='':
                        dt='N/A'
                        lst2.append(dt)   
                    else:
                        lst2.append(dt) 

                    lt=len(lst2)
                    if lt==0:                           
                        lst.append('N/A')
                    else:
                        vr=lst2[0]
                        lst.append(vr)

                    lst3=[]
                    st1=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div/div/div/div/div/div/div/div[5]/div[1]/div/div/div[2]/div[2]/div[3]/div[1]'))).text

                    if st1==' ' or st1=='':
                        st1='N/A'
                        lst3.append(st1)   
                    else:
                        lst3.append(st1) 

                    lt=len(lst3)
                    if lt==0:                           
                        lst.append('N/A')
                    else:
                        vr=lst3[0]
                        lst.append(vr)

                    lst4=[]
                    sp=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div/div/div/div/div/div/div/div[5]/div[1]/div/div/div[2]/div[2]/div[3]/div[2]'))).text

                    if sp==' ' or sp=='':
                        sp='N/A'
                        lst4.append(sp)   
                    else:
                        lst4.append(sp) 

                    lt=len(lst4)
                    if lt==0:                           
                        lst.append('N/A')
                    else:
                        vr=lst4[0]
                        lst.append(vr)

                    lst5=[]
                    rrb=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div/div/div/div/div/div/div/div[5]/div[1]/div/div/div[2]/div[2]/div[3]/div[3]'))).text

                    if rrb==' ' or rrb=='':
                        rrb='N/A'
                        lst5.append(rrb)   
                    else:
                        lst5.append(rrb) 

                    lt=len(lst5)
                    if lt==0:                           
                        lst.append('N/A')
                    else:
                        vr=lst5[0]
                        lst.append(vr)

                    lst6=[]
                    et=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div/div/div/div/div/div/div/div[5]/div[1]/div/div/div[2]/div[2]/div[3]/div[4]'))).text

                    if et==' ' or et=='':
                        et='N/A'
                        lst6.append(et)   
                    else:
                        lst6.append(et) 

                    lt=len(lst6)
                    if lt==0:                           
                        lst.append('N/A')
                    else:
                        vr=lst6[0]
                        lst.append(vr)

                    wb1=load_workbook(fil)
                    sheet = wb1['Medicare Revalidation']
                    column_letter = 'J'  
                    column_cells = sheet[column_letter]
                    last_row = None
                    for cell in reversed(column_cells):
                        if cell.value:
                            last_row = cell.row
                            break
                    start_column = 'C'
                    current_column_index = openpyxl.utils.column_index_from_string(start_column)
                    current_row = last_row + 1

                    for value in lst:
                        current_column = openpyxl.utils.get_column_letter(current_column_index)
                        sheet[current_column + str(current_row)] = value
                        current_column_index += 1                
                    sheet['J' + str(int(last_row + 1))]='Done'
                    wb1.save(filename=fil)
                    wb1.close() 

                    break
            except Exception as e:                
                time.sleep(1)
                counter += 1
        
    driver.quit()
    messagebox.showinfo("Process Status", "Process Completed")
    sys.exit(0)    

if __name__=="__main__":        
    process()
    