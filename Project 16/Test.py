import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter,column_index_from_string
import openpyxl

def test():
    
    fil='D:\\Project 16\\ERA - Template.xlsx'  
    down_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Downloads\\Temp\\')
    files = os.listdir(down_path)
    text_files = [file for file in files if file.endswith('.txt')]
    
    for text_file in text_files:
        file_path = os.path.join(down_path, text_file)
        delimiter = '\t'
        df = pd.read_csv(file_path, delimiter=delimiter)
        df = df.rename(columns={df.columns[0]: 'temp'})                            
        
        wb1=load_workbook(fil)

        indices_to_remove = df[df['temp'].str.contains('={30,}|-{30,}|^\s*$')].index
        df = df.drop(indices_to_remove)        
        check_indices = df[df['temp'].str.contains('Check#')].index
        
        fis=check_indices[0]
        hed=check_indices[1]
        hed_row = df.loc[fis : hed,'temp']
        
        hed_row = pd.DataFrame(hed_row)
        hed_row = hed_row.reset_index(drop=True)         
        hed_row_ck = hed_row[hed_row['temp'].str.contains('Check#')].index  
        
        lst = []
        lst1 = []

        if len(hed_row) == 3:
            data_row = hed_row.loc[1,'temp']
            date_pattern = r'\d{2}/\d{2}/\d{4}'
            matches = re.findall(date_pattern, data_row)
            if matches:
                date = matches[0]
            else:
                date = 'N/A'                                                                             
            split_items = re.split(r'\s{2,}', data_row)
            split_items_raw = [item for item in split_items if item]            
            chk_all = split_items_raw[0]                                              
            matches = re.findall(r'\s', chk_all)
            if matches:
                split_string = chk_all.split(' ')
                split_items_raw.pop(0)
                ck=split_string[0]
                amt=split_string[1]
                split_items_raw.insert(0, ck)
                split_items_raw.insert(1, amt)                        
            indices_to_select = [0, 1, 3]
            split_items_raw1 = [split_items_raw[i] for i in indices_to_select]
            split_items_raw1.append(str(date)) 

            lst1.append(split_items_raw)    
            lst.append(split_items_raw1)
        else:
            j=hed_row_ck[0].item()            
            while j<len(hed_row)-2:               
                data_row = hed_row.loc[j + 1,'temp']
                date_pattern = r'\d{2}/\d{2}/\d{4}'
                matches = re.findall(date_pattern, data_row)
                if matches:
                    date = matches[0]                                        
                else:
                    date = 'N/A'                                                                             
                
                split_items = re.split(r'\s{2,}', data_row)
                split_items_raw = [item for item in split_items if item]
                chk_all = split_items_raw[0]                                              
                matches = re.findall(r'\s', chk_all)
                if matches:
                    split_string = chk_all.split(' ')
                    split_items_raw.pop(0)
                    ck=split_string[0]
                    amt=split_string[1]
                    split_items_raw.insert(0, ck)
                    split_items_raw.insert(1, amt)                        
                indices_to_select = [0, 1, 3]
                split_items_raw1 = [split_items_raw[i] for i in indices_to_select]
                split_items_raw1.append(str(date)) 

                indices_to_select = [0, 1, 3]
                split_items_raw1 = [split_items_raw[i] for i in indices_to_select]
                split_items_raw1.append(str(date))                 

                lst1.append(split_items_raw)    
                lst.append(split_items_raw1)            

                j=j+1   
                        
        i=1                                                  
        while i<len(check_indices):                                
                        
            check_index = check_indices[i]
            data_row = df.loc[check_index + 1,'temp']
            split_items = re.split(r'\s{2,}', data_row)
            split_items = [item for item in split_items if item]
            
            chk_all = split_items[0]                                              
            matches = re.findall(r'\s', chk_all)
            if matches:
                split_string = chk_all.split(' ')
                split_items.pop(0)
                ck=split_string[0]
                amt=split_string[1]
                split_items.insert(0, ck)
                split_items.insert(1, amt)
                                                                
            chk_num=split_items[0]    
            chk_all = split_items[1]            
        
            matches = re.findall(r'\b\d+\.\d{2}\b', chk_all)            
            matches1 = chk_all.replace(' ', '')            

            if matches:
                split_items.insert(1, 'N/A')
            else:
                if any(char == ',' for char in matches1):
                    split_items.insert(1, 'N/A') 
                                                
            split_items.pop(0)

            input_string=split_items[1]
            number_found = any(char.isdigit() for char in input_string)
            if number_found:
                index_of_first_digit = next((i for i, char in enumerate(input_string) if char.isdigit()), None)
                if index_of_first_digit is not None:
                    nm = input_string[:index_of_first_digit].strip()  
                    val = input_string[index_of_first_digit:].strip()  
                    split_items.pop(1)
                    split_items.insert(1, nm)
                    split_items.insert(2, val)
            
            input_string=split_items[4]
            alphabetic_found = any(char.isalpha() and char != ' ' for char in input_string)
            if alphabetic_found:
                parts = input_string.split(' ')
                nm = ''
                val = ''
                for part in parts:
                    if part.replace('-', '').isdigit():
                        val += part + ' '
                    else:
                        nm += part + ' '
                nm = nm.strip()
                val = val.strip()
                # nm = ''.join(char for char in input_string if char.isalpha() or char == '-')
                # val = ''.join(char for char in input_string if char.isdigit() or char == '.') 
                split_items.pop(4)
                split_items.insert(4, val)
                split_items.insert(5, nm)
                
            
            z=6         
            joined_string = ''                                         
            while z<len(split_items):   
                input_string1=split_items[z]
                joined_string += input_string1 + ' '
                z += 1
            
            z=6                                                  
            while z<len(split_items) + 1 :   
                split_items.pop(6)
                z += 1

            split_items.insert(len(split_items), joined_string.strip())
            
            try:
                sel=check_indices[i + 1]
                sel_row = df.loc[check_index : sel,'temp']

                tm_df=pd.DataFrame(sel_row)
                tm_df = tm_df.reset_index(drop=True)
                                                    
                pay_calm = tm_df[tm_df['temp'].str.contains('Payer Claim Control Number:')]
                pay_calm = '\n'.join(pay_calm.iloc[:, 0])
                pay_calm = pay_calm.strip().split(':')[1].strip()
                split_items.append(str(pay_calm))
            except Exception as e:
                last_index = df.index[-1]
                sel_row = df.loc[check_index : last_index,'temp']

                tm_df=pd.DataFrame(sel_row)
                tm_df = tm_df.reset_index(drop=True)
                                                    
                pay_calm = tm_df[tm_df['temp'].str.contains('Payer Claim Control Number:')]
                pay_calm = '\n'.join(pay_calm.iloc[:, 0])
                pay_calm = pay_calm.strip().split(':')[1].strip()
                split_items.append(str(pay_calm))

            # chk_itm=len(split_items)
            
            # if chk_itm==8:
            #     pass
            # else:
            #     input_string=split_items[1]
            #     number_found = any(char.isdigit() for char in input_string)
            #     if number_found:
            #         index_of_first_digit = next((i for i, char in enumerate(input_string) if char.isdigit()), None)
            #         if index_of_first_digit is not None:
            #             nm = input_string[:index_of_first_digit].strip()  
            #             val = input_string[index_of_first_digit:].strip()  
            #             split_items.pop(1)
            #             split_items.insert(1, nm)
            #             split_items.insert(2, val)
            #     else:
            #         input_string=split_items[4]
            #         alphabetic_found = any(char.isalpha() and char != ' ' for char in input_string)
            #         if alphabetic_found:
            #             parts = input_string.split(' ')
            #             nm = ''
            #             val = ''
            #             for part in parts:
            #                 if part.replace('-', '').isdigit():
            #                     val += part + ' '
            #                 else:
            #                     nm += part + ' '
            #             nm = nm.strip()
            #             val = val.strip()
            #             # nm = ''.join(char for char in input_string if char.isalpha() or char == '-')
            #             # val = ''.join(char for char in input_string if char.isdigit() or char == '.') 
            #             split_items.pop(5)
            #             split_items.insert(5, nm)
            #             split_items.insert(6, val)

            lin_itms = tm_df[tm_df['temp'].str.contains('Line Item:')].index                                
            
            lst7=[]
            j=0                                                  
            while j<len(lin_itms):    
                lin_itm = lin_itms[j]
                data_row = tm_df.loc[lin_itm + 1,'temp']
                line_items = re.split(r'\s{2,}', data_row)
                line_items = [item for item in line_items if item]
                
                dt_sp = line_items[0]
                matches = re.findall(r'\b\d+\.\d{2}\b', dt_sp) 
                if matches:
                    dt_sp1 = dt_sp.strip().split(' ')[0].strip()
                    cpt_sp = dt_sp.strip().split(' ')[1].strip()
                    chr_amt = dt_sp.strip().split(' ')[2].strip()

                    line_items.pop(0)
                    line_items.insert(0, dt_sp1)
                    line_items.insert(1, cpt_sp)
                    line_items.insert(2, chr_amt)
                else:
                    dt_sp1 = dt_sp.strip().split(' ')[0].strip()
                    cpt_sp = dt_sp.strip().split(' ')[1].strip()

                    line_items.pop(0)
                    line_items.insert(0, dt_sp1)
                    line_items.insert(1, cpt_sp)

                k=2
                adj = tm_df.loc[lin_itm + k,'temp']

                while 'Line Item:' not in adj:                                                     
                    lines_items = adj.strip()                                        
                    line_items.append(lines_items)                    
                    
                    k=k+1  
                    try:
                        adj = tm_df.loc[lin_itm + k,'temp']
                        if 'Check#' in adj:
                            break
                    except Exception as e:
                        break
                lst7.append(line_items)
                j=j+1
            
            lst_cnt=len(lst7)

            l=0
            while l < lst_cnt:                
                fin_lst = None
                for sublist in lst:
                    if chk_num in sublist:
                        fin_lst = sublist
                        break                
                               
                fin_lst_itm = lst7[l]
                      
                # wb1=load_workbook(fil)
                sheet = wb1['CPT']
                column_letter = 'A'  
                column_cells = sheet[column_letter]
                last_row = None
                for cell in reversed(column_cells):
                    if cell.value:
                        last_row = cell.row
                        break              
                start_column = 'C'
                current_column_index = openpyxl.utils.column_index_from_string(start_column)
                current_row = last_row + 1

                for value in fin_lst:
                    current_column = openpyxl.utils.get_column_letter(current_column_index)
                    sheet[current_column + str(current_row)] = value
                    current_column_index += 1                
                
                start_column = 'G'
                current_column_index = openpyxl.utils.column_index_from_string(start_column)                                   

                for value in split_items:
                    current_column = openpyxl.utils.get_column_letter(current_column_index)
                    sheet[current_column + str(current_row)] = value
                    current_column_index += 1 

                start_column = 'O'
                current_column_index = openpyxl.utils.column_index_from_string(start_column)                                   

                for value in fin_lst_itm:
                    current_column = openpyxl.utils.get_column_letter(current_column_index)
                    sheet[current_column + str(current_row)] = value
                    current_column_index += 1 
                
                sheet['B' + str(int(last_row + 1))]=text_file
                sheet['A' + str(int(last_row + 1))]='Done'
                # wb1.save(filename=fil)
                # wb1.close()

                l=l+1            
            i=i+1   
        
        pyr=split_items[-2]

        fin_lst11=[]
        date_pattern = r'\b\d{2}/\d{2}/\d{4}\b'

        for i in range(len(lst1)):
            date = None
            for item in lst1[i]:
                date_match = re.search(date_pattern, item)
                if date_match:
                    date = date_match.group(0)        
            filtered_list = [re.sub(date_pattern, '', item) for item in lst1[i]]
            joined_string = ' '.join(filtered_list[4:])

            updated_sublist = filtered_list[:4] + [joined_string.strip(), date.strip()]

            fin_lst11.append(updated_sublist)           

        for i in range(len(fin_lst11)):
            fin_lst11[i].append(pyr)         

        for i, sub_list in enumerate(fin_lst11):           
            sheet = wb1['Raw']
            column_letter = 'I'  
            column_cells = sheet[column_letter]
            last_row = None
            for cell in reversed(column_cells):
                if cell.value:
                    last_row = cell.row
                    break              
            start_column = 'B'
            current_column_index = openpyxl.utils.column_index_from_string(start_column)
            current_row = last_row + 1
            
            for value in sub_list:
                current_column = openpyxl.utils.get_column_letter(current_column_index)
                sheet[current_column + str(current_row)] = value
                current_column_index += 1                        

            sheet['A' + str(int(last_row + 1))]=text_file
            sheet['I' + str(int(last_row + 1))]='Done'
            
        wb1.save(filename=fil)
        wb1.close()

# import re

# def extract_patient_details(text):
#     patient_details = []

#     # Split the text into lines and process each line
#     lines = text.split('\n')
#     for line in lines:
#         # Remove extra spaces and split the line into parts
#         parts = line.strip().split()

#         # Check if the line has the required number of parts and starts with a valid Check#
#         if len(parts) == 8 and parts[0].isdigit():
#             check_num = parts[0]
#             patient_id = parts[1]
#             last_first_name = parts[2] + ", " + parts[3]
#             charge_amt = parts[4]
#             payment_amt = parts[5]
#             accnt_num = parts[6]
#             status = parts[7]
#             payer = " ".join(parts[8:])

#             # Append the extracted details to the list
#             patient_details.append({
#                 "Check#": check_num,
#                 "Patient ID": patient_id,
#                 "Last, First Name": last_first_name,
#                 "Charge Amount": charge_amt,
#                 "Payment Amount": payment_amt,
#                 "Account#": accnt_num,
#                 "Status": status,
#                 "Payer": payer
#             })

#     return patient_details

# # Load the text from the file
# file_path = "C:\\Users\\sanandrao\\Downloads\\Temp\\1088623388_ERA_STATUS_5010_20230530.txt"
# try:
#     with open(file_path, "r") as file:
#         text_data = file.read()

#         # Extract and print the patient details
#         patient_details = extract_patient_details(text_data)
#         for idx, details in enumerate(patient_details, start=1):
#             print(f"Patient Details {idx}:")
#             for key, value in details.items():
#                 print(f"{key}: {value}")
#             print("\n")

# except FileNotFoundError:
#     print("File not found. Please provide a valid file path.")
# except PermissionError:
#     print("Permission denied. Unable to read the file.")
# except Exception as e:
#     print("An error occurred while reading the file:", str(e))


if __name__=="__main__":        
    test()    
# file_path = "C:\\Users\\sanandrao\\Downloads\\Temp\\1088623388_ERA_STATUS_5010_20230530.txt"