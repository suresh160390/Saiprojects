from tkinter import BOTH, LEFT, TOP, Button, Entry, Frame, Label, PhotoImage, StringVar, Tk,Radiobutton,StringVar,IntVar,filedialog
from tkinter import ttk
from idlelib.tooltip import Hovertip
from tkinter import messagebox
import sys
import os
import warnings
from datetime import datetime, date,timedelta
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, StaleElementReferenceException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import openpyxl
import time
import pandas as pd
from openpyxl import load_workbook
from urllib.parse import urlparse
import warnings
import numpy as np
import requests
from zipfile import ZipFile
from openpyxl.utils import get_column_letter,column_index_from_string
import shutil
from os import listdir
from os.path import isfile, join
import zipfile
warnings.filterwarnings("ignore")
from tkcalendar import DateEntry
import re

global rows
global xpath
global heding
global status
global key
global nme
global driver
element_1 = None

def file_pick():
   
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
        image_path2 = os.path.join(sys._MEIPASS, 'Static', 'Mapping.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')
        image_path2 = os.path.join(os.getcwd(), 'Static', 'Mapping.png')

    root.title("Process File Picker")
    root.resizable(False,False)
   
    w = 500
    h = 160
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)    
    
    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="Browse Process - File Picker",font=("Calibri",18,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack()
          
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=500,height=50)
    
    answer=StringVar()
    answer.set("")
       
    def browse_button():
        global filename
        answer.set("")  
        filename = filedialog.askopenfilename()  
        txt.delete(0, 'end') 
        txt.insert(0, filename)

    def Click_Done():
        global ans        
        
        ans=txt.get()        
        
        if ans=="":
           answer.set("File Path Fields Empty Is Not Allowed...")        
        else:      
            root.destroy()          
            return ans
        
    txt=Entry(Frame2,font=("Calibri",12,"bold","italic"),width=50,justify="left")
    txt.grid(row=0,column=0,padx=5,pady=10,sticky="E")    

    photo1 = PhotoImage(file=image_path2)
            
    btn1=Button(Frame2,text="Browse",command=browse_button,image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=0,column=1,padx=3,pady=0,sticky="W")     
                
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=3,y=80,width=490,height=25)
    
    title_3=Label(Frame3,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=1,padx=150,pady=0,sticky="E")

    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=3,y=105,width=500,height=50)
    
    photo2 = PhotoImage(file=image_path1)

    btn2=Button(Frame4,command=Click_Done,text="Done",image=photo2,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=3,column=0,padx=125,pady=0,sticky="W")

    def Close():
        sys.exit(0)   
    
    photo3 = PhotoImage(file=image_path)   

    btn3=Button(Frame4,command=Close,text="Exit",image=photo3,borderwidth=0,bg="#2c3e50")
    btn3.grid(row=3,column=1,padx=15,pady=0,sticky="W")

    def disable_event():
        pass

    myTip = Hovertip(btn2,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn3,'Click to Exit Process',hover_delay=1000)
    myTip2 = Hovertip(btn1,'Click to Pick File',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)

    root.mainloop()

def user_pass():
    global ans1
    global ans2
    root=Tk()

    if getattr(sys, 'frozen', False):       
        image_path = os.path.join(sys._MEIPASS, 'Static', 'Close.png')
        image_path1 = os.path.join(sys._MEIPASS, 'Static', 'Mapping1.png')
    else:
        image_path = os.path.join(os.getcwd(), 'Static', 'Close.png')
        image_path1 = os.path.join(os.getcwd(), 'Static', 'Mapping1.png')

    root.title("ERA - User Login And File Details")
    root.resizable(False,False)

    w = 600
    h = 220
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    root.config(bg="#2c3e50",highlightbackground="blue",highlightthickness=1)

    Frame1=Frame(root,bg="gold")
    Frame1.pack(side=TOP,fill=BOTH)
    title=Label(Frame1,text="User Name, Password and Date Details...",font=("Calibri",20,"bold","italic"),bg="gold",fg="black",justify="center")
    title.grid(row=0,columnspan=2,padx=8,pady=8)
    title.pack() 
 
    Frame2=Frame(root,bg="#2c3e50")
    Frame2.place(x=0,y=40,width=698,height=150)    
    
    title3=Label(Frame2,text="User Name :",font=("Calibri",11,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title3.grid(row=0,column=0,padx=5,pady=5,sticky="W")
    
    txt1=Entry(Frame2,font=("Calibri",11,"bold","italic"),width=50,justify="left")
    txt1.grid(row=0,column=1,padx=30,pady=5,sticky="W")
    
    title4=Label(Frame2,text="Password :",font=("Calibri",11,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title4.grid(row=1,column=0,padx=5,pady=5,sticky="W")
    
    txt2=Entry(Frame2,font=("Calibri",11,"bold","italic"),width=50,justify="left")
    txt2.grid(row=1,column=1,padx=30,pady=5,sticky="W")

    title8=Label(Frame2,text="From Date :",font=("Calibri",11,"bold","italic"),bg="#2c3e50",fg="white",justify=LEFT)
    title8.grid(row=2,column=0,padx=5,pady=5,sticky="W")
    
    cal = DateEntry(Frame2, selectmode='day',date_pattern='mm/dd/yyyy',textvariable="",width=10,justify="center",font=("Calibri",12,"bold","italic"),bg="aquamarine")
    cal.grid(row=2,column=1,padx=30,pady=5,sticky="W")

    options = ["New", "Dupilcate"]
    combo_box = ttk.Combobox(Frame2, values=options,textvariable="",width=10,justify="center",font=("Calibri",12,"bold","italic"))
    combo_box.grid(row=2,column=1,padx=150,pady=5,sticky="W")

    answer=StringVar()
    answer.set("")

    def Click_Done():
        global ans1
        global ans2        
        global ans5
        global ans6

        ans1=txt1.get()
        ans2=txt2.get()        
        ans5=cal.get()
        ans6=combo_box.get()

        if ans1=="":
           answer.set("User Name Field Empty Is Not Allowed...")
        elif ans2=="":
            answer.set("Password Field Empty Is Not Allowed...")   
        elif ans5=="":
            answer.set("From Date Empty Is Not Allowed...")
        elif ans6=="":
            answer.set("Dropdown Empty Is Not Allowed...")
        else:    
            root.destroy()
            return ans1,ans2,ans5,ans6
                
    Frame4=Frame(root,bg="#2c3e50")
    Frame4.place(x=0,y=145,width=698,height=20)
    
    title_3=Label(Frame4,text=answer.get(),textvariable=answer,font=("Calibri",9,"bold","italic"),bg="#2c3e50",fg="Red",justify=LEFT)
    title_3.grid(row=0,column=0,columnspan=2,padx=200,pady=0,sticky="E")
    
    Frame3=Frame(root,bg="#2c3e50")
    Frame3.place(x=0,y=165,width=698,height=200)
       
    photo1 = PhotoImage(file=image_path1)
    
    btn1=Button(Frame3,command=Click_Done,text="Run",image=photo1,borderwidth=0,bg="#2c3e50")
    btn1.grid(row=2,column=0,padx=140,pady=0,sticky="W")

    def Close():
        sys.exit(0)

    photo = PhotoImage(file=image_path)    

    btn2=Button(Frame3,command=Close,text="Close",image=photo,borderwidth=0,bg="#2c3e50")
    btn2.grid(row=2,column=1,padx=100,pady=0,sticky="E")

    def disable_event():
        pass

    txt1.focus_set()

    myTip = Hovertip(btn1,'Click to Done Continue Process',hover_delay=1000)
    myTip1 = Hovertip(btn2,'Click to Exit Process',hover_delay=1000)

    root.protocol("WM_DELETE_WINDOW", disable_event)
            
    root.mainloop()

def process():        
        
        global element_1    
        global ans1
        global ans2 
        global fs
        global ans4
        global ans5
        global ans6
        global driver_path
        # Sports Shoes for Men | Shock Absorbant, Slip Resistant Walking Shoes For Men  (Black)
        # SAMSUNG Galaxy A73 5G (Awesome Gray, 128 GB)  (8 GB RAM)
        # COIRFIT BIOLIFE 7-Zone 100% Natural 8 inch King Latex Mattress| Talalay Technology|Cool Gel Mattress |10 Years Warranty, (78x72x8 inches)
        # RUSSO 100% Pure Natural Latex & Memory Foam Mattress (78 x 60 x 6 Inch)
        # SmartGRID Technology & 100% Natural Latex| 8 inch Queen Latex Foam Mattress  (L x W: 78 inch x 60 inch)
        # i select Filpkart - The Sleep Company Luxe Royale- Pro Blend of SmartGRID Technology & 100% Natural Latex| 8 inch Queen Latex Foam Mattress  (L x W: 78 inch x 60 inch)
        # https://www.woodenstreet.com/walken-bed-with-box-storage-queen-honey-finish        
        user_pass()
        file_pick()

        user_name =ans1
        password=ans2        
        fil=ans       
        dat=ans5
        nw_dup=ans6

        folder_create = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Downloads/')
        check_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Downloads/Temp')
        down_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Downloads\\Temp\\')
        file = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Downloads\\Temp\\Files\\')
        temp_file = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Downloads\\Temp\\Files\\')
        file_835 = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Downloads\\Temp\\Files\\835\\')
        file_txt = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Downloads\\Temp\\Files\\Text\\')
        file_zip = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Downloads\\Temp\\Files\\Download\\')

        fld='Temp'
        
        if not os.path.isdir(check_path):
            os.mkdir(folder_create + 'Temp')
            os.mkdir(down_path + 'Files')
            os.mkdir(file + '835')
            os.mkdir(file + 'Text')
            os.mkdir(file + 'Download')
        else:
            path1 = os.path.join(folder_create, fld)
            shutil.rmtree(path1)
            os.mkdir(folder_create + 'Temp')
            os.mkdir(down_path + 'Files')  
            os.mkdir(file + '835')
            os.mkdir(file + 'Text')
            os.mkdir(file + 'Download')

        wb3 = openpyxl.Workbook()
        ws3=wb3.active    
        ws3['A1']='File Names'
        ws3.title = 'Data'
        wb3.save(filename=temp_file + 'Error File Names.xlsx')
        wb3.close()

        try:
            options = Options()            
            options.add_argument('--ignore-certificate-errors')
            options.add_argument('--ignore-ssl-errors')
            prefs = {"download.default_directory" : down_path,
                    "download.prompt_for_download": False,
                    "download.directory_upgrade": True,
                    "safebrowsing.enabled": True}
            options.add_experimental_option("prefs",prefs)
            driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=options)
            driver.maximize_window()
            driver.get('https://cms.officeally.com/')            
        except Exception as e:
            try:
                options =  webdriver.ChromeOptions()        
                options.add_argument('--ignore-certificate-errors')
                options.add_argument('--ignore-ssl-errors')
                options.add_argument("--disable-popup-blocking")
                prefs = {"download.default_directory" : down_path,
                        "download.prompt_for_download": False,
                        "download.directory_upgrade": True,
                        "safebrowsing.enabled": True}
                options.add_experimental_option("prefs",prefs)                
                driver_path = os.path.abspath('chromedriver.exe')                
                driver = webdriver.Chrome(executable_path=driver_path,options=options)                
                driver.maximize_window()
                driver.get('https://cms.officeally.com/')                                       
            except Exception as e:
                messagebox.showinfo("Driver Problem","Pls Check Your Driver Version")
                sys.exit(0)

        
        def text_box(xpath,heding,status,key):                
            counter = 0
            while counter < 15:
                try:   
                    WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).clear()                   
                    WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath))).send_keys(key)
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding, status)
                sys.exit(0)        
                
        def click(xpath,heding,status):
            counter = 0
            while counter < 15:
                try:             
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding,status)
                sys.exit(0)                                

        def count(xpath,heding,status):
            global rows
            counter = 0
            while counter < 15:
                try:             
                    rows=len(WebDriverWait(driver, 0).until(EC.presence_of_all_elements_located((By.XPATH,xpath))))                    
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo(heding,status)
                sys.exit(0)                         

        def Alert():
            counter = 0
            while counter < 15:
                try:             
                    WebDriverWait(driver, 0).until (EC.alert_is_present())
                    a=driver.switch_to.alert
                    a.accept()
                    break
                except Exception as e:
                    time.sleep(1)
                    counter += 1
            else:
                messagebox.showinfo('Alert','Alert Not Present')
            
        xpath= "/html/body/div[7]/div[1]/div[2]/div[1]/div[2]/div[1]/div"
        heding="Login Button"
        status="Login Button Not Found"
        click(xpath,heding,status)
                      
        xpath= "/html/body/div[7]/div[1]/div[2]/div[1]/div[2]/div[1]/nav/div/div/a"
        heding="Login Count"
        status="Login Count Field Not Found"        
        count(xpath,heding,status)
              
        xpath= "/html/body/div[7]/div[1]/div[2]/div[1]/div[2]/div[1]/nav/div/div/a[{}]"
        
        j=1
        while j<rows+1:                                
            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text
            parts = cnm.split('\n')
            cnm = parts[0]
            if cnm=="Service Center":
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                            
                break
            j=j+1       
                       
        driver.switch_to.window(driver.window_handles[1])        

        xpath= '/html/body/main/section/div/div/div/form/div[2]/div/div[1]/div/input'
        heding="User Name"
        status="User Name Field Not Found"
        key=user_name
        text_box(xpath,heding,status,key)
        
        xpath= '/html/body/main/section/div/div/div/form/div[2]/div/div[2]/div/input'
        heding="Password"
        status="Password Field Not Found"
        key=password
        text_box(xpath,heding,status,key)

        xpath= "/html/body/main/section/div/div/div/form/div[3]/button"
        heding="Continue"
        status="Continue Button Not Found"
        click(xpath,heding,status)
        
        while True:
            xpath='/html/body/main/section/div/div/div/form/div[2]/div/div[2]/span'         
            try:
                element_3 = WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH, xpath))).text
                if element_3.lstrip().rstrip()=='Wrong username or password':                               
                    messagebox.showinfo('Login Status','Please Check - UserName or Password Wrong')                    
                    user_pass()
                    
                    xpath= '/html/body/main/section/div/div/div/form/div[2]/div/div[1]/div/input'
                    heding="User Name"
                    status="User Name Field Not Found"
                    key=ans1
                    text_box(xpath,heding,status,key)
                    
                    xpath= '/html/body/main/section/div/div/div/form/div[2]/div/div[2]/div/input'
                    heding="Password"
                    status="Password Field Not Found"
                    key=ans2
                    text_box(xpath,heding,status,key)

                    xpath= "/html/body/main/section/div/div/div/form/div[3]/button"
                    heding="Continue"
                    status="Continue Button Not Found"
                    click(xpath,heding,status)
                else:
                    break    
            except Exception as e:                
                    break                                 
        
        xpath= "/html/body/app-root/div[1]/oa-layout/section/header/oa-navigation/div/div[3]/button"
        heding="Return to Classic"
        status="Return to Classic Button Not Found"
        click(xpath,heding,status)
        
        xpath= "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td[1]/table/tbody/tr[2]/td[2]/table/tbody/tr"
        heding="Request Count"
        status="Request Count Not Found"        
        count(xpath,heding,status)      
       
        j=1
        while j<rows+1:
            xpath= "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td[1]/table/tbody/tr[2]/td[2]/table/tbody/tr[{}]"                       
            cnm=WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH,xpath.format(j)))).text              
            if cnm.lstrip().rstrip()=="Download EOB / ERA 835":                
                xpath= "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td[1]/table/tbody/tr[2]/td[2]/table/tbody/tr[{}]/td/a"
                WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(j)))).click()                            
                break
            j=j+1    
        
        # current_date_time = datetime.now()
        # one_day_ago = current_date_time - timedelta(days=1)
        # formatted_date = one_day_ago.strftime("%Y-%m-%d %H:%M:%S")
        # date_object = datetime.strptime(formatted_date, "%Y-%m-%d %H:%M:%S")
        original_datetime = datetime.strptime(dat, '%m/%d/%Y')
        dob2 = original_datetime.strftime('%#m/%#d/%Y')

        # dob2 = dat.strftime("%#m/%#d/%Y")
        lst=dob2.split('/')
        mn=lst[0]
        dt=lst[1]
        yy=lst[2]               

        # dt='10'
        # dob2 = '9/10/2023'

        xpath= '/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td[5]/form/table/tbody/tr/td[3]/table[2]/tbody/tr[2]/td[3]/table[1]/tbody/tr/td/input[1]'
        heding="Month"
        status="Month Field Not Found"
        key=mn
        text_box(xpath,heding,status,key)

        xpath= '/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td[5]/form/table/tbody/tr/td[3]/table[2]/tbody/tr[2]/td[3]/table[1]/tbody/tr/td/input[2]'
        heding="Date"
        status="Date Field Not Found"
        key=dt
        text_box(xpath,heding,status,key)

        xpath= '/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td[5]/form/table/tbody/tr/td[3]/table[2]/tbody/tr[2]/td[3]/table[1]/tbody/tr/td/input[3]'
        heding="Year"
        status="Year Field Not Found"
        key=yy
        text_box(xpath,heding,status,key)

        xpath= "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td[5]/form/table/tbody/tr/td[3]/table[2]/tbody/tr[3]/td[2]/input"
        heding="Go Button"
        status="Go Button Not Found"
        click(xpath,heding,status)        

        counter = 0
        while counter < 15:
            try:
                xpath= '/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td[5]/form/div[1]/table/tbody/tr[1]/td/b'
                ck_dt = WebDriverWait(driver, 0).until(EC.visibility_of_element_located((By.XPATH, xpath))).text
                if dob2 in ck_dt:
                    break
            except Exception as e:
                pass
        else:
            messagebox.showinfo('Date Check','Date Not Matched')
            sys.exit(0)
                     
        xpath= "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td[5]/form/div[1]/table/tbody/tr[2]/td/table/tbody/tr"
        heding="EOB Table Count"
        status="EOB Table Count Not Found"        
        count(xpath,heding,status)      
        
        wb1=load_workbook(fil)

        n=2
        while n<rows+1:                      
            xpath='/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td[5]/form/div[1]/table/tbody/tr[2]/td/table/tbody/tr[{}]/td[7]'                                   
            element = WebDriverWait(driver, 0).until(EC.presence_of_element_located((By.XPATH,xpath.format(n))))                                     
            bgcolor = element.value_of_css_property("BACKGROUND-COLOR")
            rgb_values = bgcolor[4:-1].split(", ")
            bgcolor = "#{:02x}{:02x}{:02x}".format(int(rgb_values[0]), int(rgb_values[1]), int(rgb_values[2]))
            
            # #ccccff - Vilote
            # #ffb9b9 - Red
            if nw_dup == 'New':
                if bgcolor.lstrip().rstrip()=='#ffb9b9':
                    xpath= "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td[5]/form/div[1]/table/tbody/tr[2]/td/table/tbody/tr[{}]/td[7]/a"
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(n)))).click()                 

                    counter=1
                    while counter < 10:
                        file1 = [f for f in listdir(down_path) if isfile(join(down_path, f))]
                        lol_string1 = ' '.join(file1)
                        lol_string2=lol_string1.split('.')[-1]   
                        # print(lol_string2)                                                                             
                        if lol_string2!='' and lol_string2!='tmp' and lol_string2 !="crdownload" and lol_string2 !="temp":
                            time.sleep(1)
                            file2 = [f for f in os.listdir(down_path) if os.path.isfile(os.path.join(down_path, f))]
                            if file2:
                                lol_string3 = os.path.join(down_path, file2[0])
                                with zipfile.ZipFile(lol_string3, 'r') as zip_ref:                            
                                    zip_ref.extractall(down_path)
                                
                                files = os.listdir(down_path)
                                text_files = [file for file in files if file.endswith('.txt')]
                                
                                for text_file in text_files:
                                    file_path = os.path.join(down_path, text_file)
                                    delimiter = '\t'
                                    df = pd.read_csv(file_path, delimiter=delimiter)
                                    df = df.rename(columns={df.columns[0]: 'temp'})                            
                                                                
                                    indices_to_remove = df[df['temp'].str.contains('={30,}|-{30,}|^\s*$')].index
                                    df = df.drop(indices_to_remove)        
                                    check_indices = df[df['temp'].str.contains('Check#')].index
                                    
                                    fis=check_indices[0]
                                    hed=check_indices[1]
                                    hed_row = df.loc[fis : hed,'temp']
                                    
                                    hed_row = pd.DataFrame(hed_row)
                                    hed_row = hed_row.reset_index(drop=True)         
                                    hed_row_ck = hed_row[hed_row['temp'].str.contains('Check#')].index  
                                    
                                    lst = []
                                    lst1 = []

                                    if len(hed_row) == 3:
                                        data_row = hed_row.loc[1,'temp']
                                        date_pattern = r'\d{2}/\d{2}/\d{4}'
                                        matches = re.findall(date_pattern, data_row)
                                        if matches:
                                            date = matches[0]
                                        else:
                                            date = 'N/A'                                                                             
                                        
                                        split_items = re.split(r'\s{2,}', data_row)
                                        split_items_raw = [item for item in split_items if item]            
                                        chk_all = split_items_raw[0]                                              
                                        matches = re.findall(r'\s', chk_all)
                                        if matches:
                                            split_string = chk_all.split(' ')
                                            split_items_raw.pop(0)
                                            ck=split_string[0]
                                            amt=split_string[1]
                                            split_items_raw.insert(0, ck)
                                            split_items_raw.insert(1, amt)                        
                                        indices_to_select = [0, 1, 3]
                                        split_items_raw1 = [split_items_raw[i] for i in indices_to_select]
                                        split_items_raw1.append(str(date))                                     

                                        lst1.append(split_items_raw)    
                                        lst.append(split_items_raw1)
                                    else:
                                        j=hed_row_ck[0].item()            
                                        while j<len(hed_row)-2:               
                                            data_row = hed_row.loc[j + 1,'temp']
                                            date_pattern = r'\d{2}/\d{2}/\d{4}'
                                            matches = re.findall(date_pattern, data_row)
                                            if matches:
                                                date = matches[0]                                        
                                            else:
                                                date = 'N/A'                                                                             
                                            
                                            split_items = re.split(r'\s{2,}', data_row)
                                            split_items_raw = [item for item in split_items if item]
                                            chk_all = split_items_raw[0]                                              
                                            matches = re.findall(r'\s', chk_all)
                                            if matches:
                                                split_string = chk_all.split(' ')
                                                split_items_raw.pop(0)
                                                ck=split_string[0]
                                                amt=split_string[1]
                                                split_items_raw.insert(0, ck)
                                                split_items_raw.insert(1, amt)                        
                                            indices_to_select = [0, 1, 3]
                                            split_items_raw1 = [split_items_raw[i] for i in indices_to_select]
                                            split_items_raw1.append(str(date)) 

                                            indices_to_select = [0, 1, 3]
                                            split_items_raw1 = [split_items_raw[i] for i in indices_to_select]
                                            split_items_raw1.append(str(date))  
                                                                                    
                                            lst1.append(split_items_raw)    
                                            lst.append(split_items_raw1)            

                                            j=j+1   
                                                    
                                    i=1                                                  
                                    while i<len(check_indices):                                
                                                    
                                        check_index = check_indices[i]
                                        data_row = df.loc[check_index + 1,'temp']
                                        split_items = re.split(r'\s{2,}', data_row)
                                        split_items = [item for item in split_items if item]
                                        
                                        chk_all = split_items[0]                                              
                                        matches = re.findall(r'\s', chk_all)
                                        if matches:
                                            split_string = chk_all.split(' ')
                                            split_items.pop(0)
                                            ck=split_string[0]
                                            amt=split_string[1]
                                            split_items.insert(0, ck)
                                            split_items.insert(1, amt)
                                                                            
                                        chk_num=split_items[0]    
                                        chk_all = split_items[1]            
                                    
                                        matches = re.findall(r'\b\d+\.\d{2}\b', chk_all)            
                                        matches1 = chk_all.replace(' ', '')            

                                        if matches:
                                            split_items.insert(1, 'N/A')
                                        else:
                                            if any(char == ',' for char in matches1):
                                                split_items.insert(1, 'N/A')                

                                        split_items.pop(0)
                                        
                                        input_string=split_items[1]
                                        number_found = any(char.isdigit() for char in input_string)
                                        if number_found:
                                            index_of_first_digit = next((i for i, char in enumerate(input_string) if char.isdigit()), None)
                                            if index_of_first_digit is not None:
                                                nm = input_string[:index_of_first_digit].strip()  
                                                val = input_string[index_of_first_digit:].strip()  
                                                split_items.pop(1)
                                                split_items.insert(1, nm)
                                                split_items.insert(2, val)

                                        input_string=split_items[4]
                                        alphabetic_found = any(char.isalpha() and char != ' ' for char in input_string)
                                        if alphabetic_found:
                                            parts = input_string.split(' ')
                                            nm = ''
                                            val = ''
                                            for part in parts:
                                                if part.replace('-', '').isdigit():
                                                    val += part + ' '
                                                else:
                                                    nm += part + ' '
                                            nm = nm.strip()
                                            val = val.strip()
                                            # nm = ''.join(char for char in input_string if char.isalpha() or char == '-')
                                            # val = ''.join(char for char in input_string if char.isdigit() or char == '.') 
                                            split_items.pop(4)
                                            split_items.insert(4, val)
                                            split_items.insert(5, nm)
                                        
                                        z=6         
                                        joined_string = ''                                         
                                        while z<len(split_items):   
                                            input_string1=split_items[z]
                                            joined_string += input_string1 + ' '
                                            z += 1
                                        
                                        z=6                                                  
                                        while z<len(split_items) + 1 :   
                                            split_items.pop(6)
                                            z += 1

                                        split_items.insert(len(split_items), joined_string.strip())
                                        
                                        try:
                                            sel=check_indices[i + 1]
                                            sel_row = df.loc[check_index : sel,'temp']

                                            tm_df=pd.DataFrame(sel_row)
                                            tm_df = tm_df.reset_index(drop=True)
                                                                                
                                            pay_calm = tm_df[tm_df['temp'].str.contains('Payer Claim Control Number:')]
                                            pay_calm = '\n'.join(pay_calm.iloc[:, 0])
                                            pay_calm = pay_calm.strip().split(':')[1].strip()
                                            split_items.append(str(pay_calm))
                                        except Exception as e:
                                            last_index = df.index[-1]
                                            sel_row = df.loc[check_index : last_index,'temp']

                                            tm_df=pd.DataFrame(sel_row)
                                            tm_df = tm_df.reset_index(drop=True)
                                                                                
                                            pay_calm = tm_df[tm_df['temp'].str.contains('Payer Claim Control Number:')]
                                            pay_calm = '\n'.join(pay_calm.iloc[:, 0])
                                            pay_calm = pay_calm.strip().split(':')[1].strip()
                                            split_items.append(str(pay_calm))
                                        
                                        lin_itms = tm_df[tm_df['temp'].str.contains('Line Item:')].index                                
                                        
                                        lst7=[]
                                        j=0                                                  
                                        while j<len(lin_itms):    
                                            lin_itm = lin_itms[j]
                                            data_row = tm_df.loc[lin_itm + 1,'temp']
                                            line_items = re.split(r'\s{2,}', data_row)
                                            line_items = [item for item in line_items if item]
                                            
                                            dt_sp = line_items[0]
                                            matches = re.findall(r'\b\d+\.\d{2}\b', dt_sp) 
                                            if matches:
                                                dt_sp1 = dt_sp.strip().split(' ')[0].strip()
                                                cpt_sp = dt_sp.strip().split(' ')[1].strip()
                                                chr_amt = dt_sp.strip().split(' ')[2].strip()

                                                line_items.pop(0)
                                                line_items.insert(0, dt_sp1)
                                                line_items.insert(1, cpt_sp)
                                                line_items.insert(2, chr_amt)
                                            else:
                                                dt_sp1 = dt_sp.strip().split(' ')[0].strip()
                                                cpt_sp = dt_sp.strip().split(' ')[1].strip()

                                                line_items.pop(0)
                                                line_items.insert(0, dt_sp1)
                                                line_items.insert(1, cpt_sp)
                                            
                                            k=2
                                            adj = tm_df.loc[lin_itm + k,'temp']

                                            while 'Line Item:' not in adj:                                                     
                                                lines_items = adj.strip()                                        
                                                line_items.append(lines_items)                    
                                                
                                                k=k+1  
                                                try:
                                                    adj = tm_df.loc[lin_itm + k,'temp']
                                                    if 'Check#' in adj:
                                                        break
                                                except Exception as e:
                                                    break
                                            lst7.append(line_items)
                                            j=j+1
                                        
                                        lst_cnt=len(lst7)

                                        l=0
                                        while l < lst_cnt:                
                                            fin_lst = None
                                            for sublist in lst:
                                                if chk_num in sublist:
                                                    fin_lst = sublist
                                                    break                
                                                        
                                            fin_lst_itm = lst7[l]
                                                
                                            # wb1=load_workbook(fil)
                                            sheet = wb1['CPT']
                                            column_letter = 'A'  
                                            column_cells = sheet[column_letter]
                                            last_row = None
                                            for cell in reversed(column_cells):
                                                if cell.value:
                                                    last_row = cell.row
                                                    break              
                                            start_column = 'C'
                                            current_column_index = openpyxl.utils.column_index_from_string(start_column)
                                            current_row = last_row + 1

                                            for value in fin_lst:
                                                current_column = openpyxl.utils.get_column_letter(current_column_index)
                                                sheet[current_column + str(current_row)] = value
                                                current_column_index += 1                
                                            
                                            start_column = 'G'
                                            current_column_index = openpyxl.utils.column_index_from_string(start_column)                                   

                                            for value in split_items:
                                                current_column = openpyxl.utils.get_column_letter(current_column_index)
                                                sheet[current_column + str(current_row)] = value
                                                current_column_index += 1 

                                            start_column = 'O'
                                            current_column_index = openpyxl.utils.column_index_from_string(start_column)                                   

                                            for value in fin_lst_itm:
                                                current_column = openpyxl.utils.get_column_letter(current_column_index)
                                                sheet[current_column + str(current_row)] = value
                                                current_column_index += 1 
                                            
                                            sheet['B' + str(int(last_row + 1))]=text_file
                                            sheet['A' + str(int(last_row + 1))]='Done'
                                            # wb1.save(filename=fil)
                                            # wb1.close()

                                            l=l+1            
                                        i=i+1
                                    
                                    pyr=split_items[-2]

                                    fin_lst11=[]
                                    date_pattern = r'\b\d{2}/\d{2}/\d{4}\b'

                                    for i in range(len(lst1)):
                                        date = None
                                        for item in lst1[i]:
                                            date_match = re.search(date_pattern, item)
                                            if date_match:
                                                date = date_match.group(0)        
                                        filtered_list = [re.sub(date_pattern, '', item) for item in lst1[i]]
                                        joined_string = ' '.join(filtered_list[4:])

                                        updated_sublist = filtered_list[:4] + [joined_string.strip(), date.strip()]

                                        fin_lst11.append(updated_sublist)     

                                    for i in range(len(fin_lst11)):
                                        fin_lst11[i].append(pyr)     
                                    
                                    # for i in range(len(lst1)):
                                    #     input_string=lst1[i][4]
                                    #     date_pattern = r'\b\d{2}/\d{2}/\d{4}\b'
                                    #     date_matches = re.findall(date_pattern, input_string)
                                    #     if len(date_matches) > 0:
                                    #         date = date_matches[0].strip()
                                    #         text_parts = re.split(date_pattern, input_string)
                                    #         text = ''.join(text_parts).strip()
                                    #         lst1[i].pop(4)
                                    #         lst1[i].insert(4,text)
                                    #         lst1[i].insert(5,date)      

                                    for i, sub_list in enumerate(fin_lst11):           
                                        sheet = wb1['Raw']
                                        column_letter = 'I'  
                                        column_cells = sheet[column_letter]
                                        last_row = None
                                        for cell in reversed(column_cells):
                                            if cell.value:
                                                last_row = cell.row
                                                break              
                                        start_column = 'B'
                                        current_column_index = openpyxl.utils.column_index_from_string(start_column)
                                        current_row = last_row + 1
                                        
                                        for value in sub_list:
                                            current_column = openpyxl.utils.get_column_letter(current_column_index)
                                            sheet[current_column + str(current_row)] = value
                                            current_column_index += 1                        

                                        sheet['A' + str(int(last_row + 1))]=text_file
                                        sheet['I' + str(int(last_row + 1))]='Done'
                                        
                                    wb1.save(filename=fil)
                                    
                                    text_files = [file for file in files if file.endswith('.txt')]
                                    zip_files = [file for file in files if file.endswith('.zip')]                            
                                    raw_files = [file for file in files if file.endswith('.835')]                            

                                    # shutil.move(os.path.abspath(down_path + zip_files),os.path.abspath(file_zip + zip_files))
                                    
                                    for text_file in text_files:
                                        source_path = os.path.join(down_path, text_file)
                                        destination_path = os.path.join(file_txt, text_file)
                                        shutil.move(source_path, destination_path)                                                        
                                    
                                    for zip_file in zip_files:
                                        source_path = os.path.join(down_path, zip_file)
                                        destination_path = os.path.join(file_zip, zip_file)                                
                                        shutil.move(source_path, destination_path)
                                                                    
                                    for raw_file in raw_files:
                                        source_path = os.path.join(down_path, raw_file)
                                        destination_path = os.path.join(file_835, raw_file)
                                        shutil.move(source_path, destination_path)

                                    files = os.listdir(down_path)

                                    for file in files:
                                        file_full_path = os.path.join(file, file)
                                        if os.path.isfile(file_full_path):
                                            os.remove(file_full_path)    
                                break
                        else:
                            time.sleep(1)
                            counter += 1
                    else:
                        files = os.listdir(down_path)

                        for file in files:
                            file_full_path = os.path.join(file, file)
                            if os.path.isfile(file_full_path):
                                os.remove(file_full_path)          
                        
                        wb1=load_workbook(filename=temp_file + 'Error File Names.xlsx')
                        sheet = wb1['Data']
                        last_row=sheet.max_row
                        sheet['A' + str(int(last_row + 1))]=file                    
                        wb1.save(temp_file + 'Error File Names.xlsx')
                        wb1.close()             
            elif nw_dup==  'Dupilcate':
                if bgcolor.lstrip().rstrip()=='#ccccff':
                    xpath= "/html/body/table[2]/tbody/tr[1]/td/table/tbody/tr/td[5]/form/div[1]/table/tbody/tr[2]/td/table/tbody/tr[{}]/td[7]/a"
                    WebDriverWait(driver, 0).until(EC.element_to_be_clickable((By.XPATH,xpath.format(n)))).click()                            

                    counter=1
                    while counter < 10:
                        file1 = [f for f in listdir(down_path) if isfile(join(down_path, f))]
                        lol_string1 = ' '.join(file1)
                        lol_string2=lol_string1.split('.')[-1]   
                        # print(lol_string2)                                                                             
                        if lol_string2!='' and lol_string2!='tmp' and lol_string2 !="crdownload" and lol_string2 !="temp":
                            time.sleep(1)
                            file2 = [f for f in os.listdir(down_path) if os.path.isfile(os.path.join(down_path, f))]
                            if file2:
                                lol_string3 = os.path.join(down_path, file2[0])
                                with zipfile.ZipFile(lol_string3, 'r') as zip_ref:                            
                                    zip_ref.extractall(down_path)
                                
                                files = os.listdir(down_path)
                                text_files = [file for file in files if file.endswith('.txt')]
                                
                                for text_file in text_files:
                                    file_path = os.path.join(down_path, text_file)
                                    delimiter = '\t'
                                    df = pd.read_csv(file_path, delimiter=delimiter)
                                    df = df.rename(columns={df.columns[0]: 'temp'})                            
                                                                
                                    indices_to_remove = df[df['temp'].str.contains('={30,}|-{30,}|^\s*$')].index
                                    df = df.drop(indices_to_remove)        
                                    check_indices = df[df['temp'].str.contains('Check#')].index
                                    
                                    fis=check_indices[0]
                                    hed=check_indices[1]
                                    hed_row = df.loc[fis : hed,'temp']
                                    
                                    hed_row = pd.DataFrame(hed_row)
                                    hed_row = hed_row.reset_index(drop=True)         
                                    hed_row_ck = hed_row[hed_row['temp'].str.contains('Check#')].index  
                                    
                                    lst = []
                                    lst1 = []

                                    if len(hed_row) == 3:
                                        data_row = hed_row.loc[1,'temp']
                                        date_pattern = r'\d{2}/\d{2}/\d{4}'
                                        matches = re.findall(date_pattern, data_row)
                                        if matches:
                                            date = matches[0]
                                        else:
                                            date = 'N/A'                                                                             
                                        
                                        split_items = re.split(r'\s{2,}', data_row)
                                        split_items_raw = [item for item in split_items if item]            
                                        chk_all = split_items_raw[0]                                              
                                        matches = re.findall(r'\s', chk_all)
                                        if matches:
                                            split_string = chk_all.split(' ')
                                            split_items_raw.pop(0)
                                            ck=split_string[0]
                                            amt=split_string[1]
                                            split_items_raw.insert(0, ck)
                                            split_items_raw.insert(1, amt)                        
                                        indices_to_select = [0, 1, 3]
                                        split_items_raw1 = [split_items_raw[i] for i in indices_to_select]
                                        split_items_raw1.append(str(date))                                     

                                        lst1.append(split_items_raw)    
                                        lst.append(split_items_raw1)
                                    else:
                                        j=hed_row_ck[0].item()            
                                        while j<len(hed_row)-2:               
                                            data_row = hed_row.loc[j + 1,'temp']
                                            date_pattern = r'\d{2}/\d{2}/\d{4}'
                                            matches = re.findall(date_pattern, data_row)
                                            if matches:
                                                date = matches[0]                                        
                                            else:
                                                date = 'N/A'                                                                             
                                            
                                            split_items = re.split(r'\s{2,}', data_row)
                                            split_items_raw = [item for item in split_items if item]
                                            chk_all = split_items_raw[0]                                              
                                            matches = re.findall(r'\s', chk_all)
                                            if matches:
                                                split_string = chk_all.split(' ')
                                                split_items_raw.pop(0)
                                                ck=split_string[0]
                                                amt=split_string[1]
                                                split_items_raw.insert(0, ck)
                                                split_items_raw.insert(1, amt)                        
                                            indices_to_select = [0, 1, 3]
                                            split_items_raw1 = [split_items_raw[i] for i in indices_to_select]
                                            split_items_raw1.append(str(date)) 

                                            indices_to_select = [0, 1, 3]
                                            split_items_raw1 = [split_items_raw[i] for i in indices_to_select]
                                            split_items_raw1.append(str(date))  
                                                                                    
                                            lst1.append(split_items_raw)    
                                            lst.append(split_items_raw1)            

                                            j=j+1   
                                                    
                                    i=1                                                  
                                    while i<len(check_indices):                                
                                                    
                                        check_index = check_indices[i]
                                        data_row = df.loc[check_index + 1,'temp']
                                        split_items = re.split(r'\s{2,}', data_row)
                                        split_items = [item for item in split_items if item]
                                        
                                        chk_all = split_items[0]                                              
                                        matches = re.findall(r'\s', chk_all)
                                        if matches:
                                            split_string = chk_all.split(' ')
                                            split_items.pop(0)
                                            ck=split_string[0]
                                            amt=split_string[1]
                                            split_items.insert(0, ck)
                                            split_items.insert(1, amt)
                                                                            
                                        chk_num=split_items[0]    
                                        chk_all = split_items[1]            
                                    
                                        matches = re.findall(r'\b\d+\.\d{2}\b', chk_all)            
                                        matches1 = chk_all.replace(' ', '')            

                                        if matches:
                                            split_items.insert(1, 'N/A')
                                        else:
                                            if any(char == ',' for char in matches1):
                                                split_items.insert(1, 'N/A')                

                                        split_items.pop(0)
                                        
                                        input_string=split_items[1]
                                        number_found = any(char.isdigit() for char in input_string)
                                        if number_found:
                                            index_of_first_digit = next((i for i, char in enumerate(input_string) if char.isdigit()), None)
                                            if index_of_first_digit is not None:
                                                nm = input_string[:index_of_first_digit].strip()  
                                                val = input_string[index_of_first_digit:].strip()  
                                                split_items.pop(1)
                                                split_items.insert(1, nm)
                                                split_items.insert(2, val)

                                        input_string=split_items[4]
                                        alphabetic_found = any(char.isalpha() and char != ' ' for char in input_string)
                                        if alphabetic_found:
                                            parts = input_string.split(' ')
                                            nm = ''
                                            val = ''
                                            for part in parts:
                                                if part.replace('-', '').isdigit():
                                                    val += part + ' '
                                                else:
                                                    nm += part + ' '
                                            nm = nm.strip()
                                            val = val.strip()
                                            # nm = ''.join(char for char in input_string if char.isalpha() or char == '-')
                                            # val = ''.join(char for char in input_string if char.isdigit() or char == '.') 
                                            split_items.pop(4)
                                            split_items.insert(4, val)
                                            split_items.insert(5, nm)
                                        
                                        z=6         
                                        joined_string = ''                                         
                                        while z<len(split_items):   
                                            input_string1=split_items[z]
                                            joined_string += input_string1 + ' '
                                            z += 1
                                        
                                        z=6                                                  
                                        while z<len(split_items) + 1 :   
                                            split_items.pop(6)
                                            z += 1

                                        split_items.insert(len(split_items), joined_string.strip())
                                        
                                        try:
                                            sel=check_indices[i + 1]
                                            sel_row = df.loc[check_index : sel,'temp']

                                            tm_df=pd.DataFrame(sel_row)
                                            tm_df = tm_df.reset_index(drop=True)
                                                                                
                                            pay_calm = tm_df[tm_df['temp'].str.contains('Payer Claim Control Number:')]
                                            pay_calm = '\n'.join(pay_calm.iloc[:, 0])
                                            pay_calm = pay_calm.strip().split(':')[1].strip()
                                            split_items.append(str(pay_calm))
                                        except Exception as e:
                                            last_index = df.index[-1]
                                            sel_row = df.loc[check_index : last_index,'temp']

                                            tm_df=pd.DataFrame(sel_row)
                                            tm_df = tm_df.reset_index(drop=True)
                                                                                
                                            pay_calm = tm_df[tm_df['temp'].str.contains('Payer Claim Control Number:')]
                                            pay_calm = '\n'.join(pay_calm.iloc[:, 0])
                                            pay_calm = pay_calm.strip().split(':')[1].strip()
                                            split_items.append(str(pay_calm))
                                        
                                        lin_itms = tm_df[tm_df['temp'].str.contains('Line Item:')].index                                
                                        
                                        lst7=[]
                                        j=0                                                  
                                        while j<len(lin_itms):    
                                            lin_itm = lin_itms[j]
                                            data_row = tm_df.loc[lin_itm + 1,'temp']
                                            line_items = re.split(r'\s{2,}', data_row)
                                            line_items = [item for item in line_items if item]
                                            
                                            dt_sp = line_items[0]
                                            matches = re.findall(r'\b\d+\.\d{2}\b', dt_sp) 
                                            if matches:
                                                dt_sp1 = dt_sp.strip().split(' ')[0].strip()
                                                cpt_sp = dt_sp.strip().split(' ')[1].strip()
                                                chr_amt = dt_sp.strip().split(' ')[2].strip()

                                                line_items.pop(0)
                                                line_items.insert(0, dt_sp1)
                                                line_items.insert(1, cpt_sp)
                                                line_items.insert(2, chr_amt)
                                            else:
                                                dt_sp1 = dt_sp.strip().split(' ')[0].strip()
                                                cpt_sp = dt_sp.strip().split(' ')[1].strip()

                                                line_items.pop(0)
                                                line_items.insert(0, dt_sp1)
                                                line_items.insert(1, cpt_sp)
                                            
                                            k=2
                                            adj = tm_df.loc[lin_itm + k,'temp']

                                            while 'Line Item:' not in adj:                                                     
                                                lines_items = adj.strip()                                        
                                                line_items.append(lines_items)                    
                                                
                                                k=k+1  
                                                try:
                                                    adj = tm_df.loc[lin_itm + k,'temp']
                                                    if 'Check#' in adj:
                                                        break
                                                except Exception as e:
                                                    break
                                            lst7.append(line_items)
                                            j=j+1
                                        
                                        lst_cnt=len(lst7)

                                        l=0
                                        while l < lst_cnt:                
                                            fin_lst = None
                                            for sublist in lst:
                                                if chk_num in sublist:
                                                    fin_lst = sublist
                                                    break                
                                                        
                                            fin_lst_itm = lst7[l]
                                                
                                            # wb1=load_workbook(fil)
                                            sheet = wb1['CPT']
                                            column_letter = 'A'  
                                            column_cells = sheet[column_letter]
                                            last_row = None
                                            for cell in reversed(column_cells):
                                                if cell.value:
                                                    last_row = cell.row
                                                    break              
                                            start_column = 'C'
                                            current_column_index = openpyxl.utils.column_index_from_string(start_column)
                                            current_row = last_row + 1

                                            for value in fin_lst:
                                                current_column = openpyxl.utils.get_column_letter(current_column_index)
                                                sheet[current_column + str(current_row)] = value
                                                current_column_index += 1                
                                            
                                            start_column = 'G'
                                            current_column_index = openpyxl.utils.column_index_from_string(start_column)                                   

                                            for value in split_items:
                                                current_column = openpyxl.utils.get_column_letter(current_column_index)
                                                sheet[current_column + str(current_row)] = value
                                                current_column_index += 1 

                                            start_column = 'O'
                                            current_column_index = openpyxl.utils.column_index_from_string(start_column)                                   

                                            for value in fin_lst_itm:
                                                current_column = openpyxl.utils.get_column_letter(current_column_index)
                                                sheet[current_column + str(current_row)] = value
                                                current_column_index += 1 
                                            
                                            sheet['B' + str(int(last_row + 1))]=text_file
                                            sheet['A' + str(int(last_row + 1))]='Done'
                                            # wb1.save(filename=fil)
                                            # wb1.close()

                                            l=l+1            
                                        i=i+1
                                    
                                    pyr=split_items[-2]

                                    fin_lst11=[]
                                    date_pattern = r'\b\d{2}/\d{2}/\d{4}\b'

                                    for i in range(len(lst1)):
                                        date = None
                                        for item in lst1[i]:
                                            date_match = re.search(date_pattern, item)
                                            if date_match:
                                                date = date_match.group(0)        
                                        filtered_list = [re.sub(date_pattern, '', item) for item in lst1[i]]
                                        joined_string = ' '.join(filtered_list[4:])

                                        updated_sublist = filtered_list[:4] + [joined_string.strip(), date.strip()]

                                        fin_lst11.append(updated_sublist)     

                                    for i in range(len(fin_lst11)):
                                        fin_lst11[i].append(pyr)     
                                    
                                    # for i in range(len(lst1)):
                                    #     input_string=lst1[i][4]
                                    #     date_pattern = r'\b\d{2}/\d{2}/\d{4}\b'
                                    #     date_matches = re.findall(date_pattern, input_string)
                                    #     if len(date_matches) > 0:
                                    #         date = date_matches[0].strip()
                                    #         text_parts = re.split(date_pattern, input_string)
                                    #         text = ''.join(text_parts).strip()
                                    #         lst1[i].pop(4)
                                    #         lst1[i].insert(4,text)
                                    #         lst1[i].insert(5,date)      

                                    for i, sub_list in enumerate(fin_lst11):           
                                        sheet = wb1['Raw']
                                        column_letter = 'I'  
                                        column_cells = sheet[column_letter]
                                        last_row = None
                                        for cell in reversed(column_cells):
                                            if cell.value:
                                                last_row = cell.row
                                                break              
                                        start_column = 'B'
                                        current_column_index = openpyxl.utils.column_index_from_string(start_column)
                                        current_row = last_row + 1
                                        
                                        for value in sub_list:
                                            current_column = openpyxl.utils.get_column_letter(current_column_index)
                                            sheet[current_column + str(current_row)] = value
                                            current_column_index += 1                        

                                        sheet['A' + str(int(last_row + 1))]=text_file
                                        sheet['I' + str(int(last_row + 1))]='Done'
                                        
                                    wb1.save(filename=fil)
                                    
                                    text_files = [file for file in files if file.endswith('.txt')]
                                    zip_files = [file for file in files if file.endswith('.zip')]                            
                                    raw_files = [file for file in files if file.endswith('.835')]                            

                                    # shutil.move(os.path.abspath(down_path + zip_files),os.path.abspath(file_zip + zip_files))
                                    
                                    for text_file in text_files:
                                        source_path = os.path.join(down_path, text_file)
                                        destination_path = os.path.join(file_txt, text_file)
                                        shutil.move(source_path, destination_path)                                                        
                                    
                                    for zip_file in zip_files:
                                        source_path = os.path.join(down_path, zip_file)
                                        destination_path = os.path.join(file_zip, zip_file)                                
                                        shutil.move(source_path, destination_path)
                                                                    
                                    for raw_file in raw_files:
                                        source_path = os.path.join(down_path, raw_file)
                                        destination_path = os.path.join(file_835, raw_file)
                                        shutil.move(source_path, destination_path)

                                    files = os.listdir(down_path)

                                    for file in files:
                                        file_full_path = os.path.join(file, file)
                                        if os.path.isfile(file_full_path):
                                            os.remove(file_full_path)    
                                break
                        else:
                            time.sleep(1)
                            counter += 1
                    else:
                        files = os.listdir(down_path)

                        for file in files:
                            file_full_path = os.path.join(file, file)
                            if os.path.isfile(file_full_path):
                                os.remove(file_full_path)          
                        
                        wb1=load_workbook(filename=temp_file + 'Error File Names.xlsx')
                        sheet = wb1['Data']
                        last_row=sheet.max_row
                        sheet['A' + str(int(last_row + 1))]=file                    
                        wb1.save(temp_file + 'Error File Names.xlsx')
                        wb1.close()                                               
            n=n+1    
              
        wb1.close()

        driver.quit()
        messagebox.showinfo("Process Status", "Process Completed")
        sys.exit(0)                    

if __name__=="__main__":        
    process()    
    # user_pass()