from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, ElementClickInterceptedException, StaleElementReferenceException
from selenium.common.exceptions import TimeoutException
from docx.shared import Inches
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.select import Select
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from selenium import webdriver

import win32com.client
import pdf2image
import datetime
import logging
import shutil
import PyPDF2
import time
import sys
import os

date_formates = ['%d %b %Y','%m/%d/%Y','%d/%m/%Y %H:%M:%S %p','%m/%d/%Y %H:%M:%S %p','%d-%b-%y','%d-%m-%y','%Y-%m-%d %H:%M:%S','%d-%b-%Y','%d%b%Y','%d/%b/%Y','%d-%m-%Y','%d/%m/%Y','%d-%m-%Y %H:%M:%S','%d-%b-%Y %H:%M:%S','%d/%m/%Y %H:%M:%S','%d/%b/%Y %H:%M:%S','%Y-%m-%d %H:%M:%S']

wait_time = 2   # In Minutes
wait_time = int((wait_time*60)/2)


def init_log(_log):
    global log
    log = _log

def init_driver(_driver):
    global driver
    driver = _driver

def get_input_data():
    
    global master_data
    master_data = []

    workbook = load_workbook(os.getcwd() + '/Inputs/INPUT.xlsx')

    worksheet = workbook.active

    for i in range(2, worksheet.max_row+1):
        data = {}
        data['matrix_Ref_num'] = retrive_value_from_excel(worksheet['A{}'.format(i)].value)
        data['location_coordinates'] = retrive_value_from_excel(worksheet['B{}'.format(i)].value)

        if data['matrix_Ref_num'] != '':
            master_data.append(data)
    
    return master_data

def listToString(s):
 
    # initialize an empty string
    str1 = ""
 
    # traverse in the string
    for ele in s:
        str1 += ele + ',' 
 
    # return string
    return str1

def add_image_to_word_document(document, pdf_file_name, matrix_ref_num):
    
    temp_driectory = 'Application_Folder/Documents/'+matrix_ref_num+'/temp'
    delete_and_create_folder(temp_driectory)
    
    if pdf_file_name.endswith('.pdf'):
        pdf_to_images(pdf_file_name, temp_driectory)
    elif pdf_file_name.endswith('.png'):
        shutil.move(pdf_file_name,temp_driectory)
    
    images_list = os.listdir(temp_driectory)
    for image in images_list:
        document.add_picture(temp_driectory+'/'+image, width=Inches(6), height=Inches(5.5))

    shutil.rmtree(temp_driectory)

def pdf_to_images(pdf_file_name, destination_directory):
    images = pdf2image.convert_from_path(pdf_file_name, 500, poppler_path='Application_Folder/poppler-0.68.0/bin')
    for i in range(len(images)):
        images[i].save(destination_directory+'/image'+ str(i) +'.jpg', 'JPEG')

def convert_word_to_pdf(word_file_name, pdf_file_name):
    
    print('\t- Start Converting into PDF')
    # Load word document
    wdFormatPDF = 17

    word = win32com.client.Dispatch('Word.Application')
    doc = word.Documents.Open(word_file_name)
    time.sleep(1)
    doc.SaveAs(pdf_file_name, FileFormat=wdFormatPDF)
    doc.Close()
    word.Quit()

def clean_drop_down(data):
    
    if 'select' in data.lower():
        return 'Not Provied'
    else:
        return data

def scrape_data_before_print(x_path):
    if is_element_exists(x_path):
        try:
            data = driver.find_element(By.XPATH, x_path).get_attribute('value').strip()
        except:
            data = 'Not Provied'
    else:
        data = 'Not Provied'

    if data.strip() != '':
        return data
    else:
        return 'Not Provided'

def scrap_data_from_options(x_path1):
    
    data = ''
    index = 1
    while True:
        x_path = x_path1 + '/option['+str(index)+']'
        if is_element_exists(x_path):
            if driver.find_element(By.XPATH,x_path).get_attribute('selected'):
                data = driver.find_element(By.XPATH,x_path).text
                if '- Select -' in data:
                    return ''
                break
            else:
                index += 1
        else:
            break

    return data

def scrape_data_basedon_attribute(x_path,attribute_value):
    
    if is_element_exists(x_path):
        try:
            data = driver.find_element(By.XPATH,x_path).get_attribute(attribute_value)
        except:
            data = ''
    else:
        data = ''

    return data

def scrape_data_before_print_by_text(x_path):
    if is_element_exists(x_path):
        try:
            data = driver.find_element(By.XPATH, x_path).text
        except:
            data = ''
    else:
        data = ''

    if data.strip() != '':
        return data
    else:
        return 'Not Provided'

def handle_report_download(download_xpath, main_directory, file_download_path):
    loading = 1
    wait_time = 10
    wait_and_click(download_xpath)
    while True:
        try:
            WebDriverWait(driver, 10).until(EC.alert_is_present())
            alert = driver.switch_to.alert
            alert_text = alert.text
            if alert_text == 'Unable to download file, Please restore the file':
                driver.switch_to.alert.accept()
                return [False,'Unable to download file']
            driver.switch_to.alert.accept()
            return [False,'No PDF File Available']
        except TimeoutException:
            break 

    while True:
        try:
            os.chdir(main_directory+'/Application_Folder/Downloads')
            files = os.listdir(os.getcwd())
            if len(files) >= 1:
                wait_time = 60
                latest_file = max(files, key=os.path.getctime)
                ext = latest_file.split('.')[-1]
                if ext != 'crdownload' and ext != 'temp' and ext != 'tmp':
                    os.chdir(main_directory)
                    shutil.move(os.path.abspath('Application_Folder/Downloads/' +
                                latest_file), os.path.abspath(file_download_path)+'/'+latest_file)
                    return [True, latest_file]

            os.chdir(main_directory)
        except Exception as e:
            # print(e)
            log.exception(e)
            if loading >= wait_time:
                return [False, 'Issue while downloading']

        if loading >= wait_time:
            return [False, 'Issue while downloading']

        time.sleep(2)
        loading += 1

def wait_till_new_tab_open(x_path):
    old_window = driver.window_handles
    wait_and_click(x_path)
    while True:
        window = driver.window_handles
        if len(window) > len(old_window):
            driver.switch_to.window(window[-1])
            return True
        else:
            time.sleep(2)

def wait_and_click(x_path):
    wait_till_element_exist(x_path)
    i = 1
    while True:
        try:
            driver.find_element(By.XPATH, x_path).click()
            return True
            # driver.find_element(By.XPATH, x_path).send_keys(Keys.ENTER)
        except StaleElementReferenceException:
            # print(x_path, '  StaleElementReferenceException')
            time.sleep(2)
        except ElementClickInterceptedException:
            # print(x_path, '  ElementClickInterceptedException')
            time.sleep(2)
        except ElementNotInteractableException:
            # print(x_path, '  ElementClickInterceptedException')
            time.sleep(2)

        if i >= wait_time:
            return False

        i += 1

def wait_and_enter(x_path, key):
    wait_till_element_exist(x_path)
    i = 1
    while True:
        try:
            driver.find_element(By.XPATH, x_path).send_keys(key)
            return True
        except StaleElementReferenceException:
            # print(x_path, '  StaleElementReferenceException')
            time.sleep(2)
        except ElementNotInteractableException:
            # print(x_path, '  ElementClickInterceptedException')
            time.sleep(2)

        if i >= wait_time:
            return False

        i += 1

def wait_till_element_exist(x_path):
    i = 0
    while True:
        if is_element_exists(x_path):
            return True
        else:
            time.sleep(2)
            if i >= wait_time:
                # print('Element ', x_path, ' does not exists')
                return False
        i += 1

def is_element_exists(x_path):
    try:
        driver.find_element(By.XPATH, x_path)
        # print(x_path,'  True')
        return True
    except NoSuchElementException:
        # print(x_path,'  NoSuchElementException')
        return False

def wait_till_loading():
    return driver.execute_script('return document.readyState;')

def retrive_value_from_excel(data):
    
    retrieve_data = ''
    if data == None:
        return retrieve_data
    else:
        retrieve_data = data
        return retrieve_data

def starting_info(boolean, today_date=None, employee_id=None):
    
    print()
    print('****************************************************************')
    print('|Tool Started                                                  |')
    print('|Date and time : {}                                    |'.format(today_date))
    if boolean:
        print('|Started From Scarch                                           |')
    else:
        print('|Continues from Employee Id : {}                          |'.format(
            employee_id))
    print('****************************************************************')
    print()

def returning_date_with_valid_format(date_given,return_format):
    date_of_get = ''
    for date_format in date_formates:
        try:
            date_of_get = datetime.datetime.strptime(date_given, date_format)
            return date_of_get.strftime(return_format)
        except:
            pass

def returning_date_with_valid_datetime_format(date_given):
    date_of_get = ''
    for date_format in date_formates:
        try:
            date_of_get = datetime.datetime.strptime(date_given, date_format)
            return date_of_get
        except:
            pass

def scrape_data_from_innerttext(x_path):
    if is_element_exists(x_path):
        try:
            data = driver.find_element(By.XPATH,x_path).get_attribute('innerText')
        except:
            data = ''
    else:
        data = ''

    if data.strip() != '':
        return data
    else:
        return 'Not Provided'

def close_all_the_tabs_except_first_two():
    windows = driver.window_handles
    if len(windows) > 2:
        for i in range(2,len(windows)):
            driver.switch_to.window(windows[i])
            driver.close()
        driver.switch_to.window(windows[1])
    
def delete_and_create_folder(directory):
    if os.path.exists(directory):
        shutil.rmtree(directory)
    
    os.mkdir(directory)  